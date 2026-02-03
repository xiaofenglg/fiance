# -*- coding: utf-8 -*-
"""
银行理财产品V5.4策略 - 历史回测系统 (V3: T+1交易机制 + 假期处理)

回测区间: 2025-06-01 ~ 2026-01-29
初始资金: 1亿元
持仓规则: 最多10个产品，每个10%仓位（1000万）

交易机制（符合银行理财实际规则）:
  - 申购(买入): 当日提交, T+1日确认份额并开始计算收益
  - 赎回(卖出): 当日提交, T+1日确认并到账
  - 周末无收益: 周五申购 → 周一确认, 周末不计收益
  - 中国假期: 假期前申购/赎回无法在假期内确认, 资金空转无收益
  - 节前禁止申购: 下一交易日间隔>3天(长假)时不提交申购
  - 持有天数按交易日计: 排除周末和假期, 公平反映实际收益天数
  - 最短持有: 确认后至少持有5个交易日才可赎回
  - 买入条件: 日年化收益>3.5%（无需突破, 库内产品即可买入）
  - 卖出条件: 日年化收益≤2.5% 或 持有≥30个交易日

严格禁止前瞻偏差: 产品库仅用模拟日期之前的历史数据构建
"""

import os, sys, re, warnings, logging, bisect, math, ctypes
import pandas as pd
import numpy as np
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

# ============================================================
# 配置
# ============================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "净值数据库.xlsx")
OUTPUT_PATH = os.path.join(BASE_DIR, "回测报告_V5.xlsx")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('backtest_v5.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# ============================================================
# 策略参数 V5.27 - 6仓集中 + 高买入门槛(3.5%) + 快速轮动(20天) + 轮动卖出
# 核心: 6×1667万集中 + 只买3.5%+强信号 + 20天强制退出 + return×avg排名 + 轮动
# 回测最优: 年化 5.60%
# ============================================================
信号阈值_买入 = 3.5       # 买入: 当日年化>3.5%（高质量信号）
信号阈值_卖出 = 2.0       # 卖出: 当日年化≤2.0%
信号阈值_库 = 3.0          # 产品库信号检测（宽入库）
需要突破信号 = False       # False=只要>买入阈值就可买
持有天数_评估 = 7
成功标准 = 2.5             # 成功定义: 7天平均>2.5%
最低成功率 = 30.0          # 入库门槛: 30%（宽入，用排名区分优劣）
最少历史信号 = 2           # 最少2次信号
最短持续天数 = 2           # 持续≥2天
最大持有天数_评估 = 30     # 评估窗口
最长赎回天数 = 14

# 回测参数
回测开始日 = '2025-06-01'
回测结束日 = '2026-01-30'   # exclusive: 最后交易日为 2026-01-29
初始资金 = 100_000_000       # 1亿
最大持仓数 = 6               # 6仓集中（最优）
单仓金额 = 初始资金 / 6      # ≈1667万/仓
产品库重建间隔 = 7            # 每7天重建

# T+1 交易机制参数
最短持有交易日_赎回 = 3   # 3个交易日后可赎回
最大持有交易日 = 20        # 最多20个交易日（快速轮动）
长假阈值_天 = 3            # 下一交易日间隔>此值视为长假, 不提交申购


# ============================================================
# 工具函数
# ============================================================

def is_date_col(col_name):
    if not isinstance(col_name, str):
        return False
    return len(col_name) == 10 and col_name[4] == '-' and col_name[7] == '-'


def 判断产品流动性(name):
    if '日开' in name or '天天' in name or '每日' in name:
        return '日开', 0
    m = re.search(r'(\d+)\s*天持有期', name)
    if m:
        return f'{int(m.group(1))}天持有期', int(m.group(1))
    m = re.search(r'最短持有\s*(\d+)\s*天', name)
    if m:
        return f'{int(m.group(1))}天持有', int(m.group(1))
    m = re.search(r'周期\s*(\d+)\s*天', name)
    if m:
        return f'{int(m.group(1))}天周期', int(m.group(1))
    m = re.search(r'(\d+)\s*个?月', name)
    if m and '半年' not in name:
        months = int(m.group(1))
        if months <= 6:
            return f'{months}个月', months * 30
    if '周开' in name: return '周开', 7
    if '日申季赎' in name: return '季赎', 90
    if '月开' in name: return '月开', 30
    if '季开' in name or '季度开' in name: return '季开', 90
    if '半年' in name: return '半年', 180
    if '年开' in name: return '年开', 365
    if '封闭' in name: return '封闭', 999
    if '定开' in name: return '定开', 999
    m = re.search(r'(\d+)\s*Y\s*持有', name)
    if m: return f'{int(m.group(1))}年持有', int(m.group(1)) * 365
    if re.search(r'20[2-5]\d', name) and '信颐' in name: return '目标日期', 999
    return '未知', 999


# ============================================================
# 数据结构
# ============================================================

class ProductData:
    """单个产品的全部数据"""
    __slots__ = ['bank', 'code', 'name', 'liquidity', 'redeem_days',
                 'nav_dates', 'navs', 'ret_dates', 'rets',
                 'ret_date_idx', 'signals']

    def __init__(self, bank, code, name, liquidity, redeem_days):
        self.bank = bank
        self.code = code
        self.name = name
        self.liquidity = liquidity
        self.redeem_days = redeem_days
        self.nav_dates = []       # sorted
        self.navs = {}            # {date: float}
        self.ret_dates = []       # sorted
        self.rets = {}            # {date: annualized_return}
        self.ret_date_idx = {}    # {date: index} for O(1) lookup
        self.signals = []         # precomputed historical signals


class Position:
    """一笔持仓（含T+1确认信息 + 入场信号强度）"""
    __slots__ = ['product_key', 'signal_date', 'confirm_date', 'buy_nav',
                 'amount', 'entry_return', 'alpha_score',
                 'sell_date', 'sell_nav', 'pnl', 'sell_reason']

    def __init__(self, product_key, signal_date, confirm_date, buy_nav, amount,
                 entry_return=0.0, alpha_score=0.0):
        self.product_key = product_key
        self.signal_date = signal_date
        self.confirm_date = confirm_date
        self.buy_nav = buy_nav
        self.amount = amount
        self.entry_return = entry_return     # 入场时的年化收益率
        self.alpha_score = alpha_score       # 入场时的Alpha评分
        self.sell_date = None
        self.sell_nav = None
        self.pnl = 0.0
        self.sell_reason = ''


# ============================================================
# 回测引擎
# ============================================================

class BacktestEngine:

    def __init__(self):
        self.products = {}
        self.all_sim_dates = []
        self.library = {}
        self.library_date = None
        self.positions = {}            # {key: Position} 已确认的持仓
        self.pending_buys = []         # [(key, signal_date, reason)] 待T+1确认的买入
        self.pending_sells = []        # [(key, signal_date, reason)] 待T+1确认的卖出
        self.pending_buy_keys = set()  # 正在申购中的产品（防重复）
        self.pending_sell_keys = set() # 正在赎回中的产品（防重复卖出）
        self.date_to_idx = {}         # {date_str: index} 用于交易日天数计算
        self.closed_trades = []
        self.cash = 初始资金
        self.daily_values = []
        self.trade_log = []

    # ------ 数据加载 ------

    def load_data(self):
        logger.info("加载净值数据库...")
        xlsx = pd.ExcelFile(DB_PATH)
        all_dates = set()
        total = 0

        for sheet in xlsx.sheet_names:
            logger.info(f"  加载 [{sheet}]...")
            df = pd.read_excel(xlsx, sheet_name=sheet, dtype=str)
            df.columns = [str(c).strip() for c in df.columns]
            if 'level_0' in df.columns:
                df = df.rename(columns={'level_0': '产品代码', 'level_1': '产品名称'})

            date_cols = sorted([c for c in df.columns if is_date_col(c)])
            if len(date_cols) < 10:
                continue
            all_dates.update(date_cols)
            count = 0

            for _, row in df.iterrows():
                code = row.get('产品代码')
                name = row.get('产品名称', '')
                if pd.isna(code):
                    continue
                code = str(code).strip()
                name = str(name).strip() if pd.notna(name) else ''

                liq_type, redeem_days = 判断产品流动性(name)
                if redeem_days > 最长赎回天数:
                    continue

                nav_points = []
                for d in date_cols:
                    try:
                        v = float(row[d])
                        if not np.isnan(v) and v > 0:
                            nav_points.append((d, v))
                    except (ValueError, TypeError):
                        pass

                if len(nav_points) < 10:
                    continue

                key = (sheet, code)
                p = ProductData(sheet, code, name, liq_type, redeem_days)
                p.nav_dates = [x[0] for x in nav_points]
                p.navs = {x[0]: x[1] for x in nav_points}

                for i in range(1, len(nav_points)):
                    d0, n0 = nav_points[i - 1]
                    d1, n1 = nav_points[i]
                    gap = (pd.Timestamp(d1) - pd.Timestamp(d0)).days
                    if 0 < gap <= 60 and n0 > 0:
                        p.rets[d1] = (n1 / n0 - 1) / gap * 365 * 100

                p.ret_dates = sorted(p.rets.keys())
                p.ret_date_idx = {d: i for i, d in enumerate(p.ret_dates)}

                if len(p.ret_dates) < 10:
                    continue

                self._precompute_signals(p)
                self.products[key] = p
                count += 1

            total += count
            logger.info(f"    有效产品: {count}")

        self.all_sim_dates = sorted(
            d for d in all_dates if 回测开始日 <= d < 回测结束日
        )
        logger.info(f"加载完成: {total}个产品, "
                     f"{len(self.all_sim_dates)}个回测日 "
                     f"({self.all_sim_dates[0]}~{self.all_sim_dates[-1]})")

    def _precompute_signals(self, product):
        """预计算全部历史信号（用于快速构建产品库，不涉及前瞻）
        注: 这里使用买入阈值作为信号检测标准
        """
        rd = product.ret_dates
        rets = product.rets
        n = len(rd)
        signals = []

        for i in range(1, n - 持有天数_评估):
            if rets[rd[i]] > 信号阈值_库 and rets[rd[i - 1]] <= 信号阈值_库:
                hold_rets = [rets[rd[i + 1 + k]] for k in range(持有天数_评估)]
                avg_ret = float(np.mean(hold_rets))
                eval_end = rd[i + 持有天数_评估]

                persist = 0
                for j in range(i, min(i + 20, n)):
                    if rets[rd[j]] > 信号阈值_库 * 0.8:
                        persist += 1
                    else:
                        break
                cal_days = ((pd.Timestamp(rd[i + persist - 1])
                             - pd.Timestamp(rd[i])).days + 1) if persist > 0 else 0

                signals.append({
                    'date': rd[i],
                    'eval_end': eval_end,
                    'avg_return': avg_ret,
                    'success': avg_ret > 成功标准,
                    'persist_days': cal_days,
                })

        product.signals = signals

    # ------ 产品库构建（无前瞻） ------

    def build_library(self, as_of_date):
        """构建产品库：仅使用 eval_end < as_of_date 的信号（禁止前瞻）
        信号检测使用 信号阈值_库(3.0%) — 宽入；买入决策使用 信号阈值_买入(4.0%) — 严买
        """
        lib = {}
        for key, product in self.products.items():
            valid = [s for s in product.signals if s['eval_end'] < as_of_date]
            if len(valid) < 最少历史信号:
                continue
            success_n = sum(1 for s in valid if s['success'])
            rate = success_n / len(valid) * 100
            if rate < 最低成功率:
                continue
            avg_persist = float(np.mean([s['persist_days'] for s in valid]))
            if avg_persist < 最短持续天数 or avg_persist > 最大持有天数_评估:
                continue
            avg_ret = float(np.mean([s['avg_return'] for s in valid]))
            std_ret = float(np.std([s['avg_return'] for s in valid])) if len(valid) > 1 else 1.0
            sharpe_like = avg_ret / max(std_ret, 0.5)  # 收益/波动率 (信号级Sharpe)
            lib[key] = {
                'success_rate': round(rate, 1),
                'signal_count': len(valid),
                'avg_return': round(avg_ret, 2),
                'avg_persist': round(avg_persist, 1),
                'sharpe_like': round(sharpe_like, 2),
            }
        self.library = lib
        self.library_date = as_of_date
        return lib

    # ------ NAV 查询 ------

    def get_latest_nav(self, product, as_of_date):
        idx = bisect.bisect_right(product.nav_dates, as_of_date) - 1
        if idx >= 0:
            d = product.nav_dates[idx]
            return d, product.navs[d]
        return None, None

    # ------ T+1 交易机制 ------

    def _process_pending_buys(self, date, buy_count):
        """处理昨日提交的申购单（T+1确认）
        date: 今日日期（确认日）
        返回: 新增买入数
        """
        new_buys = 0
        remaining = []
        for key, signal_date, reason in self.pending_buys:
            if key in self.positions:
                # 已持有（不应发生，但防御）
                self.pending_buy_keys.discard(key)
                continue

            product = self.products[key]
            # T+1 确认: 用今日(确认日)净值作为买入价
            buy_nav = product.navs.get(date)
            if buy_nav is None:
                _, buy_nav = self.get_latest_nav(product, date)
            if buy_nav is None:
                self.pending_buy_keys.discard(key)
                continue

            # 动态仓位: 按当前组合价值计算仓位大小（复利效应）
            current_value = self.cash + sum(
                pos.amount * (self.products[k].navs.get(date, pos.buy_nav) / pos.buy_nav)
                for k, pos in self.positions.items()
            )
            dynamic_size = current_value / 最大持仓数
            amount = min(dynamic_size, self.cash)
            if amount < 单仓金额 * 0.5:
                self.pending_buy_keys.discard(key)
                continue

            entry_ret = product.rets.get(signal_date, 信号阈值_买入)
            self.cash -= amount
            self.positions[key] = Position(key, signal_date, date, buy_nav, amount,
                                           entry_return=entry_ret)
            self.pending_buy_keys.discard(key)
            new_buys += 1
            self.trade_log.append({
                'date': date, 'action': '买入确认',
                'bank': product.bank, 'code': product.code,
                'name': product.name, 'nav': buy_nav,
                'amount': amount, 'pnl': 0,
                'hold_days': 0,
                'reason': f'T+1确认(信号日{signal_date}) {reason}',
            })

        self.pending_buys = remaining  # should be empty
        return new_buys

    def _process_pending_sells(self, date, sell_count):
        """处理昨日提交的赎回单（T+1确认到账）
        date: 今日日期（到账日）
        返回: 新增卖出数
        """
        new_sells = 0
        for key, signal_date, reason in self.pending_sells:
            if key not in self.positions:
                self.pending_sell_keys.discard(key)
                continue

            pos = self.positions.pop(key)
            product = self.products[key]

            # T+1 到账: 用今日(到账日)净值作为赎回价
            sell_nav = product.navs.get(date)
            if sell_nav is None:
                _, sell_nav = self.get_latest_nav(product, date)
            if sell_nav is None:
                sell_nav = pos.buy_nav

            pos.sell_date = date
            pos.sell_nav = sell_nav
            pos.pnl = pos.amount * (sell_nav / pos.buy_nav - 1)
            pos.sell_reason = reason
            self.cash += pos.amount + pos.pnl
            self.closed_trades.append(pos)
            self.pending_sell_keys.discard(key)
            new_sells += 1

            hold_td = self._trading_days_held(pos.confirm_date, date)
            self.trade_log.append({
                'date': date, 'action': '赎回到账',
                'bank': product.bank, 'code': product.code,
                'name': product.name, 'nav': sell_nav,
                'amount': pos.amount + pos.pnl,
                'pnl': pos.pnl,
                'hold_days': hold_td,
                'reason': f'T+1到账(提交日{signal_date}) {reason}',
            })

        self.pending_sells = []
        return new_sells

    def _trading_days_held(self, confirm_date, current_date):
        """计算从确认日到当前日的交易日数（排除周末和假期）"""
        ci = self.date_to_idx.get(confirm_date)
        di = self.date_to_idx.get(current_date)
        if ci is not None and di is not None:
            return di - ci
        # fallback: 用自然日近似（不应到这里）
        return (pd.Timestamp(current_date) - pd.Timestamp(confirm_date)).days

    def _is_pre_holiday(self, di):
        """判断当前日期是否在长假前（下一交易日间隔>长假阈值天）"""
        if di + 1 >= len(self.all_sim_dates):
            return True  # 回测末尾，不再买入
        next_date = self.all_sim_dates[di + 1]
        gap = (pd.Timestamp(next_date) - pd.Timestamp(self.all_sim_dates[di])).days
        return gap > 长假阈值_天

    def _check_sells(self, date):
        """卖出决策（两条件并行）
        ① 硬阈值: 日年化 ≤ 信号阈值_卖出 → 立即卖出
        ② 最大持有: 持有 ≥ 最大持有交易日 → 强制卖出
        """
        for key in list(self.positions):
            if key in self.pending_sell_keys:
                continue
            pos = self.positions[key]
            p = self.products[key]

            hold_td = self._trading_days_held(pos.confirm_date, date)

            if hold_td < 最短持有交易日_赎回:
                continue

            ret = p.rets.get(date)

            # ① 硬阈值卖出
            if ret is not None and ret <= 信号阈值_卖出:
                self.pending_sells.append((key, date, f'收益{ret:.1f}%≤{信号阈值_卖出}%(持有{hold_td}交易日)'))
                self.pending_sell_keys.add(key)
                continue

            # ② 最大持有卖出
            if hold_td >= 最大持有交易日:
                self.pending_sells.append((key, date, f'持有{hold_td}交易日达上限'))
                self.pending_sell_keys.add(key)
                continue

    def _rotation_sell(self, date, di):
        """轮动卖出: 满仓且有更好产品时, 卖出最弱持仓腾出仓位
        仅在非长假前执行, 需满足最短持有期
        """
        if self._is_pre_holiday(di):
            return 0

        used = len(self.positions) + len(self.pending_buys)
        if used < 最大持仓数:
            return 0  # 有空仓位, 不需要轮动

        # 计算所有可卖持仓的当前收益
        sellable = []
        for key in self.positions:
            if key in self.pending_sell_keys:
                continue
            pos = self.positions[key]
            hold_td = self._trading_days_held(pos.confirm_date, date)
            if hold_td < 最短持有交易日_赎回:
                continue
            p = self.products[key]
            ret = p.rets.get(date)
            if ret is not None:
                sellable.append((key, ret, hold_td))

        if not sellable:
            return 0

        # 找出收益最低的持仓
        sellable.sort(key=lambda x: x[1])
        worst_key, worst_ret, worst_hold = sellable[0]

        # 如果最弱持仓已低于买入阈值, 检查是否有更好的候选
        if worst_ret >= 信号阈值_买入:
            return 0  # 最弱持仓仍在买入阈值以上, 不换

        # 查找最佳候选（不在持仓、不在待买、不在待卖）
        best_candidate = None
        best_ret = worst_ret
        for key, info in self.library.items():
            if key in self.positions or key in self.pending_buy_keys or key in self.pending_sell_keys:
                continue
            p = self.products[key]
            r = p.rets.get(date)
            if r is not None and r > 信号阈值_买入 and r > best_ret + 1.0:
                # 候选收益需超出最弱持仓1%以上, 避免频繁换手
                best_candidate = (key, r, info['success_rate'])
                best_ret = r

        if best_candidate is None:
            return 0

        # 执行轮动: 卖出最弱, 空出仓位给更强产品
        self.pending_sells.append((worst_key, date,
            f'轮动卖出(收益{worst_ret:.1f}%→换{best_ret:.1f}%持有{worst_hold}交易日)'))
        self.pending_sell_keys.add(worst_key)
        return 1

    def _find_and_queue_buys(self, date, slots, di):
        """收益率复合排名: 当日收益率 × 历史平均收益率（回测验证最优）
        """
        if self._is_pre_holiday(di):
            return 0

        signals = []
        for key, info in self.library.items():
            if key in self.positions or key in self.pending_buy_keys or key in self.pending_sell_keys:
                continue

            p = self.products[key]
            r = p.rets.get(date)
            if r is None or r <= 信号阈值_买入:
                continue

            if 需要突破信号:
                idx = p.ret_date_idx.get(date)
                if idx is None or idx == 0:
                    continue
                prev_r = p.rets[p.ret_dates[idx - 1]]
                if prev_r > 信号阈值_买入:
                    continue

            signals.append((key, info['success_rate'], r, info['avg_return'],
                            info.get('sharpe_like', 1.0)))

        # 收益率复合排名: 当日收益率 × 历史平均收益率
        signals.sort(key=lambda x: (-(x[2] * x[3]), -x[1]))

        queued = 0
        for key, sr, ret, _avg, _sh in signals[:slots]:
            reason = f'成功率{sr}% 收益{ret:.1f}%'
            self.pending_buys.append((key, date, reason))
            self.pending_buy_keys.add(key)
            queued += 1
            self.trade_log.append({
                'date': date, 'action': '提交申购',
                'bank': self.products[key].bank,
                'code': self.products[key].code,
                'name': self.products[key].name,
                'nav': 0, 'amount': 0, 'pnl': 0,
                'hold_days': 0,
                'reason': f'T日提交 {reason}',
            })
        return queued

    def _record_value(self, date):
        pos_val = 0
        for key, pos in self.positions.items():
            p = self.products[key]
            _, nav = self.get_latest_nav(p, date)
            if nav is not None:
                pos_val += pos.amount * (nav / pos.buy_nav)
            else:
                pos_val += pos.amount
        # 待确认的申购金额仍在现金中（T+1扣款）
        total = self.cash + pos_val
        if self.daily_values and self.daily_values[-1][0] == date:
            self.daily_values[-1] = (date, total, self.cash, pos_val)
        else:
            self.daily_values.append((date, total, self.cash, pos_val))

    # ------ 主回测循环 ------

    def run(self):
        self.load_data()

        sep = '=' * 70
        print(f"\n{sep}")
        print("   V5.27 6仓集中+高买入+快速轮动 回测 (T+1 + 假期处理)")
        print(f"   区间: {回测开始日} ~ {回测结束日}")
        print(f"   资金: {初始资金 / 1e8:.1f}亿 | "
              f"仓位: {最大持仓数}×{单仓金额 / 1e4:.0f}万")
        print(f"   买入>={信号阈值_买入}% | 卖出<={信号阈值_卖出}% | "
              f"持有{最短持有交易日_赎回}~{最大持有交易日}交易日")
        print(sep)

        # 构建交易日索引（用于交易日天数计算）
        self.date_to_idx = {d: i for i, d in enumerate(self.all_sim_dates)}

        self.build_library(回测开始日)
        logger.info(f"初始产品库: {len(self.library)}个")

        n_days = len(self.all_sim_dates)
        buy_n = sell_n = 0

        for di, date in enumerate(self.all_sim_dates):
            # ① 定期重建产品库
            if (self.library_date is None or
                    (pd.Timestamp(date)
                     - pd.Timestamp(self.library_date)).days >= 产品库重建间隔):
                old = len(self.library)
                self.build_library(date)
                if len(self.library) != old or di == 0:
                    logger.info(f"  [{date}] 产品库: {len(self.library)}个")

            # ② 处理昨日待确认的申购（T+1确认份额）
            nb = self._process_pending_buys(date, buy_n)
            buy_n += nb

            # ③ 处理昨日待确认的赎回（T+1到账）
            ns = self._process_pending_sells(date, sell_n)
            sell_n += ns

            # ④ 检查赎回条件（满足最短持有期的持仓）
            self._check_sells(date)

            # ⑤ 扫描买入信号（T日提交申购，长假前不提交）
            # 可用仓位 = 最大持仓数 - 已持仓 - 待确认申购
            used = len(self.positions) + len(self.pending_buys)
            slots = 最大持仓数 - used
            if slots > 0:
                nq = self._find_and_queue_buys(date, slots, di)

            # ⑤b 满仓轮动: 替换最弱持仓
            self._rotation_sell(date, di)

            # ⑥ 记录每日市值
            self._record_value(date)

            if (di + 1) % 50 == 0 or di == n_days - 1:
                v = self.daily_values[-1][1]
                r = (v / 初始资金 - 1) * 100
                logger.info(f"  [{date}] {di + 1}/{n_days} | "
                            f"持仓{len(self.positions)} 待买{len(self.pending_buys)} "
                            f"待卖{len(self.pending_sells)} | "
                            f"买{buy_n} 卖{sell_n} | "
                            f"{v / 1e4:,.0f}万({r:+.2f}%)")

        # 回测结束：强制平仓
        last = self.all_sim_dates[-1]
        for key in list(self.positions):
            pos = self.positions.pop(key)
            p = self.products[key]
            _, nav = self.get_latest_nav(p, last)
            if nav is None:
                nav = pos.buy_nav
            pos.sell_date = last
            pos.sell_nav = nav
            pos.pnl = pos.amount * (nav / pos.buy_nav - 1)
            pos.sell_reason = '回测结束平仓'
            self.cash += pos.amount + pos.pnl
            self.closed_trades.append(pos)
            hold_td = self._trading_days_held(pos.confirm_date, last)
            self.trade_log.append({
                'date': last, 'action': '强制平仓',
                'bank': p.bank, 'code': p.code, 'name': p.name,
                'nav': nav, 'amount': pos.amount + pos.pnl,
                'pnl': pos.pnl, 'hold_days': hold_td,
                'reason': '回测结束平仓',
            })
        self._record_value(last)

        self._print_summary()
        self._generate_report()

    # ------ 汇总输出 ------

    def _monthly_returns(self):
        monthly = {}
        for date, value, _, _ in self.daily_values:
            monthly[date[:7]] = value
        result = []
        prev = 初始资金
        for m in sorted(monthly):
            cur = monthly[m]
            result.append((m, (cur / prev - 1) * 100))
            prev = cur
        return result

    def _print_summary(self):
        sep = '=' * 70
        if not self.daily_values:
            print("无回测数据")
            return

        fv = self.daily_values[-1][1]
        total_ret = (fv / 初始资金 - 1) * 100
        t0 = pd.Timestamp(self.all_sim_dates[0])
        t1 = pd.Timestamp(self.all_sim_dates[-1])
        yrs = (t1 - t0).days / 365
        ann_ret = ((fv / 初始资金) ** (1 / yrs) - 1) * 100 if yrs > 0 else 0

        peak = 初始资金
        max_dd = 0
        dd_date = ''
        for d, v, _, _ in self.daily_values:
            if v > peak:
                peak = v
            dd = (peak - v) / peak * 100
            if dd > max_dd:
                max_dd = dd
                dd_date = d

        nt = len(self.closed_trades)
        nw = sum(1 for t in self.closed_trades if t.pnl > 0)
        nl = nt - nw
        wr = nw / nt * 100 if nt else 0
        total_pnl = sum(t.pnl for t in self.closed_trades)
        avg_pnl = total_pnl / nt if nt else 0

        hold_days_list = [self._trading_days_held(t.confirm_date, t.sell_date)
                          for t in self.closed_trades]
        avg_hold = float(np.mean(hold_days_list)) if hold_days_list else 0

        avg_win = float(np.mean([t.pnl for t in self.closed_trades if t.pnl > 0])) if nw else 0
        avg_loss = float(np.mean([t.pnl for t in self.closed_trades if t.pnl <= 0])) if nl else 0

        gp = sum(t.pnl for t in self.closed_trades if t.pnl > 0)
        gl = abs(sum(t.pnl for t in self.closed_trades if t.pnl <= 0))
        pf = gp / gl if gl > 0 else float('inf')

        daily_rets = []
        for i in range(1, len(self.daily_values)):
            pv = self.daily_values[i - 1][1]
            cv = self.daily_values[i][1]
            if pv > 0:
                daily_rets.append(cv / pv - 1)
        if daily_rets:
            rf = 0.025 / 252
            excess = np.array(daily_rets) - rf
            sharpe = float(np.mean(excess) / np.std(excess) * np.sqrt(252)) if np.std(excess) > 0 else 0
        else:
            sharpe = 0

        bm_val = 初始资金 * (1 + 0.025 * yrs)
        bm_ret = (bm_val / 初始资金 - 1) * 100

        print(f"\n{sep}")
        print("          V5.4策略回测结果（T+1 + 中国假期处理）")
        print(sep)
        print(f"\n  【基本信息】")
        print(f"  回测区间:       {self.all_sim_dates[0]} ~ {self.all_sim_dates[-1]}")
        print(f"  交易日数:       {len(self.all_sim_dates)}")
        print(f"  初始资金:       {初始资金 / 1e4:,.0f}万")
        print(f"  最终市值:       {fv / 1e4:,.0f}万")

        print(f"\n  【T+1交易规则 + 假期处理】")
        print(f"  申购确认:       T+1日确认份额（周五→周一, 节前→节后）")
        print(f"  赎回到账:       T+1日到账")
        print(f"  最短持有:       确认后{最短持有交易日_赎回}个交易日")
        print(f"  最大持有:       {最大持有交易日}个交易日")
        print(f"  节前限制:       长假前(间隔>{长假阈值_天}天)不提交申购")

        print(f"\n  【收益指标】")
        print(f"  总收益率:       {total_ret:+.4f}%")
        print(f"  年化收益率:     {ann_ret:+.4f}%")
        print(f"  总盈亏:         {total_pnl / 1e4:+,.2f}万")
        print(f"  基准(存款2.5%): {bm_ret:+.4f}% → {bm_val / 1e4:,.0f}万")
        print(f"  超额收益:       {(total_ret - bm_ret):+.4f}%")

        print(f"\n  【风险指标】")
        print(f"  最大回撤:       {max_dd:.4f}% ({dd_date})")
        print(f"  夏普比率:       {sharpe:.4f}")
        print(f"  盈亏比:         {pf:.2f}")

        print(f"\n  【交易统计】")
        print(f"  总交易次数:     {nt}")
        print(f"  盈利/亏损:      {nw}/{nl}")
        print(f"  胜率:           {wr:.1f}%")
        print(f"  平均持有交易日: {avg_hold:.1f}（T+1确认后起算, 排除周末假期）")
        print(f"  平均盈利:       {avg_win / 1e4:+,.2f}万")
        print(f"  平均亏损:       {avg_loss / 1e4:+,.2f}万")
        print(f"  平均单笔盈亏:   {avg_pnl / 1e4:+,.2f}万")

        print(f"\n  【月度收益率】")
        for m, r in self._monthly_returns():
            bar = '+' * max(0, int(r * 20)) if r >= 0 else '-' * min(20, int(abs(r) * 20))
            print(f"  {m}: {r:+.4f}%  {bar}")

        if self.closed_trades:
            st = sorted(self.closed_trades, key=lambda t: t.pnl, reverse=True)
            print(f"\n  【最佳交易 Top 5】")
            for t in st[:5]:
                p = self.products[t.product_key]
                td = self._trading_days_held(t.confirm_date, t.sell_date)
                print(f"  {p.code} | {p.name[:20]} | "
                      f"{t.confirm_date}→{t.sell_date}({td}交易日) | "
                      f"盈亏{t.pnl / 1e4:+,.2f}万")

            print(f"\n  【最差交易 Top 5】")
            for t in st[-5:]:
                p = self.products[t.product_key]
                td = self._trading_days_held(t.confirm_date, t.sell_date)
                print(f"  {p.code} | {p.name[:20]} | "
                      f"{t.confirm_date}→{t.sell_date}({td}交易日) | "
                      f"盈亏{t.pnl / 1e4:+,.2f}万")

        bank_stats = defaultdict(lambda: {'n': 0, 'pnl': 0, 'wins': 0})
        for t in self.closed_trades:
            b = self.products[t.product_key].bank
            bank_stats[b]['n'] += 1
            bank_stats[b]['pnl'] += t.pnl
            if t.pnl > 0:
                bank_stats[b]['wins'] += 1
        print(f"\n  【按银行统计】")
        for b, s in sorted(bank_stats.items()):
            wr2 = s['wins'] / s['n'] * 100 if s['n'] else 0
            print(f"  {b}: {s['n']}笔 | 盈亏{s['pnl'] / 1e4:+,.2f}万 | 胜率{wr2:.0f}%")

        print(f"\n  报告文件: {OUTPUT_PATH}")
        print(sep)

    # ------ Excel 报告 ------

    def _generate_report(self):
        logger.info("生成Excel回测报告...")
        wb = Workbook()

        hf = Font(bold=True, color="FFFFFF", size=11)
        tf = Font(bold=True, size=14)
        gf = PatternFill("solid", fgColor="2E7D32")
        bf = PatternFill("solid", fgColor="1565C0")
        df_ = PatternFill("solid", fgColor="37474F")
        rf = Font(color="CC0000", bold=True)
        gnf = Font(color="2E7D32", bold=True)
        ct = Alignment(horizontal="center", vertical="center")

        fv = self.daily_values[-1][1] if self.daily_values else 初始资金
        total_ret = (fv / 初始资金 - 1) * 100
        t0 = pd.Timestamp(self.all_sim_dates[0])
        t1 = pd.Timestamp(self.all_sim_dates[-1])
        yrs = (t1 - t0).days / 365
        ann_ret = ((fv / 初始资金) ** (1 / yrs) - 1) * 100 if yrs > 0 else 0

        peak = 初始资金
        max_dd = 0
        for _, v, _, _ in self.daily_values:
            if v > peak: peak = v
            dd = (peak - v) / peak * 100
            if dd > max_dd: max_dd = dd

        nt = len(self.closed_trades)
        nw = sum(1 for t in self.closed_trades if t.pnl > 0)
        wr = nw / nt * 100 if nt else 0
        total_pnl = sum(t.pnl for t in self.closed_trades)

        # Sheet 1: 汇总
        ws1 = wb.active
        ws1.title = "回测汇总"
        rows = [
            ["V5.4策略历史回测报告 (T+1 + 假期处理)", ""],
            ["", ""],
            ["回测参数", ""],
            ["回测区间", f"{self.all_sim_dates[0]} ~ {self.all_sim_dates[-1]}"],
            ["初始资金", f"{初始资金 / 1e4:,.0f}万"],
            ["持仓规则", f"最多{最大持仓数}个，每个{单仓金额 / 1e4:.0f}万"],
            ["买入信号", f"年化>{信号阈值_买入}%{'(需突破)' if 需要突破信号 else '(无需突破)'}"],
            ["卖出信号", f"年化≤{信号阈值_卖出}%"],
            ["成功定义", f"持有{持有天数_评估}天平均年化>{成功标准}%"],
            ["最低成功率", f"{最低成功率}%"],
            ["产品库重建", f"每{产品库重建间隔}天"],
            ["", ""],
            ["T+1交易规则 + 假期处理", ""],
            ["申购确认", "T+1日确认(周末/假期顺延)"],
            ["赎回到账", "T+1日到账(周末/假期顺延)"],
            ["最短持有", f"确认后{最短持有交易日_赎回}个交易日"],
            ["最大持有", f"{最大持有交易日}个交易日"],
            ["节前限制", f"长假前(间隔>{长假阈值_天}天)不提交申购"],
            ["卖出条件", f"收益≤{信号阈值_卖出}% 或 持有≥{最大持有交易日}交易日"],
            ["", ""],
            ["收益指标", ""],
            ["最终市值", f"{fv / 1e4:,.0f}万"],
            ["总收益率", f"{total_ret:+.4f}%"],
            ["年化收益率", f"{ann_ret:+.4f}%"],
            ["总盈亏", f"{total_pnl / 1e4:+,.2f}万"],
            ["基准(存款2.5%年化)", f"{初始资金 * (1 + 0.025 * yrs) / 1e4:,.0f}万"],
            ["", ""],
            ["风险指标", ""],
            ["最大回撤", f"{max_dd:.4f}%"],
            ["", ""],
            ["交易统计", ""],
            ["总交易次数", f"{nt}"],
            ["胜率", f"{wr:.1f}%"],
            ["盈利笔数", f"{nw}"],
            ["亏损笔数", f"{nt - nw}"],
        ]
        for r in rows:
            ws1.append(r)
        ws1['A1'].font = tf
        ws1.column_dimensions['A'].width = 22
        ws1.column_dimensions['B'].width = 40

        # Sheet 2: 交易明细
        ws2 = wb.create_sheet("交易明细")
        h2 = ['日期', '操作', '银行', '产品代码', '产品名称',
              '净值', '金额(万)', '盈亏(万)', '持有天数', '原因']
        ws2.append(h2)
        for c in ws2[1]:
            c.font = hf
            c.fill = bf
            c.alignment = ct
        for t in self.trade_log:
            ws2.append([
                t['date'], t['action'], t['bank'], t['code'],
                t['name'][:25], round(t['nav'], 6) if t['nav'] else '',
                round(t['amount'] / 1e4, 2) if t['amount'] else '',
                round(t['pnl'] / 1e4, 2),
                t.get('hold_days', ''), t['reason']
            ])
            rn = ws2.max_row
            pc = ws2.cell(row=rn, column=8)
            if isinstance(pc.value, (int, float)):
                pc.font = gnf if pc.value > 0 else (rf if pc.value < 0 else None)
        for i, w in enumerate([12, 10, 10, 16, 28, 12, 12, 12, 8, 45], 1):
            ws2.column_dimensions[get_column_letter(i)].width = w
        ws2.freeze_panes = 'A2'

        # Sheet 3: 月度收益
        ws3 = wb.create_sheet("月度收益")
        ws3.append(['月份', '收益率%', '累计市值(万)'])
        for c in ws3[1]:
            c.font = hf
            c.fill = gf
            c.alignment = ct
        cum = 初始资金
        for m, r in self._monthly_returns():
            cum = cum * (1 + r / 100)
            ws3.append([m, round(r, 4), round(cum / 1e4, 0)])
        ws3.column_dimensions['A'].width = 12
        ws3.column_dimensions['B'].width = 12
        ws3.column_dimensions['C'].width = 16

        # Sheet 4: 每日市值
        ws4 = wb.create_sheet("每日市值")
        ws4.append(['日期', '组合市值(万)', '现金(万)', '持仓市值(万)', '收益率%'])
        for c in ws4[1]:
            c.font = hf
            c.fill = df_
            c.alignment = ct
        for d, tot, cash, pv in self.daily_values:
            ws4.append([d, round(tot / 1e4, 2), round(cash / 1e4, 2),
                        round(pv / 1e4, 2), round((tot / 初始资金 - 1) * 100, 4)])
        for i, w in enumerate([12, 14, 14, 14, 12], 1):
            ws4.column_dimensions[get_column_letter(i)].width = w
        ws4.freeze_panes = 'A2'

        if len(self.daily_values) > 5:
            chart = LineChart()
            chart.title = "组合市值曲线 (T+1+假期处理)"
            chart.style = 10
            chart.y_axis.title = "市值(万)"
            chart.width = 30
            chart.height = 15
            data_ref = Reference(ws4, min_col=2, min_row=1,
                                 max_row=len(self.daily_values) + 1)
            cats = Reference(ws4, min_col=1, min_row=2,
                             max_row=len(self.daily_values) + 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            chart.series[0].graphicalProperties.line.width = 20000
            ws4.add_chart(chart, "G2")

        try:
            wb.save(OUTPUT_PATH)
            logger.info(f"报告已保存: {OUTPUT_PATH}")
        except Exception as e:
            bk = os.path.join(BASE_DIR,
                              f"回测报告_{datetime.now().strftime('%H%M%S')}.xlsx")
            wb.save(bk)
            logger.info(f"已保存到: {bk}")


if __name__ == '__main__':
    # 阻止系统休眠（运行期间保持唤醒）
    ES_CONTINUOUS = 0x80000000
    ES_SYSTEM_REQUIRED = 0x00000001
    ctypes.windll.kernel32.SetThreadExecutionState(ES_CONTINUOUS | ES_SYSTEM_REQUIRED)
    try:
        engine = BacktestEngine()
        engine.run()
    finally:
        # 恢复默认休眠策略
        ctypes.windll.kernel32.SetThreadExecutionState(ES_CONTINUOUS)
