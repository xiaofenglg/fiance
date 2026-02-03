# -*- coding: utf-8 -*-
"""
银行理财产品收益释放分析程序

功能：
1. 读取净值数据库，计算每日年化收益率
2. 识别"收益释放"信号（日年化收益率 > 4.5%）
3. 标记信号后10个交易日为"可买入"窗口期
4. 跟踪窗口期内实际收益，判断卖出时机
5. 输出分析结果到 bank_product_analysis.xlsx（5个工作表）

使用方法：
    python bank_product_analyse.py

输出工作表：
    1. 年化收益计算结果 - 各产品最近日期的日年化收益率
    2. 高收益标记 - 所有"收益释放"信号事件
    3. 可买入产品清单 - 当前处于可买入窗口的产品
    4. 跟踪期实际收益 - 每个信号事件的逐日跟踪数据
    5. 高收益事件明细 - 所有信号事件的汇总统计
"""

import os
import sys
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging

# ============================================================
# 日志配置
# ============================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('bank_product_analyse.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
日志 = logging.getLogger(__name__)

# ============================================================
# 参数配置
# ============================================================
净值数据库路径 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "净值数据库.xlsx")
输出文件路径 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "bank_product_analysis.xlsx")

收益释放阈值 = 4.5        # 年化收益率阈值（%），超过视为"收益释放"
可买入窗口天数 = 10        # 信号后可买入的交易日数
卖出触发比例 = 0.9         # 跟踪期平均收益 < 信号收益 × 此比例时触发卖出
收益率显示天数 = 30        # Sheet1显示的最近日期数量
信号分析天数 = 30          # 信号检测的回溯交易日数
最大输出行数 = 200000      # 每个工作表最大输出行数


# ============================================================
# 工具函数
# ============================================================

def 是否日期列(列名):
    """判断列名是否为日期格式 YYYY-MM-DD"""
    if not isinstance(列名, str):
        return False
    if len(列名) == 10 and 列名[4] == '-' and 列名[7] == '-':
        try:
            datetime.strptime(列名, '%Y-%m-%d')
            return True
        except ValueError:
            return False
    return False


def 识别产品类别(产品名称):
    """
    根据产品名称识别产品类别

    规则（按优先级判断）:
    1. 直接包含R1-R5标识：按标识分类
    2. 含权益类产品关键词（R3及以上）：
       - 增强、混合、权益、股票、可转债、FOF、偏债、灵活配置、进取、成长
    3. 不含权益类产品关键词（R1/R2）：
       - 纯债、货币、稳健（不含增强）、固收（不含增强/混合）
    4. 无法识别: 未知类别

    返回: '含权益类' 或 '不含权益类' 或 '未知类别'
    """
    if not isinstance(产品名称, str):
        产品名称 = str(产品名称)

    产品名称原始 = 产品名称
    产品名称 = 产品名称.upper()  # 统一转大写处理

    # 优先级1: 检查是否直接包含R1-R5标识
    if 'R3' in 产品名称 or 'R4' in 产品名称 or 'R5' in 产品名称:
        return '含权益类'
    if 'R1' in 产品名称 or 'R2' in 产品名称:
        return '不含权益类'

    # 优先级2: 含权益类产品关键词（这些关键词表示产品含有权益类资产）
    含权益关键词 = ['增强', '混合', '权益', '股票', '可转债', 'FOF', '偏债',
                    '灵活配置', '进取', '成长', '积极']
    for 关键词 in 含权益关键词:
        if 关键词 in 产品名称原始:
            return '含权益类'

    # 优先级3: 不含权益类产品关键词（纯固收产品）
    # 注意：这里需要排除含有"增强"、"混合"的情况（已在上面处理）
    不含权益关键词 = ['纯债', '货币', '稳健', '固收']
    for 关键词 in 不含权益关键词:
        if 关键词 in 产品名称原始:
            return '不含权益类'

    # 无法识别
    return '未知类别'


# ============================================================
# 数据加载
# ============================================================

def 加载净值数据库():
    """
    加载净值数据库Excel文件

    返回: {银行名称: DataFrame} 字典
    """
    if not os.path.exists(净值数据库路径):
        raise FileNotFoundError(f"净值数据库文件不存在: {净值数据库路径}")

    日志.info(f"正在加载净值数据库: {净值数据库路径}")
    xlsx文件 = pd.ExcelFile(净值数据库路径)
    数据库 = {}

    for 工作表名 in xlsx文件.sheet_names:
        df = pd.read_excel(xlsx文件, sheet_name=工作表名, dtype=str)
        df.columns = [str(列) for 列 in df.columns]

        # 处理可能的索引列名问题
        重命名映射 = {}
        if 'level_0' in df.columns:
            重命名映射['level_0'] = '产品代码'
        if 'level_1' in df.columns:
            重命名映射['level_1'] = '产品名称'
        if 重命名映射:
            df = df.rename(columns=重命名映射)

        if '产品代码' not in df.columns or '产品名称' not in df.columns:
            日志.warning(f"  {工作表名}: 缺少产品代码或产品名称列，跳过")
            continue

        数据库[工作表名] = df
        全部日期 = sorted([列 for 列 in df.columns if 是否日期列(列)])
        日志.info(f"  {工作表名}: {len(df)} 个产品, {len(全部日期)} 个日期 "
                  f"({全部日期[0] if 全部日期 else 'N/A'} ~ {全部日期[-1] if 全部日期 else 'N/A'})")

    return 数据库


def _加载JSON净值数据(银行名):
    """
    尝试加载银行对应的JSON爬虫数据，用于辅助产品代码匹配

    返回: {产品代码: {日期: 净值字符串}} 或 None
    """
    import json
    基础目录 = os.path.dirname(os.path.abspath(__file__))

    # 华夏银行 → huaxia_nav_history.json (dict: code → [{date, unit_nav}, ...])
    if '华夏' in 银行名:
        路径 = os.path.join(基础目录, 'huaxia_nav_history.json')
        if os.path.exists(路径):
            try:
                with open(路径, 'r', encoding='utf-8') as f:
                    数据 = json.load(f)
                结果 = {}
                for 代码, 净值列表 in 数据.items():
                    净值字典 = {}
                    for 条目 in 净值列表:
                        净值字典[条目['date']] = str(条目['unit_nav'])
                    结果[代码] = 净值字典
                日志.info(f"    已加载华夏JSON数据: {len(结果)} 个产品")
                return 结果
            except Exception as e:
                日志.warning(f"    加载华夏JSON失败: {e}")

    return None


def 修复产品代码(数据库):
    """
    修复数据库中缺失的产品代码和名称

    策略（投票匹配法）：
    1. 将有代码和无代码的行分组
    2. 构建反向索引: (日期, 净值) → 产品代码集合
    3. 对无代码行的每个净值，在反向索引中查找候选代码，按匹配日期数投票
    4. 取得票最多的代码作为匹配结果
    5. 同时尝试使用JSON爬虫数据增强匹配
    6. 合并相同代码的行（净值数据取并集）
    """
    日志.info("开始修复产品代码（投票匹配法）...")

    修复后数据库 = {}
    for 工作表名, df in 数据库.items():
        缺失数 = df['产品代码'].isna().sum()
        if 缺失数 == 0:
            日志.info(f"  {工作表名}: 无缺失，跳过")
            修复后数据库[工作表名] = df
            continue

        日志.info(f"  {工作表名}: {缺失数}/{len(df)} 行缺失产品代码，开始修复...")

        日期列 = sorted([列 for 列 in df.columns if 是否日期列(列)])
        有代码掩码 = df['产品代码'].notna()
        有代码df = df[有代码掩码].copy()
        无代码df = df[~有代码掩码].copy()

        # 构建代码→名称映射
        代码名称映射 = {}
        for _, 行 in 有代码df.iterrows():
            代码 = str(行['产品代码']).strip()
            名称 = str(行['产品名称']).strip()
            if 代码 and 名称:
                代码名称映射[代码] = 名称

        # 构建反向索引: (日期, 净值字符串) → set(产品代码)
        日志.info(f"    构建反向索引（{len(日期列)} 个日期）...")
        日期净值到代码 = {}

        # 来源1: 数据库有代码行
        for _, 行 in 有代码df.iterrows():
            代码 = str(行['产品代码']).strip()
            for 日期 in 日期列:
                值 = 行.get(日期)
                if pd.notna(值) and str(值).strip():
                    键 = (日期, str(值).strip())
                    if 键 not in 日期净值到代码:
                        日期净值到代码[键] = set()
                    日期净值到代码[键].add(代码)

        # 来源2: JSON爬虫数据（如有）
        JSON数据 = _加载JSON净值数据(工作表名)
        if JSON数据:
            for 代码, 净值字典 in JSON数据.items():
                # 同时补充代码名称映射（JSON可能有数据库没有的代码）
                for 日期 in 日期列:
                    if 日期 in 净值字典:
                        键 = (日期, 净值字典[日期])
                        if 键 not in 日期净值到代码:
                            日期净值到代码[键] = set()
                        日期净值到代码[键].add(代码)

        日志.info(f"    反向索引条目数: {len(日期净值到代码)}")

        # 投票匹配: 对每个无代码行，统计各候选代码的匹配日期数
        已匹配 = 0
        歧义匹配 = 0

        for idx, 行 in 无代码df.iterrows():
            代码投票 = {}
            for 日期 in 日期列:
                值 = 行.get(日期)
                if pd.notna(值) and str(值).strip():
                    键 = (日期, str(值).strip())
                    if 键 in 日期净值到代码:
                        for 候选代码 in 日期净值到代码[键]:
                            代码投票[候选代码] = 代码投票.get(候选代码, 0) + 1

            if not 代码投票:
                continue

            # 取得票最多的代码
            最佳代码 = max(代码投票, key=代码投票.get)
            最佳票数 = 代码投票[最佳代码]
            票数排名 = sorted(代码投票.values(), reverse=True)

            无代码df.at[idx, '产品代码'] = 最佳代码
            无代码df.at[idx, '产品名称'] = 代码名称映射.get(最佳代码, 最佳代码)

            if len(票数排名) > 1 and 票数排名[0] == 票数排名[1]:
                歧义匹配 += 1
            else:
                已匹配 += 1

        日志.info(f"    唯一匹配: {已匹配}, 歧义匹配(取最高票): {歧义匹配}, "
                  f"总匹配: {已匹配+歧义匹配}/{len(无代码df)}")

        # 为仍然无法匹配的行生成占位代码
        银行缩写 = 工作表名.replace('银行', '')
        未匹配序号 = 0
        for idx, 行 in 无代码df.iterrows():
            if pd.isna(行['产品代码']) or str(行['产品代码']).strip() == '':
                未匹配序号 += 1
                无代码df.at[idx, '产品代码'] = f'未知-{银行缩写}-{未匹配序号:04d}'
                无代码df.at[idx, '产品名称'] = f'未知产品-{银行缩写}-{未匹配序号:04d}'

        if 未匹配序号 > 0:
            日志.info(f"    生成占位代码: {未匹配序号} 个")

        # 合并: 有代码行 + 已修复的无代码行
        已修复 = pd.concat([有代码df, 无代码df], ignore_index=True)

        # 去重合并: 相同产品代码的多行合并净值数据
        日志.info(f"    合并重复行...")
        合并结果 = []
        按代码分组 = 已修复.groupby('产品代码')
        for 代码, 分组 in 按代码分组:
            if len(分组) == 1:
                合并结果.append(分组.iloc[0])
            else:
                # 多行合并: 取第一行的基本信息，净值取非空值
                合并行 = 分组.iloc[0].copy()
                for 日期 in 日期列:
                    非空值 = 分组[日期].dropna()
                    if len(非空值) > 0:
                        合并行[日期] = 非空值.iloc[0]
                合并结果.append(合并行)

        修复df = pd.DataFrame(合并结果).reset_index(drop=True)
        修复后数据库[工作表名] = 修复df

        日志.info(f"  {工作表名}: 修复完成, {len(有代码df)} + {len(无代码df)} → {len(修复df)} 行")

    日志.info("产品代码修复完成")
    return 修复后数据库


# ============================================================
# 收益率计算
# ============================================================

def 计算年化收益率(df, 银行名):
    """
    计算所有产品的每日年化收益率（向量化计算）

    公式: 年化收益率(%) = (当日净值 / 前一日净值 - 1) × 365 × 100

    返回: (收益率DataFrame, 全部日期列表, 收益率日期列表)
    """
    全部日期 = sorted([列 for 列 in df.columns if 是否日期列(列)])

    if len(全部日期) < 2:
        日志.warning(f"[{银行名}] 日期数量不足（{len(全部日期)}），无法计算收益率")
        return None, [], []

    日志.info(f"[{银行名}] 开始计算年化收益率...")

    # 提取净值矩阵，转换为浮点数
    净值矩阵 = df[全部日期].apply(pd.to_numeric, errors='coerce')

    # 向量化计算: 日收益率 = (当前列 - 前一列) / 前一列, 年化 = 日收益率 × 365 × 100
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", FutureWarning)
        收益率矩阵 = 净值矩阵.pct_change(axis=1) * 365 * 100

    # 收益率从第二个日期开始
    收益率日期 = 全部日期[1:]

    # 构建结果DataFrame（一次性创建，避免碎片化）
    收益率数据 = np.round(收益率矩阵[收益率日期].values, 4)
    结果 = pd.DataFrame(收益率数据, columns=收益率日期)
    结果.insert(0, '银行', 银行名)
    结果.insert(0, '产品名称', df['产品名称'].values)
    结果.insert(0, '产品代码', df['产品代码'].values)

    # 添加产品类别列（根据产品名称识别）
    结果['产品类别'] = df['产品名称'].apply(识别产品类别)
    # 将产品类别列移到产品名称后面
    列顺序 = ['产品代码', '产品名称', '产品类别', '银行'] + 收益率日期
    结果 = 结果[列顺序]

    # 过滤掉没有任何有效收益率的产品
    有效掩码 = 结果[收益率日期].notna().any(axis=1)
    结果 = 结果[有效掩码].reset_index(drop=True)

    # 统计产品类别分布
    类别统计 = 结果['产品类别'].value_counts()
    日志.info(f"[{银行名}] 收益率计算完成: {len(结果)} 个有效产品, {len(收益率日期)} 个日期")
    for 类别, 数量 in 类别统计.items():
        日志.info(f"  {类别}: {数量} 个产品")

    return 结果, 全部日期, 收益率日期


# ============================================================
# 信号识别与窗口分析
# ============================================================

def 识别并分析信号(收益率df, 收益率日期, 银行名):
    """
    识别收益释放信号并分析买入窗口

    返回: (信号列表, 事件明细列表, 可买入清单, 跟踪期收益数据)
    """
    if 收益率df is None or len(收益率日期) < 2:
        return [], [], [], []

    # 确定分析的日期范围
    分析起始索引 = max(0, len(收益率日期) - 信号分析天数)
    分析日期 = 收益率日期[分析起始索引:]
    最新日期索引 = len(收益率日期) - 1

    日志.info(f"[{银行名}] 信号分析回溯 {len(分析日期)} 个交易日 "
              f"({分析日期[0]} ~ {分析日期[-1]})")

    # 向量化识别信号
    信号矩阵 = 收益率df[分析日期].apply(pd.to_numeric, errors='coerce')
    信号掩码 = 信号矩阵 > 收益释放阈值
    信号总数 = int(信号掩码.sum().sum())
    日志.info(f"[{银行名}] 发现 {信号总数} 个收益释放信号")

    if 信号总数 == 0:
        return [], [], [], []

    全部信号 = []
    全部事件明细 = []
    全部跟踪收益 = []
    可买入字典 = {}

    # 遍历有信号的产品
    有信号产品掩码 = 信号掩码.any(axis=1)
    有信号索引列表 = 收益率df.index[有信号产品掩码].tolist()

    已处理 = 0
    for 行索引 in 有信号索引列表:
        产品代码 = 收益率df.at[行索引, '产品代码']
        产品名称 = 收益率df.at[行索引, '产品名称']
        产品类别 = 收益率df.at[行索引, '产品类别']
        产品行 = 收益率df.iloc[行索引 if isinstance(行索引, int) else 收益率df.index.get_loc(行索引)]

        # 该产品的信号日期
        行掩码 = 信号掩码.loc[行索引]
        信号日期列表 = [日期 for 日期 in 分析日期 if 行掩码.get(日期, False)]

        for 信号日期 in 信号日期列表:
            信号收益率 = float(信号矩阵.at[行索引, 信号日期])

            # 记录信号
            全部信号.append({
                '银行': 银行名,
                '产品代码': 产品代码,
                '产品名称': 产品名称,
                '产品类别': 产品类别,
                '信号日期': 信号日期,
                '信号年化收益率': round(信号收益率, 4),
                '收益释放标记': '是'
            })

            # 跟踪窗口期
            if 信号日期 not in 收益率日期:
                continue
            信号位置 = 收益率日期.index(信号日期)
            窗口结束 = min(信号位置 + 可买入窗口天数, len(收益率日期) - 1)

            跟踪列表 = []
            卖出已触发 = False
            卖出触发日 = ''
            卖出触发原因 = ''

            for j in range(信号位置 + 1, 窗口结束 + 1):
                跟踪日期 = 收益率日期[j]
                原始值 = 产品行.get(跟踪日期)
                if 原始值 is None or pd.isna(原始值):
                    continue
                try:
                    跟踪收益值 = float(原始值)
                except (ValueError, TypeError):
                    continue

                跟踪天数 = j - 信号位置
                历史收益 = [x['当日年化收益率'] for x in 跟踪列表] + [跟踪收益值]
                累计平均 = np.mean(历史收益)
                卖出阈值 = 信号收益率 * 卖出触发比例

                触发标记 = ''
                if not 卖出已触发 and 累计平均 < 卖出阈值:
                    卖出已触发 = True
                    卖出触发日 = 跟踪日期
                    卖出触发原因 = (f"第{跟踪天数}天均收益{累计平均:.2f}%"
                                    f"<阈值{卖出阈值:.2f}%")
                    触发标记 = '是'

                跟踪列表.append({
                    '银行': 银行名,
                    '产品代码': 产品代码,
                    '产品名称': 产品名称,
                    '产品类别': 产品类别,
                    '信号日期': 信号日期,
                    '信号年化收益率': round(信号收益率, 4),
                    '跟踪日期': 跟踪日期,
                    '跟踪天数': 跟踪天数,
                    '当日年化收益率': round(跟踪收益值, 4),
                    '累计平均收益': round(累计平均, 4),
                    '卖出阈值': round(卖出阈值, 4),
                    '是否触发卖出': 触发标记
                })

            全部跟踪收益.extend(跟踪列表)

            # 事件统计
            已过交易日 = min(最新日期索引 - 信号位置, 可买入窗口天数)
            剩余天数 = max(0, 可买入窗口天数 - (最新日期索引 - 信号位置))

            if 跟踪列表:
                收益值列表 = [x['当日年化收益率'] for x in 跟踪列表]
                窗口均值 = round(np.mean(收益值列表), 4)
                窗口最高 = round(max(收益值列表), 4)
                窗口最低 = round(min(收益值列表), 4)
            else:
                窗口均值 = 窗口最高 = 窗口最低 = None

            if 卖出已触发:
                状态 = '已触发卖出'
            elif 剩余天数 > 0:
                状态 = '可买入'
            else:
                状态 = '已过期'

            全部事件明细.append({
                '银行': 银行名, '产品代码': 产品代码, '产品名称': 产品名称, '产品类别': 产品类别,
                '信号日期': 信号日期, '信号年化收益率': round(信号收益率, 4),
                '跟踪天数': len(跟踪列表),
                '窗口期平均收益': 窗口均值, '窗口期最高收益': 窗口最高,
                '窗口期最低收益': 窗口最低, '窗口状态': 状态,
                '卖出触发日': 卖出触发日, '卖出触发原因': 卖出触发原因,
                '已过交易日': 已过交易日, '剩余窗口天数': 剩余天数
            })

            # 可买入清单（保留最新信号）
            if 状态 == '可买入':
                键 = (银行名, 产品代码)
                if 键 not in 可买入字典 or 信号日期 > 可买入字典[键]['信号日期']:
                    可买入字典[键] = {
                        '银行': 银行名, '产品代码': 产品代码, '产品名称': 产品名称, '产品类别': 产品类别,
                        '信号日期': 信号日期, '信号年化收益率': round(信号收益率, 4),
                        '已过交易日': 已过交易日, '剩余窗口天数': 剩余天数,
                        '窗口期平均收益': 窗口均值, '建议操作': '可买入'
                    }

        已处理 += 1
        if 已处理 % 500 == 0:
            日志.info(f"[{银行名}] 已分析 {已处理}/{len(有信号索引列表)} 个产品")

    可买入清单 = sorted(可买入字典.values(),
                        key=lambda x: x['信号年化收益率'], reverse=True)

    日志.info(f"[{银行名}] 分析完成: {len(全部信号)} 信号, "
              f"{len(可买入清单)} 可买入, {len(全部事件明细)} 事件")

    return 全部信号, 全部事件明细, 可买入清单, 全部跟踪收益


# ============================================================
# Excel报告生成（使用pandas批量写入 + openpyxl格式化表头）
# ============================================================


def 格式化表头(ws, 列数):
    """对已写入数据的工作表格式化表头行"""
    表头填充 = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    表头字体 = Font(color="FFFFFF", bold=True, size=11)
    居中 = Alignment(horizontal='center', vertical='center')
    边框 = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))

    for 列号 in range(1, 列数 + 1):
        单元格 = ws.cell(row=1, column=列号)
        单元格.fill = 表头填充
        单元格.font = 表头字体
        单元格.alignment = 居中
        单元格.border = 边框


def 设置列宽(ws, 列宽映射):
    """根据映射设置列宽"""
    for 列号, 宽度 in 列宽映射.items():
        ws.column_dimensions[get_column_letter(列号)].width = 宽度


def 生成分析报告(全部收益率结果, 全部信号, 全部事件明细, 全部可买入清单, 全部跟踪收益):
    """
    生成完整的Excel分析报告（5个工作表）

    使用pandas批量写入数据，再用openpyxl格式化表头，大幅提升大数据量写入性能
    """
    日志.info("正在生成分析报告...")

    # ---- Sheet1: 年化收益计算结果 ----
    日志.info("  准备Sheet1 年化收益计算结果...")
    全部日期集合 = set()
    数据帧列表 = []

    for 银行名, (收益率df, _, 收益率日期) in 全部收益率结果.items():
        if 收益率df is None:
            continue
        最近日期 = 收益率日期[-收益率显示天数:] if len(收益率日期) > 收益率显示天数 else 收益率日期
        全部日期集合.update(最近日期)
        数据帧列表.append(收益率df)

    if 数据帧列表:
        合并df = pd.concat(数据帧列表, ignore_index=True)
        日期列 = sorted(全部日期集合, reverse=True)

        if 日期列 and 日期列[0] in 合并df.columns:
            合并df['_排序键'] = pd.to_numeric(合并df[日期列[0]], errors='coerce').fillna(-9999)
            合并df = 合并df.sort_values('_排序键', ascending=False).drop('_排序键', axis=1)

        信息列 = ['银行', '产品代码', '产品名称', '产品类别']
        有效日期列 = [d for d in 日期列 if d in 合并df.columns]
        sheet1列 = 信息列 + 有效日期列
        sheet1列 = [c for c in sheet1列 if c in 合并df.columns]
        sheet1_df = 合并df[sheet1列].head(最大输出行数)
    else:
        sheet1_df = pd.DataFrame({'提示': ['无数据']})

    # ---- Sheet2: 高收益标记 ----
    日志.info("  准备Sheet2 高收益标记...")
    排序信号 = sorted(全部信号, key=lambda x: (str(x.get('信号日期', '')),
                                               float(x.get('信号年化收益率', 0))),
                      reverse=True)
    if len(排序信号) > 最大输出行数:
        排序信号 = 排序信号[:最大输出行数]
    if 排序信号:
        sheet2_df = pd.DataFrame(排序信号,
                                  columns=['银行', '产品代码', '产品名称', '产品类别', '信号日期',
                                           '信号年化收益率', '收益释放标记'])
    else:
        sheet2_df = pd.DataFrame({'提示': ['未发现收益释放信号']})

    # ---- Sheet3: 可买入产品清单 ----
    日志.info("  准备Sheet3 可买入产品清单...")
    排序可买入 = sorted(全部可买入清单, key=lambda x: x['信号年化收益率'], reverse=True)
    if 排序可买入:
        sheet3_df = pd.DataFrame(排序可买入,
                                  columns=['银行', '产品代码', '产品名称', '产品类别', '信号日期',
                                           '信号年化收益率', '已过交易日', '剩余窗口天数',
                                           '窗口期平均收益', '建议操作'])
    else:
        sheet3_df = pd.DataFrame({'提示': ['当前无可买入产品']})

    # ---- Sheet4: 跟踪期实际收益 ----
    日志.info("  准备Sheet4 跟踪期实际收益...")
    排序跟踪 = sorted(全部跟踪收益,
                      key=lambda x: (str(x.get('信号日期', '')),
                                     str(x.get('产品代码', '')),
                                     int(x.get('跟踪天数', 0))),
                      reverse=True)
    if len(排序跟踪) > 最大输出行数:
        排序跟踪 = 排序跟踪[:最大输出行数]
        日志.warning(f"  Sheet4 行数过多，截取前 {最大输出行数} 行")
    if 排序跟踪:
        sheet4_df = pd.DataFrame(排序跟踪,
                                  columns=['银行', '产品代码', '产品名称', '产品类别', '信号日期',
                                           '信号年化收益率', '跟踪日期', '跟踪天数',
                                           '当日年化收益率', '累计平均收益', '卖出阈值',
                                           '是否触发卖出'])
    else:
        sheet4_df = pd.DataFrame({'提示': ['无跟踪期数据']})

    # ---- Sheet5: 高收益事件明细 ----
    日志.info("  准备Sheet5 高收益事件明细...")
    排序事件 = sorted(全部事件明细,
                      key=lambda x: (str(x.get('信号日期', '')),
                                     float(x.get('信号年化收益率', 0))),
                      reverse=True)
    if len(排序事件) > 最大输出行数:
        排序事件 = 排序事件[:最大输出行数]
    if 排序事件:
        sheet5_df = pd.DataFrame(排序事件,
                                  columns=['银行', '产品代码', '产品名称', '产品类别', '信号日期',
                                           '信号年化收益率', '跟踪天数', '窗口期平均收益',
                                           '窗口期最高收益', '窗口期最低收益', '窗口状态',
                                           '卖出触发日', '卖出触发原因', '已过交易日',
                                           '剩余窗口天数'])
    else:
        sheet5_df = pd.DataFrame({'提示': ['无事件明细']})

    # ---- 批量写入Excel ----
    日志.info("  写入Excel文件（批量模式）...")
    实际输出路径 = 输出文件路径
    try:
        # 尝试打开文件检查是否被占用
        with open(输出文件路径, 'a'):
            pass
    except PermissionError:
        实际输出路径 = 输出文件路径.replace('.xlsx', f'_{datetime.now().strftime("%H%M%S")}.xlsx')
        日志.warning(f"  原文件被占用，改为输出到: {实际输出路径}")

    with pd.ExcelWriter(实际输出路径, engine='openpyxl') as writer:
        sheet1_df.to_excel(writer, sheet_name='年化收益计算结果', index=False)
        sheet2_df.to_excel(writer, sheet_name='高收益标记', index=False)
        sheet3_df.to_excel(writer, sheet_name='可买入产品清单', index=False)
        sheet4_df.to_excel(writer, sheet_name='跟踪期实际收益', index=False)
        sheet5_df.to_excel(writer, sheet_name='高收益事件明细', index=False)

        wb = writer.book

        # 格式化各工作表表头和列宽
        # 处理可能为空的工作表（动态列数）
        sheet2_列数 = len(sheet2_df.columns)
        sheet2_列宽映射 = {1: 30} if sheet2_列数 == 1 else {1: 12, 2: 22, 3: 40, 4: 14, 5: 14, 6: 16, 7: 14}

        sheet3_列数 = len(sheet3_df.columns)
        sheet3_列宽映射 = {1: 30} if sheet3_列数 == 1 else {1: 12, 2: 22, 3: 40, 4: 14, 5: 14, 6: 16, 7: 12, 8: 14, 9: 16, 10: 12}

        sheet4_列数 = len(sheet4_df.columns)
        sheet4_列宽映射 = {1: 30} if sheet4_列数 == 1 else {1: 12, 2: 22, 3: 40, 4: 14, 5: 14, 6: 16, 7: 14, 8: 10, 9: 16, 10: 14, 11: 12, 12: 14}

        sheet5_列数 = len(sheet5_df.columns)
        sheet5_列宽映射 = {1: 30} if sheet5_列数 == 1 else {1: 12, 2: 22, 3: 40, 4: 14, 5: 14, 6: 16, 7: 10, 8: 16, 9: 16, 10: 16, 11: 14, 12: 14, 13: 36, 14: 12, 15: 14}

        for 工作表名, 列数, 列宽映射, 冻结 in [
            ('年化收益计算结果', len(sheet1_df.columns),
             {1: 12, 2: 22, 3: 40, 4: 14}, 'E2'),  # 新增产品类别列
            ('高收益标记', sheet2_列数, sheet2_列宽映射, 'A2'),  # 动态处理列数
            ('可买入产品清单', sheet3_列数, sheet3_列宽映射, 'A2'),  # 动态处理列数
            ('跟踪期实际收益', sheet4_列数, sheet4_列宽映射, 'A2'),  # 动态处理列数
            ('高收益事件明细', sheet5_列数, sheet5_列宽映射, 'A2'),  # 动态处理列数
        ]:
            ws = wb[工作表名]
            格式化表头(ws, 列数)
            设置列宽(ws, 列宽映射)
            # 对日期列设置默认宽度
            if 列宽映射:
                起始列号 = max(列宽映射.keys()) + 1
                if 起始列号 <= 列数:
                    for 列号 in range(起始列号, 列数 + 1):
                        ws.column_dimensions[get_column_letter(列号)].width = 13
            ws.freeze_panes = 冻结

    日志.info(f"分析报告已保存: {实际输出路径}")
    日志.info(f"  Sheet1 年化收益计算结果: {len(sheet1_df)} 行")
    日志.info(f"  Sheet2 高收益标记: {len(sheet2_df)} 行")
    日志.info(f"  Sheet3 可买入产品清单: {len(sheet3_df)} 行")
    日志.info(f"  Sheet4 跟踪期实际收益: {len(sheet4_df)} 行")
    日志.info(f"  Sheet5 高收益事件明细: {len(sheet5_df)} 行")


# ============================================================
# 主函数
# ============================================================

def main():
    """主入口"""
    print("=" * 70)
    print("           银行理财产品收益释放分析程序")
    print("=" * 70)
    print(f"运行时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"数据来源: {净值数据库路径}")
    print(f"输出文件: {输出文件路径}")
    print(f"收益释放阈值: {收益释放阈值}%")
    print(f"可买入窗口: {可买入窗口天数} 个交易日")
    print(f"卖出触发: 平均收益 < 信号收益 × {卖出触发比例*100:.0f}%")
    print(f"信号回溯: {信号分析天数} 个交易日")
    print()

    # 第1步: 加载数据
    print("【第1步】加载净值数据库...")
    净值数据 = 加载净值数据库()

    # 第1.5步: 修复缺失的产品代码
    print("\n【第1.5步】修复产品代码...")
    净值数据 = 修复产品代码(净值数据)

    # 第2步: 计算收益率
    print("\n【第2步】计算年化收益率...")
    全部收益率结果 = {}

    for 银行名, df in 净值数据.items():
        收益率df, 全部日期, 收益率日期 = 计算年化收益率(df, 银行名)
        if 收益率df is not None:
            全部收益率结果[银行名] = (收益率df, 全部日期, 收益率日期)

    # 第3步: 识别信号并分析
    print("\n【第3步】识别收益释放信号...")
    汇总信号 = []
    汇总事件 = []
    汇总可买入 = []
    汇总跟踪 = []

    for 银行名, (收益率df, _, 收益率日期) in 全部收益率结果.items():
        信号, 事件, 可买入, 跟踪 = 识别并分析信号(收益率df, 收益率日期, 银行名)
        汇总信号.extend(信号)
        汇总事件.extend(事件)
        汇总可买入.extend(可买入)
        汇总跟踪.extend(跟踪)

    # 第4步: 生成报告
    print("\n【第4步】生成分析报告...")
    生成分析报告(全部收益率结果, 汇总信号, 汇总事件, 汇总可买入, 汇总跟踪)

    # 汇总统计
    print("\n" + "=" * 70)
    print("                    分析完成")
    print("=" * 70)
    print(f"收益释放信号总数: {len(汇总信号)}")
    print(f"当前可买入产品数: {len(汇总可买入)}")
    print(f"高收益事件总数:   {len(汇总事件)}")
    print(f"跟踪期记录总数:   {len(汇总跟踪)}")

    # 按银行统计
    for 银行名 in 全部收益率结果:
        银行信号数 = sum(1 for s in 汇总信号 if s['银行'] == 银行名)
        银行可买入 = sum(1 for s in 汇总可买入 if s['银行'] == 银行名)
        print(f"  {银行名}: {银行信号数} 个信号, {银行可买入} 个可买入")

    # 按产品类别统计
    print("\n产品类别统计:")
    if 汇总信号:
        类别统计 = {}
        for 信号 in 汇总信号:
            类别 = 信号.get('产品类别', '未知类别')
            类别统计[类别] = 类别统计.get(类别, 0) + 1
        for 类别, 数量 in sorted(类别统计.items()):
            print(f"  {类别}: {数量} 个信号")

    if 汇总可买入:
        可买入类别统计 = {}
        for 产品 in 汇总可买入:
            类别 = 产品.get('产品类别', '未知类别')
            可买入类别统计[类别] = 可买入类别统计.get(类别, 0) + 1
        print("\n可买入产品类别分布:")
        for 类别, 数量 in sorted(可买入类别统计.items()):
            print(f"  {类别}: {数量} 个")

    print(f"\n输出文件: {输出文件路径}")
    print("=" * 70)

    return 输出文件路径


if __name__ == "__main__":
    main()
