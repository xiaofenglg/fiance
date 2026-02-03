# -*- coding: utf-8 -*-
"""
收益释放时间规律分析程序 V2 — 阶段分组 + 节假日窗口 + 季节自适应

分析范围: 2023~2025年三年净值数据
阶段分组: 月初(1-7) / 月上旬(8-14) / 月中(15-21) / 月底(22-31)
节假日窗口: 春节前一周 / 端午前三天 / 国庆前五天 / 清明前后两天
季节调节: 夏秋(6-9月)门槛降至涨差(>0%); 冬春用涨差³加权
权重因子: 各季度历史平均释放次数

数据源: 净值数据库.xlsx (民生银行, 中信银行, 华夏银行)
输出:   收益释放规律分析.xlsx
"""

import os
import re
import sys
import warnings
import numpy as np
import pandas as pd
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Windows 终端 UTF-8
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

# ============================================================
# 配置
# ============================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "净值数据库.xlsx")
OUTPUT_PATH = os.path.join(BASE_DIR, "收益释放规律分析.xlsx")

DATE_START = '2023-01-01'
DATE_END = '2025-12-31'
HOLD_DAYS = 7          # 赎回评估持有期
长间隔保护天数 = 60
最长赎回天数 = 14

# --- 阶段定义 ---
MONTHLY_PHASES = [
    ('月初(1-7号)',    1,  7),
    ('月上旬(8-14号)', 8, 14),
    ('月中(15-21号)', 15, 21),
    ('月底(22-31号)', 22, 31),
]

# --- 节假日窗口 (2023-2025 农历/阳历) ---
# 春节除夕: 2023-01-21, 2024-02-09, 2025-01-28  → 前7天
# 端午:     2023-06-22, 2024-06-10, 2025-05-31  → 前3天
# 国庆:     每年10-01                            → 前5天
# 清明:     2023-04-05, 2024-04-04, 2025-04-04  → 前后各2天
def _holiday_ranges():
    ranges = defaultdict(list)
    for eve in ['2023-01-21', '2024-02-09', '2025-01-28']:
        end = pd.Timestamp(eve) - pd.Timedelta(days=1)
        start = end - pd.Timedelta(days=6)
        ranges['春节前一周'].append((start, end))
    for dt in ['2023-06-22', '2024-06-10', '2025-05-31']:
        end = pd.Timestamp(dt) - pd.Timedelta(days=1)
        start = end - pd.Timedelta(days=2)
        ranges['端午前三天'].append((start, end))
    for y in [2023, 2024, 2025]:
        end = pd.Timestamp(f'{y}-09-30')
        start = end - pd.Timedelta(days=4)
        ranges['国庆前五天'].append((start, end))
    for dt in ['2023-04-05', '2024-04-04', '2025-04-04']:
        c = pd.Timestamp(dt)
        ranges['清明前后两天'].append((c - pd.Timedelta(days=2),
                                       c + pd.Timedelta(days=2)))
    return dict(ranges)

HOLIDAY_RANGES = _holiday_ranges()
HOLIDAY_NAMES = list(HOLIDAY_RANGES.keys())

# ============================================================
# 复用核心函数
# ============================================================
def is_date_col(col_name):
    if not isinstance(col_name, str):
        return False
    return len(col_name) == 10 and col_name[4] == '-' and col_name[7] == '-'


def 判断产品流动性(name):
    if '日开' in name or '天天' in name or '每日' in name:
        return '日开', 0
    m = re.search(r'(\d+)\s*天持有期', name)
    if m: return f'{int(m.group(1))}天持有期', int(m.group(1))
    m = re.search(r'最短持有\s*(\d+)\s*天', name)
    if m: return f'{int(m.group(1))}天持有', int(m.group(1))
    m = re.search(r'周期\s*(\d+)\s*天', name)
    if m: return f'{int(m.group(1))}天周期', int(m.group(1))
    m = re.search(r'(\d+)\s*个?月', name)
    if m and '半年' not in name:
        months = int(m.group(1))
        if months <= 6: return f'{months}个月', months * 30
    for kw, lbl, d in [('周开','周开',7),('日申季赎','季赎',90),('月开','月开',30),
                        ('季开','季开',90),('季度开','季开',90),('半年','半年',180),
                        ('年开','年开',365),('封闭','封闭',999),('定开','定开',999)]:
        if kw in name: return lbl, d
    m = re.search(r'(\d+)\s*Y\s*持有', name)
    if m: return f'{int(m.group(1))}年持有', int(m.group(1)) * 365
    if re.search(r'20[2-5]\d', name) and '信颐' in name:
        return '目标日期', 999
    return '未知', 999


def get_product_returns(row, date_cols):
    nav_points = []
    for d in date_cols:
        try:
            v = float(row[d])
            if not np.isnan(v) and v > 0:
                nav_points.append((d, v))
        except (ValueError, TypeError):
            pass
    if len(nav_points) < 2:
        return {}
    rets = {}
    for i in range(1, len(nav_points)):
        d0, n0 = nav_points[i - 1]
        d1, n1 = nav_points[i]
        gap = (pd.Timestamp(d1) - pd.Timestamp(d0)).days
        if gap > 长间隔保护天数 or gap <= 0 or n0 <= 0:
            continue
        rets[d1] = (n1 / n0 - 1) / gap * 365 * 100
    return rets


def 智能合并数据(df, date_cols):
    有代码 = df[df['产品代码'].notna()].copy()
    无代码 = df[df['产品代码'].isna()].copy()
    if len(无代码) == 0:
        return 有代码
    有净值 = 有代码[date_cols].apply(pd.to_numeric, errors='coerce')
    无净值 = 无代码[date_cols].apply(pd.to_numeric, errors='coerce')
    重叠 = [d for d in date_cols if 有净值[d].notna().sum() > 0 and 无净值[d].notna().sum() > 0]
    if len(重叠) < 3:
        return 有代码
    匹配日 = 重叠[-20:] if len(重叠) > 20 else 重叠
    idx_map = {}
    for idx, row in 有代码.iterrows():
        kv = []
        for d in 匹配日:
            val = row.get(d)
            if pd.notna(val):
                try: kv.append(str(round(float(val), 4)))
                except: pass
        if len(kv) >= 3: idx_map['|'.join(kv)] = idx
    res = {idx: row.to_dict() for idx, row in 有代码.iterrows()}
    for idx, row in 无代码.iterrows():
        kv = []
        for d in 匹配日:
            val = row.get(d)
            if pd.notna(val):
                try: kv.append(str(round(float(val), 4)))
                except: pass
        if len(kv) >= 3:
            key = '|'.join(kv)
            if key in idx_map:
                orig = idx_map[key]
                for d in date_cols:
                    if pd.notna(row.get(d)):
                        res[orig][d] = row[d]
    return pd.DataFrame(list(res.values()))


# ============================================================
# 季节阈值与权重
# ============================================================
def is_summer_autumn(month):
    return 6 <= month <= 9


def seasonal_threshold(month):
    """夏秋>0%涨差; 冬春>3.0%"""
    return 0.0 if is_summer_autumn(month) else 3.0


def seasonal_weight(ret_pct, month):
    """冬春: (min(ret/3, 3))³ 上限27倍; 夏秋: 等权1.0"""
    if is_summer_autumn(month):
        return 1.0
    # 冬春涨差³ — 归一化到阈值, 上限3倍防止极端值主导
    ratio = min(ret_pct / 3.0, 3.0)
    return max(ratio, 0.01) ** 3


# ============================================================
# 数据加载与事件识别
# ============================================================
def load_and_identify():
    """
    返回: events, bank_summary, bank_data_months
      bank_data_months: {bank: set((year,month), ...)}  实际数据覆盖月份
    """
    print("=" * 72)
    print("  收益释放规律分析 V2 — 阶段分组 · 节假日窗口 · 季节自适应")
    print("=" * 72)
    print(f"\n  数据源 : {DB_PATH}")
    print(f"  区间   : {DATE_START} ~ {DATE_END}")
    print(f"  夏秋(6-9月): 阈值>0%等权   冬春: 阈值>3.0%涨差³加权")
    print(f"  流动性 : 赎回≤{最长赎回天数}天\n")

    xlsx = pd.ExcelFile(DB_PATH)
    ts_start = pd.Timestamp(DATE_START)
    ts_end = pd.Timestamp(DATE_END)

    all_events = []
    bank_summary = {}
    bank_data_months = {}   # bank -> set of (year, month)

    for sheet in xlsx.sheet_names:
        print(f"  [{sheet}] ", end="", flush=True)
        df = pd.read_excel(xlsx, sheet_name=sheet, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]
        if 'level_0' in df.columns:
            df = df.rename(columns={'level_0': '产品代码', 'level_1': '产品名称'})

        date_cols = sorted([c for c in df.columns if is_date_col(c)])
        if len(date_cols) < 10:
            print("日期列不足, 跳过")
            continue

        # ★ 数据月份: 从列标题统计真实覆盖范围
        data_months = set()
        for d in date_cols:
            ts = pd.Timestamp(d)
            if ts_start <= ts <= ts_end:
                data_months.add((ts.year, ts.month))
        bank_data_months[sheet] = data_months

        df = 智能合并数据(df, date_cols)
        n_total = n_tradable = n_with_events = 0

        for _, row in df.iterrows():
            code = row.get('产品代码')
            name = row.get('产品名称')
            if pd.isna(code):
                continue
            n_total += 1
            _, redeem_days = 判断产品流动性(str(name) if pd.notna(name) else '')
            if redeem_days > 最长赎回天数:
                continue
            n_tradable += 1

            rets = get_product_returns(row, date_cols)
            if len(rets) < 5:
                continue

            sorted_dates = sorted(rets.keys())
            has_event = False

            for di, date_str in enumerate(sorted_dates):
                ts = pd.Timestamp(date_str)
                if ts < ts_start or ts > ts_end:
                    continue
                ret_val = rets[date_str]
                month = ts.month
                if ret_val <= seasonal_threshold(month):
                    continue

                # 赎回成功: 后续 HOLD_DAYS 个数据点平均>0
                hold_rets = []
                for j in range(di + 1, min(di + 1 + HOLD_DAYS, len(sorted_dates))):
                    hold_rets.append(rets[sorted_dates[j]])
                hold_success = (np.mean(hold_rets) > 0) if hold_rets else False

                all_events.append({
                    'date':  date_str,
                    'ts':    ts,
                    'bank':  sheet,
                    'code':  str(code),
                    'name':  str(name) if pd.notna(name) else '',
                    'ret':   ret_val,
                    'month': month,
                    'day':   ts.day,
                    'quarter': (month - 1) // 3 + 1,
                    'hold_success': hold_success,
                    'sw': seasonal_weight(ret_val, month),
                })
                has_event = True

            if has_event:
                n_with_events += 1

        bank_n = sum(1 for e in all_events if e['bank'] == sheet)
        bank_summary[sheet] = {
            '总产品': n_total, '可交易': n_tradable,
            '有释放产品': n_with_events, '释放事件': bank_n,
        }
        print(f"产品{n_total} / 可交易{n_tradable} / "
              f"有释放{n_with_events} / 事件{bank_n}  "
              f"(数据覆盖{len(data_months)}个月)")

    print(f"\n  合计事件: {len(all_events)}")
    return all_events, bank_summary, bank_data_months


# ============================================================
# 季度权重因子
# ============================================================
def compute_quarterly_weights(events):
    """各季度历史平均释放次数 → 归一化权重"""
    q_counts = defaultdict(int)
    q_years = defaultdict(set)
    for e in events:
        q_counts[e['quarter']] += 1
        q_years[e['quarter']].add(e['ts'].year)

    q_avg = {}
    for q in range(1, 5):
        ny = max(len(q_years[q]), 1)
        q_avg[q] = q_counts[q] / ny

    overall = np.mean([v for v in q_avg.values() if v > 0]) if q_avg else 1.0
    if overall == 0:
        overall = 1.0
    return {q: q_avg.get(q, 0) / overall for q in range(1, 5)}


# ============================================================
# 分组辅助
# ============================================================
def classify_monthly_phase(day):
    for name, lo, hi in MONTHLY_PHASES:
        if lo <= day <= hi:
            return name
    return MONTHLY_PHASES[-1][0]


def classify_holiday(ts):
    matched = []
    for hname, ranges in HOLIDAY_RANGES.items():
        for (s, e) in ranges:
            if s <= ts <= e:
                matched.append(hname)
                break
    return matched


# ============================================================
# 统计计算
# ============================================================
def compute_phase_stats(phase_events, n_periods, qw):
    """
    返回: {平均释放次数, 平均幅度(bp), 波动率(bp), 赎回成功率(%), 阶段概率(%)}
    n_periods = 该阶段实际出现的期数 (月度用数据月数, 节假日用年数)
    """
    empty = {'平均释放次数': 0, '平均幅度(bp)': 0, '波动率(bp)': 0,
             '赎回成功率(%)': 0, '阶段概率(%)': 0, '事件数': 0}
    if not phase_events:
        return empty

    n = len(phase_events)
    rets = np.array([e['ret'] for e in phase_events])

    # 综合权重 = 季节权重 × 季度权重
    weights = np.array([e['sw'] * qw.get(e['quarter'], 1.0)
                        for e in phase_events])
    wsum = weights.sum()
    if wsum == 0:
        wsum = 1.0

    avg_ret = np.dot(rets, weights) / wsum                # 加权均值 (%)
    var = np.dot(weights, (rets - avg_ret) ** 2) / wsum   # 加权方差
    avg_bp = round(avg_ret * 100, 1)                      # → bp
    vol_bp = round(np.sqrt(max(var, 0)) * 100, 1)

    succ = np.array([1.0 if e['hold_success'] else 0.0 for e in phase_events])
    succ_rate = round(np.dot(succ, weights) / wsum * 100, 1)

    avg_count = round(n / max(n_periods, 1), 1)

    period_keys = set()
    for e in phase_events:
        period_keys.add((e['ts'].year, e['ts'].month))
    prob = round(len(period_keys) / max(n_periods, 1) * 100, 1)

    return {
        '平均释放次数': avg_count,
        '平均幅度(bp)': avg_bp,
        '波动率(bp)': vol_bp,
        '赎回成功率(%)': succ_rate,
        '阶段概率(%)': min(prob, 100.0),
        '事件数': n,
    }


# ============================================================
# 主分析
# ============================================================
def run_analysis(events, bank_summary, bank_data_months):
    banks = sorted(bank_summary.keys())
    qw = compute_quarterly_weights(events)

    print("\n  季度权重因子:")
    for q in range(1, 5):
        print(f"    Q{q}: {qw[q]:.3f}")

    # 全局数据月数 = 所有银行月份的并集
    global_dm = set()
    for s in bank_data_months.values():
        global_dm |= s
    global_n = len(global_dm)

    print(f"  全局数据月数: {global_n}")
    for b in banks:
        print(f"    {b}: {len(bank_data_months.get(b, set()))}个月")

    # 分桶
    monthly_buckets = defaultdict(lambda: defaultdict(list))
    holiday_buckets = defaultdict(lambda: defaultdict(list))

    for e in events:
        phase = classify_monthly_phase(e['day'])
        monthly_buckets[phase][e['bank']].append(e)
        monthly_buckets[phase]['全局'].append(e)
        for hname in classify_holiday(e['ts']):
            holiday_buckets[hname][e['bank']].append(e)
            holiday_buckets[hname]['全局'].append(e)

    # 月度统计
    monthly_results = {}
    for pname, _, _ in MONTHLY_PHASES:
        for scope in ['全局'] + banks:
            ev = monthly_buckets[pname].get(scope, [])
            np_ = global_n if scope == '全局' else len(bank_data_months.get(scope, set()))
            monthly_results[(pname, scope)] = compute_phase_stats(ev, np_, qw)

    # 节假日统计
    holiday_results = {}
    # 每个银行实际覆盖的年数
    bank_years = {}
    for b in banks:
        bank_years[b] = len(set(ym[0] for ym in bank_data_months.get(b, set())))
    global_years = len(set(ym[0] for ym in global_dm))

    for hname in HOLIDAY_NAMES:
        for scope in ['全局'] + banks:
            ev = holiday_buckets[hname].get(scope, [])
            ny = global_years if scope == '全局' else bank_years.get(scope, 3)
            holiday_results[(hname, scope)] = compute_phase_stats(ev, ny, qw)

    return monthly_results, holiday_results, qw, global_n, bank_data_months


# ============================================================
# 控制台输出
# ============================================================
def print_table(title, phase_names, results, scopes):
    print(f"\n{'─' * 72}")
    print(f"  {title}")
    print(f"{'─' * 72}")

    col_w = [18, 8, 8, 10, 10, 8, 8]
    header = ['阶段', '事件数', '平均释放', '幅度bp', '波动bp', '成功率%', '概率%']
    for scope in scopes:
        print(f"\n  ▸ {scope}")
        print("  " + "".join(h.ljust(w) for h, w in zip(header, col_w)))
        print("  " + "─" * sum(col_w))
        for pname in phase_names:
            s = results.get((pname, scope))
            if s is None:
                continue
            row = [
                pname,
                str(s['事件数']),
                f"{s['平均释放次数']:.1f}",
                f"{s['平均幅度(bp)']:.0f}",
                f"{s['波动率(bp)']:.0f}",
                f"{s['赎回成功率(%)']:.1f}",
                f"{s['阶段概率(%)']:.1f}",
            ]
            print("  " + "".join(v.ljust(w) for v, w in zip(row, col_w)))


def print_results(monthly_results, holiday_results, bank_summary, qw,
                  global_n, bank_data_months):
    banks = sorted(bank_summary.keys())
    scopes = ['全局'] + banks

    print(f"\n{'─' * 72}")
    print("  银行总览")
    print(f"{'─' * 72}")
    for b, s in bank_summary.items():
        nm = len(bank_data_months.get(b, set()))
        print(f"  {b}: 产品{s['总产品']} / 可交易{s['可交易']} / "
              f"有释放{s['有释放产品']} / 事件{s['释放事件']}  ({nm}个月)")

    phase_names = [p[0] for p in MONTHLY_PHASES]
    print_table("月度阶段分析", phase_names, monthly_results, scopes)
    print_table("节假日窗口分析", HOLIDAY_NAMES, holiday_results, scopes)

    # 关键发现
    print(f"\n{'─' * 72}")
    print("  关键发现")
    print(f"{'─' * 72}")

    best_mag = max(phase_names, key=lambda p: monthly_results.get((p, '全局'), {}).get('平均幅度(bp)', 0))
    best_suc = max(phase_names, key=lambda p: monthly_results.get((p, '全局'), {}).get('赎回成功率(%)', 0))
    best_prob = max(phase_names, key=lambda p: monthly_results.get((p, '全局'), {}).get('阶段概率(%)', 0))

    bm = monthly_results[(best_mag, '全局')]
    bs = monthly_results[(best_suc, '全局')]
    bp_ = monthly_results[(best_prob, '全局')]

    print(f"  幅度最高阶段 : {best_mag} → {bm['平均幅度(bp)']:.0f}bp")
    print(f"  成功率最高   : {best_suc} → {bs['赎回成功率(%)']:.1f}%")
    print(f"  概率最高     : {best_prob} → {bp_['阶段概率(%)']:.1f}%")

    if HOLIDAY_NAMES:
        valid_h = [h for h in HOLIDAY_NAMES
                   if holiday_results.get((h, '全局'), {}).get('事件数', 0) > 0]
        if valid_h:
            best_h = max(valid_h,
                         key=lambda h: holiday_results[(h, '全局')].get('平均幅度(bp)', 0))
            bh = holiday_results[(best_h, '全局')]
            print(f"  最佳节假日   : {best_h} → {bh['平均幅度(bp)']:.0f}bp / "
                  f"成功率{bh['赎回成功率(%)']:.1f}%")

    for b in banks:
        best_p = max(phase_names,
                     key=lambda p: monthly_results.get((p, b), {}).get('赎回成功率(%)', 0))
        st = monthly_results[(best_p, b)]
        print(f"  {b}最佳: {best_p} → "
              f"成功率{st['赎回成功率(%)']:.1f}% / {st['平均幅度(bp)']:.0f}bp")


# ============================================================
# Excel 输出
# ============================================================
def write_excel(monthly_results, holiday_results, bank_summary, qw,
                global_n, bank_data_months, events):
    wb = Workbook()
    hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    bold = Font(bold=True, size=10)
    thin = Border(*(Side(style='thin'),) * 4)
    ctr = Alignment(horizontal='center', vertical='center')
    gold = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    def whdr(ws, headers, r=1):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.font, cell.fill, cell.alignment, cell.border = hdr_font, hdr_fill, ctr, thin

    def autow(ws):
        for col in ws.columns:
            mx = 0
            lt = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    mx = max(mx, sum(2 if ord(ch) > 127 else 1 for ch in str(cell.value)))
            ws.column_dimensions[lt].width = min(mx + 4, 40)

    def write_stats(ws, phase_names, results, scopes, sr=1):
        headers = ['分析范围', '阶段', '事件数', '平均释放次数',
                    '平均幅度(bp)', '波动率(bp)', '赎回成功率(%)', '阶段概率(%)']
        whdr(ws, headers, sr)
        r = sr + 1
        for scope in scopes:
            for pname in phase_names:
                s = results.get((pname, scope))
                if s is None: continue
                vals = [scope, pname, s['事件数'], s['平均释放次数'],
                        s['平均幅度(bp)'], s['波动率(bp)'],
                        s['赎回成功率(%)'], s['阶段概率(%)']]
                for c, v in enumerate(vals, 1):
                    cell = ws.cell(row=r, column=c, value=v)
                    cell.border, cell.alignment = thin, ctr
                if s['赎回成功率(%)'] >= 75:
                    ws.cell(row=r, column=7).fill = gold
                r += 1
        return r

    banks = sorted(bank_summary.keys())
    scopes = ['全局'] + banks

    # Sheet 1: 月度阶段分析
    ws1 = wb.active
    ws1.title = "月度阶段分析"
    write_stats(ws1, [p[0] for p in MONTHLY_PHASES], monthly_results, scopes)
    autow(ws1)

    # Sheet 2: 节假日窗口
    ws2 = wb.create_sheet("节假日窗口")
    write_stats(ws2, HOLIDAY_NAMES, holiday_results, scopes)
    autow(ws2)

    # Sheet 3: 季节权重
    ws3 = wb.create_sheet("季节权重")
    whdr(ws3, ['季度', '权重因子', '说明'])
    labels = {1: 'Q1(1-3月)', 2: 'Q2(4-6月)', 3: 'Q3(7-9月)', 4: 'Q4(10-12月)'}
    notes = {1: '冬春: 阈值>3%涨差³加权', 2: '混合(4-5冬春/6夏秋)',
             3: '夏秋: 阈值>0%等权', 4: '秋冬: 阈值>3%涨差³加权'}
    for i, q in enumerate(range(1, 5), 2):
        ws3.cell(row=i, column=1, value=labels[q]).border = thin
        ws3.cell(row=i, column=2, value=round(qw[q], 4)).border = thin
        ws3.cell(row=i, column=3, value=notes[q]).border = thin
    r = 7
    ws3.cell(row=r, column=1, value='分析参数').font = bold
    for txt in [f'分析区间: {DATE_START} ~ {DATE_END}',
                f'夏秋(6-9月): 阈值>0%, 等权(1.0)',
                f'冬春(其余): 阈值>3.0%, 权重=(min(ret/3,3))³  上限27',
                f'赎回评估: 持有{HOLD_DAYS}天均值>0%为成功',
                f'流动性: 赎回≤{最长赎回天数}天',
                f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M")}']:
        r += 1
        ws3.cell(row=r, column=1, value=txt)
    autow(ws3)

    # Sheet 4: 总览
    ws4 = wb.create_sheet("总览")
    whdr(ws4, ['银行', '总产品', '可交易', '有释放产品', '释放事件', '数据月数'])
    r = 2
    for b in banks:
        s = bank_summary[b]
        for c, v in enumerate([b, s['总产品'], s['可交易'], s['有释放产品'],
                                s['释放事件'], len(bank_data_months.get(b, set()))], 1):
            ws4.cell(row=r, column=c, value=v).border = thin
        r += 1
    ws4.cell(row=r, column=1, value='合计').font = bold
    ws4.cell(row=r, column=1).border = thin
    for ci, key in enumerate(['总产品', '可交易', '有释放产品', '释放事件'], 2):
        ws4.cell(row=r, column=ci,
                 value=sum(s[key] for s in bank_summary.values())).border = thin
    ws4.cell(row=r, column=6, value=global_n).border = thin
    autow(ws4)

    # Sheet 5: 释放事件明细（完整输出）
    ws5 = wb.create_sheet("释放事件明细")
    whdr(ws5, ['日期', '银行', '产品代码', '产品名称',
                '年化收益(%)', '月份阶段', '节假日窗口',
                '季节权重', '季度权重', '赎回成功'])
    sorted_ev = sorted(events, key=lambda e: e['date'])
    for i, e in enumerate(sorted_ev, 2):
        phase = classify_monthly_phase(e['day'])
        hstr = ','.join(classify_holiday(e['ts'])) or '-'
        vals = [e['date'], e['bank'], e['code'], e['name'],
                round(e['ret'], 2), phase, hstr,
                round(e['sw'], 2), round(qw.get(e['quarter'], 1.0), 4),
                '是' if e['hold_success'] else '否']
        for c, v in enumerate(vals, 1):
            ws5.cell(row=i, column=c, value=v).border = thin
    autow(ws5)

    # 保存 — 如果文件被占用, 加后缀
    path = OUTPUT_PATH
    try:
        wb.save(path)
    except PermissionError:
        path = OUTPUT_PATH.replace('.xlsx', f'_{datetime.now():%H%M%S}.xlsx')
        wb.save(path)

    print(f"\n  已保存: {path}")
    print(f"  工作表: {', '.join(wb.sheetnames)}")


# ============================================================
# 主函数
# ============================================================
def main():
    warnings.filterwarnings('ignore')

    if not os.path.exists(DB_PATH):
        print(f"错误: 找不到 {DB_PATH}")
        sys.exit(1)

    events, bank_summary, bank_data_months = load_and_identify()
    if not events:
        print("未找到释放事件, 退出")
        sys.exit(0)

    # 全局月数
    global_dm = set()
    for s in bank_data_months.values():
        global_dm |= s
    global_n = len(global_dm)

    monthly_results, holiday_results, qw, _, _ = run_analysis(
        events, bank_summary, bank_data_months)

    print_results(monthly_results, holiday_results, bank_summary, qw,
                  global_n, bank_data_months)

    write_excel(monthly_results, holiday_results, bank_summary, qw,
                global_n, bank_data_months, events)

    print("\n  分析完成。")


if __name__ == '__main__':
    main()
