# -*- coding: utf-8 -*-
"""
计算净值数据库的每日年化收益率

功能：
1. 读取净值数据库
2. 计算每日年化收益率 = 日收益率 * 365
3. 按最近一日收益率从高到低排序
4. 只显示最近20个净值日期
5. 单日收益超过4.5%标红
6. 生成Excel文件
"""

import os
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# 文件路径
NAV_DB_FILE = os.path.join(os.path.dirname(__file__), "净值数据库.xlsx")
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "年化收益率排名_全部.xlsx")

# 红色标记阈值
RED_THRESHOLD = 4.5  # 单日收益超过4.5%


def load_nav_database():
    """加载净值数据库"""
    if not os.path.exists(NAV_DB_FILE):
        raise FileNotFoundError(f"净值数据库不存在: {NAV_DB_FILE}")

    xlsx = pd.ExcelFile(NAV_DB_FILE)
    data = {}
    for sheet_name in xlsx.sheet_names:
        df = pd.read_excel(xlsx, sheet_name=sheet_name, dtype=str)
        # 确保列名是字符串格式
        df.columns = [str(c) if not isinstance(c, str) else c for c in df.columns]

        # 处理可能的索引重置后的列名
        rename_map = {}
        if 'level_0' in df.columns:
            rename_map['level_0'] = '产品代码'
        if 'level_1' in df.columns:
            rename_map['level_1'] = '产品名称'
        if rename_map:
            df = df.rename(columns=rename_map)

        data[sheet_name] = df
        logger.info(f"加载 {sheet_name}: {len(df)} 个产品, 列: {list(df.columns)[:5]}")
        # 打印日期列信息用于调试
        date_cols = sorted([c for c in df.columns if is_date_column(c)], reverse=True)
        logger.info(f"  日期列数量: {len(date_cols)}, 最新: {date_cols[0] if date_cols else 'N/A'}")
    return data


def is_date_column(col_name):
    """判断是否是日期列"""
    if not isinstance(col_name, str):
        return False
    if len(col_name) == 10 and col_name[4] == '-' and col_name[7] == '-':
        try:
            datetime.strptime(col_name, '%Y-%m-%d')
            return True
        except:
            return False
    return False


def calculate_daily_return(df, num_dates=20):
    """
    计算每日年化收益率

    Args:
        df: 原始数据DataFrame
        num_dates: 显示的日期数量

    Returns:
        result_df: 包含年化收益率的DataFrame
        date_columns: 使用的日期列表（最近的在前）
    """
    # 获取日期列并排序（从新到旧）
    date_cols = sorted([c for c in df.columns if is_date_column(c)], reverse=True)

    if len(date_cols) < 2:
        logger.warning("日期数量不足，无法计算收益率")
        return None, []

    # 只取最近的 num_dates + 1 个日期（需要多一个来计算收益率）
    selected_dates = date_cols[:num_dates + 1]

    # 产品基本信息列
    info_cols = ['产品代码', '产品名称']
    info_cols = [c for c in info_cols if c in df.columns]

    if not info_cols:
        logger.warning(f"未找到产品信息列，可用列: {list(df.columns)[:10]}")

    # 创建结果DataFrame
    result_data = []

    for idx, row in df.iterrows():
        product_info = {col: row[col] for col in info_cols}

        # 计算每日年化收益率
        daily_returns = {}
        valid_return_count = 0

        for i in range(len(selected_dates) - 1):
            current_date = selected_dates[i]
            prev_date = selected_dates[i + 1]

            try:
                current_nav = float(row[current_date]) if pd.notna(row.get(current_date)) and str(row.get(current_date)).strip() else None
                prev_nav = float(row[prev_date]) if pd.notna(row.get(prev_date)) and str(row.get(prev_date)).strip() else None

                if current_nav and prev_nav and prev_nav > 0:
                    # 日收益率
                    daily_return = (current_nav - prev_nav) / prev_nav
                    # 年化收益率 = 日收益率 * 365 * 100 (转为百分比)
                    annual_return = daily_return * 365 * 100
                    daily_returns[current_date] = round(annual_return, 2)
                    valid_return_count += 1
                else:
                    daily_returns[current_date] = None
            except (ValueError, TypeError):
                daily_returns[current_date] = None

        # 只保留至少有一个有效收益率的产品
        if valid_return_count > 0:
            product_info.update(daily_returns)
            result_data.append(product_info)

    if not result_data:
        return None, []

    result_df = pd.DataFrame(result_data)

    # 返回的日期列（不包含最后一个用于计算的日期）
    return_dates = selected_dates[:-1]

    return result_df, return_dates


def create_excel_with_formatting(all_results, output_file):
    """
    创建带格式的Excel文件（所有银行合并到一个Sheet）

    Args:
        all_results: {银行名: (df, date_cols)} 的字典
        output_file: 输出文件路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "年化收益率排名"

    # 定义样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    red_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')

    # 合并所有银行数据
    combined_data = []
    all_date_cols = set()

    for bank_name, (df, date_cols) in all_results.items():
        if df is None or df.empty:
            continue

        # 添加银行列
        df = df.copy()
        df['银行'] = bank_name

        # 收集所有日期列
        all_date_cols.update(date_cols)

        combined_data.append((df, date_cols))

    if not combined_data:
        logger.warning("没有数据可以写入")
        return

    # 获取统一的日期列（从新到旧排序）
    all_date_cols = sorted(list(all_date_cols), reverse=True)[:20]

    # 合并所有DataFrame
    all_dfs = []
    for df, date_cols in combined_data:
        all_dfs.append(df)

    combined_df = pd.concat(all_dfs, ignore_index=True)

    # 按最近一日收益率排序（从高到低）
    latest_date = all_date_cols[0] if all_date_cols else None
    if latest_date and latest_date in combined_df.columns:
        combined_df['_sort_key'] = combined_df[latest_date].apply(
            lambda x: float(x) if pd.notna(x) and x is not None else float('-inf')
        )
        combined_df = combined_df.sort_values('_sort_key', ascending=False)
        combined_df = combined_df.drop('_sort_key', axis=1)

    # 构建列顺序：银行、产品代码、产品名称、日期列（从新到旧）
    info_cols = ['银行', '产品代码', '产品名称']
    info_cols = [c for c in info_cols if c in combined_df.columns]
    ordered_cols = info_cols + list(all_date_cols)

    # 调试信息
    logger.info(f"合并后DataFrame列: {list(combined_df.columns)[:10]}...")
    logger.info(f"日期列: {list(all_date_cols)[:5]}...")

    # 只保留存在的列
    ordered_cols = [c for c in ordered_cols if c in combined_df.columns]
    logger.info(f"最终列数: {len(ordered_cols)}, 包含: {ordered_cols[:5]}...")
    df_output = combined_df[ordered_cols]

    # 写入表头
    for col_idx, col_name in enumerate(ordered_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_alignment

    # 写入数据
    info_col_count = len(info_cols)
    for row_idx, row in enumerate(df_output.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            # 数值列居中对齐
            if col_idx > info_col_count:
                cell.alignment = center_alignment

                # 检查是否需要标红（单日年化收益超过阈值）
                try:
                    if value is not None and float(value) > RED_THRESHOLD:
                        cell.fill = red_fill
                        cell.font = red_font
                except (ValueError, TypeError):
                    pass

    # 调整列宽
    for col_idx, col_name in enumerate(ordered_cols, 1):
        if col_name == '产品名称':
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 40
        elif col_name == '产品代码':
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20
        elif col_name == '银行':
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 12
        else:
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 12

    # 冻结首行和前三列（银行、产品代码、产品名称）
    ws.freeze_panes = 'D2'

    logger.info(f"合计写入 {len(df_output)} 条数据")

    # 保存文件
    wb.save(output_file)
    logger.info(f"文件已保存到: {output_file}")


def main():
    """主函数"""
    print("=" * 60)
    print("         净值数据库年化收益率计算")
    print("=" * 60)
    print(f"开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"数据来源: {NAV_DB_FILE}")
    print(f"输出文件: {OUTPUT_FILE}")
    print(f"红色标记阈值: 单日年化收益率 > {RED_THRESHOLD}%")
    print()

    # 加载数据
    nav_data = load_nav_database()

    # 计算各银行的收益率
    all_results = {}

    for bank_name, df in nav_data.items():
        print(f"\n处理 {bank_name}...")
        result_df, date_cols = calculate_daily_return(df, num_dates=20)

        if result_df is not None:
            all_results[bank_name] = (result_df, date_cols)
            print(f"  - 有效产品数: {len(result_df)}")
            print(f"  - 日期范围: {date_cols[-1] if date_cols else 'N/A'} ~ {date_cols[0] if date_cols else 'N/A'}")
        else:
            print(f"  - 无有效数据")

    # 生成Excel
    if all_results:
        create_excel_with_formatting(all_results, OUTPUT_FILE)
        print("\n" + "=" * 60)
        print("处理完成!")
        print(f"输出文件: {OUTPUT_FILE}")
        print("=" * 60)
    else:
        print("没有有效数据可处理")

    return OUTPUT_FILE


if __name__ == "__main__":
    main()
