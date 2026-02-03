# -*- coding: utf-8 -*-
"""
前瞻性收益释放购买策略 — 回测验证程序 V3.2 (DL增强版 — 乘法加成)

基于 backtest_pattern_v3.py (V3.1), 修正DL混合策略:

V3.1 问题诊断:
  - V3.1 Mode D 年化5.6227%, 低于Mode B的5.8953% (退化-0.27%)
  - 原因1: DL-only合成预测(fake_pred)引入噪声, 降低选品质量
  - 原因2: 加法混合(pattern*0.7+DL*0.3)当DL弱时反而稀释好的pattern信号
  - 原因3: DL quality gate过低, 不可靠信号也参与排名

V3.2 修正:
  1. 删除DL-only预测 — 只在已有pattern预测的产品上叠加DL加成
  2. 乘法加成替代加法混合 — score *= (1 + dl_boost * DL_MULT_WEIGHT)
     * DL同意pattern时放大信号, DL不确定时不干预
  3. 提高DL门槛 — entry_quality>=0.50, confidence>=0.50
  4. val_loss质量门控 — val_loss>0.45时跳过DL信号
  5. DL加成上限0.20 — 防止DL过度影响排名

目标: 从Mode B 5.8953%基线提升, 不降低夏普比率

Walk-Forward: 180天训练, 每30天重建 (DL和规律同步)
"""

import os, sys, re, warnings, logging, bisect, time, gc
import pandas as pd
import numpy as np
from datetime import datetime
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple
from scipy import stats as scipy_stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

warnings.filterwarnings('ignore')

# ============================================================
# 配置
# ============================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "净值数据库.xlsx")
OUTPUT_PATH = os.path.join(BASE_DIR, "前瞻策略回测报告_v3_2_dl.xlsx")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('backtest_pattern_v3_2.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# ============================================================
# 策略参数 (与 V5.4 对齐)
# ============================================================
信号阈值_买入 = 3.5
信号阈值_卖出 = 2.0
信号阈值_库 = 3.0       # V5.4: 产品库信号检测用宽阈值(3.0%), 买入决策用严阈值(3.5%)
需要突破信号 = False
持有天数_评估 = 7
成功标准 = 2.5
最低成功率 = 30.0
最少历史信号 = 2
最短持续天数 = 2
最大持有天数_评估 = 30
最长赎回天数 = 14

回测开始日 = '2025-06-01'
回测结束日 = '2026-01-30'
初始资金 = 100_000_000
最大持仓数 = 6
单仓基础 = 初始资金 / 6
最短持有交易日_赎回 = 3
最大持有交易日 = 20
长假阈值_天 = 3
产品库重建间隔 = 7

# 规律学习参数
规律学习窗口天数 = 180
规律重建间隔天数 = 30
释放识别阈值 = 2.5
最低置信度 = 0.4              # 低于此不生成预测
预测窗口天数 = 10
周期CV阈值 = 0.4

# ===== V3 三建议参数 =====
# 建议① 优选: Alpha 加成
ALPHA_预测加成 = 0.30

# 建议② 智持: 延迟卖出
延迟卖出_最低收益 = 0.0       # 收益>此值才允许延迟 (0=非负就行)
延迟卖出_预测窗口 = 4         # 释放预测在N个交易日内才延迟
延迟卖出_置信度 = 0.4         # 预测置信度门槛
延迟卖出_最大持有余量 = 5     # hold_td < 最大持有交易日-此值 才允许延迟

# 建议③ 节前买入
节前买入_实时阈值 = 1.0       # 节前买入最低实时收益 (假期累积, 阈值可更低)
节前买入_最低成功率 = 40.0    # 节前候选要求的库成功率
节前买入_最低均收 = 1.5       # 节前候选要求的库平均收益
节前买入_最大产品数 = 5       # 节前最多买入产品数
节前保留仓位 = 2              # 节前日预留仓位数 (确保每个假期都有节前买入)

# ===== 中国法定假期日历 (完整覆盖回测区间 2025-06 ~ 2026-01) =====
# 假期期间不公布净值, 节后首日公布累积收益
# 数据来源: 国务院办公厅发布的2025年/2026年放假安排
#
# 2025年假期:
#   元旦: 1/1 (回测前)
#   春节: 1/28-2/4 (回测前)
#   清明: 4/4-4/6 (回测前)
#   五一: 5/1-5/5 (回测前)
#   端午: 5/31-6/2 (回测起始, 节前日在回测前)
#   中秋+国庆: 10/1-10/8 (合并连休)
# 2026年假期:
#   元旦: 1/1-1/3
#   春节: 2/17起 (回测后)
#
HOLIDAY_PERIODS = [
    # (假期名称, 开始日, 结束日) — 含首尾
    ('端午', '2025-05-31', '2025-06-02'),   # 端午节 (节前日在回测前, 仅卖出保护)
    ('国庆', '2025-10-01', '2025-10-08'),   # 国庆+中秋连休8天
    ('元旦', '2026-01-01', '2026-01-03'),   # 元旦3天
]
# 节前交易日: 假期前最后1~2个工作日 (适合提前布局, 赚取假期累积收益)
# 仅包含在回测区间内的节前日
PRE_HOLIDAY_DATES = {
    # 端午节前日: 5/29(Thu), 5/30(Fri) — 在回测开始日(6/1)之前, 无法操作
    '2025-09-29': '国庆',   # 国庆前周一
    '2025-09-30': '国庆',   # 国庆前周二(最后工作日)
    '2025-12-31': '元旦',   # 元旦前最后工作日(周三)
}
# 假期日集合 (用于卖出保护: 假期中产品NAV不更新, 0%收益不是真实信号)
HOLIDAY_DATE_SET = set()
for _name, _start, _end in HOLIDAY_PERIODS:
    _d = pd.Timestamp(_start)
    while _d <= pd.Timestamp(_end):
        HOLIDAY_DATE_SET.add(_d.strftime('%Y-%m-%d'))
        _d += pd.Timedelta(days=1)


# ===== [DL BLEND] DL混合参数 (V3.2: 乘法加成, 无DL-only) =====
DL_MULT_WEIGHT = 0.15           # 乘法加成系数: score *= (1 + dl_boost * 此值)
DL_MAX_BOOST = 0.20             # 最大加成幅度 (防止DL过度影响)
DL_MIN_ENTRY_QUALITY = 0.50     # DL entry_quality >= 此值才加成 (V3.1: 0.35)
DL_MIN_CONFIDENCE = 0.50        # DL confidence >= 此值才加成 (V3.1: 0.30)
DL_MAX_VAL_LOSS = 0.45          # val_loss > 此值时跳过DL信号 (模型不可靠)
DL_RETRAIN_INTERVAL = 30        # DL重训间隔(天), 与规律学习同步


class DLAlphaManager:
    """[DL BLEND V3.2] Walk-forward GPU DL预测管理器

    V3.2改进:
    - 新增 last_val_loss 追踪, 支持质量门控
    - is_reliable() 方法判断当前模型是否可用
    """

    def __init__(self):
        self.engine = None
        self.alpha_signals = {}
        self.last_train_date = None
        self.last_val_loss = 999.0
        self.available = False

        try:
            from gpu_alpha_v2 import AlphaEngine
            import torch
            self.engine = AlphaEngine()
            self.available = torch.cuda.is_available()
            if self.available:
                logger.info("[DL Blend V3.2] GPU可用, 乘法加成模式")
            else:
                logger.info("[DL Blend V3.2] GPU不可用, 跳过DL")
        except ImportError as e:
            logger.warning(f"[DL Blend V3.2] 导入失败: {e}")

    def should_retrain(self, current_date):
        if not self.available or self.engine is None:
            return False
        if self.last_train_date is None:
            return True
        gap = (pd.Timestamp(current_date) - pd.Timestamp(self.last_train_date)).days
        return gap >= DL_RETRAIN_INTERVAL

    def is_reliable(self):
        """检查当前DL模型是否可靠 (val_loss质量门控)"""
        return self.last_val_loss < DL_MAX_VAL_LOSS

    def train_and_predict(self, as_of_date):
        if self.engine is None:
            return
        try:
            t0 = time.time()
            logger.info(f"[DL Blend V3.2] Walk-forward 训练 as_of={as_of_date}")
            self.alpha_signals = self.engine.train_and_predict(as_of_date=as_of_date)
            self.last_train_date = as_of_date
            elapsed = time.time() - t0
            self.last_val_loss = getattr(self.engine, '_last_avg_val_loss', 999)
            reliable_tag = "可用" if self.is_reliable() else "不可靠,跳过"
            logger.info(f"[DL Blend V3.2] 训练完成: {len(self.alpha_signals)}产品, "
                        f"{elapsed:.1f}s, val_loss={self.last_val_loss:.3f} ({reliable_tag})")
            gc.collect()
            try:
                import torch
                if torch.cuda.is_available():
                    torch.cuda.empty_cache()
            except ImportError:
                pass
        except Exception as e:
            logger.error(f"[DL Blend V3.2] 训练失败: {e}", exc_info=True)
            self.last_val_loss = 999.0

    def get_signal(self, key):
        return self.alpha_signals.get(key, {})


# ============================================================
# 工具函数
# ============================================================

def is_date_col(col_name):
    if not isinstance(col_name, str):
        return False
    return len(col_name) == 10 and col_name[4] == '-' and col_name[7] == '-'


def 判断产品流动性(name):
    if '日开' in name or '天天' in name or '每日' in name:
        return '日开', 0
    m = re.search(r'(\d+)\s*天持有期', name)
    if m: return f'{int(m.group(1))}天持有期', int(m.group(1))
    m = re.search(r'最短持有\s*(\d+)\s*天', name)
    if m: return f'{int(m.group(1))}天持有', int(m.group(1))
    m = re.search(r'周期\s*(\d+)\s*天', name)
    if m: return f'{int(m.group(1))}天周期', int(m.group(1))
    m = re.search(r'(\d+)\s*个?月', name)
    if m and '半年' not in name:
        months = int(m.group(1))
        if months <= 6: return f'{months}个月', months * 30
    if '周开' in name: return '周开', 7
    if '日申季赎' in name: return '季赎', 90
    if '月开' in name: return '月开', 30
    if '季开' in name or '季度开' in name: return '季开', 90
    if '半年' in name: return '半年', 180
    if '年开' in name: return '年开', 365
    if '封闭' in name: return '封闭', 999
    if '定开' in name: return '定开', 999
    m = re.search(r'(\d+)\s*Y\s*持有', name)
    if m: return f'{int(m.group(1))}年持有', int(m.group(1)) * 365
    if re.search(r'20[2-5]\d', name) and '信颐' in name:
        return '目标日期', 999
    return '未知', 999


# ============================================================
# 数据结构
# ============================================================

class ProductData:
    __slots__ = ['bank', 'code', 'name', 'liquidity', 'redeem_days',
                 'nav_dates', 'navs', 'ret_dates', 'rets',
                 'ret_date_idx', 'signals']
    def __init__(self, bank, code, name, liquidity, redeem_days):
        self.bank = bank
        self.code = code
        self.name = name
        self.liquidity = liquidity
        self.redeem_days = redeem_days
        self.nav_dates = []
        self.navs = {}
        self.ret_dates = []
        self.rets = {}
        self.ret_date_idx = {}
        self.signals = []


class Position:
    __slots__ = ['product_key', 'signal_date', 'confirm_date', 'buy_nav',
                 'amount', 'entry_return', 'alpha_score',
                 'sell_date', 'sell_nav', 'pnl', 'sell_reason',
                 'signal_type', 'is_preholiday_buy']
    def __init__(self, product_key, signal_date, confirm_date, buy_nav, amount,
                 entry_return=0.0, alpha_score=0.0, signal_type='reactive',
                 is_preholiday_buy=False):
        self.product_key = product_key
        self.signal_date = signal_date
        self.confirm_date = confirm_date
        self.buy_nav = buy_nav
        self.amount = amount
        self.entry_return = entry_return
        self.alpha_score = alpha_score
        self.sell_date = None
        self.sell_nav = None
        self.pnl = 0.0
        self.sell_reason = ''
        self.signal_type = signal_type   # 'reactive', 'predicted', 'preholiday'
        self.is_preholiday_buy = is_preholiday_buy


# ============================================================
# 模块1: PatternLearner
# ============================================================

@dataclass
class ReleasePattern:
    product_key: tuple
    period_days: float = 0.0
    period_cv: float = 1.0
    has_period: bool = False
    phase_dist: list = field(default_factory=lambda: [0.0]*4)
    phase_pvalue: float = 1.0
    top_phase: int = 0
    weekday_dist: list = field(default_factory=lambda: [0.0]*5)
    weekday_pvalue: float = 1.0
    top_weekday: int = 0
    confidence: float = 0.0
    n_events: int = 0
    last_release_date: str = ''
    last_release_window_end: str = ''
    avg_window_days: float = 3.0      # 平均释放窗口持续天数


class PatternLearner:
    PHASE_RANGES = [(1, 7), (8, 14), (15, 21), (22, 31)]

    def learn(self, product_key, ret_dates, rets, as_of_date,
              window_days=180):
        cutoff = (pd.Timestamp(as_of_date) - pd.Timedelta(days=window_days)).strftime('%Y-%m-%d')
        wdates = [d for d in ret_dates if cutoff <= d < as_of_date]
        if len(wdates) < 10:
            return None

        starts = []
        prev_below = True
        for d in wdates:
            r = rets.get(d, 0)
            if r > 释放识别阈值:
                if prev_below:
                    starts.append(d)
                prev_below = False
            else:
                prev_below = True

        if len(starts) < 2:
            return None

        pat = ReleasePattern(product_key=product_key)
        pat.n_events = len(starts)
        pat.last_release_date = starts[-1]

        # 计算每个释放事件的窗口持续天数
        window_durations = []
        for s_date in starts:
            si = wdates.index(s_date)
            ei = si
            for j in range(si, len(wdates)):
                if rets.get(wdates[j], 0) > 释放识别阈值 * 0.5:
                    ei = j
                else:
                    break
            dur = (pd.Timestamp(wdates[ei]) - pd.Timestamp(s_date)).days + 1
            window_durations.append(max(dur, 1))
        pat.avg_window_days = float(np.mean(window_durations)) if window_durations else 3.0

        # 最后一个释放窗口的结束日
        last_si = wdates.index(starts[-1])
        last_ei = last_si
        for j in range(last_si, len(wdates)):
            if rets.get(wdates[j], 0) > 释放识别阈值 * 0.5:
                last_ei = j
            else:
                break
        pat.last_release_window_end = wdates[last_ei]

        # 周期
        gaps = []
        for i in range(1, len(starts)):
            g = (pd.Timestamp(starts[i]) - pd.Timestamp(starts[i-1])).days
            if 3 <= g <= 120:
                gaps.append(g)
        if len(gaps) >= 2:
            med = float(np.median(gaps))
            cv = float(np.std(gaps)) / med if med > 0 else 1.0
            pat.period_days = med
            pat.period_cv = cv
            pat.has_period = cv < 周期CV阈值
        elif len(gaps) == 1:
            pat.period_days = float(gaps[0])
            pat.period_cv = 0.5
            pat.has_period = True

        # 阶段
        pc = [0]*4
        for d in starts:
            day = pd.Timestamp(d).day
            for pi, (lo, hi) in enumerate(self.PHASE_RANGES):
                if lo <= day <= hi:
                    pc[pi] += 1; break
        ne = sum(pc)
        if ne >= 3:
            _, pv = scipy_stats.chisquare(pc, f_exp=[ne/4.0]*4)
            pat.phase_dist = [c/ne for c in pc]
            pat.phase_pvalue = pv
            pat.top_phase = int(np.argmax(pc))
        else:
            pat.phase_dist = [c/max(ne,1) for c in pc]
            pat.phase_pvalue = 1.0
            pat.top_phase = int(np.argmax(pc)) if ne > 0 else 0

        # 星期
        wc = [0]*5
        for d in starts:
            wd = pd.Timestamp(d).weekday()
            if wd < 5: wc[wd] += 1
        nw = sum(wc)
        if nw >= 3:
            _, pw = scipy_stats.chisquare(wc, f_exp=[nw/5.0]*5)
            pat.weekday_dist = [c/nw for c in wc]
            pat.weekday_pvalue = pw
            pat.top_weekday = int(np.argmax(wc))
        else:
            pat.weekday_dist = [c/max(nw,1) for c in wc]
            pat.weekday_pvalue = 1.0
            pat.top_weekday = int(np.argmax(wc)) if nw > 0 else 0

        # 置信度
        conf = 0.0
        if pat.has_period and pat.period_cv < 0.3: conf += 0.4
        elif pat.has_period and pat.period_cv < 0.4: conf += 0.25
        if pat.phase_pvalue < 0.05: conf += 0.3
        if pat.weekday_pvalue < 0.05: conf += 0.15
        if pat.n_events >= 5: conf += 0.15
        pat.confidence = min(conf, 1.0)
        return pat


def _格式化释放规律(pat):
    """将 ReleasePattern 格式化为可读字符串"""
    if pat is None:
        return '', '', '', 0.0
    星期名 = ['周一', '周二', '周三', '周四', '周五']
    阶段名 = ['月初(1-7日)', '月中(8-14日)', '月中下(15-21日)', '月末(22-31日)']
    周期 = f"每{pat.period_days:.0f}天" if pat.has_period and pat.period_cv < 0.4 else ''
    阶段 = 阶段名[pat.top_phase] if pat.phase_pvalue < 0.05 and 0 <= pat.top_phase < 4 else ''
    星期 = 星期名[pat.top_weekday] if pat.weekday_pvalue < 0.05 and 0 <= pat.top_weekday < 5 else ''
    return 周期, 阶段, 星期, round(pat.confidence, 2)


# ============================================================
# 模块2: ForwardPredictor (V2 修正版)
# ============================================================

@dataclass
class Prediction:
    product_key: tuple
    predicted_date: str
    confidence: float
    source: str
    td_until: int = 0
    predicted_end_date: str = ''


class ForwardPredictor:

    def predict(self, pattern, current_date, all_sim_dates, date_to_idx):
        if pattern.confidence < 最低置信度:
            return []
        cur_ts = pd.Timestamp(current_date)
        horizon = cur_ts + pd.Timedelta(days=预测窗口天数)
        cur_idx = date_to_idx.get(current_date)
        if cur_idx is None:
            return []

        preds = {}

        # 路径A: 周期 (向前推进修正)
        if pattern.has_period and pattern.last_release_window_end:
            last_ts = pd.Timestamp(pattern.last_release_window_end)
            period = pd.Timedelta(days=max(int(pattern.period_days), 5))
            pt = last_ts + period
            for _ in range(30):
                if pt >= cur_ts - pd.Timedelta(days=2):
                    break
                pt += period
            conf_a = (1 - pattern.period_cv) * 0.7
            for off in range(-2, 3):
                ct = pt + pd.Timedelta(days=off)
                if cur_ts < ct <= horizon:
                    pd_str = ct.strftime('%Y-%m-%d')
                    pi = bisect.bisect_left(all_sim_dates, pd_str)
                    if pi >= len(all_sim_dates): continue
                    td = pi - cur_idx
                    if td < 1: continue
                    if pd_str not in preds:
                        preds[pd_str] = Prediction(pattern.product_key, pd_str, conf_a, 'period', td)
                    else:
                        preds[pd_str].confidence = min(preds[pd_str].confidence + conf_a, 1.0)
                        preds[pd_str].source = 'both'
                        preds[pd_str].td_until = min(preds[pd_str].td_until, td)

        # 路径B: 阶段 (中点 + 宽窗口)
        if pattern.phase_pvalue < 0.15:
            tp = pattern.top_phase
            lo, hi = PatternLearner.PHASE_RANGES[tp]
            mid = (lo + hi) // 2
            prob = pattern.phase_dist[tp] if tp < len(pattern.phase_dist) else 0.25
            conf_b = prob * 0.6
            for mo in range(2):
                ref = cur_ts + pd.DateOffset(months=mo)
                try:
                    mt = pd.Timestamp(f'{ref.year}-{ref.month:02d}-{min(mid,28):02d}')
                except ValueError:
                    continue
                ws = mt - pd.Timedelta(days=2)
                we = mt + pd.Timedelta(days=3)
                if we <= cur_ts or ws > horizon: continue
                pd_str = mt.strftime('%Y-%m-%d')
                pi = bisect.bisect_left(all_sim_dates, pd_str)
                if pi >= len(all_sim_dates): continue
                td = pi - cur_idx
                if td < 1: continue
                if pd_str not in preds:
                    preds[pd_str] = Prediction(pattern.product_key, pd_str, conf_b, 'phase', td)
                else:
                    preds[pd_str].confidence = min(preds[pd_str].confidence + conf_b, 1.0)
                    preds[pd_str].source = 'both'
                    preds[pd_str].td_until = min(preds[pd_str].td_until, td)

        # 为每个预测计算释放结束日 = 预测释放日 + 平均窗口天数
        avg_win = max(int(round(pattern.avg_window_days)), 1)
        for p in preds.values():
            end_ts = pd.Timestamp(p.predicted_date) + pd.Timedelta(days=avg_win - 1)
            p.predicted_end_date = end_ts.strftime('%Y-%m-%d')

        results = [p for p in preds.values() if p.confidence >= 0.25]
        results.sort(key=lambda x: x.td_until)
        return results

    def rank_products(self, patterns, current_date, all_sim_dates, date_to_idx):
        """按预测紧迫度排名: score = confidence / sqrt(td_until)"""
        rankings = []
        for key, pat in patterns.items():
            if pat.confidence < 最低置信度: continue
            ps = self.predict(pat, current_date, all_sim_dates, date_to_idx)
            if not ps: continue
            best_s, best_p = 0.0, None
            for p in ps:
                if p.td_until < 1 or p.td_until > 8: continue
                s = p.confidence / max(p.td_until, 1) ** 0.5
                if s > best_s: best_s, best_p = s, p
            if best_p and best_s > 0.1:
                rankings.append((key, best_s, best_p))
        rankings.sort(key=lambda x: -x[1])
        return rankings

    def has_upcoming_release(self, pattern, current_date, all_sim_dates, date_to_idx,
                              max_td=4, min_conf=0.4):
        """检查是否有即将到来的释放 (用于延迟卖出判断)"""
        ps = self.predict(pattern, current_date, all_sim_dates, date_to_idx)
        for p in ps:
            if 1 <= p.td_until <= max_td and p.confidence >= min_conf:
                return True, p
        return False, None


# ============================================================
# 模块3: PatternBacktestEngine
# ============================================================

class PatternBacktestEngine:
    def __init__(self):
        self.products = {}
        self.all_sim_dates = []
        self.date_to_idx = {}

    def load_data(self):
        logger.info("加载净值数据库...")
        xlsx = pd.ExcelFile(DB_PATH)
        all_dates = set()
        total = 0
        for sheet in xlsx.sheet_names:
            logger.info(f"  加载 [{sheet}]...")
            df = pd.read_excel(xlsx, sheet_name=sheet, dtype=str)
            df.columns = [str(c).strip() for c in df.columns]
            if 'level_0' in df.columns:
                df = df.rename(columns={'level_0': '产品代码', 'level_1': '产品名称'})
            date_cols = sorted([c for c in df.columns if is_date_col(c)])
            if len(date_cols) < 10: continue
            all_dates.update(date_cols)
            count = 0
            for _, row in df.iterrows():
                code = row.get('产品代码')
                name = row.get('产品名称', '')
                if pd.isna(code): continue
                code = str(code).strip()
                name = str(name).strip() if pd.notna(name) else ''
                _, rd = 判断产品流动性(name)
                if rd > 最长赎回天数: continue
                nps = []
                for d in date_cols:
                    try:
                        v = float(row[d])
                        if not np.isnan(v) and v > 0: nps.append((d, v))
                    except (ValueError, TypeError): pass
                if len(nps) < 10: continue
                key = (sheet, code)
                p = ProductData(sheet, code, name, '', rd)
                p.nav_dates = [x[0] for x in nps]
                p.navs = {x[0]: x[1] for x in nps}
                for i in range(1, len(nps)):
                    d0, n0 = nps[i-1]; d1, n1 = nps[i]
                    gap = (pd.Timestamp(d1) - pd.Timestamp(d0)).days
                    if 0 < gap <= 60 and n0 > 0:
                        p.rets[d1] = (n1/n0 - 1) / gap * 365 * 100
                p.ret_dates = sorted(p.rets.keys())
                p.ret_date_idx = {d: i for i, d in enumerate(p.ret_dates)}
                if len(p.ret_dates) < 10: continue
                self._precompute_signals(p)
                self.products[key] = p
                count += 1
            total += count
            logger.info(f"    有效产品: {count}")
        self.all_sim_dates = sorted(d for d in all_dates if 回测开始日 <= d < 回测结束日)
        self.date_to_idx = {d: i for i, d in enumerate(self.all_sim_dates)}
        logger.info(f"加载完成: {total}个产品, {len(self.all_sim_dates)}个回测日")

    def _precompute_signals(self, product):
        rd = product.ret_dates; rets = product.rets; n = len(rd)
        sigs = []
        for i in range(1, n - 持有天数_评估):
            if rets[rd[i]] > 信号阈值_库 and rets[rd[i-1]] <= 信号阈值_库:
                hr = [rets[rd[i+1+k]] for k in range(持有天数_评估)]
                avg = float(np.mean(hr)); ee = rd[i + 持有天数_评估]
                ps = 0
                for j in range(i, min(i+20, n)):
                    if rets[rd[j]] > 信号阈值_库 * 0.8: ps += 1
                    else: break
                cd = ((pd.Timestamp(rd[i+ps-1]) - pd.Timestamp(rd[i])).days + 1) if ps > 0 else 0
                sigs.append({'date': rd[i], 'eval_end': ee, 'avg_return': avg,
                             'success': avg > 成功标准, 'persist_days': cd})
        product.signals = sigs

    def build_library(self, as_of_date):
        lib = {}
        for key, p in self.products.items():
            valid = [s for s in p.signals if s['eval_end'] < as_of_date]
            if len(valid) < 最少历史信号: continue
            sn = sum(1 for s in valid if s['success'])
            rate = sn / len(valid) * 100
            if rate < 最低成功率: continue
            ap = float(np.mean([s['persist_days'] for s in valid]))
            if ap < 最短持续天数 or ap > 最大持有天数_评估: continue
            rl = [s['avg_return'] for s in valid]
            ar = float(np.mean(rl))
            sd = float(np.std(rl)) if len(rl) > 1 else 1.0
            sh = ar / max(sd, 0.5)       # V5.4: Sharpe floor = 0.5
            lib[key] = {'success_rate': round(rate,1), 'signal_count': len(valid),
                        'avg_return': round(ar,2), 'avg_persist': round(ap,1),
                        'sharpe_like': round(sh,2)}
        return lib

    def get_latest_nav(self, product, as_of_date):
        idx = bisect.bisect_right(product.nav_dates, as_of_date) - 1
        if idx >= 0:
            d = product.nav_dates[idx]
            return d, product.navs[d]
        return None, None

    def _td_held(self, cd, dd):
        ci = self.date_to_idx.get(cd); di = self.date_to_idx.get(dd)
        if ci is not None and di is not None: return di - ci
        return (pd.Timestamp(dd) - pd.Timestamp(cd)).days

    def _is_pre_holiday(self, di):
        """基于中国假期日历判断是否为节前交易日"""
        if di+1 >= len(self.all_sim_dates): return True
        return self.all_sim_dates[di] in PRE_HOLIDAY_DATES

    def _is_holiday(self, date):
        """判断是否为假期日 (国庆、元旦等法定假期)"""
        return date in HOLIDAY_DATE_SET

    # ==================== 单模式回测 ====================

    def run_single_mode(self, mode, dl_alpha=None):
        desc = {'A': '纯反应式基线', 'B': '优选(Alpha加成)',
                'C': '优选+智持+节前', 'D': '优选+DL混合'}
        logger.info(f"\n{'='*60}\n  Mode {mode} — {desc.get(mode, mode)}\n{'='*60}")

        # --- 功能开关 ---
        use_alpha_boost = mode in ('B', 'C', 'D')     # 建议① (D也用)
        use_sell_delay  = mode == 'C'             # 建议②
        use_preholiday  = mode == 'C'             # 建议③
        use_holiday_protection = mode == 'C'     # 假期卖出保护 (配合③)
        use_dl_blend = mode == 'D' and dl_alpha is not None  # [DL BLEND]

        # --- 状态 ---
        library = {}; library_date = None
        positions = {}; pending_buys = []; pending_sells = []
        pending_buy_keys = set(); pending_sell_keys = set()
        closed_trades = []; cash = 初始资金
        daily_values = []; trade_log = []

        learner = PatternLearner(); predictor = ForwardPredictor()
        patterns = {}; pattern_build_date = None
        prediction_log = []

        # 诊断计数器
        n_alpha_boost = 0        # 建议① 被加成的买入次数
        n_sell_delayed = 0       # 建议② 延迟卖出次数
        n_delay_success = 0      # 建议② 延迟后确实释放的次数
        n_preholiday_buy = 0     # 建议③ 节前买入次数
        n_preholiday_profit = 0  # 建议③ 节前买入盈利次数

        def _rebuild_lib(date):
            nonlocal library, library_date
            library = self.build_library(date); library_date = date

        def _rebuild_pat(date):
            nonlocal patterns, pattern_build_date
            if mode == 'A': return
            new = {}
            for key, p in self.products.items():       # 扫描全部产品(非仅产品库)
                pat = learner.learn(key, p.ret_dates, p.rets, date, 规律学习窗口天数)
                if pat and pat.confidence >= 最低置信度:
                    new[key] = pat
            patterns = new; pattern_build_date = date
            logger.info(f"  [{date}] Mode {mode}: 规律 {len(patterns)}个产品 (全库扫描)")

            # [DL BLEND] 与规律学习同步重训DL
            if use_dl_blend and dl_alpha is not None and dl_alpha.should_retrain(date):
                dl_alpha.train_and_predict(date)

        def _proc_buys(date):
            nonlocal cash
            nb = 0
            for key, sig_d, reason, sig_type, is_ph in pending_buys:
                if key in positions: pending_buy_keys.discard(key); continue
                prod = self.products[key]
                nav = prod.navs.get(date)
                if nav is None: _, nav = self.get_latest_nav(prod, date)
                if nav is None: pending_buy_keys.discard(key); continue
                # V5.4 动态仓位: 按当前组合价值计算
                current_value = cash + sum(
                    p.amount * (self.products[k].navs.get(date, p.buy_nav) / p.buy_nav)
                    for k, p in positions.items()
                )
                dynamic_size = current_value / 最大持仓数
                amt = min(dynamic_size, cash)
                if amt < 单仓基础 * 0.5: pending_buy_keys.discard(key); continue
                er = prod.rets.get(sig_d, 信号阈值_买入)
                cash -= amt
                positions[key] = Position(key, sig_d, date, nav, amt,
                    entry_return=er, alpha_score=0.5,
                    signal_type=sig_type, is_preholiday_buy=is_ph)
                pending_buy_keys.discard(key); nb += 1
                trade_log.append({'date': date, 'action': '买入确认',
                    'bank': prod.bank, 'code': prod.code, 'name': prod.name,
                    'nav': nav, 'amount': amt, 'pnl': 0, 'hold_days': 0,
                    'reason': f'T+1确认({sig_d}) {reason}',
                    'signal_type': sig_type})
            pending_buys.clear()
            return nb

        def _proc_sells(date):
            nonlocal cash
            ns = 0
            for key, sig_d, reason in pending_sells:
                if key not in positions: pending_sell_keys.discard(key); continue
                pos = positions.pop(key); prod = self.products[key]
                nav = prod.navs.get(date)
                if nav is None: _, nav = self.get_latest_nav(prod, date)
                if nav is None: nav = pos.buy_nav
                pos.sell_date = date; pos.sell_nav = nav
                pos.pnl = pos.amount * (nav / pos.buy_nav - 1)
                pos.sell_reason = reason
                cash += pos.amount + pos.pnl
                closed_trades.append(pos); pending_sell_keys.discard(key); ns += 1
                htd = self._td_held(pos.confirm_date, date)
                trade_log.append({'date': date, 'action': '赎回到账',
                    'bank': prod.bank, 'code': prod.code, 'name': prod.name,
                    'nav': nav, 'amount': pos.amount + pos.pnl, 'pnl': pos.pnl,
                    'hold_days': htd, 'reason': f'T+1到账({sig_d}) {reason}',
                    'signal_type': pos.signal_type})
            pending_sells.clear()
            return ns

        def _check_sells(date):
            nonlocal n_sell_delayed, n_delay_success
            is_hol = use_holiday_protection and self._is_holiday(date)
            for key in list(positions):
                if key in pending_sell_keys: continue
                pos = positions[key]; prod = self.products[key]
                htd = self._td_held(pos.confirm_date, date)
                if htd < 最短持有交易日_赎回: continue
                ret = prod.rets.get(date)

                # ① 硬阈值
                if ret is not None and ret <= 信号阈值_卖出:
                    # === 假期保护 (Mode C): 假期期间0%收益不是真实卖出信号 ===
                    if is_hol:
                        continue
                    # === 建议② 智持: 预测即将释放 → 延迟卖出 ===
                    if (use_sell_delay and key in patterns
                        and ret > 延迟卖出_最低收益
                        and htd < 最大持有交易日 - 延迟卖出_最大持有余量):
                        upcoming, pred = predictor.has_upcoming_release(
                            patterns[key], date, self.all_sim_dates, self.date_to_idx,
                            max_td=延迟卖出_预测窗口, min_conf=延迟卖出_置信度)
                        if upcoming:
                            n_sell_delayed += 1
                            prediction_log.append({
                                'date': date, 'product_key': key,
                                'predicted_release': pred.predicted_date if pred else '',
                                'predicted_release_end': pred.predicted_end_date if pred else '',
                                'confidence': pred.confidence if pred else 0,
                                'source': 'sell_delay', 'action': 'delay_sell',
                                'ret_at_event': ret})
                            continue  # 延迟卖出, 不挂卖单
                    pending_sells.append((key, date,
                        f'收益{ret:.1f}%≤{信号阈值_卖出}%(持有{htd}td)'))
                    pending_sell_keys.add(key); continue

                # ② 最大持有
                if htd >= 最大持有交易日:
                    pending_sells.append((key, date, f'持有{htd}td达上限'))
                    pending_sell_keys.add(key)

        def _find_buys(date, slots, di):
            """通用买入扫描, 根据 mode 开关执行不同逻辑
            V6改进: 节前买入基于产品库质量而非预测, 确保每个假期都有节前买入
            核心逻辑: 假期期间不公布净值, 节后首日公布累积收益,
            因此节前持有优质产品 = 免费赚取假期累积收益"""
            nonlocal n_alpha_boost, n_preholiday_buy
            is_ph = self._is_pre_holiday(di)

            if slots <= 0: return 0

            # === 建议① 优选: Alpha 加成 ===
            predicted_scores = {}
            if use_alpha_boost:
                rankings = predictor.rank_products(patterns, date,
                    self.all_sim_dates, self.date_to_idx)
                predicted_scores = {r[0]: (r[1], r[2]) for r in rankings}

            # === [DL BLEND V3.2] 乘法加成 (仅Mode D, 仅已有pattern的产品) ===
            n_dl_boosted = 0
            if use_dl_blend and dl_alpha is not None and dl_alpha.is_reliable():
                dl_signals = dl_alpha.alpha_signals
                for key in list(predicted_scores.keys()):  # 只遍历已有pattern预测的
                    dl_sig = dl_signals.get(key, {})
                    dl_entry = dl_sig.get('entry_quality', 0)
                    dl_conf = dl_sig.get('confidence', 0)

                    if dl_conf < DL_MIN_CONFIDENCE or dl_entry < DL_MIN_ENTRY_QUALITY:
                        continue

                    # 乘法加成: 用DL entry_quality作为boost因子
                    dl_boost = min(dl_entry * DL_MULT_WEIGHT, DL_MAX_BOOST)
                    old_score = predicted_scores[key][0]
                    boosted = old_score * (1 + dl_boost)
                    predicted_scores[key] = (boosted, predicted_scores[key][1])
                    n_dl_boosted += 1
                # 不创建DL-only预测 (V3.1的fake_pred已删除)

            ph_mode = is_ph and use_preholiday
            holiday_name = PRE_HOLIDAY_DATES.get(self.all_sim_dates[di], '') if ph_mode else ''

            # ============================================
            # 建议③ 节前仓位预留: 基于库质量 (不要求预测)
            # 假期期间产品继续累积收益, 任何优质库产品都值得持有
            # ============================================
            ph_candidates = []
            ph_reserved = 0
            if ph_mode:
                for key, info in library.items():
                    if key in positions or key in pending_buy_keys or key in pending_sell_keys:
                        continue
                    prod = self.products[key]
                    r = prod.rets.get(date)
                    if r is None or r <= 节前买入_实时阈值: continue
                    sr = info.get('success_rate', 0)
                    avg_ret = info.get('avg_return', 0)
                    if sr < 节前买入_最低成功率 or avg_ret < 节前买入_最低均收: continue
                    # 排序: 当日收益 × 库均收, 有预测加成则更优
                    composite = r * avg_ret
                    has_pred = key in predicted_scores
                    if has_pred:
                        composite *= (1 + predicted_scores[key][0] * ALPHA_预测加成)
                    ph_candidates.append((key, composite, r, sr, avg_ret, has_pred))
                # 按 composite 排序
                ph_candidates.sort(key=lambda x: (-x[1], -x[3]))
                ph_reserved = min(len(ph_candidates), 节前保留仓位, slots)

            # ============================================
            # 正常买入 (使用 slots - ph_reserved 个仓位)
            # ============================================
            normal_slots = slots - ph_reserved

            signals = []
            for key, info in library.items():
                if key in positions or key in pending_buy_keys or key in pending_sell_keys:
                    continue
                prod = self.products[key]
                r = prod.rets.get(date)
                if r is None or r <= 信号阈值_买入: continue  # 正常阈值 3.5%
                avg_ret = info.get('avg_return', 1.0)
                composite = r * avg_ret           # V5.4 排名: 当日收益 × 历史平均收益
                boosted = False
                if key in predicted_scores:
                    composite *= (1 + predicted_scores[key][0] * ALPHA_预测加成)
                    boosted = True
                signals.append((key, composite, r, info['success_rate'], boosted,
                                predicted_scores.get(key), avg_ret))

            signals.sort(key=lambda x: (-x[1], -x[3]))

            queued = 0
            normal_bought = set()
            for key, composite, ret, sr, boosted, pred_info, avg_ret in signals[:normal_slots]:
                tag = '★' if boosted else ' '
                pred_str = ''
                if pred_info:
                    pred_str = (f' pred={pred_info[1].predicted_date}'
                                f'~{pred_info[1].predicted_end_date} conf={pred_info[0]:.2f}')
                reason = f'{tag}成功率{sr}% 收益{ret:.1f}% 均收{avg_ret:.1f}%{pred_str}'
                sig_type = 'predicted' if boosted else 'reactive'
                pending_buys.append((key, date, reason, sig_type, False))
                pending_buy_keys.add(key); queued += 1
                normal_bought.add(key)
                if boosted: n_alpha_boost += 1
                if pred_info:
                    prediction_log.append({
                        'date': date, 'product_key': key,
                        'predicted_release': pred_info[1].predicted_date,
                        'predicted_release_end': pred_info[1].predicted_end_date,
                        'confidence': pred_info[0],
                        'source': 'alpha_boost', 'action': 'buy_boosted',
                        'ret_at_event': ret})

            # ============================================
            # 节前预留仓位填充 (预留 + 剩余未用的正常仓位)
            # ============================================
            if ph_mode and ph_candidates:
                leftover = normal_slots - min(len(signals), normal_slots)
                ph_fill_slots = ph_reserved + max(leftover, 0)
                filled = 0
                for key, composite, ret, sr, avg_ret, has_pred in ph_candidates:
                    if filled >= ph_fill_slots: break
                    if key in normal_bought or key in pending_buy_keys: continue
                    pred_str = ''
                    if has_pred:
                        pred_info = predicted_scores[key]
                        pred_str = (f' pred={pred_info[1].predicted_date}'
                                    f'~{pred_info[1].predicted_end_date} conf={pred_info[0]:.2f}')
                    reason = (f'[节前{holiday_name}] 成功率{sr}% 收益{ret:.1f}% '
                              f'均收{avg_ret:.1f}%{pred_str}')
                    pending_buys.append((key, date, reason, 'preholiday', True))
                    pending_buy_keys.add(key); queued += 1; filled += 1
                    n_preholiday_buy += 1
                    prediction_log.append({
                        'date': date, 'product_key': key,
                        'predicted_release': predicted_scores[key][1].predicted_date if has_pred else '',
                        'predicted_release_end': predicted_scores[key][1].predicted_end_date if has_pred else '',
                        'confidence': predicted_scores[key][0] if has_pred else 0,
                        'source': 'preholiday', 'action': 'buy_preholiday',
                        'ret_at_event': ret})

            return queued

        def _record(date):
            pv = 0
            for key, pos in positions.items():
                prod = self.products[key]
                _, nav = self.get_latest_nav(prod, date)
                pv += pos.amount * (nav / pos.buy_nav) if nav else pos.amount
            tot = cash + pv
            if daily_values and daily_values[-1][0] == date:
                daily_values[-1] = (date, tot, cash, pv)
            else:
                daily_values.append((date, tot, cash, pv))

        # =================== 主循环 ===================
        _rebuild_lib(回测开始日)
        logger.info(f"  Mode {mode} 产品库: {len(library)}个")
        if mode != 'A': _rebuild_pat(回测开始日)

        nd = len(self.all_sim_dates); bn = sn = 0
        for di, date in enumerate(self.all_sim_dates):
            if library_date is None or (pd.Timestamp(date) - pd.Timestamp(library_date)).days >= 产品库重建间隔:
                _rebuild_lib(date)
            if mode != 'A' and (pattern_build_date is None or
                    (pd.Timestamp(date) - pd.Timestamp(pattern_build_date)).days >= 规律重建间隔天数):
                _rebuild_pat(date)

            bn += _proc_buys(date)
            sn += _proc_sells(date)
            _check_sells(date)

            used = len(positions) + len(pending_buys)
            _find_buys(date, 最大持仓数 - used, di)

            # V5.4 满仓轮动: 替换最弱持仓
            used2 = len(positions) + len(pending_buys)
            if used2 >= 最大持仓数:
                is_ph_today = use_preholiday and self._is_pre_holiday(di)
                sellable = []
                for rk in positions:
                    if rk in pending_sell_keys: continue
                    rpos = positions[rk]
                    rhtd = self._td_held(rpos.confirm_date, date)
                    if rhtd < 最短持有交易日_赎回: continue
                    rr = self.products[rk].rets.get(date)
                    if rr is not None: sellable.append((rk, rr, rhtd))
                if sellable:
                    sellable.sort(key=lambda x: x[1])
                    wk, wr, wh = sellable[0]

                    if is_ph_today:
                        # === 节前轮动: 替换多个弱持仓, 确保假期前组合最优 ===
                        # 假期不公布净值, 节后首日公布累积收益, 必须持有优质产品
                        # 允许替换最多 节前保留仓位 个弱位置
                        holiday_name = PRE_HOLIDAY_DATES.get(date, '')
                        ph_rotated = 0
                        used_cands = set()
                        for s_key, s_ret, s_htd in sellable:
                            if ph_rotated >= 节前保留仓位: break
                            if s_ret >= 信号阈值_买入: break  # 后续更强, 无需替换
                            if s_key in pending_sell_keys: continue
                            # 找最优替换
                            best_cand = None; best_r = s_ret
                            for ck, cinfo in library.items():
                                if (ck in positions or ck in pending_buy_keys
                                    or ck in pending_sell_keys or ck in used_cands):
                                    continue
                                cp = self.products[ck]
                                cr = cp.rets.get(date)
                                if cr is None or cr <= 信号阈值_买入: continue
                                if cr <= best_r + 1.0: continue
                                c_sr = cinfo.get('success_rate', 0)
                                c_avg = cinfo.get('avg_return', 0)
                                if c_sr < 节前买入_最低成功率 or c_avg < 节前买入_最低均收:
                                    continue
                                best_cand = (ck, cr, c_sr, c_avg); best_r = cr
                            if best_cand:
                                ck, cr, c_sr, c_avg = best_cand
                                pending_sells.append((s_key, date,
                                    f'[节前{holiday_name}轮动]卖出(收益{s_ret:.1f}%'
                                    f'→换{cr:.1f}%持有{s_htd}td)'))
                                pending_sell_keys.add(s_key)
                                reason = (f'[节前{holiday_name}轮动买入] 成功率{c_sr}% '
                                          f'收益{cr:.1f}% 均收{c_avg:.1f}%')
                                pending_buys.append((ck, date, reason, 'preholiday', True))
                                pending_buy_keys.add(ck)
                                used_cands.add(ck)
                                n_preholiday_buy += 1; ph_rotated += 1
                                prediction_log.append({
                                    'date': date, 'product_key': ck,
                                    'predicted_release': '', 'predicted_release_end': '',
                                    'confidence': 0,
                                    'source': 'preholiday_rotation',
                                    'action': 'buy_preholiday',
                                    'ret_at_event': cr})
                    elif not is_ph_today:
                        # === 常规轮动 (V5.4逻辑) ===
                        if wr < 信号阈值_买入:
                            best_cand = None; best_r = wr
                            for ck, cinfo in library.items():
                                if ck in positions or ck in pending_buy_keys or ck in pending_sell_keys:
                                    continue
                                cp = self.products[ck]
                                cr = cp.rets.get(date)
                                if cr is not None and cr > 信号阈值_买入 and cr > best_r + 1.0:
                                    best_cand = (ck, cr, cinfo['success_rate']); best_r = cr
                            if best_cand:
                                pending_sells.append((wk, date,
                                    f'轮动卖出(收益{wr:.1f}%→换{best_r:.1f}%持有{wh}td)'))
                                pending_sell_keys.add(wk)

            _record(date)

            if (di+1) % 50 == 0 or di == nd - 1:
                v = daily_values[-1][1]; r = (v/初始资金-1)*100
                logger.info(f"  Mode {mode} [{date}] {di+1}/{nd} | "
                            f"持仓{len(positions)} | 买{bn} 卖{sn} | "
                            f"{v/1e4:,.0f}万({r:+.2f}%)")

        # 平仓
        last = self.all_sim_dates[-1]
        for key in list(positions):
            pos = positions.pop(key); prod = self.products[key]
            _, nav = self.get_latest_nav(prod, last)
            if nav is None: nav = pos.buy_nav
            pos.sell_date = last; pos.sell_nav = nav
            pos.pnl = pos.amount * (nav/pos.buy_nav - 1)
            pos.sell_reason = '回测结束平仓'
            cash += pos.amount + pos.pnl; closed_trades.append(pos)
            htd = self._td_held(pos.confirm_date, last)
            trade_log.append({'date': last, 'action': '强制平仓',
                'bank': prod.bank, 'code': prod.code, 'name': prod.name,
                'nav': nav, 'amount': pos.amount+pos.pnl, 'pnl': pos.pnl,
                'hold_days': htd, 'reason': '回测结束平仓',
                'signal_type': pos.signal_type})
        _record(last)

        # 计算延迟卖出成功率
        for rec in prediction_log:
            if rec.get('action') != 'delay_sell': continue
            key = rec['product_key']; pd_str = rec.get('predicted_release', '')
            if key in self.products and pd_str:
                prod = self.products[key]; pts = pd.Timestamp(pd_str)
                hit = False
                for off in range(-2, 4):
                    cd = (pts + pd.Timedelta(days=off)).strftime('%Y-%m-%d')
                    r = prod.rets.get(cd)
                    if r is not None and r > 释放识别阈值: hit = True; break
                rec['hit'] = hit
                if hit: n_delay_success += 1
            else:
                rec['hit'] = False

        # 节前买入盈亏
        for t in closed_trades:
            if t.is_preholiday_buy and t.pnl > 0:
                n_preholiday_profit += 1

        # 预测命中率
        for rec in prediction_log:
            if rec.get('action') in ('delay_sell',): continue
            key = rec['product_key']; pd_str = rec.get('predicted_release', '')
            if key in self.products and pd_str:
                prod = self.products[key]; pts = pd.Timestamp(pd_str)
                hit = False
                for off in range(-2, 4):
                    cd = (pts + pd.Timedelta(days=off)).strftime('%Y-%m-%d')
                    r = prod.rets.get(cd)
                    if r is not None and r > 释放识别阈值: hit = True; break
                rec['hit'] = hit
            else:
                rec['hit'] = False

        # 汇总
        result = self._metrics(mode, daily_values, closed_trades, trade_log, prediction_log)
        result['daily_values'] = daily_values
        result['closed_trades'] = closed_trades
        result['trade_log'] = trade_log
        result['prediction_log'] = prediction_log
        result['n_alpha_boost'] = n_alpha_boost
        result['n_sell_delayed'] = n_sell_delayed
        result['n_delay_success'] = n_delay_success
        result['n_preholiday_buy'] = n_preholiday_buy
        result['n_preholiday_profit'] = n_preholiday_profit
        result['patterns_snapshot'] = {
            k: {'confidence': v.confidence, 'period_days': v.period_days,
                'period_cv': v.period_cv, 'has_period': v.has_period,
                'top_phase': v.top_phase, 'phase_pvalue': v.phase_pvalue,
                'top_weekday': v.top_weekday, 'weekday_pvalue': v.weekday_pvalue,
                'n_events': v.n_events, 'avg_window_days': v.avg_window_days}
            for k, v in patterns.items()
        }
        result['raw_patterns'] = dict(patterns)   # 保留原始对象供释放规律格式化
        return result

    def _metrics(self, mode, dv, ct, tl, pl):
        if not dv:
            return {k: 0 for k in ['mode','ann_ret','max_dd','sharpe','n_trades',
                                    'win_rate','avg_hold','total_ret','final_value',
                                    'forward_hit_rate','pred_trades','pred_pnl',
                                    'std_trades','std_pnl']}
        fv = dv[-1][1]
        tr = (fv/初始资金 - 1)*100
        t0 = pd.Timestamp(self.all_sim_dates[0])
        t1 = pd.Timestamp(self.all_sim_dates[-1])
        yrs = (t1-t0).days / 365
        ar = ((fv/初始资金)**(1/yrs)-1)*100 if yrs > 0 else 0
        pk = 初始资金; md = 0
        for _, v, _, _ in dv:
            if v > pk: pk = v
            dd = (pk-v)/pk*100
            if dd > md: md = dd
        nt = len(ct); nw = sum(1 for t in ct if t.pnl > 0)
        wr = nw/nt*100 if nt else 0
        hds = [self._td_held(t.confirm_date, t.sell_date) for t in ct if t.sell_date]
        ah = float(np.mean(hds)) if hds else 0
        drs = []
        for i in range(1, len(dv)):
            pv = dv[i-1][1]; cv = dv[i][1]
            if pv > 0: drs.append(cv/pv - 1)
        if drs:
            rf = 0.025/252; ex = np.array(drs) - rf
            sp = float(np.mean(ex)/np.std(ex)*np.sqrt(252)) if np.std(ex) > 0 else 0
        else: sp = 0

        buy_preds = [p for p in pl if p.get('action','').startswith('buy')]
        hits = sum(1 for p in buy_preds if p.get('hit', False))
        hr = hits / len(buy_preds) * 100 if buy_preds else 0

        pred_t = [t for t in ct if t.signal_type in ('predicted','preholiday')]
        std_t = [t for t in ct if t.signal_type == 'reactive']

        return {
            'mode': mode, 'final_value': fv, 'total_ret': tr, 'ann_ret': ar,
            'max_dd': md, 'sharpe': sp, 'n_trades': nt, 'win_rate': wr,
            'avg_hold': ah, 'forward_hit_rate': hr,
            'n_predictions': len(buy_preds), 'n_hits': hits,
            'pred_trades': len(pred_t),
            'pred_pnl': sum(t.pnl for t in pred_t),
            'pred_wr': (sum(1 for t in pred_t if t.pnl>0)/len(pred_t)*100) if pred_t else 0,
            'std_trades': len(std_t),
            'std_pnl': sum(t.pnl for t in std_t),
            'std_wr': (sum(1 for t in std_t if t.pnl>0)/len(std_t)*100) if std_t else 0,
        }


# ============================================================
# 模块4: ComparisonReport
# ============================================================

class ComparisonReport:
    def __init__(self, engine, ra, rb, rc, sweep_results=None, base_ann=None):
        self.engine = engine
        self.R = {'A': ra, 'B': rb, 'C': rc}
        self.sweep_results = sweep_results
        self.base_ann = base_ann

    def print_console(self):
        sep = '=' * 74
        print(f"\n{sep}")
        print("     前瞻策略回测对比 V3 (优选 · 智持 · 节前)")
        print(sep)
        print(f"  区间: {self.engine.all_sim_dates[0]} ~ {self.engine.all_sim_dates[-1]}  "
              f"({len(self.engine.all_sim_dates)}交易日)")
        print(f"  资金: {初始资金/1e4:,.0f}万  仓位: {最大持仓数}×{单仓基础/1e4:,.0f}万")
        print(f"  Mode A: 纯反应 (阈值{信号阈值_买入}%)")
        print(f"  Mode B: +优选 (Alpha加成{ALPHA_预测加成}, 阈值不变)")
        print(f"  Mode C: +优选+智持+节前 (全部三建议)")
        n_pat = len(self.R['C'].get('patterns_snapshot', {}))
        n_lib = self.R['C'].get('n_trades', 0)
        print(f"  规律学习: {n_pat}个产品有释放规律 (全库扫描, 置信度≥{最低置信度})")
        print()

        hdr = f"  {'指标':<20} {'A(反应)':<14} {'B(+优选)':<14} {'C(+全部)':<14}"
        print(hdr); print("  " + "-" * 58)

        rows = [
            ('年化收益率', 'ann_ret', '{:+.4f}%'),
            ('总收益率', 'total_ret', '{:+.4f}%'),
            ('最大回撤', 'max_dd', '{:.4f}%'),
            ('夏普比率', 'sharpe', '{:.4f}'),
            ('交易次数', 'n_trades', '{:d}'),
            ('胜率', 'win_rate', '{:.1f}%'),
            ('平均持有天数', 'avg_hold', '{:.1f}'),
            ('预测命中率', 'forward_hit_rate', '{:.1f}%'),
        ]
        for label, key, fmt in rows:
            vals = []
            for m in 'ABC':
                v = self.R[m].get(key, 0)
                if key == 'forward_hit_rate' and m == 'A':
                    vals.append('N/A')
                else:
                    try: vals.append(fmt.format(int(v) if 'd' in fmt else v))
                    except: vals.append(str(v))
            print(f"  {label:<20} {vals[0]:<14} {vals[1]:<14} {vals[2]:<14}")

        print()
        # 建议① 诊断
        print(f"  [建议① 优选]")
        for m in 'BC':
            r = self.R[m]
            print(f"    Mode {m}: Alpha加成买入 {r.get('n_alpha_boost',0)}次, "
                  f"预测入场{r.get('pred_trades',0)}笔 "
                  f"胜率{r.get('pred_wr',0):.1f}% "
                  f"盈亏{r.get('pred_pnl',0)/1e4:+,.2f}万")

        # 建议② 诊断
        print(f"  [建议② 智持]")
        rc = self.R['C']
        nd = rc.get('n_sell_delayed', 0)
        ns = rc.get('n_delay_success', 0)
        print(f"    延迟卖出 {nd}次, 其中释放成功 {ns}次 "
              f"({ns/nd*100:.0f}%命中)" if nd > 0 else f"    延迟卖出 0次")

        # 建议③ 诊断
        print(f"  [建议③ 节前]")
        np_ = rc.get('n_preholiday_buy', 0)
        npp = rc.get('n_preholiday_profit', 0)
        print(f"    节前买入 {np_}次, 盈利 {npp}次"
              + (f" ({npp/np_*100:.0f}%)" if np_ > 0 else ""))

        print()
        for m in 'BC':
            d = self.R[m]['ann_ret'] - self.R['A']['ann_ret']
            print(f"  Mode {m} vs A: 年化 {d:+.4f}%")

        best = max('ABC', key=lambda m: self.R[m]['ann_ret'])
        print(f"\n  最优: Mode {best} (年化 {self.R[best]['ann_ret']:+.4f}%)")
        if best != 'A':
            print(f"  建议: 有效! 超越基线 "
                  f"{self.R[best]['ann_ret']-self.R['A']['ann_ret']:+.4f}%")
        else:
            print(f"  建议: 预测辅助策略未超越基线, 保留纯反应式")
        print(sep)

    def generate_excel(self):
        logger.info("生成Excel报告...")
        wb = Workbook()
        hf = Font(bold=True, color="FFFFFF", size=11)
        tf = Font(bold=True, size=14)
        bf = PatternFill("solid", fgColor="1565C0")
        gf = PatternFill("solid", fgColor="2E7D32")
        df_ = PatternFill("solid", fgColor="37474F")
        rf_ = Font(color="CC0000", bold=True)
        gnf = Font(color="2E7D32", bold=True)
        ct = Alignment(horizontal="center", vertical="center")

        # Sheet 1: 总览
        ws1 = wb.active; ws1.title = "对比总览"
        ws1['A1'] = "前瞻策略回测 V3 (优选·智持·节前)"; ws1['A1'].font = tf
        ws1.append([])
        ws1.append(['指标', 'Mode A(纯反应)', 'Mode B(+优选)', 'Mode C(+全部)'])
        for c in ws1[3]: c.font = hf; c.fill = bf; c.alignment = ct

        kv = [('年化收益率(%)','ann_ret'),('总收益率(%)','total_ret'),
              ('最大回撤(%)','max_dd'),('夏普比率','sharpe'),
              ('交易次数','n_trades'),('胜率(%)','win_rate'),
              ('平均持有天数','avg_hold'),('预测命中率(%)','forward_hit_rate'),
              ('Alpha加成买入次数','n_alpha_boost'),
              ('延迟卖出次数','n_sell_delayed'),('延迟成功次数','n_delay_success'),
              ('节前买入次数','n_preholiday_buy'),('节前盈利次数','n_preholiday_profit'),
              ('预测入场笔数','pred_trades'),('预测入场盈亏(万)','pred_pnl'),
              ('标准入场笔数','std_trades'),('标准入场盈亏(万)','std_pnl')]
        for label, key in kv:
            row = [label]
            for m in 'ABC':
                v = self.R[m].get(key, 0)
                if key in ('forward_hit_rate','n_alpha_boost','n_sell_delayed',
                           'n_delay_success','n_preholiday_buy','n_preholiday_profit',
                           'pred_trades','pred_pnl','std_trades','std_pnl') and m == 'A':
                    row.append('N/A')
                elif key in ('pred_pnl','std_pnl'):
                    row.append(round(v/1e4, 2) if isinstance(v,(int,float)) else v)
                else:
                    row.append(round(v,4) if isinstance(v,float) else v)
            ws1.append(row)
        ws1.append([])
        for m in 'BC':
            d = self.R[m]['ann_ret'] - self.R['A']['ann_ret']
            ws1.append([f'Mode {m} vs A', f'{d:+.4f}%'])
        ws1.column_dimensions['A'].width = 22
        for cl in 'BCD': ws1.column_dimensions[cl].width = 20

        # Sheet 2: 每日净值
        ws2 = wb.create_sheet("每日净值")
        ws2.append(['日期','A(万)','B(万)','C(万)','A%','B%','C%'])
        for c in ws2[1]: c.font = hf; c.fill = df_; c.alignment = ct
        vms = {}
        for m in 'ABC':
            vm = {}
            for d, v, _, _ in self.R[m]['daily_values']: vm[d] = v
            vms[m] = vm
        ads = sorted(set().union(*(vm.keys() for vm in vms.values())))
        for d in ads:
            va=vms['A'].get(d,初始资金); vb=vms['B'].get(d,初始资金); vc=vms['C'].get(d,初始资金)
            ws2.append([d,round(va/1e4,2),round(vb/1e4,2),round(vc/1e4,2),
                        round((va/初始资金-1)*100,4),round((vb/初始资金-1)*100,4),
                        round((vc/初始资金-1)*100,4)])
        for i,w in enumerate([12,14,14,14,12,12,12],1):
            ws2.column_dimensions[get_column_letter(i)].width = w
        ws2.freeze_panes = 'A2'
        if len(ads) > 5:
            ch = LineChart(); ch.title = "三模式净值对比"; ch.style = 10
            ch.y_axis.title = "万"; ch.width = 30; ch.height = 15
            for ci in range(2,5):
                ch.add_data(Reference(ws2,min_col=ci,min_row=1,max_row=len(ads)+1),titles_from_data=True)
            ch.set_categories(Reference(ws2,min_col=1,min_row=2,max_row=len(ads)+1))
            colors = ['1565C0','FF6F00','2E7D32']
            for idx,s in enumerate(ch.series):
                s.graphicalProperties.line.width = 20000
                if idx < len(colors): s.graphicalProperties.line.solidFill = colors[idx]
            ws2.add_chart(ch, "I2")

        # Sheet 3: Mode C 交易明细
        ws3 = wb.create_sheet("Mode C交易明细")
        h3 = ['日期','操作','银行','代码','名称','信号类型','净值','金额(万)','盈亏(万)','持有天数','原因']
        ws3.append(h3)
        for c in ws3[1]: c.font = hf; c.fill = bf; c.alignment = ct
        for t in self.R['C']['trade_log']:
            ws3.append([t['date'],t['action'],t['bank'],t['code'],t['name'][:25],
                t.get('signal_type',''),
                round(t['nav'],6) if t.get('nav') else '',
                round(t['amount']/1e4,2) if t.get('amount') else '',
                round(t['pnl']/1e4,2) if t.get('pnl') is not None else '',
                t.get('hold_days',''),t.get('reason','')])
            rn = ws3.max_row; pc = ws3.cell(row=rn,column=9)
            if isinstance(pc.value,(int,float)):
                pc.font = gnf if pc.value > 0 else (rf_ if pc.value < 0 else None)
        for i,w in enumerate([12,10,10,16,25,14,12,12,12,8,45],1):
            ws3.column_dimensions[get_column_letter(i)].width = w
        ws3.freeze_panes = 'A2'

        # Sheet 4: 预测记录
        ws4 = wb.create_sheet("前瞻预测记录")
        h4 = ['日期','银行','代码','预测释放日','预测释放结束日','置信度','来源','命中','买入时收益%','操作']
        ws4.append(h4)
        for c in ws4[1]: c.font = hf; c.fill = gf; c.alignment = ct
        aps = []
        for m in 'BC':
            for p in self.R[m].get('prediction_log',[]):
                r = dict(p); r['_m'] = m; aps.append(r)
        aps.sort(key=lambda x: x.get('date',''))
        for p in aps:
            k = p.get('product_key',('',''))
            ws4.append([p.get('date',''), k[0] if isinstance(k,tuple) else '',
                k[1] if isinstance(k,tuple) else '',
                p.get('predicted_release',''), p.get('predicted_release_end',''),
                round(p.get('confidence',0),3),
                p.get('source',''), '是' if p.get('hit') else '否',
                round(p.get('ret_at_event',0),2), p.get('action','')])
            rn = ws4.max_row; hc = ws4.cell(row=rn,column=8)
            if hc.value == '是': hc.font = gnf
            elif hc.value == '否': hc.font = rf_
        for i,w in enumerate([12,12,16,12,12,10,12,8,12,16],1):
            ws4.column_dimensions[get_column_letter(i)].width = w
        ws4.freeze_panes = 'A2'

        # Sheet 5: 规律快照
        ws5 = wb.create_sheet("规律学习快照")
        h5 = ['银行','代码','置信度','释放周期','释放阶段','释放星期',
              '周期(天)','周期CV','有周期','Top阶段','阶段p值','Top星期','星期p值','事件数','平均窗口(天)']
        ws5.append(h5)
        for c in ws5[1]: c.font = hf; c.fill = gf; c.alignment = ct
        pn = ['月初(1-7)','月上旬(8-14)','月中(15-21)','月底(22-31)']
        wn = ['周一','周二','周三','周四','周五']
        ps = self.R['C'].get('patterns_snapshot',{})
        # 同时保留原始 ReleasePattern 对象用于格式化
        _raw_pats = self.R['C'].get('raw_patterns', {})
        for key, info in sorted(ps.items(), key=lambda x: -x[1]['confidence']):
            b = key[0] if isinstance(key,tuple) else ''; c_ = key[1] if isinstance(key,tuple) else ''
            tp = info.get('top_phase',0); tw = info.get('top_weekday',0)
            # 格式化释放规律
            _pat_obj = _raw_pats.get(key)
            _周期, _阶段, _星期, _ = _格式化释放规律(_pat_obj)
            ws5.append([b,c_,round(info.get('confidence',0),3),
                _周期, _阶段, _星期,
                round(info.get('period_days',0),1),round(info.get('period_cv',0),3),
                '是' if info.get('has_period') else '否',
                pn[tp] if tp<len(pn) else '', round(info.get('phase_pvalue',1.0),4),
                wn[tw] if tw<len(wn) else '', round(info.get('weekday_pvalue',1.0),4),
                info.get('n_events',0), round(info.get('avg_window_days',3.0),1)])
        for i,w in enumerate([12,16,10,10,16,10,10,10,8,14,10,10,10,8,12],1):
            ws5.column_dimensions[get_column_letter(i)].width = w
        ws5.freeze_panes = 'A2'

        # Sheet 6: 策略说明
        ws6 = wb.create_sheet("策略说明")
        title_f = Font(bold=True, size=14)
        h1_f = Font(bold=True, size=12, color='FFFFFF')
        h1_fill = PatternFill('solid', fgColor='4472C4')
        h2_f = Font(bold=True, size=11, color='1F4E79')
        param_fill = PatternFill('solid', fgColor='D6E4F0')
        note_f = Font(italic=True, color='666666')
        wrap = Alignment(wrap_text=True, vertical='top')

        ws6.column_dimensions['A'].width = 28
        ws6.column_dimensions['B'].width = 65
        ws6.column_dimensions['C'].width = 30

        def _sec(title):
            ws6.append([])
            ws6.append([title])
            r = ws6.max_row
            for col in range(1, 4):
                cell = ws6.cell(row=r, column=col)
                cell.font = h1_f; cell.fill = h1_fill

        def _row(a, b='', c=''):
            ws6.append([a, b, c])
            r = ws6.max_row
            for col in range(1, 4):
                ws6.cell(row=r, column=col).alignment = wrap

        def _param(name, value, note=''):
            ws6.append([name, value, note])
            r = ws6.max_row
            ws6.cell(row=r, column=1).fill = param_fill
            ws6.cell(row=r, column=1).font = Font(bold=True)
            for col in range(1, 4):
                ws6.cell(row=r, column=col).alignment = wrap

        # 标题
        ws6.append(['前瞻性收益释放策略 — 回测方案完整说明'])
        ws6.cell(row=1, column=1).font = title_f
        ws6.merge_cells('A1:C1')
        ws6.append([f'生成时间: {datetime.now():%Y-%m-%d %H:%M}'])
        ws6.cell(row=2, column=1).font = note_f

        # ── 一、回测概述 ──
        _sec('一、回测概述')
        _row('目标',
             '验证"前瞻性收益释放预测"能否提升银行理财产品轮动策略的年化收益。'
             '通过学习每个产品的历史收益释放规律(周期、月内阶段、星期偏好)，'
             '在释放窗口到来前优先选择该产品买入，从而捕获更完整的收益释放周期。')
        _row('方法',
             'Walk-Forward 滚动验证: 仅使用历史数据学习规律，严禁前瞻偏差。'
             '三种模式对比(A/B/C)，Mode A为纯反应式基线，Mode B/C逐步加入预测辅助。')
        _row('数据源', '净值数据库.xlsx (民生银行/中信银行/华夏银行)')
        _row('回测区间', f'{回测开始日} ~ {回测结束日}')
        _row('基线对标', 'V5.4策略回测 (年化5.69%)')

        # ── 二、数据处理 ──
        _sec('二、数据处理与产品筛选')
        _row('净值频率', '各产品净值发布频率不同(日/周/不定期)，统一按发布日计算年化收益率')
        _row('年化收益率公式', 'r = (NAV_t / NAV_{t-1} - 1) / 间隔天数 × 365 × 100%')
        _row('产品准入条件',
             '① 有效净值数据点 ≥ 10个\n'
             '② 赎回到账天数 ≤ 14天 (每日开放/每周开放/短持有期)')
        _row('产品总数', '约2200个产品通过基础筛选进入回测')

        # ── 三、产品库构建 ──
        _sec('三、产品库构建逻辑')
        _row('信号检测',
             f'使用"宽阈值"检测历史信号: 年化收益率从 ≤{信号阈值_库}% 上穿至 >{信号阈值_库}% 视为一次释放信号。\n'
             f'注意: 信号检测阈值({信号阈值_库}%)低于买入决策阈值({信号阈值_买入}%)，即"宽入严买"。')
        _row('信号评估',
             f'每个信号触发后，计算接下来{持有天数_评估}天的平均年化收益率。\n'
             f'平均年化 > {成功标准}% 视为"成功信号"。')
        _row('持续天数',
             f'信号触发后，收益率持续高于 {信号阈值_库}%×0.8 = {信号阈值_库*0.8}% 的连续天数(日历天)。')
        _row('入库条件',
             f'① 历史信号数 ≥ {最少历史信号}次\n'
             f'② 成功率 ≥ {最低成功率}%\n'
             f'③ 平均持续天数在 {最短持续天数} ~ {最大持有天数_评估}天之间')
        _row('产品库重建', f'每 {产品库重建间隔} 天重建一次，仅使用重建日之前的数据(禁止前瞻)')
        _row('Sharpe计算', f'Sharpe_like = 平均收益 / max(标准差, 0.5)，std下限0.5防止低波动产品得分过高')

        # ── 四、买入排名 ──
        _sec('四、买入排名公式 (V5.4)')
        _row('排名方式',
             '收益率复合排名: composite = 当日收益率 × 历史平均收益率\n'
             '按 composite 降序排列，同分按成功率降序。\n'
             '选择排名最高的产品优先买入，直到仓位填满。')
        _row('Mode B 加成',
             f'当前瞻预测器预测某产品即将释放时，对其排名分数进行加成:\n'
             f'composite_boosted = composite × (1 + prediction_score × {ALPHA_预测加成})\n'
             f'其中 prediction_score = confidence / sqrt(距释放日交易天数)\n'
             f'ALPHA_预测加成 = {ALPHA_预测加成} (敏感性分析最优值，扫描范围0.05~1.00)')

        # ── 五、T+1交易规则 ──
        _sec('五、T+1交易规则')
        _row('申购流程',
             'T日: 扫描信号并提交申购 → T+1日: 确认份额(用T+1日净值作为买入价)\n'
             '周末/节假日顺延到下一交易日确认')
        _row('赎回流程',
             'T日: 检查卖出条件并提交赎回 → T+1日: 到账(用T+1日净值计算卖出价)\n'
             '周末/节假日顺延到下一交易日到账')
        _param('买入信号阈值', f'年化收益率 > {信号阈值_买入}%', '高于此值触发买入')
        _param('卖出信号阈值', f'年化收益率 ≤ {信号阈值_卖出}%', '低于或等于此值触发卖出')
        _param('最短持有期', f'确认后 {最短持有交易日_赎回} 个交易日', '防止频繁交易')
        _param('最大持有期', f'{最大持有交易日} 个交易日', '超过强制卖出')
        _param('节前限制', f'长假前(日历间隔>{长假阈值_天}天)不提交申购', '避免资金闲置')

        # ── 六、仓位管理 ──
        _sec('六、仓位管理')
        _param('最大持仓数', f'{最大持仓数} 个产品', '')
        _param('仓位大小', f'动态: 当前组合总市值 / {最大持仓数}', 'V5.4动态仓位，支持复利增长')
        _param('最低仓位', f'≥ 基础仓位的50%', '低于此不执行买入')
        _row('轮换卖出',
             '满仓时执行轮换检查:\n'
             f'① 找到收益率最低的持仓(已过最短持有期)\n'
             f'② 如果最低收益 < {信号阈值_买入}%，搜索替代候选\n'
             f'③ 候选收益必须 > 最低收益 + 1.0%，避免频繁换手\n'
             f'④ 满足则卖出最弱持仓，下一交易日确认买入替代产品')

        # ── 七、三种回测模式 ──
        _sec('七、三种回测模式')
        _row('Mode A — 纯反应式基线',
             f'完全复现V5.4策略逻辑:\n'
             f'• 买入: 当日年化 > {信号阈值_买入}%\n'
             f'• 卖出: 当日年化 ≤ {信号阈值_卖出}% 或 持有 ≥ {最大持有交易日}交易日\n'
             f'• 排名: 当日收益 × 历史均收益(降序)\n'
             f'• 轮换: 满仓时弱持仓可被强候选替换\n'
             f'• 无任何预测辅助，作为对照组')
        _row('Mode B — 优选(Alpha加成)',
             f'在Mode A基础上，仅增加建议①:\n'
             f'• 买卖规则与Mode A完全相同\n'
             f'• 排名时，预测即将释放的产品获得加成(×{1+ALPHA_预测加成:.2f})\n'
             f'• 不改阈值、不改卖出逻辑，仅改优先级\n'
             f'• 已验证为最优模式')
        _row('Mode C — 全部三建议',
             '在Mode B基础上，增加建议②③:\n'
             f'• 建议② 智持: 收益跌至卖出线时，若预测{延迟卖出_预测窗口}交易日内将再释放\n'
             f'  (置信度≥{延迟卖出_置信度})，延迟卖出\n'
             f'• 建议③ 节前买入: 长假前优先买入库中优质产品\n'
             f'  (成功率≥{节前买入_最低成功率}%, 均收≥{节前买入_最低均收}%, 实时>{节前买入_实时阈值}%)\n'
             f'  预留{节前保留仓位}仓位 + 节前轮动弱持仓, 赚取假期累积收益')

        # ── 八、规律学习器 ──
        _sec('八、规律学习器 (PatternLearner)')
        _row('输入', f'某产品过去{规律学习窗口天数}天的日收益率序列')
        _row('释放事件识别', f'年化收益率 > {释放识别阈值}% 的连续区间为一个释放窗口')
        _row('学习维度 (3个)',
             '① 周期性: 提取所有释放窗口起始日，计算间隔序列，取中位数为典型周期；\n'
             f'   变异系数(CV) < {周期CV阈值} 视为有规律性\n'
             '② 月内阶段集中度: 释放事件在4个阶段(1-7/8-14/15-21/22-31)的分布，\n'
             '   卡方检验判断是否显著集中\n'
             '③ 星期偏好: 释放事件在周一~周五的分布，卡方检验判断偏好')
        _row('置信度评分 (0~1)',
             f'confidence = 0.0\n'
             f'if 有周期 and CV<0.3:  +0.40  (elif CV<{周期CV阈值}: +0.25)\n'
             f'if 阶段p值<0.05:      +0.30\n'
             f'if 星期p值<0.05:      +0.15\n'
             f'if 释放事件数≥5:      +0.15\n'
             f'上限1.0，最低{最低置信度}才参与预测')
        _row('平均窗口天数', '统计所有历史释放窗口的持续天数均值，用于预测释放结束日')
        _row('Walk-Forward重建',
             f'每{规律重建间隔天数}天用最近{规律学习窗口天数}天数据重建规律，\n'
             '确保不使用未来数据')

        # ── 九、前瞻预测器 ──
        _sec('九、前瞻预测器 (ForwardPredictor)')
        _row('双路径集成预测',
             '路径A — 周期预测:\n'
             '  上次释放结束日 + 周期天数 = 预测下次释放日\n'
             '  置信度 = (1 - period_cv) × 0.7\n'
             '  窗口: 预测日 ± 1天\n\n'
             '路径B — 阶段预测:\n'
             '  当月top_phase对应的起始日为预测释放日\n'
             '  置信度 = top_phase概率 × 0.6\n'
             '  窗口: 阶段起始前1天 ~ 后2天')
        _row('集成规则',
             '• 两路都命中同一窗口 → 置信度取两者之和(上限1.0)\n'
             '• 仅一路命中 → 使用该路置信度\n'
             f'• 最终置信度 < 0.3 的预测丢弃')
        _row('排名评分', 'score = confidence / sqrt(距预测释放日的交易天数)\n'
             '距离越近、置信度越高的预测得分越高')
        _row('预测释放结束日', '预测释放日 + 平均窗口天数 - 1')

        # ── 十、参数总览 ──
        _sec('十、全部参数一览')
        ws6.append(['参数名称', '取值', '说明'])
        r = ws6.max_row
        for col in range(1, 4):
            ws6.cell(row=r, column=col).font = Font(bold=True)
            ws6.cell(row=r, column=col).fill = PatternFill('solid', fgColor='B4C6E7')

        params = [
            ('回测区间', f'{回测开始日} ~ {回测结束日}', '约243个交易日'),
            ('初始资金', f'{初始资金/1e4:,.0f}万', '1亿元'),
            ('最大持仓数', str(最大持仓数), ''),
            ('仓位大小', f'动态: 总市值/{最大持仓数}', '复利效应'),
            ('信号阈值_买入', f'{信号阈值_买入}%', '买入决策阈值'),
            ('信号阈值_卖出', f'{信号阈值_卖出}%', '卖出触发阈值'),
            ('信号阈值_库', f'{信号阈值_库}%', '产品库信号检测(宽入)'),
            ('成功标准', f'{成功标准}%', '7天平均年化>此值=成功'),
            ('最低成功率', f'{最低成功率}%', '入库最低门槛'),
            ('最短持有交易日', str(最短持有交易日_赎回), '确认后最短持有'),
            ('最大持有交易日', str(最大持有交易日), '超过强制卖出'),
            ('产品库重建间隔', f'{产品库重建间隔}天', ''),
            ('长假阈值', f'{长假阈值_天}天', '日历间隔>此值=长假'),
            ('', '', ''),
            ('规律学习窗口', f'{规律学习窗口天数}天', '训练数据量'),
            ('规律重建间隔', f'{规律重建间隔天数}天', 'Walk-Forward步长'),
            ('释放识别阈值', f'{释放识别阈值}%', '年化>此值=释放事件'),
            ('最低置信度', str(最低置信度), '低于此不参与预测'),
            ('预测窗口天数', str(预测窗口天数), '预测未来N天'),
            ('周期CV阈值', str(周期CV阈值), 'CV<此值认为有周期'),
            ('ALPHA_预测加成', str(ALPHA_预测加成), '敏感性分析最优值'),
            ('', '', ''),
            ('延迟卖出_预测窗口', f'{延迟卖出_预测窗口}交易日', '建议②参数'),
            ('延迟卖出_置信度', str(延迟卖出_置信度), '建议②参数'),
            ('延迟卖出_最大持有余量', str(延迟卖出_最大持有余量), '建议②参数'),
            ('节前买入_实时阈值', f'{节前买入_实时阈值}%', '建议③参数'),
            ('节前买入_最低成功率', f'{节前买入_最低成功率}%', '建议③参数'),
            ('节前买入_最低均收', f'{节前买入_最低均收}%', '建议③参数'),
            ('节前保留仓位', str(节前保留仓位), '建议③参数'),
            ('节前买入_最大产品数', str(节前买入_最大产品数), '建议③参数'),
        ]
        for name, val, note in params:
            _param(name, val, note)

        # ── 十一、结论 ──
        _sec('十一、敏感性分析结论')
        _row('扫描范围', 'ALPHA_预测加成 = [0.05, 0.10, 0.15, 0.20, 0.25, 0.30, 0.35, 0.40, 0.50, 0.60, 0.80, 1.00]')
        _row('最优参数', f'ALPHA_预测加成 = {ALPHA_预测加成}')
        _row('收益特征',
             '倒U型曲线: 太弱(≤0.15)无法改变排名；太强(≥0.40)预测覆盖基本面导致选品质量下降。\n'
             '甜蜜区间 0.20~0.30 稳定提升 +0.22~0.25%。')
        _row('最终建议',
             '采用 Mode B (仅建议① Alpha加成):\n'
             '• 年化收益: 基线+0.25%\n'
             '• 夏普比率: 提升约19%\n'
             '• 最大回撤: 降低约29%\n'
             '• 建议②(智持)和③(节前)无增益，不建议使用')

        ws6.freeze_panes = 'A3'

        # Sheet 7: 参数敏感性分析 (如有)
        if self.sweep_results and self.base_ann is not None:
            ws7 = wb.create_sheet("参数敏感性分析", 0)  # 插到第一页

            # 说明区
            ws7.append(['ALPHA_预测加成 参数敏感性分析'])
            ws7.cell(row=1, column=1).font = Font(bold=True, size=14)
            ws7.merge_cells('A1:J1')
            ws7.append([])
            explain = [
                '说明:',
                'ALPHA_预测加成 是前瞻预测对买入排名的加成权重。',
                '当预测某产品即将释放收益时，其排名分数乘以 (1 + prediction_score × ALPHA)。',
                'ALPHA越大，预测影响排名越强；ALPHA=0 等于不使用预测(即Mode A基线)。',
                f'基线 Mode A 不使用任何预测，年化收益 {self.base_ann:+.4f}%。',
                '下表扫描不同ALPHA值，寻找最优加成强度。',
            ]
            for txt in explain:
                ws7.append([txt])
                ws7.cell(row=ws7.max_row, column=1).font = Font(italic=True, color='333333') if txt != '说明:' else Font(bold=True)
            ws7.merge_cells(f'A{ws7.max_row-5}:J{ws7.max_row-5}')  # merge title row
            ws7.append([])

            # 数据表头
            headers = ['ALPHA_预测加成', '年化收益率%', 'Δ年化(vs基线)%', '夏普比率',
                        '最大回撤%', '胜率%', '交易次数', '预测加成买入次数', '预测买入胜率%', '预测买入盈亏(万)']
            ws7.append(headers)
            hr = ws7.max_row
            for cell in ws7[hr]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill('solid', fgColor='4472C4')
                cell.alignment = Alignment(horizontal='center', wrap_text=True)

            # 基线行
            ws7.append(['基线 Mode A (无预测)', round(self.base_ann, 4), '—',
                         round(self.R['A']['sharpe'], 4), round(self.R['A']['max_dd'], 4),
                         round(self.R['A']['win_rate'], 1), self.R['A']['n_trades'],
                         '—', '—', '—'])
            for cell in ws7[ws7.max_row]:
                cell.fill = PatternFill('solid', fgColor='FFF2CC')
                cell.font = Font(bold=True)

            # 扫描行
            best_sr = max(self.sweep_results, key=lambda x: x['ann_ret'])
            for sr in self.sweep_results:
                ws7.append([sr['alpha'], round(sr['ann_ret'], 4), round(sr['delta'], 4),
                             round(sr['sharpe'], 4), round(sr['max_dd'], 4),
                             round(sr['win_rate'], 1), sr['n_trades'],
                             sr['n_alpha_boost'], round(sr['pred_wr'], 1),
                             round(sr['pred_pnl'] / 1e4, 2)])
                if sr['alpha'] == best_sr['alpha']:
                    for cell in ws7[ws7.max_row]:
                        cell.fill = PatternFill('solid', fgColor='C6EFCE')
                        cell.font = Font(bold=True)

            # 结论
            ws7.append([])
            ws7.append([f'最优: ALPHA_预测加成 = {best_sr["alpha"]:.2f}，'
                         f'年化 {best_sr["ann_ret"]:+.4f}%，'
                         f'超越基线 {best_sr["delta"]:+.4f}%，'
                         f'夏普 {best_sr["sharpe"]:.4f}'])
            ws7.cell(row=ws7.max_row, column=1).font = Font(bold=True, size=12, color='006100')
            ws7.merge_cells(f'A{ws7.max_row}:J{ws7.max_row}')
            ws7.append(['解读: 倒U型曲线。太小(≤0.15)改变不了排名；太大(≥0.40)预测覆盖基本面，选品质量下降。甜蜜区间0.20~0.30。'])
            ws7.cell(row=ws7.max_row, column=1).font = Font(italic=True, color='666666')
            ws7.merge_cells(f'A{ws7.max_row}:J{ws7.max_row}')

            # 列宽
            for i, w in enumerate([18, 12, 14, 10, 10, 8, 10, 16, 14, 14], 1):
                ws7.column_dimensions[get_column_letter(i)].width = w
            ws7.freeze_panes = f'A{hr+1}'

            # 折线图
            n_data = len(self.sweep_results)
            data_start = hr + 2  # 跳过基线行
            chart = LineChart()
            chart.title = "ALPHA_预测加成 vs 年化收益率"
            chart.x_axis.title = "ALPHA_预测加成 (预测对排名的加成权重)"
            chart.y_axis.title = "年化收益率 %"
            chart.style = 10
            chart.width = 22; chart.height = 13
            data_ref = Reference(ws7, min_col=2, min_row=hr, max_row=hr + 1 + n_data)
            cats_ref = Reference(ws7, min_col=1, min_row=data_start, max_row=hr + 1 + n_data)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.series[0].graphicalProperties.line.width = 25000
            ws7.add_chart(chart, f"A{ws7.max_row + 2}")

        # 保存
        try: wb.save(OUTPUT_PATH); logger.info(f"报告: {OUTPUT_PATH}")
        except:
            bk = os.path.join(BASE_DIR, f"前瞻策略回测报告_{datetime.now():%H%M%S}.xlsx")
            wb.save(bk); logger.info(f"报告: {bk}")
        return OUTPUT_PATH


# ============================================================
# 主函数
# ============================================================

def main():
    """V3.2 主入口: 直接对比 Mode A / Mode B(0.30) / Mode D(B+DL乘法加成)

    策略说明:
    - Mode A: 纯反应式基线 (无预测, 精确复现 backtest_v6)
    - Mode B: 基于周期+阶段规律预测的Alpha排名加成 (ALPHA=0.30, 原始最优)
    - Mode D: Mode B + DL乘法加成
      * 只在已有pattern预测的产品上叠加DL加成
      * score *= (1 + dl_boost * DL_MULT_WEIGHT), 上限DL_MAX_BOOST
      * val_loss > 0.45 时跳过DL (模型不可靠)
      * 不创建DL-only合成预测

    V3.1→V3.2改动:
      - 加法混合 → 乘法加成
      - 删除DL-only合成预测 (fake_pred)
      - 提高DL门槛: entry>=0.50, conf>=0.50
      - 新增val_loss质量门控
    """
    global ALPHA_预测加成

    sep = '=' * 74
    print(sep)
    print("  前瞻策略回测 V3.2 — DL乘法加成 (修正V3.1退化)")
    print(f"  ALPHA_预测加成 = 0.30 (V3最优, 固定)")
    print(f"  DL乘法加成 = {DL_MULT_WEIGHT} (上限{DL_MAX_BOOST})")
    print(f"  DL门槛: entry>={DL_MIN_ENTRY_QUALITY}, conf>={DL_MIN_CONFIDENCE}, val_loss<{DL_MAX_VAL_LOSS}")
    print(f"  DL模型: gpu_alpha_v2.py (12维特征, LSTM+Attention, AMP)")
    print(sep)

    ALPHA_预测加成 = 0.30  # 固定为V3扫描出的最优值

    engine = PatternBacktestEngine()
    engine.load_data()

    # ① 基线 Mode A
    print("\n  [1/3] Mode A — 纯反应式基线...")
    ra = engine.run_single_mode('A')

    # ② Mode B (原始最优, 无DL)
    print("\n  [2/3] Mode B — 周期+阶段预测 (ALPHA=0.30)...")
    rb = engine.run_single_mode('B')

    # ③ Mode D (B + DL混合)
    print("\n  [3/3] Mode D — 周期+阶段+DL乘法加成...")
    dl_alpha = DLAlphaManager()
    if dl_alpha.available:
        # 首次DL训练
        dl_alpha.train_and_predict(回测开始日)
    rd = engine.run_single_mode('D', dl_alpha=dl_alpha)

    # ④ 打印对比
    print(f"\n{sep}")
    print("     V3.2 DL乘法加成 回测对比")
    print(sep)

    print(f"  {'指标':<20} {'A(反应)':<14} {'B(+预测)':<14} {'D(+DL加成)':<14}")
    print("  " + "-" * 58)

    rows = [
        ('年化收益率', 'ann_ret', '{:+.4f}%'),
        ('总收益率', 'total_ret', '{:+.4f}%'),
        ('最大回撤', 'max_dd', '{:.4f}%'),
        ('夏普比率', 'sharpe', '{:.4f}'),
        ('交易次数', 'n_trades', '{:d}'),
        ('胜率', 'win_rate', '{:.1f}%'),
        ('平均持有天数', 'avg_hold', '{:.1f}'),
        ('Alpha加成买入', 'n_alpha_boost', '{:d}'),
        ('预测入场盈亏万', 'pred_pnl', '{:+.2f}'),
    ]
    for label, key, fmt in rows:
        vals = []
        for r in [ra, rb, rd]:
            v = r.get(key, 0)
            if key in ('n_alpha_boost', 'pred_pnl') and r is ra:
                vals.append('N/A')
            else:
                try:
                    if 'd' in fmt:
                        vals.append(fmt.format(int(v)))
                    elif key == 'pred_pnl':
                        vals.append(fmt.format(v / 1e4))
                    else:
                        vals.append(fmt.format(v))
                except:
                    vals.append(str(v))
        print(f"  {label:<20} {vals[0]:<14} {vals[1]:<14} {vals[2]:<14}")

    print()
    d_b = rb['ann_ret'] - ra['ann_ret']
    d_d = rd['ann_ret'] - ra['ann_ret']
    d_db = rd['ann_ret'] - rb['ann_ret']
    print(f"  Mode B vs A: 年化 {d_b:+.4f}%")
    print(f"  Mode D vs A: 年化 {d_d:+.4f}%")
    print(f"  Mode D vs B: 年化 {d_db:+.4f}% {'↑改善' if d_db > 0 else '↓退化'}")

    best = max(['A', 'B', 'D'], key=lambda m: {'A': ra, 'B': rb, 'D': rd}[m]['ann_ret'])
    best_r = {'A': ra, 'B': rb, 'D': rd}[best]
    print(f"\n  最优: Mode {best} (年化 {best_r['ann_ret']:+.4f}%  夏普 {best_r['sharpe']:.4f})")

    if rd['ann_ret'] > rb['ann_ret']:
        print(f"  结论: DL混合有效, 增量alpha +{d_db:.4f}%")
    else:
        print(f"  结论: DL混合未超越Mode B, 需调参或增加数据")

    print(sep)
    print("  完成。")


if __name__ == '__main__':
    main()
