"""测试华夏理财净值解析并输出Excel"""
import sys
sys.path.insert(0, '.')

from bank_crawler import HuaxiaCrawler, MultiBankCrawler, BankType
import pandas as pd
from datetime import datetime
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(message)s')

def test_nav_to_excel():
    print("="*60)
    print("测试华夏理财净值解析 - 输出Excel")
    print("="*60)

    crawler = HuaxiaCrawler()

    # 获取产品列表（前20个测试）
    products = crawler.get_product_list()[:20]
    print(f"\n获取到 {len(products)} 个测试产品\n")

    results = []
    for i, product in enumerate(products):
        product_code = product.get('id', '')
        excel_url = product.get('address', '')
        title = product.get('title', '')

        if not excel_url:
            continue

        # 下载并解析Excel
        nav_data = crawler.download_nav_excel(excel_url)

        if nav_data:
            # 从标题提取产品名称
            import re
            name_match = re.search(r'^(.+?)\[', title)
            product_name = name_match.group(1).strip() if name_match else title

            row = {
                '银行': '华夏理财',
                '产品代码': product_code,
                '产品名称': product_name,
                '最新净值': nav_data[0]['unit_nav'] if nav_data else None,
                '净值日期': nav_data[0]['date'] if nav_data else None,
                '净值天数': len(nav_data)
            }

            # 添加每天的净值数据
            for j, nav in enumerate(nav_data):
                row[f'日期{j+1}'] = nav['date']
                row[f'净值{j+1}'] = nav['unit_nav']

            results.append(row)
            print(f"[{i+1}] {product_code}: {len(nav_data)}天净值")

    # 保存到Excel
    df = pd.DataFrame(results)
    filename = f'华夏理财_净值测试_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='净值数据', index=False)

        # 设置净值列格式
        try:
            from openpyxl.styles import numbers
            ws = writer.sheets['净值数据']
            for col_idx, col_cell in enumerate(ws[1], 1):
                col_name = col_cell.value
                if col_name and ('净值' in str(col_name) and '日期' not in str(col_name)):
                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value is not None:
                            cell.number_format = '0.0000'
        except Exception as e:
            print(f"格式设置失败: {e}")

    print(f"\n已保存到: {filename}")
    print(f"共 {len(results)} 个产品")
    return filename

if __name__ == "__main__":
    test_nav_to_excel()
