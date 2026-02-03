# -*- coding: utf-8 -*-
"""
理财产品综合筛选器

功能：
1. 连续N日收益上涨筛选
2. 收益波动率筛选
3. 特定收益区间筛选
4. 最大历史回撤
5. 夏普比率
6. 平均收益率/中位数收益率
7. 胜率（正收益天数比例）
8. 收益稳定性（变异系数）
9. 近期动量
10. 生成Excel报告
"""

import os
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# 文件路径
NAV_DB_FILE = os.path.join(os.path.dirname(__file__), "净值数据库.xlsx")
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "产品筛选报告_v2.xlsx")

# 无风险利率（年化，用于计算夏普比率）
RISK_FREE_RATE = 2.0  # 假设2%年化无风险利率


def load_nav_database():
    """加载净值数据库"""
    if not os.path.exists(NAV_DB_FILE):
        raise FileNotFoundError(f"净值数据库不存在: {NAV_DB_FILE}")

    xlsx = pd.ExcelFile(NAV_DB_FILE)
    data = {}
    for sheet_name in xlsx.sheet_names:
        df = pd.read_excel(xlsx, sheet_name=sheet_name, dtype=str)
        df.columns = [str(c) if not isinstance(c, str) else c for c in df.columns]

        # 处理列名
        rename_map = {}
        if 'level_0' in df.columns:
            rename_map['level_0'] = '产品代码'
        if 'level_1' in df.columns:
            rename_map['level_1'] = '产品名称'
        if rename_map:
            df = df.rename(columns=rename_map)

        data[sheet_name] = df
        logger.info(f"加载 {sheet_name}: {len(df)} 个产品")
    return data


def is_date_column(col_name):
    """判断是否是日期列"""
    if not isinstance(col_name, str):
        return False
    if len(col_name) == 10 and col_name[4] == '-' and col_name[7] == '-':
        try:
            datetime.strptime(col_name, '%Y-%m-%d')
            return True
        except:
            return False
    return False


def calculate_metrics(df, bank_name, min_dates=5):
    """
    计算所有产品的各项指标

    Args:
        df: 净值数据DataFrame
        bank_name: 银行名称
        min_dates: 最少需要的净值日期数量

    Returns:
        包含各项指标的DataFrame
    """
    # 获取日期列（从新到旧排序）
    date_cols = sorted([c for c in df.columns if is_date_column(c)], reverse=True)

    if len(date_cols) < min_dates:
        logger.warning(f"{bank_name}: 日期数量不足 ({len(date_cols)} < {min_dates})")
        return None

    results = []

    skipped_empty = 0
    for idx, row in df.iterrows():
        product_code = str(row.get('产品代码', '')).strip() if pd.notna(row.get('产品代码')) else ''
        product_name = str(row.get('产品名称', '')).strip() if pd.notna(row.get('产品名称')) else ''

        if not product_code or product_code == 'nan':
            skipped_empty += 1
            continue

        # 获取净值序列（从新到旧）
        nav_values = []
        nav_dates = []
        for date in date_cols:
            try:
                val = row.get(date)
                if pd.notna(val) and str(val).strip():
                    nav_values.append(float(val))
                    nav_dates.append(date)
            except (ValueError, TypeError):
                continue

        if len(nav_values) < min_dates:
            continue

        # 计算日收益率序列（从新到旧，即 nav_values[i] 相对于 nav_values[i+1]）
        daily_returns = []
        for i in range(len(nav_values) - 1):
            if nav_values[i+1] > 0:
                ret = (nav_values[i] - nav_values[i+1]) / nav_values[i+1] * 100  # 百分比
                daily_returns.append(ret)

        if len(daily_returns) < 2:
            continue

        # ========== 计算各项指标 ==========

        # 1. 最新日收益率
        latest_return = daily_returns[0] if daily_returns else None

        # 2. 连续上涨天数（从最新日期往前数）
        consecutive_up_days = 0
        for ret in daily_returns:
            if ret > 0:
                consecutive_up_days += 1
            else:
                break

        # 3. 是否连续两日上涨
        is_2day_up = 1 if len(daily_returns) >= 2 and daily_returns[0] > 0 and daily_returns[1] > 0 else 0

        # 4. 平均日收益率
        avg_daily_return = np.mean(daily_returns)

        # 5. 年化收益率（平均日收益率 * 365）
        annual_return = avg_daily_return * 365

        # 6. 收益率中位数
        median_return = np.median(daily_returns)

        # 7. 收益率标准差（波动率）
        volatility = np.std(daily_returns)

        # 8. 年化波动率
        annual_volatility = volatility * np.sqrt(365)

        # 9. 夏普比率 = (年化收益率 - 无风险利率) / 年化波动率
        if annual_volatility > 0:
            sharpe_ratio = (annual_return - RISK_FREE_RATE) / annual_volatility
        else:
            sharpe_ratio = 0

        # 10. 最大单日收益
        max_daily_return = max(daily_returns)

        # 11. 最小单日收益
        min_daily_return = min(daily_returns)

        # 12. 胜率（正收益天数比例）
        positive_days = sum(1 for r in daily_returns if r > 0)
        win_rate = positive_days / len(daily_returns) * 100

        # 13. 变异系数（收益稳定性，越小越稳定）
        if abs(avg_daily_return) > 0.0001:
            cv = abs(volatility / avg_daily_return)
        else:
            cv = float('inf') if volatility > 0 else 0

        # 14. 最大回撤（使用净值序列计算）
        # 净值从新到旧，需要反转为从旧到新
        nav_old_to_new = list(reversed(nav_values))
        peak = nav_old_to_new[0]
        max_drawdown = 0
        for nav in nav_old_to_new:
            if nav > peak:
                peak = nav
            drawdown = (peak - nav) / peak * 100 if peak > 0 else 0
            if drawdown > max_drawdown:
                max_drawdown = drawdown

        # 15. 近5日平均收益 vs 前5日平均收益（动量指标）
        if len(daily_returns) >= 10:
            recent_5d_avg = np.mean(daily_returns[:5])
            prev_5d_avg = np.mean(daily_returns[5:10])
            momentum = recent_5d_avg - prev_5d_avg
        elif len(daily_returns) >= 5:
            recent_5d_avg = np.mean(daily_returns[:5])
            prev_5d_avg = None
            momentum = None
        else:
            recent_5d_avg = np.mean(daily_returns)
            prev_5d_avg = None
            momentum = None

        # 16. 最新5日年化收益率
        if len(daily_returns) >= 5:
            recent_5d_annual = np.mean(daily_returns[:5]) * 365
        else:
            recent_5d_annual = np.mean(daily_returns) * 365

        # 17. 净值数据天数
        data_days = len(nav_values)

        # 18. 持有2周预期年化收益率
        # 假设今日购买明日确认，持有2周（约10个交易日）
        # 使用历史10日滚动收益率的平均值来预估
        holding_days = 10  # 2周约10个交易日
        if len(daily_returns) >= holding_days:
            # 计算滚动10日累计收益率
            rolling_2week_returns = []
            for i in range(len(daily_returns) - holding_days + 1):
                # 累计收益率 = (1+r1)*(1+r2)*...*(1+r10) - 1
                cumulative = 1.0
                for j in range(holding_days):
                    cumulative *= (1 + daily_returns[i + j] / 100)
                rolling_2week_returns.append((cumulative - 1) * 100)

            # 平均2周收益率
            avg_2week_return = np.mean(rolling_2week_returns)
            # 年化: 2周收益率 * (365/14)
            expected_2week_annual = avg_2week_return * (365 / 14)

            # 最近一次2周收益率（用于参考）
            recent_2week_return = rolling_2week_returns[0] if rolling_2week_returns else None
            recent_2week_annual = recent_2week_return * (365 / 14) if recent_2week_return else None
        else:
            # 数据不足时用平均日收益估算
            avg_2week_return = avg_daily_return * holding_days
            expected_2week_annual = avg_2week_return * (365 / 14)
            recent_2week_return = None
            recent_2week_annual = None

        results.append({
            '银行': bank_name,
            '产品代码': product_code,
            '产品名称': product_name,
            '最新净值': nav_values[0],
            '最新日期': nav_dates[0],
            '数据天数': data_days,
            '最新日收益率%': round(latest_return, 4) if latest_return else None,
            '连续上涨天数': consecutive_up_days,
            '连续两日上涨': is_2day_up,
            '平均日收益率%': round(avg_daily_return, 4),
            '年化收益率%': round(annual_return, 2),
            '近5日年化%': round(recent_5d_annual, 2),
            '收益中位数%': round(median_return, 4),
            '日波动率%': round(volatility, 4),
            '年化波动率%': round(annual_volatility, 2),
            '夏普比率': round(sharpe_ratio, 2),
            '最大日收益%': round(max_daily_return, 4),
            '最小日收益%': round(min_daily_return, 4),
            '胜率%': round(win_rate, 2),
            '变异系数': round(cv, 2) if cv != float('inf') else 999,
            '最大回撤%': round(max_drawdown, 2),
            '动量': round(momentum, 4) if momentum is not None else None,
            '2周预期年化%': round(expected_2week_annual, 2),
            '近期2周年化%': round(recent_2week_annual, 2) if recent_2week_annual is not None else None,
        })

    if skipped_empty > 0:
        logger.warning(f"{bank_name}: 跳过 {skipped_empty} 行无产品代码的数据")

    if not results:
        return None

    return pd.DataFrame(results)


def create_screening_excel(all_metrics_df, output_file):
    """
    创建筛选报告Excel（多Sheet）
    """
    wb = Workbook()

    # 样式定义
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 绿色
    bad_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # 红色
    highlight_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 黄色
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')

    # ========== Sheet 1: 全部数据（按年化收益率排序）==========
    ws1 = wb.active
    ws1.title = "全部产品"
    df_all = all_metrics_df.sort_values('年化收益率%', ascending=False)
    write_dataframe_to_sheet(ws1, df_all, header_fill, header_font, thin_border, center_alignment)
    logger.info(f"全部产品: {len(df_all)} 条")

    # ========== Sheet 2: 连续两日上涨 ==========
    ws2 = wb.create_sheet("连续两日上涨")
    df_2day_up = all_metrics_df[all_metrics_df['连续两日上涨'] == 1].sort_values('年化收益率%', ascending=False)
    write_dataframe_to_sheet(ws2, df_2day_up, header_fill, header_font, thin_border, center_alignment)
    logger.info(f"连续两日上涨: {len(df_2day_up)} 条")

    # ========== Sheet 3: 高夏普比率（>1）==========
    ws3 = wb.create_sheet("高夏普比率")
    df_high_sharpe = all_metrics_df[all_metrics_df['夏普比率'] > 1].sort_values('夏普比率', ascending=False)
    write_dataframe_to_sheet(ws3, df_high_sharpe, header_fill, header_font, thin_border, center_alignment)
    logger.info(f"高夏普比率(>1): {len(df_high_sharpe)} 条")

    # ========== Sheet 4: 低波动率（年化波动率<5%）==========
    ws4 = wb.create_sheet("低波动率")
    df_low_vol = all_metrics_df[all_metrics_df['年化波动率%'] < 5].sort_values('年化收益率%', ascending=False)
    write_dataframe_to_sheet(ws4, df_low_vol, header_fill, header_font, thin_border, center_alignment)
    logger.info(f"低波动率(<5%): {len(df_low_vol)} 条")

    # ========== Sheet 5: 低回撤（最大回撤<1%）==========
    ws5 = wb.create_sheet("低回撤")
    df_low_dd = all_metrics_df[all_metrics_df['最大回撤%'] < 1].sort_values('年化收益率%', ascending=False)
    write_dataframe_to_sheet(ws5, df_low_dd, header_fill, header_font, thin_border, center_alignment)
    logger.info(f"低回撤(<1%): {len(df_low_dd)} 条")

    # ========== Sheet 6: 高胜率（>80%）==========
    ws6 = wb.create_sheet("高胜率")
    df_high_wr = all_metrics_df[all_metrics_df['胜率%'] > 80].sort_values('年化收益率%', ascending=False)
    write_dataframe_to_sheet(ws6, df_high_wr, header_fill, header_font, thin_border, center_alignment)
    logger.info(f"高胜率(>80%): {len(df_high_wr)} 条")

    # ========== Sheet 7: 稳健型（低波动+高胜率+正收益）==========
    ws7 = wb.create_sheet("稳健型产品")
    df_stable = all_metrics_df[
        (all_metrics_df['年化波动率%'] < 5) &
        (all_metrics_df['胜率%'] > 70) &
        (all_metrics_df['年化收益率%'] > 0)
    ].sort_values('夏普比率', ascending=False)
    write_dataframe_to_sheet(ws7, df_stable, header_fill, header_font, thin_border, center_alignment)
    logger.info(f"稳健型产品: {len(df_stable)} 条")

    # ========== Sheet 8: 高收益型（年化>10%）==========
    ws8 = wb.create_sheet("高收益型")
    df_high_ret = all_metrics_df[all_metrics_df['年化收益率%'] > 10].sort_values('夏普比率', ascending=False)
    write_dataframe_to_sheet(ws8, df_high_ret, header_fill, header_font, thin_border, center_alignment)
    logger.info(f"高收益型(>10%): {len(df_high_ret)} 条")

    # ========== Sheet 9: 近期强势（近5日年化>年化收益）==========
    ws9 = wb.create_sheet("近期强势")
    df_momentum = all_metrics_df[
        all_metrics_df['近5日年化%'] > all_metrics_df['年化收益率%']
    ].sort_values('近5日年化%', ascending=False)
    write_dataframe_to_sheet(ws9, df_momentum, header_fill, header_font, thin_border, center_alignment)
    logger.info(f"近期强势: {len(df_momentum)} 条")

    # ========== Sheet 10: 2周预期收益优选 ==========
    ws10 = wb.create_sheet("2周预期收益TOP")
    df_2week = all_metrics_df[
        all_metrics_df['2周预期年化%'].notna()
    ].sort_values('2周预期年化%', ascending=False).head(500)
    write_dataframe_to_sheet(ws10, df_2week, header_fill, header_font, thin_border, center_alignment)
    logger.info(f"2周预期收益TOP: {len(df_2week)} 条")

    # ========== Sheet 11: 综合优选（多条件筛选）==========
    ws11 = wb.create_sheet("综合优选")
    df_best = all_metrics_df[
        (all_metrics_df['夏普比率'] > 0.5) &
        (all_metrics_df['胜率%'] > 60) &
        (all_metrics_df['最大回撤%'] < 3) &
        (all_metrics_df['年化收益率%'] > 3)
    ].sort_values('2周预期年化%', ascending=False)
    write_dataframe_to_sheet(ws11, df_best, header_fill, header_font, thin_border, center_alignment)
    logger.info(f"综合优选: {len(df_best)} 条")

    # 保存
    wb.save(output_file)
    logger.info(f"报告已保存到: {output_file}")


def write_dataframe_to_sheet(ws, df, header_fill, header_font, thin_border, center_alignment):
    """将DataFrame写入工作表"""
    if df.empty:
        ws.cell(row=1, column=1, value="无符合条件的数据")
        return

    # 写入表头
    columns = list(df.columns)
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_alignment

    # 写入数据
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            # 数值列居中
            col_name = columns[col_idx - 1]
            if col_name not in ['产品代码', '产品名称']:
                cell.alignment = center_alignment

    # 调整列宽
    for col_idx, col_name in enumerate(columns, 1):
        if col_name == '产品名称':
            ws.column_dimensions[get_column_letter(col_idx)].width = 35
        elif col_name == '产品代码':
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
        elif col_name == '银行':
            ws.column_dimensions[get_column_letter(col_idx)].width = 10
        else:
            ws.column_dimensions[get_column_letter(col_idx)].width = 12

    # 冻结首行和前三列
    ws.freeze_panes = 'D2'


def main():
    """主函数"""
    print("=" * 70)
    print("              理财产品综合筛选器")
    print("=" * 70)
    print(f"开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"数据来源: {NAV_DB_FILE}")
    print(f"输出文件: {OUTPUT_FILE}")
    print()

    print("计算指标包括：")
    print("  - 连续上涨天数 / 连续两日上涨")
    print("  - 年化收益率 / 近5日年化收益率")
    print("  - 2周预期年化收益率 (今日买入明日确认持有2周)")
    print("  - 日波动率 / 年化波动率")
    print("  - 夏普比率 (风险调整后收益)")
    print("  - 胜率 (正收益天数占比)")
    print("  - 最大回撤")
    print("  - 变异系数 (收益稳定性)")
    print("  - 动量指标 (近期vs前期)")
    print()

    # 加载数据
    nav_data = load_nav_database()

    # 计算各银行的指标
    all_metrics = []

    for bank_name, df in nav_data.items():
        print(f"\n计算 {bank_name} 指标...")
        metrics_df = calculate_metrics(df, bank_name, min_dates=5)

        if metrics_df is not None:
            all_metrics.append(metrics_df)
            print(f"  - 有效产品数: {len(metrics_df)}")

    if not all_metrics:
        print("没有有效数据可处理")
        return

    # 合并所有数据
    all_metrics_df = pd.concat(all_metrics, ignore_index=True)
    print(f"\n合计产品数: {len(all_metrics_df)}")

    # 生成报告
    create_screening_excel(all_metrics_df, OUTPUT_FILE)

    # 打印统计摘要
    print("\n" + "=" * 70)
    print("                    筛选统计摘要")
    print("=" * 70)
    print(f"{'筛选条件':<20} {'产品数量':>10}")
    print("-" * 35)
    print(f"{'全部产品':<20} {len(all_metrics_df):>10}")
    print(f"{'连续两日上涨':<20} {len(all_metrics_df[all_metrics_df['连续两日上涨']==1]):>10}")
    print(f"{'高夏普比率(>1)':<20} {len(all_metrics_df[all_metrics_df['夏普比率']>1]):>10}")
    print(f"{'低波动率(<5%)':<20} {len(all_metrics_df[all_metrics_df['年化波动率%']<5]):>10}")
    print(f"{'低回撤(<1%)':<20} {len(all_metrics_df[all_metrics_df['最大回撤%']<1]):>10}")
    print(f"{'高胜率(>80%)':<20} {len(all_metrics_df[all_metrics_df['胜率%']>80]):>10}")
    print(f"{'高收益(>10%)':<20} {len(all_metrics_df[all_metrics_df['年化收益率%']>10]):>10}")
    print("=" * 70)
    print(f"\n报告已保存: {OUTPUT_FILE}")

    return OUTPUT_FILE


if __name__ == "__main__":
    main()
