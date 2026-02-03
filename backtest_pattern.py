# -*- coding: utf-8 -*-
"""
前瞻性收益释放购买策略 — 回测验证程序 V3

核心思路转变: 预测不用于降低买入阈值, 而用于辅助决策——
  ① 优选: 多个3%+信号竞争有限仓位时, 优先选预测即将释放的产品
  ② 智持: 收益回落到卖出线时, 若预测几天后将再释放, 延迟卖出
  ③ 节前: 长假前通常不买入, 但若预测节后立即释放, 允许提前布局

三模式:
  Mode A — 纯反应式基线 (精确复现 backtest_v6)
  Mode B — 仅优选 (建议①, 不改阈值, 只改排名)
  Mode C — 全部三策略 (建议①+②+③)

Walk-Forward: 180天训练, 每30天重建
输出: 控制台对比 + 前瞻策略回测报告.xlsx (5个sheet)
"""

import os, sys, re, warnings, logging, bisect
import pandas as pd
import numpy as np
from datetime import datetime
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple
from scipy import stats as scipy_stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

warnings.filterwarnings('ignore')

# ============================================================
# 配置
# ============================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "净值数据库.xlsx")
OUTPUT_PATH = os.path.join(BASE_DIR, "前瞻策略回测报告.xlsx")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('backtest_pattern.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# ============================================================
# 策略参数 (与 V5.4 对齐)
# ============================================================
信号阈值_买入 = 3.5
信号阈值_卖出 = 2.0
信号阈值_库 = 3.0       # V5.4: 产品库信号检测用宽阈值(3.0%), 买入决策用严阈值(3.5%)
需要突破信号 = False
持有天数_评估 = 7
成功标准 = 2.5
最低成功率 = 30.0
最少历史信号 = 2
最短持续天数 = 2
最大持有天数_评估 = 30
最长赎回天数 = 14

回测开始日 = '2025-06-01'
回测结束日 = '2026-01-30'
初始资金 = 100_000_000
最大持仓数 = 6
单仓基础 = 初始资金 / 6
最短持有交易日_赎回 = 3
最大持有交易日 = 20
长假阈值_天 = 3
产品库重建间隔 = 7

# 规律学习参数
规律学习窗口天数 = 180
规律重建间隔天数 = 30
释放识别阈值 = 2.5
最低置信度 = 0.4
预测窗口天数 = 10
周期CV阈值 = 0.4

# ===== V3 三建议参数 =====
# 建议① 优选: Alpha 加成
ALPHA_预测加成 = 0.30

# 建议② 智持: 延迟卖出
延迟卖出_最低收益 = 0.0       # 收益>此值才允许延迟 (0=非负就行)
延迟卖出_预测窗口 = 4         # 释放预测在N个交易日内才延迟
延迟卖出_置信度 = 0.4         # 预测置信度门槛
延迟卖出_最大持有余量 = 5     # hold_td < 最大持有交易日-此值 才允许延迟

# 建议③ 节前买入
节前买入_实时阈值 = 2.0       # 节前买入要求的最低实时收益
节前买入_置信度 = 0.5         # 节前买入预测置信度
节前买入_最大产品数 = 5       # 节前最多买入产品数


# ============================================================
# 工具函数
# ============================================================

def is_date_col(col_name):
    if not isinstance(col_name, str):
        return False
    return len(col_name) == 10 and col_name[4] == '-' and col_name[7] == '-'


def 判断产品流动性(name):
    if '日开' in name or '天天' in name or '每日' in name:
        return '日开', 0
    m = re.search(r'(\d+)\s*天持有期', name)
    if m: return f'{int(m.group(1))}天持有期', int(m.group(1))
    m = re.search(r'最短持有\s*(\d+)\s*天', name)
    if m: return f'{int(m.group(1))}天持有', int(m.group(1))
    m = re.search(r'周期\s*(\d+)\s*天', name)
    if m: return f'{int(m.group(1))}天周期', int(m.group(1))
    m = re.search(r'(\d+)\s*个?月', name)
    if m and '半年' not in name:
        months = int(m.group(1))
        if months <= 6: return f'{months}个月', months * 30
    if '周开' in name: return '周开', 7
    if '日申季赎' in name: return '季赎', 90
    if '月开' in name: return '月开', 30
    if '季开' in name or '季度开' in name: return '季开', 90
    if '半年' in name: return '半年', 180
    if '年开' in name: return '年开', 365
    if '封闭' in name: return '封闭', 999
    if '定开' in name: return '定开', 999
    m = re.search(r'(\d+)\s*Y\s*持有', name)
    if m: return f'{int(m.group(1))}年持有', int(m.group(1)) * 365
    if re.search(r'20[2-5]\d', name) and '信颐' in name:
        return '目标日期', 999
    return '未知', 999


# ============================================================
# 数据结构
# ============================================================

class ProductData:
    __slots__ = ['bank', 'code', 'name', 'liquidity', 'redeem_days',
                 'nav_dates', 'navs', 'ret_dates', 'rets',
                 'ret_date_idx', 'signals']
    def __init__(self, bank, code, name, liquidity, redeem_days):
        self.bank = bank
        self.code = code
        self.name = name
        self.liquidity = liquidity
        self.redeem_days = redeem_days
        self.nav_dates = []
        self.navs = {}
        self.ret_dates = []
        self.rets = {}
        self.ret_date_idx = {}
        self.signals = []


class Position:
    __slots__ = ['product_key', 'signal_date', 'confirm_date', 'buy_nav',
                 'amount', 'entry_return', 'alpha_score',
                 'sell_date', 'sell_nav', 'pnl', 'sell_reason',
                 'signal_type', 'is_preholiday_buy']
    def __init__(self, product_key, signal_date, confirm_date, buy_nav, amount,
                 entry_return=0.0, alpha_score=0.0, signal_type='reactive',
                 is_preholiday_buy=False):
        self.product_key = product_key
        self.signal_date = signal_date
        self.confirm_date = confirm_date
        self.buy_nav = buy_nav
        self.amount = amount
        self.entry_return = entry_return
        self.alpha_score = alpha_score
        self.sell_date = None
        self.sell_nav = None
        self.pnl = 0.0
        self.sell_reason = ''
        self.signal_type = signal_type   # 'reactive', 'predicted', 'preholiday'
        self.is_preholiday_buy = is_preholiday_buy


# ============================================================
# 模块1: PatternLearner
# ============================================================

@dataclass
class ReleasePattern:
    product_key: tuple
    period_days: float = 0.0
    period_cv: float = 1.0
    has_period: bool = False
    phase_dist: list = field(default_factory=lambda: [0.0]*4)
    phase_pvalue: float = 1.0
    top_phase: int = 0
    weekday_dist: list = field(default_factory=lambda: [0.0]*5)
    weekday_pvalue: float = 1.0
    top_weekday: int = 0
    confidence: float = 0.0
    n_events: int = 0
    last_release_date: str = ''
    last_release_window_end: str = ''
    avg_window_days: float = 3.0      # 平均释放窗口持续天数


class PatternLearner:
    PHASE_RANGES = [(1, 7), (8, 14), (15, 21), (22, 31)]

    def learn(self, product_key, ret_dates, rets, as_of_date,
              window_days=180):
        cutoff = (pd.Timestamp(as_of_date) - pd.Timedelta(days=window_days)).strftime('%Y-%m-%d')
        wdates = [d for d in ret_dates if cutoff <= d < as_of_date]
        if len(wdates) < 10:
            return None

        starts = []
        prev_below = True
        for d in wdates:
            r = rets.get(d, 0)
            if r > 释放识别阈值:
                if prev_below:
                    starts.append(d)
                prev_below = False
            else:
                prev_below = True

        if len(starts) < 2:
            return None

        pat = ReleasePattern(product_key=product_key)
        pat.n_events = len(starts)
        pat.last_release_date = starts[-1]

        # 计算每个释放事件的窗口持续天数
        window_durations = []
        for s_date in starts:
            si = wdates.index(s_date)
            ei = si
            for j in range(si, len(wdates)):
                if rets.get(wdates[j], 0) > 释放识别阈值 * 0.5:
                    ei = j
                else:
                    break
            dur = (pd.Timestamp(wdates[ei]) - pd.Timestamp(s_date)).days + 1
            window_durations.append(max(dur, 1))
        pat.avg_window_days = float(np.mean(window_durations)) if window_durations else 3.0

        # 最后一个释放窗口的结束日
        last_si = wdates.index(starts[-1])
        last_ei = last_si
        for j in range(last_si, len(wdates)):
            if rets.get(wdates[j], 0) > 释放识别阈值 * 0.5:
                last_ei = j
            else:
                break
        pat.last_release_window_end = wdates[last_ei]

        # 周期
        gaps = []
        for i in range(1, len(starts)):
            g = (pd.Timestamp(starts[i]) - pd.Timestamp(starts[i-1])).days
            if 3 <= g <= 120:
                gaps.append(g)
        if len(gaps) >= 2:
            med = float(np.median(gaps))
            cv = float(np.std(gaps)) / med if med > 0 else 1.0
            pat.period_days = med
            pat.period_cv = cv
            pat.has_period = cv < 周期CV阈值
        elif len(gaps) == 1:
            pat.period_days = float(gaps[0])
            pat.period_cv = 0.5
            pat.has_period = True

        # 阶段
        pc = [0]*4
        for d in starts:
            day = pd.Timestamp(d).day
            for pi, (lo, hi) in enumerate(self.PHASE_RANGES):
                if lo <= day <= hi:
                    pc[pi] += 1; break
        ne = sum(pc)
        if ne >= 3:
            _, pv = scipy_stats.chisquare(pc, f_exp=[ne/4.0]*4)
            pat.phase_dist = [c/ne for c in pc]
            pat.phase_pvalue = pv
            pat.top_phase = int(np.argmax(pc))
        else:
            pat.phase_dist = [c/max(ne,1) for c in pc]
            pat.phase_pvalue = 1.0
            pat.top_phase = int(np.argmax(pc)) if ne > 0 else 0

        # 星期
        wc = [0]*5
        for d in starts:
            wd = pd.Timestamp(d).weekday()
            if wd < 5: wc[wd] += 1
        nw = sum(wc)
        if nw >= 3:
            _, pw = scipy_stats.chisquare(wc, f_exp=[nw/5.0]*5)
            pat.weekday_dist = [c/nw for c in wc]
            pat.weekday_pvalue = pw
            pat.top_weekday = int(np.argmax(wc))
        else:
            pat.weekday_dist = [c/max(nw,1) for c in wc]
            pat.weekday_pvalue = 1.0
            pat.top_weekday = int(np.argmax(wc)) if nw > 0 else 0

        # 置信度
        conf = 0.0
        if pat.has_period and pat.period_cv < 0.3: conf += 0.4
        elif pat.has_period and pat.period_cv < 0.4: conf += 0.25
        if pat.phase_pvalue < 0.05: conf += 0.3
        if pat.weekday_pvalue < 0.05: conf += 0.15
        if pat.n_events >= 5: conf += 0.15
        pat.confidence = min(conf, 1.0)
        return pat


# ============================================================
# 模块2: ForwardPredictor (V2 修正版)
# ============================================================

@dataclass
class Prediction:
    product_key: tuple
    predicted_date: str
    confidence: float
    source: str
    td_until: int = 0
    predicted_end_date: str = ''


class ForwardPredictor:

    def predict(self, pattern, current_date, all_sim_dates, date_to_idx):
        if pattern.confidence < 最低置信度:
            return []
        cur_ts = pd.Timestamp(current_date)
        horizon = cur_ts + pd.Timedelta(days=预测窗口天数)
        cur_idx = date_to_idx.get(current_date)
        if cur_idx is None:
            return []

        preds = {}

        # 路径A: 周期 (向前推进修正)
        if pattern.has_period and pattern.last_release_window_end:
            last_ts = pd.Timestamp(pattern.last_release_window_end)
            period = pd.Timedelta(days=max(int(pattern.period_days), 5))
            pt = last_ts + period
            for _ in range(30):
                if pt >= cur_ts - pd.Timedelta(days=2):
                    break
                pt += period
            conf_a = (1 - pattern.period_cv) * 0.7
            for off in range(-2, 3):
                ct = pt + pd.Timedelta(days=off)
                if cur_ts < ct <= horizon:
                    pd_str = ct.strftime('%Y-%m-%d')
                    pi = bisect.bisect_left(all_sim_dates, pd_str)
                    if pi >= len(all_sim_dates): continue
                    td = pi - cur_idx
                    if td < 1: continue
                    if pd_str not in preds:
                        preds[pd_str] = Prediction(pattern.product_key, pd_str, conf_a, 'period', td)
                    else:
                        preds[pd_str].confidence = min(preds[pd_str].confidence + conf_a, 1.0)
                        preds[pd_str].source = 'both'
                        preds[pd_str].td_until = min(preds[pd_str].td_until, td)

        # 路径B: 阶段 (中点 + 宽窗口)
        if pattern.phase_pvalue < 0.15:
            tp = pattern.top_phase
            lo, hi = PatternLearner.PHASE_RANGES[tp]
            mid = (lo + hi) // 2
            prob = pattern.phase_dist[tp] if tp < len(pattern.phase_dist) else 0.25
            conf_b = prob * 0.6
            for mo in range(2):
                ref = cur_ts + pd.DateOffset(months=mo)
                try:
                    mt = pd.Timestamp(f'{ref.year}-{ref.month:02d}-{min(mid,28):02d}')
                except ValueError:
                    continue
                ws = mt - pd.Timedelta(days=2)
                we = mt + pd.Timedelta(days=3)
                if we <= cur_ts or ws > horizon: continue
                pd_str = mt.strftime('%Y-%m-%d')
                pi = bisect.bisect_left(all_sim_dates, pd_str)
                if pi >= len(all_sim_dates): continue
                td = pi - cur_idx
                if td < 1: continue
                if pd_str not in preds:
                    preds[pd_str] = Prediction(pattern.product_key, pd_str, conf_b, 'phase', td)
                else:
                    preds[pd_str].confidence = min(preds[pd_str].confidence + conf_b, 1.0)
                    preds[pd_str].source = 'both'
                    preds[pd_str].td_until = min(preds[pd_str].td_until, td)

        # 为每个预测计算释放结束日 = 预测释放日 + 平均窗口天数
        avg_win = max(int(round(pattern.avg_window_days)), 1)
        for p in preds.values():
            end_ts = pd.Timestamp(p.predicted_date) + pd.Timedelta(days=avg_win - 1)
            p.predicted_end_date = end_ts.strftime('%Y-%m-%d')

        results = [p for p in preds.values() if p.confidence >= 0.25]
        results.sort(key=lambda x: x.td_until)
        return results

    def rank_products(self, patterns, current_date, all_sim_dates, date_to_idx):
        """按预测紧迫度排名: score = confidence / sqrt(td_until)"""
        rankings = []
        for key, pat in patterns.items():
            if pat.confidence < 最低置信度: continue
            ps = self.predict(pat, current_date, all_sim_dates, date_to_idx)
            if not ps: continue
            best_s, best_p = 0.0, None
            for p in ps:
                if p.td_until < 1 or p.td_until > 8: continue
                s = p.confidence / max(p.td_until, 1) ** 0.5
                if s > best_s: best_s, best_p = s, p
            if best_p and best_s > 0.1:
                rankings.append((key, best_s, best_p))
        rankings.sort(key=lambda x: -x[1])
        return rankings

    def has_upcoming_release(self, pattern, current_date, all_sim_dates, date_to_idx,
                              max_td=4, min_conf=0.4):
        """检查是否有即将到来的释放 (用于延迟卖出判断)"""
        ps = self.predict(pattern, current_date, all_sim_dates, date_to_idx)
        for p in ps:
            if 1 <= p.td_until <= max_td and p.confidence >= min_conf:
                return True, p
        return False, None


# ============================================================
# 模块3: PatternBacktestEngine
# ============================================================

class PatternBacktestEngine:
    def __init__(self):
        self.products = {}
        self.all_sim_dates = []
        self.date_to_idx = {}

    def load_data(self):
        logger.info("加载净值数据库...")
        xlsx = pd.ExcelFile(DB_PATH)
        all_dates = set()
        total = 0
        for sheet in xlsx.sheet_names:
            logger.info(f"  加载 [{sheet}]...")
            df = pd.read_excel(xlsx, sheet_name=sheet, dtype=str)
            df.columns = [str(c).strip() for c in df.columns]
            if 'level_0' in df.columns:
                df = df.rename(columns={'level_0': '产品代码', 'level_1': '产品名称'})
            date_cols = sorted([c for c in df.columns if is_date_col(c)])
            if len(date_cols) < 10: continue
            all_dates.update(date_cols)
            count = 0
            for _, row in df.iterrows():
                code = row.get('产品代码')
                name = row.get('产品名称', '')
                if pd.isna(code): continue
                code = str(code).strip()
                name = str(name).strip() if pd.notna(name) else ''
                _, rd = 判断产品流动性(name)
                if rd > 最长赎回天数: continue
                nps = []
                for d in date_cols:
                    try:
                        v = float(row[d])
                        if not np.isnan(v) and v > 0: nps.append((d, v))
                    except (ValueError, TypeError): pass
                if len(nps) < 10: continue
                key = (sheet, code)
                p = ProductData(sheet, code, name, '', rd)
                p.nav_dates = [x[0] for x in nps]
                p.navs = {x[0]: x[1] for x in nps}
                for i in range(1, len(nps)):
                    d0, n0 = nps[i-1]; d1, n1 = nps[i]
                    gap = (pd.Timestamp(d1) - pd.Timestamp(d0)).days
                    if 0 < gap <= 60 and n0 > 0:
                        p.rets[d1] = (n1/n0 - 1) / gap * 365 * 100
                p.ret_dates = sorted(p.rets.keys())
                p.ret_date_idx = {d: i for i, d in enumerate(p.ret_dates)}
                if len(p.ret_dates) < 10: continue
                self._precompute_signals(p)
                self.products[key] = p
                count += 1
            total += count
            logger.info(f"    有效产品: {count}")
        self.all_sim_dates = sorted(d for d in all_dates if 回测开始日 <= d < 回测结束日)
        self.date_to_idx = {d: i for i, d in enumerate(self.all_sim_dates)}
        logger.info(f"加载完成: {total}个产品, {len(self.all_sim_dates)}个回测日")

    def _precompute_signals(self, product):
        rd = product.ret_dates; rets = product.rets; n = len(rd)
        sigs = []
        for i in range(1, n - 持有天数_评估):
            if rets[rd[i]] > 信号阈值_库 and rets[rd[i-1]] <= 信号阈值_库:
                hr = [rets[rd[i+1+k]] for k in range(持有天数_评估)]
                avg = float(np.mean(hr)); ee = rd[i + 持有天数_评估]
                ps = 0
                for j in range(i, min(i+20, n)):
                    if rets[rd[j]] > 信号阈值_库 * 0.8: ps += 1
                    else: break
                cd = ((pd.Timestamp(rd[i+ps-1]) - pd.Timestamp(rd[i])).days + 1) if ps > 0 else 0
                sigs.append({'date': rd[i], 'eval_end': ee, 'avg_return': avg,
                             'success': avg > 成功标准, 'persist_days': cd})
        product.signals = sigs

    def build_library(self, as_of_date):
        lib = {}
        for key, p in self.products.items():
            valid = [s for s in p.signals if s['eval_end'] < as_of_date]
            if len(valid) < 最少历史信号: continue
            sn = sum(1 for s in valid if s['success'])
            rate = sn / len(valid) * 100
            if rate < 最低成功率: continue
            ap = float(np.mean([s['persist_days'] for s in valid]))
            if ap < 最短持续天数 or ap > 最大持有天数_评估: continue
            rl = [s['avg_return'] for s in valid]
            ar = float(np.mean(rl))
            sd = float(np.std(rl)) if len(rl) > 1 else 1.0
            sh = ar / max(sd, 0.5)       # V5.4: Sharpe floor = 0.5
            lib[key] = {'success_rate': round(rate,1), 'signal_count': len(valid),
                        'avg_return': round(ar,2), 'avg_persist': round(ap,1),
                        'sharpe_like': round(sh,2)}
        return lib

    def get_latest_nav(self, product, as_of_date):
        idx = bisect.bisect_right(product.nav_dates, as_of_date) - 1
        if idx >= 0:
            d = product.nav_dates[idx]
            return d, product.navs[d]
        return None, None

    def _td_held(self, cd, dd):
        ci = self.date_to_idx.get(cd); di = self.date_to_idx.get(dd)
        if ci is not None and di is not None: return di - ci
        return (pd.Timestamp(dd) - pd.Timestamp(cd)).days

    def _is_pre_holiday(self, di):
        if di+1 >= len(self.all_sim_dates): return True
        gap = (pd.Timestamp(self.all_sim_dates[di+1]) - pd.Timestamp(self.all_sim_dates[di])).days
        return gap > 长假阈值_天

    # ==================== 单模式回测 ====================

    def run_single_mode(self, mode):
        desc = {'A': '纯反应式基线', 'B': '优选(Alpha加成)', 'C': '优选+智持+节前'}
        logger.info(f"\n{'='*60}\n  Mode {mode} — {desc[mode]}\n{'='*60}")

        # --- 功能开关 ---
        use_alpha_boost = mode in ('B', 'C')     # 建议①
        use_sell_delay  = mode == 'C'             # 建议②
        use_preholiday  = mode == 'C'             # 建议③

        # --- 状态 ---
        library = {}; library_date = None
        positions = {}; pending_buys = []; pending_sells = []
        pending_buy_keys = set(); pending_sell_keys = set()
        closed_trades = []; cash = 初始资金
        daily_values = []; trade_log = []

        learner = PatternLearner(); predictor = ForwardPredictor()
        patterns = {}; pattern_build_date = None
        prediction_log = []

        # 诊断计数器
        n_alpha_boost = 0        # 建议① 被加成的买入次数
        n_sell_delayed = 0       # 建议② 延迟卖出次数
        n_delay_success = 0      # 建议② 延迟后确实释放的次数
        n_preholiday_buy = 0     # 建议③ 节前买入次数
        n_preholiday_profit = 0  # 建议③ 节前买入盈利次数

        def _rebuild_lib(date):
            nonlocal library, library_date
            library = self.build_library(date); library_date = date

        def _rebuild_pat(date):
            nonlocal patterns, pattern_build_date
            if mode == 'A': return
            new = {}
            for key in library:
                p = self.products[key]
                pat = learner.learn(key, p.ret_dates, p.rets, date, 规律学习窗口天数)
                if pat and pat.confidence >= 最低置信度:
                    new[key] = pat
            patterns = new; pattern_build_date = date
            logger.info(f"  [{date}] Mode {mode}: 规律 {len(patterns)}个产品")

        def _proc_buys(date):
            nonlocal cash
            nb = 0
            for key, sig_d, reason, sig_type, is_ph in pending_buys:
                if key in positions: pending_buy_keys.discard(key); continue
                prod = self.products[key]
                nav = prod.navs.get(date)
                if nav is None: _, nav = self.get_latest_nav(prod, date)
                if nav is None: pending_buy_keys.discard(key); continue
                # V5.4 动态仓位: 按当前组合价值计算
                current_value = cash + sum(
                    p.amount * (self.products[k].navs.get(date, p.buy_nav) / p.buy_nav)
                    for k, p in positions.items()
                )
                dynamic_size = current_value / 最大持仓数
                amt = min(dynamic_size, cash)
                if amt < 单仓基础 * 0.5: pending_buy_keys.discard(key); continue
                er = prod.rets.get(sig_d, 信号阈值_买入)
                cash -= amt
                positions[key] = Position(key, sig_d, date, nav, amt,
                    entry_return=er, alpha_score=0.5,
                    signal_type=sig_type, is_preholiday_buy=is_ph)
                pending_buy_keys.discard(key); nb += 1
                trade_log.append({'date': date, 'action': '买入确认',
                    'bank': prod.bank, 'code': prod.code, 'name': prod.name,
                    'nav': nav, 'amount': amt, 'pnl': 0, 'hold_days': 0,
                    'reason': f'T+1确认({sig_d}) {reason}',
                    'signal_type': sig_type})
            pending_buys.clear()
            return nb

        def _proc_sells(date):
            nonlocal cash
            ns = 0
            for key, sig_d, reason in pending_sells:
                if key not in positions: pending_sell_keys.discard(key); continue
                pos = positions.pop(key); prod = self.products[key]
                nav = prod.navs.get(date)
                if nav is None: _, nav = self.get_latest_nav(prod, date)
                if nav is None: nav = pos.buy_nav
                pos.sell_date = date; pos.sell_nav = nav
                pos.pnl = pos.amount * (nav / pos.buy_nav - 1)
                pos.sell_reason = reason
                cash += pos.amount + pos.pnl
                closed_trades.append(pos); pending_sell_keys.discard(key); ns += 1
                htd = self._td_held(pos.confirm_date, date)
                trade_log.append({'date': date, 'action': '赎回到账',
                    'bank': prod.bank, 'code': prod.code, 'name': prod.name,
                    'nav': nav, 'amount': pos.amount + pos.pnl, 'pnl': pos.pnl,
                    'hold_days': htd, 'reason': f'T+1到账({sig_d}) {reason}',
                    'signal_type': pos.signal_type})
            pending_sells.clear()
            return ns

        def _check_sells(date):
            nonlocal n_sell_delayed, n_delay_success
            for key in list(positions):
                if key in pending_sell_keys: continue
                pos = positions[key]; prod = self.products[key]
                htd = self._td_held(pos.confirm_date, date)
                if htd < 最短持有交易日_赎回: continue
                ret = prod.rets.get(date)

                # ① 硬阈值
                if ret is not None and ret <= 信号阈值_卖出:
                    # === 建议② 智持: 预测即将释放 → 延迟卖出 ===
                    if (use_sell_delay and key in patterns
                        and ret > 延迟卖出_最低收益
                        and htd < 最大持有交易日 - 延迟卖出_最大持有余量):
                        upcoming, pred = predictor.has_upcoming_release(
                            patterns[key], date, self.all_sim_dates, self.date_to_idx,
                            max_td=延迟卖出_预测窗口, min_conf=延迟卖出_置信度)
                        if upcoming:
                            n_sell_delayed += 1
                            prediction_log.append({
                                'date': date, 'product_key': key,
                                'predicted_release': pred.predicted_date if pred else '',
                                'predicted_release_end': pred.predicted_end_date if pred else '',
                                'confidence': pred.confidence if pred else 0,
                                'source': 'sell_delay', 'action': 'delay_sell',
                                'ret_at_event': ret})
                            continue  # 延迟卖出, 不挂卖单
                    pending_sells.append((key, date,
                        f'收益{ret:.1f}%≤{信号阈值_卖出}%(持有{htd}td)'))
                    pending_sell_keys.add(key); continue

                # ② 最大持有
                if htd >= 最大持有交易日:
                    pending_sells.append((key, date, f'持有{htd}td达上限'))
                    pending_sell_keys.add(key)

        def _find_buys(date, slots, di):
            """通用买入扫描, 根据 mode 开关执行不同逻辑"""
            nonlocal n_alpha_boost, n_preholiday_buy
            is_ph = self._is_pre_holiday(di)

            # === 建议③ 节前买入 ===
            if is_ph:
                if not use_preholiday:
                    return 0  # Mode A/B: 节前不买
                # Mode C: 节前允许买入预测即将释放的产品
                rankings = predictor.rank_products(patterns, date,
                    self.all_sim_dates, self.date_to_idx)
                queued = 0
                for key, score, pred in rankings[:节前买入_最大产品数]:
                    if queued >= slots: break
                    if key in positions or key in pending_buy_keys or key in pending_sell_keys:
                        continue
                    if key not in library: continue
                    prod = self.products[key]
                    r = prod.rets.get(date)
                    if r is None or r <= 节前买入_实时阈值: continue
                    if pred.confidence < 节前买入_置信度: continue
                    avg_ret = library[key].get('avg_return', 1.0)
                    reason = (f'[节前] 均收{avg_ret:.1f}% 收益{r:.1f}% '
                              f'conf={pred.confidence:.2f} pred={pred.predicted_date}~{pred.predicted_end_date}')
                    pending_buys.append((key, date, reason, 'preholiday', True))
                    pending_buy_keys.add(key); queued += 1; n_preholiday_buy += 1
                    prediction_log.append({
                        'date': date, 'product_key': key,
                        'predicted_release': pred.predicted_date,
                        'predicted_release_end': pred.predicted_end_date,
                        'confidence': pred.confidence,
                        'source': 'preholiday', 'action': 'buy_preholiday',
                        'ret_at_event': r})
                return queued

            if slots <= 0: return 0

            # === 建议① 优选: Alpha 加成 ===
            predicted_scores = {}
            if use_alpha_boost:
                rankings = predictor.rank_products(patterns, date,
                    self.all_sim_dates, self.date_to_idx)
                predicted_scores = {r[0]: (r[1], r[2]) for r in rankings}

            # 标准扫描 — 买入阈值 3.5%, 排名用 当日收益×历史平均收益 (V5.4)
            signals = []
            for key, info in library.items():
                if key in positions or key in pending_buy_keys or key in pending_sell_keys:
                    continue
                prod = self.products[key]
                r = prod.rets.get(date)
                if r is None or r <= 信号阈值_买入: continue
                avg_ret = info.get('avg_return', 1.0)
                composite = r * avg_ret           # V5.4 排名: 当日收益 × 历史平均收益
                boosted = False
                if key in predicted_scores:
                    composite *= (1 + predicted_scores[key][0] * ALPHA_预测加成)
                    boosted = True
                signals.append((key, composite, r, info['success_rate'], boosted,
                                predicted_scores.get(key), avg_ret))

            # V5.4 排序: composite 降序, 同分按 success_rate 降序
            signals.sort(key=lambda x: (-x[1], -x[3]))

            queued = 0
            for key, composite, ret, sr, boosted, pred_info, avg_ret in signals[:slots]:
                tag = '★' if boosted else ' '
                pred_str = ''
                if pred_info:
                    pred_str = (f' pred={pred_info[1].predicted_date}'
                                f'~{pred_info[1].predicted_end_date} conf={pred_info[0]:.2f}')
                reason = f'{tag}成功率{sr}% 收益{ret:.1f}% 均收{avg_ret:.1f}%{pred_str}'
                sig_type = 'predicted' if boosted else 'reactive'
                pending_buys.append((key, date, reason, sig_type, False))
                pending_buy_keys.add(key); queued += 1
                if boosted: n_alpha_boost += 1
                if pred_info:
                    prediction_log.append({
                        'date': date, 'product_key': key,
                        'predicted_release': pred_info[1].predicted_date,
                        'predicted_release_end': pred_info[1].predicted_end_date,
                        'confidence': pred_info[0],
                        'source': 'alpha_boost', 'action': 'buy_boosted',
                        'ret_at_event': ret})
            return queued

        def _record(date):
            pv = 0
            for key, pos in positions.items():
                prod = self.products[key]
                _, nav = self.get_latest_nav(prod, date)
                pv += pos.amount * (nav / pos.buy_nav) if nav else pos.amount
            tot = cash + pv
            if daily_values and daily_values[-1][0] == date:
                daily_values[-1] = (date, tot, cash, pv)
            else:
                daily_values.append((date, tot, cash, pv))

        # =================== 主循环 ===================
        _rebuild_lib(回测开始日)
        logger.info(f"  Mode {mode} 产品库: {len(library)}个")
        if mode != 'A': _rebuild_pat(回测开始日)

        nd = len(self.all_sim_dates); bn = sn = 0
        for di, date in enumerate(self.all_sim_dates):
            if library_date is None or (pd.Timestamp(date) - pd.Timestamp(library_date)).days >= 产品库重建间隔:
                _rebuild_lib(date)
            if mode != 'A' and (pattern_build_date is None or
                    (pd.Timestamp(date) - pd.Timestamp(pattern_build_date)).days >= 规律重建间隔天数):
                _rebuild_pat(date)

            bn += _proc_buys(date)
            sn += _proc_sells(date)
            _check_sells(date)

            used = len(positions) + len(pending_buys)
            _find_buys(date, 最大持仓数 - used, di)

            # V5.4 满仓轮动: 替换最弱持仓
            if not self._is_pre_holiday(di):
                used2 = len(positions) + len(pending_buys)
                if used2 >= 最大持仓数:
                    sellable = []
                    for rk in positions:
                        if rk in pending_sell_keys: continue
                        rpos = positions[rk]
                        rhtd = self._td_held(rpos.confirm_date, date)
                        if rhtd < 最短持有交易日_赎回: continue
                        rr = self.products[rk].rets.get(date)
                        if rr is not None: sellable.append((rk, rr, rhtd))
                    if sellable:
                        sellable.sort(key=lambda x: x[1])
                        wk, wr, wh = sellable[0]
                        if wr < 信号阈值_买入:
                            best_cand = None; best_r = wr
                            for ck, cinfo in library.items():
                                if ck in positions or ck in pending_buy_keys or ck in pending_sell_keys:
                                    continue
                                cp = self.products[ck]
                                cr = cp.rets.get(date)
                                if cr is not None and cr > 信号阈值_买入 and cr > best_r + 1.0:
                                    best_cand = (ck, cr, cinfo['success_rate']); best_r = cr
                            if best_cand:
                                pending_sells.append((wk, date,
                                    f'轮动卖出(收益{wr:.1f}%→换{best_r:.1f}%持有{wh}td)'))
                                pending_sell_keys.add(wk)

            _record(date)

            if (di+1) % 50 == 0 or di == nd - 1:
                v = daily_values[-1][1]; r = (v/初始资金-1)*100
                logger.info(f"  Mode {mode} [{date}] {di+1}/{nd} | "
                            f"持仓{len(positions)} | 买{bn} 卖{sn} | "
                            f"{v/1e4:,.0f}万({r:+.2f}%)")

        # 平仓
        last = self.all_sim_dates[-1]
        for key in list(positions):
            pos = positions.pop(key); prod = self.products[key]
            _, nav = self.get_latest_nav(prod, last)
            if nav is None: nav = pos.buy_nav
            pos.sell_date = last; pos.sell_nav = nav
            pos.pnl = pos.amount * (nav/pos.buy_nav - 1)
            pos.sell_reason = '回测结束平仓'
            cash += pos.amount + pos.pnl; closed_trades.append(pos)
            htd = self._td_held(pos.confirm_date, last)
            trade_log.append({'date': last, 'action': '强制平仓',
                'bank': prod.bank, 'code': prod.code, 'name': prod.name,
                'nav': nav, 'amount': pos.amount+pos.pnl, 'pnl': pos.pnl,
                'hold_days': htd, 'reason': '回测结束平仓',
                'signal_type': pos.signal_type})
        _record(last)

        # 计算延迟卖出成功率
        for rec in prediction_log:
            if rec.get('action') != 'delay_sell': continue
            key = rec['product_key']; pd_str = rec.get('predicted_release', '')
            if key in self.products and pd_str:
                prod = self.products[key]; pts = pd.Timestamp(pd_str)
                hit = False
                for off in range(-2, 4):
                    cd = (pts + pd.Timedelta(days=off)).strftime('%Y-%m-%d')
                    r = prod.rets.get(cd)
                    if r is not None and r > 释放识别阈值: hit = True; break
                rec['hit'] = hit
                if hit: n_delay_success += 1
            else:
                rec['hit'] = False

        # 节前买入盈亏
        for t in closed_trades:
            if t.is_preholiday_buy and t.pnl > 0:
                n_preholiday_profit += 1

        # 预测命中率
        for rec in prediction_log:
            if rec.get('action') in ('delay_sell',): continue
            key = rec['product_key']; pd_str = rec.get('predicted_release', '')
            if key in self.products and pd_str:
                prod = self.products[key]; pts = pd.Timestamp(pd_str)
                hit = False
                for off in range(-2, 4):
                    cd = (pts + pd.Timedelta(days=off)).strftime('%Y-%m-%d')
                    r = prod.rets.get(cd)
                    if r is not None and r > 释放识别阈值: hit = True; break
                rec['hit'] = hit
            else:
                rec['hit'] = False

        # 汇总
        result = self._metrics(mode, daily_values, closed_trades, trade_log, prediction_log)
        result['daily_values'] = daily_values
        result['closed_trades'] = closed_trades
        result['trade_log'] = trade_log
        result['prediction_log'] = prediction_log
        result['n_alpha_boost'] = n_alpha_boost
        result['n_sell_delayed'] = n_sell_delayed
        result['n_delay_success'] = n_delay_success
        result['n_preholiday_buy'] = n_preholiday_buy
        result['n_preholiday_profit'] = n_preholiday_profit
        result['patterns_snapshot'] = {
            k: {'confidence': v.confidence, 'period_days': v.period_days,
                'period_cv': v.period_cv, 'has_period': v.has_period,
                'top_phase': v.top_phase, 'phase_pvalue': v.phase_pvalue,
                'top_weekday': v.top_weekday, 'weekday_pvalue': v.weekday_pvalue,
                'n_events': v.n_events, 'avg_window_days': v.avg_window_days}
            for k, v in patterns.items()
        }
        return result

    def _metrics(self, mode, dv, ct, tl, pl):
        if not dv:
            return {k: 0 for k in ['mode','ann_ret','max_dd','sharpe','n_trades',
                                    'win_rate','avg_hold','total_ret','final_value',
                                    'forward_hit_rate','pred_trades','pred_pnl',
                                    'std_trades','std_pnl']}
        fv = dv[-1][1]
        tr = (fv/初始资金 - 1)*100
        t0 = pd.Timestamp(self.all_sim_dates[0])
        t1 = pd.Timestamp(self.all_sim_dates[-1])
        yrs = (t1-t0).days / 365
        ar = ((fv/初始资金)**(1/yrs)-1)*100 if yrs > 0 else 0
        pk = 初始资金; md = 0
        for _, v, _, _ in dv:
            if v > pk: pk = v
            dd = (pk-v)/pk*100
            if dd > md: md = dd
        nt = len(ct); nw = sum(1 for t in ct if t.pnl > 0)
        wr = nw/nt*100 if nt else 0
        hds = [self._td_held(t.confirm_date, t.sell_date) for t in ct if t.sell_date]
        ah = float(np.mean(hds)) if hds else 0
        drs = []
        for i in range(1, len(dv)):
            pv = dv[i-1][1]; cv = dv[i][1]
            if pv > 0: drs.append(cv/pv - 1)
        if drs:
            rf = 0.025/252; ex = np.array(drs) - rf
            sp = float(np.mean(ex)/np.std(ex)*np.sqrt(252)) if np.std(ex) > 0 else 0
        else: sp = 0

        buy_preds = [p for p in pl if p.get('action','').startswith('buy')]
        hits = sum(1 for p in buy_preds if p.get('hit', False))
        hr = hits / len(buy_preds) * 100 if buy_preds else 0

        pred_t = [t for t in ct if t.signal_type in ('predicted','preholiday')]
        std_t = [t for t in ct if t.signal_type == 'reactive']

        return {
            'mode': mode, 'final_value': fv, 'total_ret': tr, 'ann_ret': ar,
            'max_dd': md, 'sharpe': sp, 'n_trades': nt, 'win_rate': wr,
            'avg_hold': ah, 'forward_hit_rate': hr,
            'n_predictions': len(buy_preds), 'n_hits': hits,
            'pred_trades': len(pred_t),
            'pred_pnl': sum(t.pnl for t in pred_t),
            'pred_wr': (sum(1 for t in pred_t if t.pnl>0)/len(pred_t)*100) if pred_t else 0,
            'std_trades': len(std_t),
            'std_pnl': sum(t.pnl for t in std_t),
            'std_wr': (sum(1 for t in std_t if t.pnl>0)/len(std_t)*100) if std_t else 0,
        }


# ============================================================
# 模块4: ComparisonReport
# ============================================================

class ComparisonReport:
    def __init__(self, engine, ra, rb, rc):
        self.engine = engine
        self.R = {'A': ra, 'B': rb, 'C': rc}

    def print_console(self):
        sep = '=' * 74
        print(f"\n{sep}")
        print("     前瞻策略回测对比 V3 (优选 · 智持 · 节前)")
        print(sep)
        print(f"  区间: {self.engine.all_sim_dates[0]} ~ {self.engine.all_sim_dates[-1]}  "
              f"({len(self.engine.all_sim_dates)}交易日)")
        print(f"  资金: {初始资金/1e4:,.0f}万  仓位: {最大持仓数}×{单仓基础/1e4:,.0f}万")
        print(f"  Mode A: 纯反应 (阈值{信号阈值_买入}%)")
        print(f"  Mode B: +优选 (Alpha加成{ALPHA_预测加成}, 阈值不变)")
        print(f"  Mode C: +优选+智持+节前 (全部三建议)")
        print()

        hdr = f"  {'指标':<20} {'A(反应)':<14} {'B(+优选)':<14} {'C(+全部)':<14}"
        print(hdr); print("  " + "-" * 58)

        rows = [
            ('年化收益率', 'ann_ret', '{:+.4f}%'),
            ('总收益率', 'total_ret', '{:+.4f}%'),
            ('最大回撤', 'max_dd', '{:.4f}%'),
            ('夏普比率', 'sharpe', '{:.4f}'),
            ('交易次数', 'n_trades', '{:d}'),
            ('胜率', 'win_rate', '{:.1f}%'),
            ('平均持有天数', 'avg_hold', '{:.1f}'),
            ('预测命中率', 'forward_hit_rate', '{:.1f}%'),
        ]
        for label, key, fmt in rows:
            vals = []
            for m in 'ABC':
                v = self.R[m].get(key, 0)
                if key == 'forward_hit_rate' and m == 'A':
                    vals.append('N/A')
                else:
                    try: vals.append(fmt.format(int(v) if 'd' in fmt else v))
                    except: vals.append(str(v))
            print(f"  {label:<20} {vals[0]:<14} {vals[1]:<14} {vals[2]:<14}")

        print()
        # 建议① 诊断
        print(f"  [建议① 优选]")
        for m in 'BC':
            r = self.R[m]
            print(f"    Mode {m}: Alpha加成买入 {r.get('n_alpha_boost',0)}次, "
                  f"预测入场{r.get('pred_trades',0)}笔 "
                  f"胜率{r.get('pred_wr',0):.1f}% "
                  f"盈亏{r.get('pred_pnl',0)/1e4:+,.2f}万")

        # 建议② 诊断
        print(f"  [建议② 智持]")
        rc = self.R['C']
        nd = rc.get('n_sell_delayed', 0)
        ns = rc.get('n_delay_success', 0)
        print(f"    延迟卖出 {nd}次, 其中释放成功 {ns}次 "
              f"({ns/nd*100:.0f}%命中)" if nd > 0 else f"    延迟卖出 0次")

        # 建议③ 诊断
        print(f"  [建议③ 节前]")
        np_ = rc.get('n_preholiday_buy', 0)
        npp = rc.get('n_preholiday_profit', 0)
        print(f"    节前买入 {np_}次, 盈利 {npp}次"
              + (f" ({npp/np_*100:.0f}%)" if np_ > 0 else ""))

        print()
        for m in 'BC':
            d = self.R[m]['ann_ret'] - self.R['A']['ann_ret']
            print(f"  Mode {m} vs A: 年化 {d:+.4f}%")

        best = max('ABC', key=lambda m: self.R[m]['ann_ret'])
        print(f"\n  最优: Mode {best} (年化 {self.R[best]['ann_ret']:+.4f}%)")
        if best != 'A':
            print(f"  建议: 有效! 超越基线 "
                  f"{self.R[best]['ann_ret']-self.R['A']['ann_ret']:+.4f}%")
        else:
            print(f"  建议: 预测辅助策略未超越基线, 保留纯反应式")
        print(sep)

    def generate_excel(self):
        logger.info("生成Excel报告...")
        wb = Workbook()
        hf = Font(bold=True, color="FFFFFF", size=11)
        tf = Font(bold=True, size=14)
        bf = PatternFill("solid", fgColor="1565C0")
        gf = PatternFill("solid", fgColor="2E7D32")
        df_ = PatternFill("solid", fgColor="37474F")
        rf_ = Font(color="CC0000", bold=True)
        gnf = Font(color="2E7D32", bold=True)
        ct = Alignment(horizontal="center", vertical="center")

        # Sheet 1: 总览
        ws1 = wb.active; ws1.title = "对比总览"
        ws1['A1'] = "前瞻策略回测 V3 (优选·智持·节前)"; ws1['A1'].font = tf
        ws1.append([])
        ws1.append(['指标', 'Mode A(纯反应)', 'Mode B(+优选)', 'Mode C(+全部)'])
        for c in ws1[3]: c.font = hf; c.fill = bf; c.alignment = ct

        kv = [('年化收益率(%)','ann_ret'),('总收益率(%)','total_ret'),
              ('最大回撤(%)','max_dd'),('夏普比率','sharpe'),
              ('交易次数','n_trades'),('胜率(%)','win_rate'),
              ('平均持有天数','avg_hold'),('预测命中率(%)','forward_hit_rate'),
              ('Alpha加成买入次数','n_alpha_boost'),
              ('延迟卖出次数','n_sell_delayed'),('延迟成功次数','n_delay_success'),
              ('节前买入次数','n_preholiday_buy'),('节前盈利次数','n_preholiday_profit'),
              ('预测入场笔数','pred_trades'),('预测入场盈亏(万)','pred_pnl'),
              ('标准入场笔数','std_trades'),('标准入场盈亏(万)','std_pnl')]
        for label, key in kv:
            row = [label]
            for m in 'ABC':
                v = self.R[m].get(key, 0)
                if key in ('forward_hit_rate','n_alpha_boost','n_sell_delayed',
                           'n_delay_success','n_preholiday_buy','n_preholiday_profit',
                           'pred_trades','pred_pnl','std_trades','std_pnl') and m == 'A':
                    row.append('N/A')
                elif key in ('pred_pnl','std_pnl'):
                    row.append(round(v/1e4, 2) if isinstance(v,(int,float)) else v)
                else:
                    row.append(round(v,4) if isinstance(v,float) else v)
            ws1.append(row)
        ws1.append([])
        for m in 'BC':
            d = self.R[m]['ann_ret'] - self.R['A']['ann_ret']
            ws1.append([f'Mode {m} vs A', f'{d:+.4f}%'])
        ws1.column_dimensions['A'].width = 22
        for cl in 'BCD': ws1.column_dimensions[cl].width = 20

        # Sheet 2: 每日净值
        ws2 = wb.create_sheet("每日净值")
        ws2.append(['日期','A(万)','B(万)','C(万)','A%','B%','C%'])
        for c in ws2[1]: c.font = hf; c.fill = df_; c.alignment = ct
        vms = {}
        for m in 'ABC':
            vm = {}
            for d, v, _, _ in self.R[m]['daily_values']: vm[d] = v
            vms[m] = vm
        ads = sorted(set().union(*(vm.keys() for vm in vms.values())))
        for d in ads:
            va=vms['A'].get(d,初始资金); vb=vms['B'].get(d,初始资金); vc=vms['C'].get(d,初始资金)
            ws2.append([d,round(va/1e4,2),round(vb/1e4,2),round(vc/1e4,2),
                        round((va/初始资金-1)*100,4),round((vb/初始资金-1)*100,4),
                        round((vc/初始资金-1)*100,4)])
        for i,w in enumerate([12,14,14,14,12,12,12],1):
            ws2.column_dimensions[get_column_letter(i)].width = w
        ws2.freeze_panes = 'A2'
        if len(ads) > 5:
            ch = LineChart(); ch.title = "三模式净值对比"; ch.style = 10
            ch.y_axis.title = "万"; ch.width = 30; ch.height = 15
            for ci in range(2,5):
                ch.add_data(Reference(ws2,min_col=ci,min_row=1,max_row=len(ads)+1),titles_from_data=True)
            ch.set_categories(Reference(ws2,min_col=1,min_row=2,max_row=len(ads)+1))
            colors = ['1565C0','FF6F00','2E7D32']
            for idx,s in enumerate(ch.series):
                s.graphicalProperties.line.width = 20000
                if idx < len(colors): s.graphicalProperties.line.solidFill = colors[idx]
            ws2.add_chart(ch, "I2")

        # Sheet 3: Mode C 交易明细
        ws3 = wb.create_sheet("Mode C交易明细")
        h3 = ['日期','操作','银行','代码','名称','信号类型','净值','金额(万)','盈亏(万)','持有天数','原因']
        ws3.append(h3)
        for c in ws3[1]: c.font = hf; c.fill = bf; c.alignment = ct
        for t in self.R['C']['trade_log']:
            ws3.append([t['date'],t['action'],t['bank'],t['code'],t['name'][:25],
                t.get('signal_type',''),
                round(t['nav'],6) if t.get('nav') else '',
                round(t['amount']/1e4,2) if t.get('amount') else '',
                round(t['pnl']/1e4,2) if t.get('pnl') is not None else '',
                t.get('hold_days',''),t.get('reason','')])
            rn = ws3.max_row; pc = ws3.cell(row=rn,column=9)
            if isinstance(pc.value,(int,float)):
                pc.font = gnf if pc.value > 0 else (rf_ if pc.value < 0 else None)
        for i,w in enumerate([12,10,10,16,25,14,12,12,12,8,45],1):
            ws3.column_dimensions[get_column_letter(i)].width = w
        ws3.freeze_panes = 'A2'

        # Sheet 4: 预测记录
        ws4 = wb.create_sheet("前瞻预测记录")
        h4 = ['日期','银行','代码','预测释放日','预测释放结束日','置信度','来源','命中','买入时收益%','操作']
        ws4.append(h4)
        for c in ws4[1]: c.font = hf; c.fill = gf; c.alignment = ct
        aps = []
        for m in 'BC':
            for p in self.R[m].get('prediction_log',[]):
                r = dict(p); r['_m'] = m; aps.append(r)
        aps.sort(key=lambda x: x.get('date',''))
        for p in aps:
            k = p.get('product_key',('',''))
            ws4.append([p.get('date',''), k[0] if isinstance(k,tuple) else '',
                k[1] if isinstance(k,tuple) else '',
                p.get('predicted_release',''), p.get('predicted_release_end',''),
                round(p.get('confidence',0),3),
                p.get('source',''), '是' if p.get('hit') else '否',
                round(p.get('ret_at_event',0),2), p.get('action','')])
            rn = ws4.max_row; hc = ws4.cell(row=rn,column=8)
            if hc.value == '是': hc.font = gnf
            elif hc.value == '否': hc.font = rf_
        for i,w in enumerate([12,12,16,12,12,10,12,8,12,16],1):
            ws4.column_dimensions[get_column_letter(i)].width = w
        ws4.freeze_panes = 'A2'

        # Sheet 5: 规律快照
        ws5 = wb.create_sheet("规律学习快照")
        h5 = ['银行','代码','置信度','周期(天)','周期CV','有周期','Top阶段','阶段p值','Top星期','星期p值','事件数','平均窗口(天)']
        ws5.append(h5)
        for c in ws5[1]: c.font = hf; c.fill = gf; c.alignment = ct
        pn = ['月初(1-7)','月上旬(8-14)','月中(15-21)','月底(22-31)']
        wn = ['周一','周二','周三','周四','周五']
        ps = self.R['C'].get('patterns_snapshot',{})
        for key, info in sorted(ps.items(), key=lambda x: -x[1]['confidence']):
            b = key[0] if isinstance(key,tuple) else ''; c_ = key[1] if isinstance(key,tuple) else ''
            tp = info.get('top_phase',0); tw = info.get('top_weekday',0)
            ws5.append([b,c_,round(info.get('confidence',0),3),
                round(info.get('period_days',0),1),round(info.get('period_cv',0),3),
                '是' if info.get('has_period') else '否',
                pn[tp] if tp<len(pn) else '', round(info.get('phase_pvalue',1.0),4),
                wn[tw] if tw<len(wn) else '', round(info.get('weekday_pvalue',1.0),4),
                info.get('n_events',0), round(info.get('avg_window_days',3.0),1)])
        for i,w in enumerate([12,16,10,10,10,8,14,10,10,10,8,12],1):
            ws5.column_dimensions[get_column_letter(i)].width = w
        ws5.freeze_panes = 'A2'

        try: wb.save(OUTPUT_PATH); logger.info(f"报告: {OUTPUT_PATH}")
        except:
            bk = os.path.join(BASE_DIR, f"前瞻策略回测报告_{datetime.now():%H%M%S}.xlsx")
            wb.save(bk); logger.info(f"报告: {bk}")
        return OUTPUT_PATH


# ============================================================
# 主函数
# ============================================================

def main():
    global ALPHA_预测加成

    sweep_values = [0.05, 0.10, 0.15, 0.20, 0.25, 0.30, 0.35, 0.40, 0.50, 0.60, 0.80, 1.00]

    sep = '=' * 74
    print(sep)
    print("  前瞻策略回测 V3 — ALPHA_预测加成 参数敏感性分析")
    print(f"  扫描值: {sweep_values}")
    print(f"  基线: Mode A (纯反应, 无预测)")
    print(sep)

    engine = PatternBacktestEngine()
    engine.load_data()

    # ① 基线 Mode A (只需跑一次)
    ra = engine.run_single_mode('A')
    base_ann = ra['ann_ret']

    # ② 扫描 Mode B 不同 ALPHA 值
    sweep_results = []
    for val in sweep_values:
        ALPHA_预测加成 = val
        logger.info(f"\n  >>> ALPHA_预测加成 = {val} <<<")
        rb = engine.run_single_mode('B')
        sweep_results.append({
            'alpha': val,
            'ann_ret': rb['ann_ret'],
            'total_ret': rb['total_ret'],
            'sharpe': rb['sharpe'],
            'max_dd': rb['max_dd'],
            'win_rate': rb['win_rate'],
            'n_trades': rb['n_trades'],
            'n_alpha_boost': rb.get('n_alpha_boost', 0),
            'pred_wr': rb.get('pred_wr', 0),
            'pred_pnl': rb.get('pred_pnl', 0),
            'delta': rb['ann_ret'] - base_ann,
        })

    # ③ 打印对比表
    print(f"\n{sep}")
    print("     ALPHA_预测加成 参数敏感性分析结果")
    print(sep)
    print(f"  基线 Mode A: 年化 {base_ann:+.4f}%  夏普 {ra['sharpe']:.4f}  "
          f"最大回撤 {ra['max_dd']:.4f}%")
    print()
    print(f"  {'ALPHA':<8} {'年化%':>8} {'Δ年化%':>8} {'夏普':>8} {'回撤%':>8} "
          f"{'胜率%':>7} {'交易':>5} {'加成':>5} {'预测胜率':>8} {'预测盈亏万':>10}")
    print("  " + "-" * 90)

    best = None
    for r in sweep_results:
        tag = ''
        if best is None or r['ann_ret'] > best['ann_ret']:
            best = r
        print(f"  {r['alpha']:<8.2f} {r['ann_ret']:>+8.4f} {r['delta']:>+8.4f} "
              f"{r['sharpe']:>8.4f} {r['max_dd']:>8.4f} "
              f"{r['win_rate']:>7.1f} {r['n_trades']:>5d} {r['n_alpha_boost']:>5d} "
              f"{r['pred_wr']:>8.1f} {r['pred_pnl']/1e4:>+10.2f}")

    print("  " + "-" * 90)
    print(f"\n  最优 ALPHA = {best['alpha']:.2f}  年化 {best['ann_ret']:+.4f}%  "
          f"vs 基线 {best['delta']:+.4f}%  夏普 {best['sharpe']:.4f}")

    # ④ 用最优 ALPHA 跑 Mode C
    ALPHA_预测加成 = best['alpha']
    logger.info(f"\n  >>> 最优 ALPHA={best['alpha']} 下跑 Mode C <<<")
    rc = engine.run_single_mode('C')
    print(f"  最优ALPHA下 Mode C: 年化 {rc['ann_ret']:+.4f}%  "
          f"vs 基线 {rc['ann_ret']-base_ann:+.4f}%  夏普 {rc['sharpe']:.4f}")

    # ⑤ 生成 Excel 报告 (用最优参数的 B/C)
    rb_best = None
    for val in sweep_values:
        if val == best['alpha']:
            ALPHA_预测加成 = val
            rb_best = engine.run_single_mode('B')
            break
    if rb_best is None:
        rb_best = engine.run_single_mode('B')

    report = ComparisonReport(engine, ra, rb_best, rc)
    report.print_console()

    # 额外写入敏感性分析sheet
    report.generate_excel()

    # 补充写入敏感性分析到Excel
    from openpyxl import load_workbook
    try:
        wb = load_workbook(OUTPUT_PATH)
    except:
        bk_files = [f for f in os.listdir(BASE_DIR) if f.startswith('前瞻策略回测报告_') and f.endswith('.xlsx')]
        if bk_files:
            wb = load_workbook(os.path.join(BASE_DIR, sorted(bk_files)[-1]))
        else:
            print("  警告: 无法打开报告文件写入敏感性分析")
            return

    ws = wb.create_sheet("参数敏感性分析", 0)
    headers = ['ALPHA_预测加成', '年化收益率%', 'Δ年化%', '夏普比率', '最大回撤%',
               '胜率%', '交易次数', '加成次数', '预测胜率%', '预测盈亏(万)']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='4472C4')
        cell.font = Font(bold=True, color='FFFFFF')
    # 基线行
    ws.append(['基线(A)', round(base_ann,4), 0, round(ra['sharpe'],4),
               round(ra['max_dd'],4), round(ra['win_rate'],1), ra['n_trades'],
               'N/A', 'N/A', 'N/A'])
    for cell in ws[2]:
        cell.fill = PatternFill('solid', fgColor='FFF2CC')
    # 扫描行
    best_row = None
    for i, r in enumerate(sweep_results):
        row = [r['alpha'], round(r['ann_ret'],4), round(r['delta'],4),
               round(r['sharpe'],4), round(r['max_dd'],4), round(r['win_rate'],1),
               r['n_trades'], r['n_alpha_boost'], round(r['pred_wr'],1),
               round(r['pred_pnl']/1e4,2)]
        ws.append(row)
        if r['alpha'] == best['alpha']:
            best_row = i + 3  # 1-based, header=1, baseline=2
    # 高亮最优行
    if best_row:
        for cell in ws[best_row]:
            cell.fill = PatternFill('solid', fgColor='C6EFCE')
            cell.font = Font(bold=True)

    for i, w in enumerate([16,12,10,10,10,8,8,8,10,12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    # 添加年化收益折线图
    chart = LineChart()
    chart.title = "ALPHA_预测加成 vs 年化收益率"
    chart.x_axis.title = "ALPHA_预测加成"
    chart.y_axis.title = "年化收益率%"
    chart.style = 10
    chart.width = 20; chart.height = 12
    n_data = len(sweep_results)
    data = Reference(ws, min_col=2, min_row=2, max_row=2+n_data)
    cats = Reference(ws, min_col=1, min_row=3, max_row=2+n_data)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.line.width = 25000
    ws.add_chart(chart, "A" + str(4 + n_data))

    try:
        wb.save(OUTPUT_PATH)
        print(f"\n  报告(含敏感性分析): {OUTPUT_PATH}")
    except:
        bk = os.path.join(BASE_DIR, f"前瞻策略回测报告_{datetime.now():%H%M%S}.xlsx")
        wb.save(bk)
        print(f"\n  报告(含敏感性分析): {bk}")
    print("  完成。")


if __name__ == '__main__':
    main()
