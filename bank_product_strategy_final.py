# -*- coding: utf-8 -*-
"""
银行理财产品"收益释放"实时交易决策系统 (V4.0 - 重构版)

核心理念：
这是一个实时交易决策系统，而不是历史分析系统。

三大核心模块：
1. 【实时机会】当前可操作的买入机会（最近1-3天的新信号）
2. 【持仓管理】已买入产品的持有/卖出决策
3. 【策略验证】历史回测证明策略有效性

关键区别：
- V3.0: 分析过去30天信号（已过期，无法操作）
- V4.0: 只推荐最近1-3天信号（现在可以操作）

作者：AI-FINANCE
版本：V4.0
日期：2026-01-28
"""

import os
import sys
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import logging
import warnings
import json
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

# 中文字体设置
plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei']
plt.rcParams['axes.unicode_minus'] = False

# ============================================================
# 日志配置
# ============================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('bank_product_strategy.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
日志 = logging.getLogger(__name__)

# ============================================================
# 策略参数
# ============================================================
净值数据库路径 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "净值数据库.xlsx")
输出文件路径 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "实时交易决策报告_V4.xlsx")
图表输出目录 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "charts")

# 信号识别参数
收益拉升阈值 = 5.0         # %
收益下滑阈值 = 2.0         # %

# 周期约束
最短持有天数 = 7
最长持有天数 = 28
默认预测周期 = 14

# 实时决策参数（将根据数据质量动态调整）
默认实时信号窗口天数 = 3   # 数据充足时使用
默认历史回测天数 = 90      # 数据充足时使用
最大持仓数量 = 10          # 同时最多持有10个产品

# 数据质量阈值
优秀数据天数阈值 = 180     # >= 180天视为优秀
良好数据天数阈值 = 60      # >= 60天视为良好

# 风险控制参数
最大回撤止损阈值 = 30.0
预期收益折扣系数 = 0.6

# 回测参数
初始资金 = 100000.0
单笔投资金额 = 10000.0     # 固定每笔10000元（更合理）
无风险收益率 = 0.025


# ============================================================
# 工具函数（复用）
# ============================================================

def 是否日期列(列名):
    if not isinstance(列名, str):
        return False
    if len(列名) == 10 and 列名[4] == '-' and 列名[7] == '-':
        try:
            datetime.strptime(列名, '%Y-%m-%d')
            return True
        except ValueError:
            return False
    return False


def 识别产品类别(产品名称):
    if not isinstance(产品名称, str):
        产品名称 = str(产品名称)

    产品名称原始 = 产品名称
    产品名称 = 产品名称.upper()

    if 'R3' in 产品名称 or 'R4' in 产品名称 or 'R5' in 产品名称:
        return '含权益类'
    if 'R1' in 产品名称 or 'R2' in 产品名称:
        return '不含权益类'

    含权益关键词 = ['增强', '混合', '权益', '股票', '可转债', 'FOF', '偏债',
                    '灵活配置', '进取', '成长', '积极']
    for 关键词 in 含权益关键词:
        if 关键词 in 产品名称原始:
            return '含权益类'

    不含权益关键词 = ['纯债', '货币', '稳健', '固收']
    for 关键词 in 不含权益关键词:
        if 关键词 in 产品名称原始:
            return '不含权益类'

    return '未知类别'


def 计算最大回撤(收益序列):
    if not 收益序列 or len(收益序列) == 0:
        return 0.0, 0

    累计收益 = np.cumsum(收益序列)
    最高点 = np.maximum.accumulate(累计收益)
    回撤 = 最高点 - 累计收益
    最大回撤 = np.max(回撤)
    最大回撤百分比 = (最大回撤 / np.max(最高点) * 100) if np.max(最高点) > 0 else 0

    return 最大回撤百分比, np.argmax(回撤)


def 评估数据质量并获取参数(收益率日期列表):
    """
    根据数据量动态调整策略参数
    返回: (实时窗口天数, 回测天数, 数据质量等级, 数据跨度天数)
    """
    数据天数 = len(收益率日期列表)

    if 数据天数 >= 优秀数据天数阈值:
        质量等级 = "优秀"
        实时窗口 = 3
        回测天数 = min(90, 数据天数 - 10)  # 留10天给实时分析
    elif 数据天数 >= 良好数据天数阈值:
        质量等级 = "良好"
        实时窗口 = 2
        回测天数 = min(数据天数 // 2, 数据天数 - 10)
    else:
        质量等级 = "不足"
        实时窗口 = 1
        回测天数 = max(7, 数据天数 - 10)

    return 实时窗口, 回测天数, 质量等级, 数据天数


# ============================================================
# 数据加载（复用，简化版）
# ============================================================

def _加载JSON净值数据(银行名):
    基础目录 = os.path.dirname(os.path.abspath(__file__))
    if '华夏' in 银行名:
        路径 = os.path.join(基础目录, 'huaxia_nav_history.json')
        if os.path.exists(路径):
            try:
                with open(路径, 'r', encoding='utf-8') as f:
                    数据 = json.load(f)
                结果 = {}
                for 代码, 净值列表 in 数据.items():
                    净值字典 = {}
                    for 条目 in 净值列表:
                        净值字典[条目['date']] = str(条目['unit_nav'])
                    结果[代码] = 净值字典
                return 结果
            except:
                pass
    return None


def 智能修复产品代码(数据库):
    """
    智能修复产品代码：
    1. 通过净值序列精确匹配将无代码行与有代码行关联
    2. 匹配失败的行，如果有最新数据，也尝试保留（但优先级低）
    """
    日志.info("智能修复产品代码...")
    修复后数据库 = {}

    for 工作表名, df in 数据库.items():
        缺失数 = df['产品代码'].isna().sum()
        if 缺失数 == 0:
            修复后数据库[工作表名] = df
            日志.info(f"  {工作表名}: 无需修复，{len(df)}行")
            continue

        日期列 = sorted([列 for 列 in df.columns if 是否日期列(列)])
        有代码掩码 = df['产品代码'].notna()
        有代码df = df[有代码掩码].copy()
        无代码df = df[~有代码掩码].copy()

        日志.info(f"  {工作表名}: {len(有代码df)}行有代码, {len(无代码df)}行无代码")

        if len(无代码df) == 0:
            修复后数据库[工作表名] = 有代码df
            continue

        # 找出两者都有数据的日期范围
        有代码净值 = 有代码df[日期列].apply(pd.to_numeric, errors='coerce')
        无代码净值 = 无代码df[日期列].apply(pd.to_numeric, errors='coerce')

        重叠日期 = []
        for d in 日期列:
            有代码有值 = 有代码净值[d].notna().sum() > 0
            无代码有值 = 无代码净值[d].notna().sum() > 0
            if 有代码有值 and 无代码有值:
                重叠日期.append(d)

        日志.info(f"    重叠日期数: {len(重叠日期)}")

        # 使用重叠日期构建精确匹配索引（使用字符串key更精确）
        匹配索引 = {}
        匹配日期 = 重叠日期[-20:] if len(重叠日期) > 20 else 重叠日期

        if len(匹配日期) >= 3:
            for idx, row in 有代码df.iterrows():
                产品代码 = row['产品代码']
                产品名称 = row['产品名称']
                key_vals = []
                for d in 匹配日期:
                    val = row.get(d)
                    if pd.notna(val):
                        try:
                            key_vals.append(str(round(float(val), 4)))
                        except:
                            pass
                if len(key_vals) >= 3:
                    key = '|'.join(key_vals)
                    匹配索引[key] = (产品代码, 产品名称)

        日志.info(f"    构建匹配索引: {len(匹配索引)}个")

        # 匹配无代码行
        匹配成功 = 0
        合并数据 = []
        最新日期 = 日期列[-5:]  # 最近5天

        for idx, row in 无代码df.iterrows():
            # 构建匹配key
            key_vals = []
            for d in 匹配日期:
                val = row.get(d)
                if pd.notna(val):
                    try:
                        key_vals.append(str(round(float(val), 4)))
                    except:
                        pass

            已匹配 = False
            if len(key_vals) >= 3:
                key = '|'.join(key_vals)
                if key in 匹配索引:
                    产品代码, 产品名称 = 匹配索引[key]
                    新行 = row.copy()
                    新行['产品代码'] = 产品代码
                    新行['产品名称'] = 产品名称
                    合并数据.append(新行)
                    匹配成功 += 1
                    已匹配 = True

        日志.info(f"    匹配成功: {匹配成功}/{len(无代码df)}")

        # 合并所有数据
        结果字典 = {}  # 用字典按产品代码合并

        # 1. 添加原有代码行
        for idx, row in 有代码df.iterrows():
            代码 = row['产品代码']
            if 代码 not in 结果字典:
                结果字典[代码] = row.to_dict()
            else:
                # 合并净值
                for d in 日期列:
                    if pd.isna(结果字典[代码].get(d)) and pd.notna(row.get(d)):
                        结果字典[代码][d] = row[d]

        # 2. 合并匹配成功的行
        for 新行 in 合并数据:
            代码 = 新行['产品代码']
            if 代码 in 结果字典:
                # 合并净值数据（新数据优先）
                for d in 日期列:
                    if pd.notna(新行.get(d)):
                        结果字典[代码][d] = 新行[d]
            else:
                结果字典[代码] = 新行.to_dict()

        修复后数据库[工作表名] = pd.DataFrame(list(结果字典.values()))
        日志.info(f"    最终: {len(修复后数据库[工作表名])}行")

    return 修复后数据库


def 加载净值数据库():
    if not os.path.exists(净值数据库路径):
        日志.error(f"找不到文件: {净值数据库路径}")
        return {}

    日志.info("正在加载净值数据库...")
    xlsx = pd.ExcelFile(净值数据库路径)
    数据库 = {}

    for sheet in xlsx.sheet_names:
        df = pd.read_excel(xlsx, sheet_name=sheet, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]

        重命名映射 = {}
        if 'level_0' in df.columns:
            重命名映射['level_0'] = '产品代码'
        if 'level_1' in df.columns:
            重命名映射['level_1'] = '产品名称'
        if 重命名映射:
            df = df.rename(columns=重命名映射)

        if '产品代码' in df.columns and '产品名称' in df.columns:
            数据库[sheet] = df

    return 数据库


def 计算基础收益率(df, 银行名):
    全部日期 = sorted([列 for 列 in df.columns if 是否日期列(列)])

    if len(全部日期) < 2:
        return None, [], []

    净值矩阵 = df[全部日期].apply(pd.to_numeric, errors='coerce')

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", FutureWarning)
        收益率矩阵 = 净值矩阵.pct_change(axis=1) * 365 * 100

    收益率日期 = 全部日期[1:]

    收益率数据 = np.round(收益率矩阵[收益率日期].values, 4)
    结果 = pd.DataFrame(收益率数据, columns=收益率日期)
    结果.insert(0, '产品代码', df['产品代码'].values)
    结果.insert(1, '产品名称', df['产品名称'].values)
    结果.insert(2, '银行', 银行名)
    结果.insert(3, '产品类别', df['产品名称'].apply(识别产品类别))

    有效掩码 = 结果[收益率日期].notna().any(axis=1)
    结果 = 结果[有效掩码].reset_index(drop=True)

    return 结果, 全部日期, 收益率日期


# ============================================================
# 历史规律分析（复用）
# ============================================================

def 分析历史拉升规律(收益率序列, 当前日期, 收益率日期列表):
    try:
        if 当前日期 not in 收益率日期列表:
            return 默认预测周期, 收益拉升阈值, 0

        当前索引 = 收益率日期列表.index(当前日期)
        历史日期 = 收益率日期列表[:当前索引]

        if len(历史日期) < 30:
            return 默认预测周期, 收益拉升阈值, 0

        历史收益 = [收益率序列.get(日期, 0) for 日期 in 历史日期]

        片段列表 = []
        当前片段 = []

        for 收益 in 历史收益:
            if 收益 > (收益拉升阈值 - 0.5):
                当前片段.append(收益)
            else:
                if len(当前片段) >= 3:
                    片段列表.append(当前片段[:])
                当前片段 = []

        if len(当前片段) >= 3:
            片段列表.append(当前片段)

        if not 片段列表:
            return 默认预测周期, 收益拉升阈值, 0

        持续天数集 = [len(p) for p in 片段列表]
        平均收益集 = [np.mean(p) for p in 片段列表]

        预期天数 = int(np.mean(持续天数集))
        预期天数 = max(最短持有天数, min(预期天数, 最长持有天数))
        预期收益 = np.mean(平均收益集)

        return 预期天数, 预期收益, len(片段列表)

    except:
        return 默认预测周期, 收益拉升阈值, 0


# ============================================================
# 核心模块1：实时交易机会识别（重构）
# ============================================================

def 计算产品潜力评分(收益序列, 收益率日期, 当前日期, 当前收益, 历史次数, 预期收益):
    """
    计算产品的综合潜力评分（0-100分）
    综合考虑：启动强度、历史可靠性、收益稳定性、数据充足度
    """
    评分 = 0

    # 1. 启动强度评分（0-25分）：当前收益越高越好，但超高可能有风险
    if 当前收益 >= 15:
        评分 += 20  # 超高收益略减分（可能是异常）
    elif 当前收益 >= 10:
        评分 += 25  # 高收益满分
    elif 当前收益 >= 7:
        评分 += 20
    elif 当前收益 >= 5:
        评分 += 15
    else:
        评分 += 10

    # 2. 历史可靠性评分（0-30分）：历史拉升次数越多越可靠
    if 历史次数 >= 5:
        评分 += 30
    elif 历史次数 >= 3:
        评分 += 25
    elif 历史次数 >= 2:
        评分 += 20
    elif 历史次数 >= 1:
        评分 += 15
    else:
        评分 += 5  # 无历史记录，给基础分

    # 3. 预期收益评分（0-25分）
    if 预期收益 >= 10:
        评分 += 25
    elif 预期收益 >= 7:
        评分 += 20
    elif 预期收益 >= 5:
        评分 += 15
    else:
        评分 += 10

    # 4. 数据充足度评分（0-20分）：检查最近30天的数据完整度
    当前索引 = 收益率日期.index(当前日期) if 当前日期 in 收益率日期 else len(收益率日期) - 1
    回溯天数 = min(30, 当前索引)
    有效数据天数 = 0
    for i in range(当前索引 - 回溯天数, 当前索引):
        if i >= 0:
            日期 = 收益率日期[i]
            if 收益序列.get(日期, 0) != 0:
                有效数据天数 += 1

    数据完整率 = 有效数据天数 / max(回溯天数, 1)
    评分 += int(数据完整率 * 20)

    return min(评分, 100)


def 识别实时交易机会(收益率df, 全部日期, 收益率日期, 银行名):
    """
    关键改进V4.1：
    1. 只返回最近1-3天出现的新信号
    2. 智能筛选：综合评分系统，只保留最优潜力产品
    3. 限制数量：每个银行最多返回20个最优机会
    """
    # 动态获取参数
    实时窗口, _, 质量等级, 数据天数 = 评估数据质量并获取参数(收益率日期)

    if len(收益率日期) < 实时窗口:
        日志.warning(f"[{银行名}] 数据不足，无法分析实时机会")
        return []

    # 只分析最近N天
    最新日期 = 收益率日期[-1]
    实时窗口日期 = 收益率日期[-实时窗口:]

    日志.info(f"[{银行名}] 数据质量:{质量等级}（{数据天数}天），实时窗口:{实时窗口}天")
    日志.info(f"[{银行名}] 扫描实时机会（{实时窗口日期[0]} ~ {最新日期}）")

    原始机会列表 = []

    for idx, row in 收益率df.iterrows():
        产品代码 = row['产品代码']
        产品名称 = row['产品名称']
        产品类别 = row['产品类别']

        收益序列 = {}
        for 日期 in 收益率日期:
            try:
                收益 = pd.to_numeric(row[日期], errors='coerce')
                收益序列[日期] = 收益 if pd.notna(收益) else 0
            except:
                收益序列[日期] = 0

        # 检查实时窗口内是否有新信号
        for i, 当前日期 in enumerate(实时窗口日期):
            当前收益 = 收益序列[当前日期]

            # 判断是否是新启动信号
            当前索引 = 收益率日期.index(当前日期)
            if 当前索引 > 0:
                昨日日期 = 收益率日期[当前索引 - 1]
                昨日收益 = 收益序列[昨日日期]
                是新信号 = (当前收益 > 收益拉升阈值) and (昨日收益 <= 收益拉升阈值)
            else:
                是新信号 = 当前收益 > 收益拉升阈值

            if not 是新信号:
                continue

            # 分析历史规律
            预期天数, 预期收益, 历史次数 = 分析历史拉升规律(收益序列, 当前日期, 收益率日期)

            # 计算该信号到现在的天数
            信号到现在天数 = len(收益率日期) - 1 - 当前索引

            # 计算当前应持有的天数（如果买入了）
            持有期收益列表 = []
            for j in range(当前索引 + 1, len(收益率日期)):
                持有期收益列表.append(收益序列[收益率日期[j]])

            当前持有天数 = len(持有期收益列表)
            当前持有年化 = np.mean(持有期收益列表) if 持有期收益列表 else 当前收益
            当前回撤, _ = 计算最大回撤(持有期收益列表) if 持有期收益列表 else (0, 0)

            # 计算综合潜力评分
            潜力评分 = 计算产品潜力评分(收益序列, 收益率日期, 当前日期, 当前收益, 历史次数, 预期收益)

            # 判断操作建议（结合评分）
            if 信号到现在天数 == 0:
                if 潜力评分 >= 70:
                    操作建议 = "[!]强烈推荐买入"
                elif 潜力评分 >= 50:
                    操作建议 = "[!]建议买入"
                else:
                    操作建议 = "可考虑买入"
            elif 信号到现在天数 <= 2:
                if 潜力评分 >= 70:
                    操作建议 = "推荐买入（信号新鲜）"
                else:
                    操作建议 = "可买入（信号新鲜）"
            else:
                操作建议 = "观望（信号较旧）"

            # 风险评估
            风险等级 = "低风险"
            if 当前回撤 > 15:
                风险等级 = "高风险"
            elif 当前回撤 > 8:
                风险等级 = "中风险"

            原始机会列表.append({
                '银行': 银行名,
                '产品代码': 产品代码,
                '产品名称': 产品名称,
                '产品类别': 产品类别,
                '信号日期': 当前日期,
                '信号距今天数': 信号到现在天数,
                '启动收益率%': round(当前收益, 2),
                '历史拉升次数': 历史次数,
                '预期维持天数': 预期天数,
                '预期年化收益%': round(预期收益, 2),
                '如已买入持有天数': 当前持有天数,
                '如已买入当前年化%': round(当前持有年化, 2),
                '当前回撤%': round(当前回撤, 2),
                '风险等级': 风险等级,
                '潜力评分': 潜力评分,
                '操作建议': 操作建议
            })

    日志.info(f"[{银行名}] 原始信号: {len(原始机会列表)} 个")

    # 智能筛选：按潜力评分排序，只保留前20个最优机会
    if len(原始机会列表) > 20:
        原始机会列表.sort(key=lambda x: (-x['潜力评分'], -x['启动收益率%']))
        实时机会列表 = 原始机会列表[:20]
        日志.info(f"[{银行名}] 筛选后保留 Top 20 最优机会")
    else:
        实时机会列表 = 原始机会列表

    日志.info(f"[{银行名}] 最终推荐: {len(实时机会列表)} 个实时交易机会")
    return 实时机会列表


# ============================================================
# 核心模块2：持仓管理（新增）
# ============================================================

def 创建持仓模板文件():
    """创建持仓配置模板文件"""
    持仓配置路径 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "我的持仓_模板.xlsx")

    # 创建示例数据
    示例数据 = pd.DataFrame({
        '银行': ['华夏银行', '民生银行'],
        '产品代码': ['示例产品代码001', '示例产品代码002'],
        '买入日期': ['2025-01-15', '2025-01-20'],
        '备注': ['请删除示例数据后填入实际持仓', '日期格式: YYYY-MM-DD']
    })

    try:
        示例数据.to_excel(持仓配置路径, index=False)
        日志.info(f"已创建持仓模板文件: {持仓配置路径}")
        return 持仓配置路径
    except Exception as e:
        日志.error(f"创建持仓模板失败: {e}")
        return None


def 加载持仓配置():
    """
    加载用户的持仓配置文件
    返回: [{'银行': 'xxx', '产品代码': 'xxx', '买入日期': 'yyyy-mm-dd'}, ...]
    """
    持仓配置路径 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "我的持仓.xlsx")

    if not os.path.exists(持仓配置路径):
        日志.info("未找到持仓配置文件，默认空仓")
        # 创建模板文件供用户参考
        模板路径 = 创建持仓模板文件()
        if 模板路径:
            日志.info(f"已自动创建持仓模板文件: {模板路径}")
            日志.info("如需记录持仓，请复制模板文件为'我的持仓.xlsx'并填入实际数据")
        return []

    try:
        df = pd.read_excel(持仓配置路径)
        必需列 = ['银行', '产品代码', '买入日期']

        if not all(col in df.columns for col in 必需列):
            日志.warning(f"持仓配置文件格式错误，需要列：{必需列}")
            return []

        持仓列表 = df.to_dict('records')
        日志.info(f"已加载 {len(持仓列表)} 个持仓记录")
        return 持仓列表
    except Exception as e:
        日志.error(f"加载持仓配置失败: {e}")
        return []


def 生成持仓管理建议(收益率df, 全部日期, 收益率日期, 银行名, 用户持仓列表):
    """
    根据用户提供的持仓信息，生成卖出建议
    如果用户持仓为空，返回空列表
    """
    if not 用户持仓列表:
        return []

    # 筛选当前银行的持仓
    银行持仓 = [p for p in 用户持仓列表 if p.get('银行') == 银行名]

    if not 银行持仓:
        return []

    日志.info(f"[{银行名}] 分析 {len(银行持仓)} 个持仓产品")

    持仓管理列表 = []

    # 遍历用户的持仓（而不是所有产品）
    for 持仓记录 in 银行持仓:
        产品代码 = str(持仓记录.get('产品代码', '')).strip()
        买入日期 = str(持仓记录.get('买入日期', '')).strip()

        # 在收益率df中查找该产品
        产品行 = 收益率df[收益率df['产品代码'] == 产品代码]

        if 产品行.empty:
            日志.warning(f"[{银行名}] 持仓产品 {产品代码} 未在数据库中找到")
            continue

        row = 产品行.iloc[0]
        产品名称 = row['产品名称']
        产品类别 = row['产品类别']

        # 构建收益序列
        收益序列 = {}
        for 日期 in 收益率日期:
            try:
                收益 = pd.to_numeric(row[日期], errors='coerce')
                收益序列[日期] = 收益 if pd.notna(收益) else 0
            except:
                收益序列[日期] = 0

        # 检查买入日期是否在数据范围内
        if 买入日期 not in 收益率日期:
            日志.warning(f"[{银行名}] 产品 {产品代码} 的买入日期 {买入日期} 不在数据范围内")
            continue

        买入索引 = 收益率日期.index(买入日期)

        # 计算买入时的收益率（启动信号）
        买入时收益 = 收益序列[买入日期]

        # 分析历史规律，获取预期
        预期天数, 预期收益, 历史次数 = 分析历史拉升规律(收益序列, 买入日期, 收益率日期)

        # 计算持有期表现（从买入次日到最新日期）
        持有期收益列表 = []
        for j in range(买入索引 + 1, len(收益率日期)):
            持有期收益列表.append(收益序列[收益率日期[j]])

        if not 持有期收益列表:
            日志.warning(f"[{银行名}] 产品 {产品代码} 买入后无可用数据")
            continue

        持有天数 = len(持有期收益列表)
        持有年化 = np.mean(持有期收益列表)
        累计收益 = sum(持有期收益列表)
        最大回撤, _ = 计算最大回撤(持有期收益列表)

        # 决策逻辑
        卖出建议 = "持有"
        卖出原因 = ""

        if 持有天数 >= 预期天数:
            卖出建议 = "[!]建议卖出"
            卖出原因 = "已达预期周期"
        elif 最大回撤 > 最大回撤止损阈值:
            卖出建议 = "[!]建议卖出"
            卖出原因 = f"回撤过大({最大回撤:.1f}%)"
        elif 持有年化 < (预期收益 * 预期收益折扣系数):
            卖出建议 = "[!]建议卖出"
            卖出原因 = "收益不达预期"
        elif 持有天数 < 最短持有天数:
            卖出建议 = "继续持有"
            卖出原因 = "未达最短持有期"
        else:
            if 持有年化 >= 预期收益:
                卖出建议 = "表现良好，建议持有"
                卖出原因 = f"收益达标({持有年化:.2f}% >= {预期收益:.2f}%)"
            else:
                卖出建议 = "观察中"
                卖出原因 = "暂未达预期但仍在观察期"

        持仓管理列表.append({
            '银行': 银行名,
            '产品代码': 产品代码,
            '产品名称': 产品名称,
            '产品类别': 产品类别,
            '买入日期': 买入日期,
            '买入时收益率%': round(买入时收益, 2),
            '持有天数': 持有天数,
            '预期持有天数': 预期天数,
            '预期年化收益%': round(预期收益, 2),
            '实际持有年化%': round(持有年化, 2),
            '累计收益%': round(累计收益, 2),
            '最大回撤%': round(最大回撤, 2),
            '操作建议': 卖出建议,
            '原因': 卖出原因
        })

    日志.info(f"[{银行名}] 生成 {len(持仓管理列表)} 个持仓管理记录")
    return 持仓管理列表


# ============================================================
# 核心模块3：历史回测验证（简化版）
# ============================================================

def 执行历史回测(收益率df, 全部日期, 收益率日期, 银行名):
    """
    用过去N天的数据验证策略有效性
    N根据数据质量动态调整
    """
    # 动态获取参数
    _, 回测天数, 质量等级, 数据天数 = 评估数据质量并获取参数(收益率日期)

    if len(收益率日期) < 回测天数:
        日志.warning(f"[{银行名}] 数据不足，无法执行回测")
        return {}

    回测起始位 = len(收益率日期) - 回测天数
    回测日期 = 收益率日期[回测起始位:]

    日志.info(f"[{银行名}] 数据质量:{质量等级}（{数据天数}天），回测天数:{回测天数}天")
    日志.info(f"[{银行名}] 执行历史回测（{回测日期[0]} ~ {回测日期[-1]}）")

    成功交易 = 0
    失败交易 = 0
    总收益 = 0
    回撤列表 = []

    for idx, row in 收益率df.iterrows():
        收益序列 = {}
        for 日期 in 收益率日期:
            try:
                收益 = pd.to_numeric(row[日期], errors='coerce')
                收益序列[日期] = 收益 if pd.notna(收益) else 0
            except:
                收益序列[日期] = 0

        # 在回测窗口内寻找信号
        for i, 当前日期 in enumerate(回测日期[:-7]):  # 至少留7天观察期
            当前收益 = 收益序列[当前日期]
            当前索引 = 收益率日期.index(当前日期)

            if 当前索引 > 0:
                昨日收益 = 收益序列[收益率日期[当前索引 - 1]]
                是信号 = (当前收益 > 收益拉升阈值) and (昨日收益 <= 收益拉升阈值)
            else:
                是信号 = 当前收益 > 收益拉升阈值

            if not 是信号:
                continue

            预期天数, 预期收益, _ = 分析历史拉升规律(收益序列, 当前日期, 收益率日期)

            # 模拟持有
            持有期收益 = []
            for j in range(当前索引 + 1, min(当前索引 + 预期天数 + 1, len(收益率日期))):
                持有期收益.append(收益序列[收益率日期[j]])

            if not 持有期收益:
                continue

            持有年化 = np.mean(持有期收益)
            回撤, _ = 计算最大回撤(持有期收益)

            总收益 += 持有年化
            回撤列表.append(回撤)

            if 持有年化 >= 预期收益:
                成功交易 += 1
            else:
                失败交易 += 1

    总交易数 = 成功交易 + 失败交易
    成功率 = (成功交易 / 总交易数 * 100) if 总交易数 > 0 else 0
    平均收益 = (总收益 / 总交易数) if 总交易数 > 0 else 0
    平均回撤 = np.mean(回撤列表) if 回撤列表 else 0

    回测统计 = {
        '数据质量': 质量等级,
        '数据天数': 数据天数,
        '回测天数': 回测天数,
        '回测起始日': 回测日期[0],
        '回测结束日': 回测日期[-1],
        '总交易数': 总交易数,
        '成功交易数': 成功交易,
        '失败交易数': 失败交易,
        '成功率%': round(成功率, 2),
        '平均收益%': round(平均收益, 2),
        '平均回撤%': round(平均回撤, 2)
    }

    日志.info(f"[{银行名}] 回测完成: 成功率{成功率:.1f}%, 平均收益{平均收益:.2f}%")
    return 回测统计


# ============================================================
# 可视化（简化版）
# ============================================================

def 生成简化图表(实时机会, 持仓管理, 回测统计列表):
    """生成核心图表"""
    if not os.path.exists(图表输出目录):
        os.makedirs(图表输出目录)

    图表列表 = []

    # 图表1: 实时机会收益分布
    if 实时机会:
        df = pd.DataFrame(实时机会)
        plt.figure(figsize=(10, 6))
        plt.hist(df['启动收益率%'], bins=20, color='green', alpha=0.7, edgecolor='black')
        plt.axvline(df['启动收益率%'].mean(), color='red', linestyle='--', linewidth=2)
        plt.xlabel('启动收益率 (%)', fontsize=12)
        plt.ylabel('频数', fontsize=12)
        plt.title('实时交易机会 - 启动收益率分布', fontsize=14, fontweight='bold')
        plt.grid(axis='y', alpha=0.3)
        路径1 = os.path.join(图表输出目录, '实时机会分布.png')
        plt.savefig(路径1, dpi=150, bbox_inches='tight')
        plt.close()
        图表列表.append(路径1)

    # 图表2: 回测成功率对比
    if 回测统计列表:
        银行名称 = [x['银行'] for x in 回测统计列表]
        成功率 = [x['成功率%'] for x in 回测统计列表]

        plt.figure(figsize=(10, 6))
        bars = plt.bar(银行名称, 成功率, color=['#3498db', '#2ecc71', '#e74c3c'], alpha=0.8)
        plt.ylabel('成功率 (%)', fontsize=12)
        plt.title('历史回测 - 策略成功率对比', fontsize=14, fontweight='bold')
        plt.ylim(0, 100)
        plt.grid(axis='y', alpha=0.3)

        # 在柱子上标注数值
        for bar in bars:
            height = bar.get_height()
            plt.text(bar.get_x() + bar.get_width()/2., height,
                    f'{height:.1f}%', ha='center', va='bottom', fontsize=11)

        路径2 = os.path.join(图表输出目录, '回测成功率.png')
        plt.savefig(路径2, dpi=150, bbox_inches='tight')
        plt.close()
        图表列表.append(路径2)

    日志.info(f"已生成 {len(图表列表)} 个图表")
    return 图表列表


# ============================================================
# Excel报告生成（重构版）
# ============================================================

def 生成决策报告(实时机会, 持仓管理, 回测统计列表, 图表列表):
    """生成实时交易决策报告"""
    日志.info("正在生成实时交易决策报告...")

    wb = Workbook()

    # 样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="4F81BD")
    align_center = Alignment(horizontal="center", vertical="center")

    def 设置表头(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center

    def 设置列宽(ws, 列宽字典):
        for 列号, 宽度 in 列宽字典.items():
            ws.column_dimensions[get_column_letter(列号)].width = 宽度

    # Sheet 1: 实时交易机会（最重要）
    ws1 = wb.active
    ws1.title = "实时交易机会"

    if 实时机会:
        headers1 = ['银行', '产品代码', '产品名称', '产品类别', '信号日期', '信号距今天数',
                    '启动收益率%', '历史拉升次数', '预期维持天数', '预期年化收益%',
                    '潜力评分', '如已买入持有天数', '如已买入当前年化%', '当前回撤%', '风险等级', '操作建议']
        设置表头(ws1, headers1)

        # 按潜力评分和操作建议排序
        实时机会 = sorted(实时机会, key=lambda x: (
            0 if '强烈推荐' in x['操作建议'] else 1 if '建议买入' in x['操作建议'] else 2 if '推荐买入' in x['操作建议'] else 3,
            -x.get('潜力评分', 0),
            -x['启动收益率%']
        ))

        for row in 实时机会:
            ws1.append([row.get(k, '') for k in headers1])

        设置列宽(ws1, {
            1: 12, 2: 22, 3: 40, 4: 14, 5: 14, 6: 14,
            7: 14, 8: 14, 9: 14, 10: 16, 11: 10, 12: 16, 13: 18, 14: 12, 15: 12, 16: 22
        })
    else:
        ws1['A1'] = '当前没有实时交易机会'
        ws1['A1'].font = Font(bold=True, size=12)

    ws1.freeze_panes = 'A2'

    # Sheet 2: 持仓管理建议
    ws2 = wb.create_sheet("持仓管理建议")

    if 持仓管理:
        headers2 = ['银行', '产品代码', '产品名称', '产品类别', '买入日期', '买入时收益率%',
                    '持有天数', '预期持有天数', '预期年化收益%', '实际持有年化%', '累计收益%',
                    '最大回撤%', '操作建议', '原因']
        设置表头(ws2, headers2)

        # 按操作建议排序（建议卖出优先）
        持仓管理 = sorted(持仓管理, key=lambda x: (
            0 if '建议卖出' in x['操作建议'] else 1,
            -x['实际持有年化%']
        ))

        for row in 持仓管理:
            ws2.append([row[k] for k in headers2])

        设置列宽(ws2, {
            1: 12, 2: 22, 3: 40, 4: 14, 5: 14, 6: 14,
            7: 12, 8: 14, 9: 16, 10: 16, 11: 14, 12: 12, 13: 22, 14: 22
        })
    else:
        ws2['A1'] = '当前没有持仓记录'
        ws2['A1'].font = Font(bold=True, size=12)

    ws2.freeze_panes = 'A2'

    # Sheet 3: 策略回测验证
    ws3 = wb.create_sheet("策略回测验证")

    if 回测统计列表:
        ws3['A1'] = '银行理财产品套利策略 - 历史回测验证'
        ws3['A1'].font = Font(bold=True, size=14)
        ws3.merge_cells('A1:H1')

        行号 = 3
        for 回测统计 in 回测统计列表:
            ws3[f'A{行号}'] = f"=== {回测统计['银行']} ==="
            ws3[f'A{行号}'].font = Font(bold=True, size=12)
            行号 += 1

            for 键, 值 in 回测统计.items():
                if 键 != '银行':
                    ws3[f'A{行号}'] = 键
                    ws3[f'B{行号}'] = 值
                    行号 += 1

            行号 += 1

        设置列宽(ws3, {1: 20, 2: 20})

    # Sheet 4: 可视化图表
    if 图表列表:
        ws4 = wb.create_sheet("可视化图表")
        行位置 = 1

        for 图表路径 in 图表列表:
            if os.path.exists(图表路径):
                try:
                    img = XLImage(图表路径)
                    img.width = 600
                    img.height = 360
                    ws4.add_image(img, f'A{行位置}')
                    行位置 += 20
                except:
                    pass

    # 保存
    try:
        wb.save(输出文件路径)
        日志.info(f"报告已生成: {输出文件路径}")
        日志.info(f"  Sheet1 实时交易机会: {len(实时机会)} 行")
        日志.info(f"  Sheet2 持仓管理建议: {len(持仓管理)} 行")
        日志.info(f"  Sheet3 策略回测验证: {len(回测统计列表)} 个银行")
        if 图表列表:
            日志.info(f"  Sheet4 可视化图表: {len(图表列表)} 个")
    except Exception as e:
        日志.error(f"保存文件失败: {e}")


# ============================================================
# 主程序（重构版）
# ============================================================

def main():
    print("=" * 70)
    print("   银行理财产品收益释放 - 实时交易决策系统 V4.0")
    print("=" * 70)
    print(f"运行时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"数据来源: {净值数据库路径}")
    print(f"输出文件: {输出文件路径}")
    print()
    print("【核心改进】")
    print(f"1. 实时机会：只推荐最近1-3天内的新信号（根据数据质量动态调整）")
    print(f"2. 持仓管理：假设已买入，现在该如何操作")
    print(f"3. 策略验证：根据数据质量动态调整回测天数")
    print(f"4. 数据质量：自动评估每个银行的数据充足度")
    print()

    # 步骤1: 加载数据
    print("【第1步】加载净值数据库...")
    数据库 = 加载净值数据库()
    if not 数据库:
        return

    # 步骤2: 智能修复产品代码
    print("\n【第2步】智能修复产品代码...")
    数据库 = 智能修复产品代码(数据库)

    # 步骤3: 加载用户持仓
    print("\n【第3步】加载用户持仓配置...")
    用户持仓列表 = 加载持仓配置()
    if 用户持仓列表:
        print(f"  已加载 {len(用户持仓列表)} 个持仓记录")
    else:
        print("  未找到持仓配置或为空仓，将只显示实时交易机会")

    # 步骤4: 三大模块分析
    print("\n【第4步】执行三大核心模块...")

    所有实时机会 = []
    所有持仓管理 = []
    所有回测统计 = []

    for 银行名, df in 数据库.items():
        收益率df, 全部日期, 收益率日期 = 计算基础收益率(df, 银行名)
        if 收益率df is None:
            continue

        # 模块1: 实时交易机会
        实时机会 = 识别实时交易机会(收益率df, 全部日期, 收益率日期, 银行名)
        所有实时机会.extend(实时机会)

        # 模块2: 持仓管理（传入用户持仓）
        持仓管理 = 生成持仓管理建议(收益率df, 全部日期, 收益率日期, 银行名, 用户持仓列表)
        所有持仓管理.extend(持仓管理)

        # 模块3: 策略回测
        回测统计 = 执行历史回测(收益率df, 全部日期, 收益率日期, 银行名)
        if 回测统计:
            回测统计['银行'] = 银行名
            所有回测统计.append(回测统计)

    # 步骤5: 生成图表
    print("\n【第5步】生成可视化图表...")
    图表列表 = 生成简化图表(所有实时机会, 所有持仓管理, 所有回测统计)

    # 步骤6: 生成报告
    print("\n【第6步】生成决策报告...")
    生成决策报告(所有实时机会, 所有持仓管理, 所有回测统计, 图表列表)

    # 汇总
    print("\n" + "=" * 70)
    print("                    分析完成")
    print("=" * 70)
    print(f"[*] 实时交易机会: {len(所有实时机会)} 个（现在可操作）")
    print(f"[*] 持仓管理建议: {len(所有持仓管理)} 个（如已买入）")
    print(f"[*] 策略验证: {len(所有回测统计)} 个银行")

    if 所有实时机会:
        立即买入 = sum(1 for x in 所有实时机会 if '立即买入' in x['操作建议'])
        可买入 = sum(1 for x in 所有实时机会 if '可买入' in x['操作建议'])
        print(f"\n  其中：")
        print(f"    [!] 立即买入: {立即买入} 个")
        print(f"    [√] 可买入: {可买入} 个")

    if 所有回测统计:
        print(f"\n策略历史表现:")
        for 统计 in 所有回测统计:
            print(f"  {统计['银行']}: 成功率 {统计['成功率%']:.1f}%, "
                  f"平均收益 {统计['平均收益%']:.2f}%")

    print(f"\n输出文件: {输出文件路径}")
    print(f"图表目录: {图表输出目录}")
    print("=" * 70)


if __name__ == '__main__':
    main()
