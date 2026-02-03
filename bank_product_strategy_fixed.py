# -*- coding: utf-8 -*-
"""
银行理财产品"收益释放"操纵行为监测与套利分析系统 (V2.0 - 修正版)

设计思路：
针对理财子公司短期（1-4周）拉高收益（5%-8%）吸引客户，随后收益回落（1%-1.8%）的现象。
本系统旨在：
1. 捕捉"收益拉升"的启动信号。
2. 基于历史数据分析该产品的"拉升惯性"（平均持续多久，平均收益多少）。
3. 模拟买入持有，实时监控"实际持有收益"与"预期收益"的偏离。
4. 在收益下滑初期或周期结束时发出卖出警报。

修正内容（V2.0）：
- 修复数据加载和产品代码修复逻辑
- 修复收益率序列索引问题
- 修复历史规律分析中的数组访问问题
- 恢复完整的产品类别识别逻辑
- 增强Excel格式化
- 添加空数据保护
"""

import os
import sys
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging
import warnings
import json

# ============================================================
# 日志配置
# ============================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('bank_product_strategy.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
日志 = logging.getLogger(__name__)

# ============================================================
# 核心策略参数
# ============================================================
净值数据库路径 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "净值数据库.xlsx")
输出文件路径 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "高收益捕捉分析报告.xlsx")

# 信号识别参数
收益拉升阈值 = 5.0         # %，只有突破这个值才被视为"拉升行为"
收益下滑阈值 = 2.0         # %，拉升后如果日收益跌破此值，视为"收割开始"

# 周期约束
最短持有天数 = 7           # 天
最长持有天数 = 28          # 天 (4周)
默认预测周期 = 14          # 天，如果没有历史规律，默认预估的拉升时长

# 分析范围
信号回溯天数 = 30          # 只关注最近30天内出现的信号
历史分析回溯 = 365         # 分析过去一年的数据来寻找规律


# ============================================================
# 工具函数
# ============================================================

def 是否日期列(列名):
    """判断列名是否为日期格式 YYYY-MM-DD"""
    if not isinstance(列名, str):
        return False
    if len(列名) == 10 and 列名[4] == '-' and 列名[7] == '-':
        try:
            datetime.strptime(列名, '%Y-%m-%d')
            return True
        except ValueError:
            return False
    return False


def 识别产品类别(产品名称):
    """
    根据产品名称识别产品类别（完整版）

    规则（按优先级判断）:
    1. 直接包含R1-R5标识：按标识分类
    2. 含权益类产品关键词（R3及以上）：
       - 增强、混合、权益、股票、可转债、FOF、偏债、灵活配置、进取、成长
    3. 不含权益类产品关键词（R1/R2）：
       - 纯债、货币、稳健（不含增强）、固收（不含增强/混合）
    4. 无法识别: 未知类别

    返回: '含权益类' 或 '不含权益类' 或 '未知类别'
    """
    if not isinstance(产品名称, str):
        产品名称 = str(产品名称)

    产品名称原始 = 产品名称
    产品名称 = 产品名称.upper()  # 统一转大写处理

    # 优先级1: 检查是否直接包含R1-R5标识
    if 'R3' in 产品名称 or 'R4' in 产品名称 or 'R5' in 产品名称:
        return '含权益类'
    if 'R1' in 产品名称 or 'R2' in 产品名称:
        return '不含权益类'

    # 优先级2: 含权益类产品关键词（这些关键词表示产品含有权益类资产）
    含权益关键词 = ['增强', '混合', '权益', '股票', '可转债', 'FOF', '偏债',
                    '灵活配置', '进取', '成长', '积极']
    for 关键词 in 含权益关键词:
        if 关键词 in 产品名称原始:
            return '含权益类'

    # 优先级3: 不含权益类产品关键词（纯固收产品）
    # 注意：这里需要排除含有"增强"、"混合"的情况（已在上面处理）
    不含权益关键词 = ['纯债', '货币', '稳健', '固收']
    for 关键词 in 不含权益关键词:
        if 关键词 in 产品名称原始:
            return '不含权益类'

    # 无法识别
    return '未知类别'


# ============================================================
# 数据加载与预处理
# ============================================================

def _加载JSON净值数据(银行名):
    """尝试加载银行对应的JSON爬虫数据"""
    基础目录 = os.path.dirname(os.path.abspath(__file__))

    if '华夏' in 银行名:
        路径 = os.path.join(基础目录, 'huaxia_nav_history.json')
        if os.path.exists(路径):
            try:
                with open(路径, 'r', encoding='utf-8') as f:
                    数据 = json.load(f)
                结果 = {}
                for 代码, 净值列表 in 数据.items():
                    净值字典 = {}
                    for 条目 in 净值列表:
                        净值字典[条目['date']] = str(条目['unit_nav'])
                    结果[代码] = 净值字典
                日志.info(f"    已加载华夏JSON数据: {len(结果)} 个产品")
                return 结果
            except Exception as e:
                日志.warning(f"    加载华夏JSON失败: {e}")

    return None


def 修复产品代码(数据库):
    """
    修复数据库中缺失的产品代码和名称（投票匹配法）
    """
    日志.info("开始修复产品代码（投票匹配法）...")

    修复后数据库 = {}
    for 工作表名, df in 数据库.items():
        缺失数 = df['产品代码'].isna().sum()
        if 缺失数 == 0:
            日志.info(f"  {工作表名}: 无缺失，跳过")
            修复后数据库[工作表名] = df
            continue

        日志.info(f"  {工作表名}: {缺失数}/{len(df)} 行缺失产品代码，开始修复...")

        日期列 = sorted([列 for 列 in df.columns if 是否日期列(列)])
        有代码掩码 = df['产品代码'].notna()
        有代码df = df[有代码掩码].copy()
        无代码df = df[~有代码掩码].copy()

        # 构建代码→名称映射
        代码名称映射 = {}
        for _, 行 in 有代码df.iterrows():
            代码 = str(行['产品代码']).strip()
            名称 = str(行['产品名称']).strip()
            if 代码 and 名称:
                代码名称映射[代码] = 名称

        # 构建反向索引
        日志.info(f"    构建反向索引（{len(日期列)} 个日期）...")
        日期净值到代码 = {}

        for _, 行 in 有代码df.iterrows():
            代码 = str(行['产品代码']).strip()
            for 日期 in 日期列:
                值 = 行.get(日期)
                if pd.notna(值) and str(值).strip():
                    键 = (日期, str(值).strip())
                    if 键 not in 日期净值到代码:
                        日期净值到代码[键] = set()
                    日期净值到代码[键].add(代码)

        # 来源2: JSON爬虫数据
        JSON数据 = _加载JSON净值数据(工作表名)
        if JSON数据:
            for 代码, 净值字典 in JSON数据.items():
                for 日期 in 日期列:
                    if 日期 in 净值字典:
                        键 = (日期, 净值字典[日期])
                        if 键 not in 日期净值到代码:
                            日期净值到代码[键] = set()
                        日期净值到代码[键].add(代码)

        # 投票匹配
        已匹配 = 0
        for idx, 行 in 无代码df.iterrows():
            代码投票 = {}
            for 日期 in 日期列:
                值 = 行.get(日期)
                if pd.notna(值) and str(值).strip():
                    键 = (日期, str(值).strip())
                    if 键 in 日期净值到代码:
                        for 候选代码 in 日期净值到代码[键]:
                            代码投票[候选代码] = 代码投票.get(候选代码, 0) + 1

            if 代码投票:
                最佳代码 = max(代码投票, key=代码投票.get)
                无代码df.at[idx, '产品代码'] = 最佳代码
                无代码df.at[idx, '产品名称'] = 代码名称映射.get(最佳代码, 最佳代码)
                已匹配 += 1

        日志.info(f"    匹配成功: {已匹配}/{len(无代码df)}")

        # 合并
        已修复 = pd.concat([有代码df, 无代码df], ignore_index=True)

        # 去重合并
        合并结果 = []
        按代码分组 = 已修复.groupby('产品代码')
        for 代码, 分组 in 按代码分组:
            if len(分组) == 1:
                合并结果.append(分组.iloc[0])
            else:
                合并行 = 分组.iloc[0].copy()
                for 日期 in 日期列:
                    非空值 = 分组[日期].dropna()
                    if len(非空值) > 0:
                        合并行[日期] = 非空值.iloc[0]
                合并结果.append(合并行)

        修复df = pd.DataFrame(合并结果).reset_index(drop=True)
        修复后数据库[工作表名] = 修复df

        日志.info(f"  {工作表名}: 修复完成，{len(修复df)} 行")

    日志.info("产品代码修复完成")
    return 修复后数据库


def 加载净值数据库():
    """加载净值数据库Excel文件"""
    if not os.path.exists(净值数据库路径):
        日志.error(f"找不到文件: {净值数据库路径}")
        return {}

    日志.info("正在加载净值数据库...")
    xlsx = pd.ExcelFile(净值数据库路径)
    数据库 = {}

    for sheet in xlsx.sheet_names:
        df = pd.read_excel(xlsx, sheet_name=sheet, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]

        # 处理索引列名
        重命名映射 = {}
        if 'level_0' in df.columns:
            重命名映射['level_0'] = '产品代码'
        if 'level_1' in df.columns:
            重命名映射['level_1'] = '产品名称'
        if 重命名映射:
            df = df.rename(columns=重命名映射)

        if '产品代码' in df.columns and '产品名称' in df.columns:
            数据库[sheet] = df
            日志.info(f"  {sheet}: {len(df)} 个产品")

    return 数据库


def 计算基础收益率(df, 银行名):
    """
    计算每日年化收益率
    返回: (收益率DataFrame, 全部日期列表, 收益率日期列表)
    """
    全部日期 = sorted([列 for 列 in df.columns if 是否日期列(列)])

    if len(全部日期) < 2:
        日志.warning(f"[{银行名}] 日期数量不足")
        return None, [], []

    # 提取净值矩阵
    净值矩阵 = df[全部日期].apply(pd.to_numeric, errors='coerce')

    # 计算年化收益率
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", FutureWarning)
        收益率矩阵 = 净值矩阵.pct_change(axis=1) * 365 * 100

    # 收益率从第二个日期开始
    收益率日期 = 全部日期[1:]

    # 构建结果DataFrame
    收益率数据 = np.round(收益率矩阵[收益率日期].values, 4)
    结果 = pd.DataFrame(收益率数据, columns=收益率日期)
    结果.insert(0, '产品代码', df['产品代码'].values)
    结果.insert(1, '产品名称', df['产品名称'].values)
    结果.insert(2, '银行', 银行名)
    结果.insert(3, '产品类别', df['产品名称'].apply(识别产品类别))

    # 过滤无效产品
    有效掩码 = 结果[收益率日期].notna().any(axis=1)
    结果 = 结果[有效掩码].reset_index(drop=True)

    日志.info(f"[{银行名}] 收益率计算完成: {len(结果)} 个有效产品")

    return 结果, 全部日期, 收益率日期


# ============================================================
# 历史规律分析模块
# ============================================================

def 分析历史拉升规律(收益率序列, 当前日期, 收益率日期列表):
    """
    回溯该产品过去的历史，寻找类似的"高收益拉升"片段。
    返回: (预期持续天数, 预期年化收益率, 历史发生次数)
    """
    try:
        # 找到当前日期在列表中的位置
        if 当前日期 not in 收益率日期列表:
            return 默认预测周期, 收益拉升阈值, 0

        当前索引 = 收益率日期列表.index(当前日期)

        # 截取历史数据
        历史日期 = 收益率日期列表[:当前索引]
        if len(历史日期) < 30:
            return 默认预测周期, 收益拉升阈值, 0

        历史收益 = [收益率序列.get(日期, 0) for 日期 in 历史日期]

        # 识别高收益片段
        片段列表 = []
        当前片段 = []

        for 收益 in 历史收益:
            if 收益 > (收益拉升阈值 - 0.5):
                当前片段.append(收益)
            else:
                if len(当前片段) >= 3:
                    片段列表.append(当前片段[:])
                当前片段 = []

        if len(当前片段) >= 3:
            片段列表.append(当前片段)

        if not 片段列表:
            return 默认预测周期, 收益拉升阈值, 0

        # 统计规律
        持续天数集 = [len(p) for p in 片段列表]
        平均收益集 = [np.mean(p) for p in 片段列表]

        预期天数 = int(np.mean(持续天数集))
        预期天数 = max(最短持有天数, min(预期天数, 最长持有天数))
        预期收益 = np.mean(平均收益集)

        return 预期天数, 预期收益, len(片段列表)

    except Exception as e:
        日志.warning(f"分析历史规律出错: {e}")
        return 默认预测周期, 收益拉升阈值, 0


# ============================================================
# 策略执行引擎
# ============================================================

def 执行策略分析(收益率df, 全部日期, 收益率日期, 银行名):
    """
    对每个产品执行：信号监测 -> 历史分析 -> 模拟持有 -> 绩效评估
    """
    分析结果 = []
    跟踪详情 = []

    # 只分析最近的信号
    分析起始位 = max(0, len(收益率日期) - 信号回溯天数)

    日志.info(f"[{银行名}] 开始策略分析（{len(收益率df)} 个产品）...")

    处理计数 = 0
    for idx, row in 收益率df.iterrows():
        产品代码 = row['产品代码']
        产品名称 = row['产品名称']
        产品类别 = row['产品类别']

        # 提取收益率序列（转为字典便于按日期访问）
        收益序列 = {}
        for 日期 in 收益率日期:
            try:
                收益 = pd.to_numeric(row[日期], errors='coerce')
                收益序列[日期] = 收益 if pd.notna(收益) else 0
            except:
                收益序列[日期] = 0

        # 遍历分析窗口寻找启动信号
        for i in range(分析起始位, len(收益率日期)):
            当前日期 = 收益率日期[i]
            当前收益 = 收益序列[当前日期]

            # 检查是否是启动信号
            if i > 0:
                昨日日期 = 收益率日期[i-1]
                昨日收益 = 收益序列[昨日日期]
                是启动信号 = (当前收益 > 收益拉升阈值) and (昨日收益 <= 收益拉升阈值)
            else:
                是启动信号 = 当前收益 > 收益拉升阈值

            if not 是启动信号:
                continue

            # 1. 分析历史规律
            预期天数, 预期收益, 历史次数 = 分析历史拉升规律(收益序列, 当前日期, 收益率日期)

            # 2. 模拟持有
            实际持有天数 = 0
            累计收益和 = 0
            卖出触发 = False
            卖出原因 = ""
            当前状态 = "持有中"

            # 向后跟踪
            for j in range(i + 1, len(收益率日期)):
                跟踪日期 = 收益率日期[j]
                跟踪收益 = 收益序列[跟踪日期]

                实际持有天数 += 1
                累计收益和 += 跟踪收益
                当前持有年化 = 累计收益和 / 实际持有天数

                # 检查卖出条件
                if 实际持有天数 >= 预期天数:
                    卖出触发 = True
                    卖出原因 = f"达到预期周期({预期天数}天)"
                    当前状态 = "建议卖出"
                elif 实际持有天数 >= 最短持有天数:
                    if 跟踪收益 < 收益下滑阈值:
                        卖出触发 = True
                        卖出原因 = f"收益下滑(当日{跟踪收益:.2f}%)"
                        当前状态 = "建议卖出"
                    elif 当前持有年化 < (预期收益 * 0.6):
                        卖出触发 = True
                        卖出原因 = "实际表现不及预期"
                        当前状态 = "建议卖出"

                # 记录跟踪详情
                跟踪详情.append({
                    '银行': 银行名,
                    '产品代码': 产品代码,
                    '产品名称': 产品名称,
                    '产品类别': 产品类别,
                    '信号日期': 当前日期,
                    '跟踪日期': 跟踪日期,
                    '持有天数': 实际持有天数,
                    '当日收益': round(跟踪收益, 2),
                    '当前持有年化': round(当前持有年化, 2),
                    '卖出触发': '是' if 卖出触发 else ''
                })

                if 卖出触发:
                    break

            # 判断最终状态
            if not 卖出触发:
                if 实际持有天数 >= 最长持有天数:
                    当前状态 = "强制卖出(超期)"
                else:
                    当前状态 = "建议持有"

            最终持有年化 = (累计收益和 / 实际持有天数) if 实际持有天数 > 0 else 0

            # 记录信号事件
            分析结果.append({
                '银行': 银行名,
                '产品代码': 产品代码,
                '产品名称': 产品名称,
                '产品类别': 产品类别,
                '信号日期': 当前日期,
                '启动收益率%': round(当前收益, 2),
                '历史拉升次数': 历史次数,
                '预期维持天数': 预期天数,
                '预期年化收益%': round(预期收益, 2),
                '最新状态': 当前状态,
                '已持有天数': 实际持有天数,
                '当前持有年化%': round(最终持有年化, 2),
                '收益差(实际-预期)': round(最终持有年化 - 预期收益, 2),
                '卖出原因': 卖出原因,
                '操作建议': '立即买入' if 实际持有天数 == 0 else 当前状态
            })

        处理计数 += 1
        if 处理计数 % 500 == 0:
            日志.info(f"  已处理 {处理计数}/{len(收益率df)} 个产品")

    日志.info(f"[{银行名}] 策略分析完成: 捕获 {len(分析结果)} 个信号")
    return 分析结果, 跟踪详情


# ============================================================
# Excel报告生成
# ============================================================

def 生成Excel报告(分析结果列表, 跟踪详情列表):
    """生成格式化的Excel报告"""
    if not 分析结果列表:
        日志.warning("没有任何信号数据，跳过报告生成")
        return

    日志.info("正在生成Excel报告...")

    wb = Workbook()

    # 样式定义
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="4F81BD")
    align_center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def 设置表头(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border

    def 设置列宽(ws, 列宽字典):
        for 列号, 宽度 in 列宽字典.items():
            ws.column_dimensions[get_column_letter(列号)].width = 宽度

    # Sheet 1: 高收益机会监控
    ws1 = wb.active
    ws1.title = "高收益机会监控"
    headers1 = ['银行', '产品代码', '产品名称', '产品类别', '信号日期', '启动收益率%',
                '历史拉升次数', '预期维持天数', '预期年化收益%',
                '最新状态', '已持有天数', '当前持有年化%', '收益差(实际-预期)', '卖出原因', '操作建议']
    设置表头(ws1, headers1)

    # 按当前持有年化倒序
    分析结果列表 = sorted(分析结果列表, key=lambda x: x['当前持有年化%'], reverse=True)

    for row in 分析结果列表:
        ws1.append([row[k] for k in headers1])

    设置列宽(ws1, {
        1: 12, 2: 22, 3: 40, 4: 14, 5: 14, 6: 14,
        7: 14, 8: 14, 9: 16, 10: 14, 11: 12,
        12: 16, 13: 16, 14: 30, 15: 14
    })
    ws1.freeze_panes = 'A2'

    # Sheet 2: 持有期逐日跟踪
    ws2 = wb.create_sheet("持有期逐日跟踪")
    headers2 = ['银行', '产品代码', '产品名称', '产品类别', '信号日期', '跟踪日期',
                '持有天数', '当日收益', '当前持有年化', '卖出触发']
    设置表头(ws2, headers2)

    for row in 跟踪详情列表:
        ws2.append([row[k] for k in headers2])

    设置列宽(ws2, {
        1: 12, 2: 22, 3: 40, 4: 14, 5: 14, 6: 14,
        7: 12, 8: 14, 9: 16, 10: 12
    })
    ws2.freeze_panes = 'A2'

    # 保存
    try:
        wb.save(输出文件路径)
        日志.info(f"报告已生成: {输出文件路径}")
        日志.info(f"  Sheet1 高收益机会监控: {len(分析结果列表)} 行")
        日志.info(f"  Sheet2 持有期逐日跟踪: {len(跟踪详情列表)} 行")
    except Exception as e:
        日志.error(f"保存文件失败: {e}")


# ============================================================
# 主程序
# ============================================================

def main():
    print("=" * 70)
    print("   银行理财产品收益释放操纵行为监测与套利分析系统 V2.0")
    print("=" * 70)
    print(f"运行时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"数据来源: {净值数据库路径}")
    print(f"输出文件: {输出文件路径}")
    print(f"收益拉升阈值: {收益拉升阈值}%")
    print(f"收益下滑阈值: {收益下滑阈值}%")
    print(f"持有期约束: {最短持有天数}-{最长持有天数}天")
    print(f"信号回溯: {信号回溯天数}天")
    print()

    # 步骤1: 加载数据
    print("【第1步】加载净值数据库...")
    数据库 = 加载净值数据库()
    if not 数据库:
        return

    # 步骤2: 修复产品代码
    print("\n【第2步】修复产品代码...")
    数据库 = 修复产品代码(数据库)

    # 步骤3: 计算收益率 + 执行策略
    print("\n【第3步】计算收益率并执行策略分析...")
    所有信号事件 = []
    所有跟踪详情 = []

    for 银行名, df in 数据库.items():
        收益率df, 全部日期, 收益率日期 = 计算基础收益率(df, 银行名)
        if 收益率df is None:
            continue

        信号, 跟踪 = 执行策略分析(收益率df, 全部日期, 收益率日期, 银行名)
        所有信号事件.extend(信号)
        所有跟踪详情.extend(跟踪)

    # 步骤4: 生成报告
    print("\n【第4步】生成分析报告...")
    生成Excel报告(所有信号事件, 所有跟踪详情)

    # 统计汇总
    print("\n" + "=" * 70)
    print("                    分析完成")
    print("=" * 70)
    print(f"捕获拉升信号总数: {len(所有信号事件)}")
    print(f"跟踪期记录总数:   {len(所有跟踪详情)}")

    # 按类别统计
    if 所有信号事件:
        类别统计 = {}
        for 信号 in 所有信号事件:
            类别 = 信号['产品类别']
            类别统计[类别] = 类别统计.get(类别, 0) + 1

        print("\n产品类别分布:")
        for 类别, 数量 in sorted(类别统计.items()):
            print(f"  {类别}: {数量} 个信号")

    print(f"\n输出文件: {输出文件路径}")
    print("=" * 70)


if __name__ == '__main__':
    main()
