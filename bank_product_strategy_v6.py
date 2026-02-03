# -*- coding: utf-8 -*-
"""
银行理财产品"收益释放"实时交易决策系统 (V8.0 - Liquidity Master + SOPM v1)

核心改进：
1. 流动性过滤：只保留赎回周期≤14天的产品（日开/7天/14天持有期）
2. 正确年化：期间收益 / 实际天数 × 365，适配不同数据频率
3. 间隔感知：只在真实净值数据点之间计算收益，消除NaN→0虚假信号
4. 高成功率产品库缓存 + 预期收益预测 + 持仓跟踪
5. 潜力观察池：信号不足或成功率待提升的产品单独跟踪，避免遗漏优质产品
6. V5.27优化：6仓集中(1667万/仓) + 高买入门槛(3.5%) + 20天快速轮动 + 动态仓位
7. V5.28: 同步回测双阈值(库3.0%+买入3.5%)，精选候选池从3.0%起步，覆盖更多银行
8. 精选推荐Top20：8维度综合评分，多维度优中选优
9. V6.0: 前瞻预测Alpha加成 — 学习产品释放规律(周期/阶段/星期)，
   预测即将释放的产品，精选推荐时加成排名（回测验证: 年化+0.12%, 夏普1.26→1.52）

作者：AI-FINANCE
版本：V8.1
日期：2026-01-30
"""

import os
import sys
import json
import re
import bisect
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple
from scipy import stats as scipy_stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import logging
import warnings
import ctypes

from aifinance_lib.database import Database
from aifinance_lib.strategy.data_loader import load_prepared_data

# V13: 使用统一费率引擎 (Single Source of Truth)
import fee_engine
from fee_engine import (
    get_redemption_rate,
    calculate_net_yield,
    calculate_fee_drag,
    check_liquidity_cost_veto,
    has_redemption_fee,
    get_fee_info,
)


# 兼容性别名：旧代码中的 calc_net_return 改为调用 fee_engine
def calc_net_return(gross_return, fee_rate, holding_days=14):
    """[兼容性封装] 调用 fee_engine.calculate_net_yield"""
    return calculate_net_yield(gross_return, holding_days, fee_rate=fee_rate)


# 兼容性别名：旧代码中的 get_fee_rate 改为调用 fee_engine
def get_fee_rate(bank, code, days_held):
    """[兼容性封装] 调用 fee_engine.get_redemption_rate"""
    return get_redemption_rate(code, days_held, bank=bank)


# ============================================================
# 配置
# ============================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('bank_product_strategy_v5.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# 文件路径
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "净值数据库.xlsx")
OUTPUT_PATH = os.path.join(BASE_DIR, "高成功率交易机会_V6.xlsx")
CACHE_PATH = os.path.join(BASE_DIR, "高成功率产品库_缓存.json")
PORTFOLIO_PATH = os.path.join(BASE_DIR, "我的持仓_模板.xlsx")

# 核心策略参数（V6.0 - 前瞻预测Alpha加成，经回测验证年化提升+0.12%）
# 核心改进: 6仓集中 + 高买入门槛(3.5%) + 20天快速轮动 + 动态仓位
# V5.28: 同步回测双阈值 — 库建设用3.0%(宽入), 买入推荐用3.5%(严选)
# V6.0: 前瞻预测Alpha加成 — 学习释放规律，预测即将释放的产品加成排名
信号阈值_买入 = 3.5          # 买入信号: 年化收益率 > 3.5%（高质量信号, ★推荐）
信号阈值_卖出 = 2.0          # 卖出信号: 年化收益率 ≤ 2.0% 时赎回
信号阈值_库 = 3.0            # 产品库/精选信号检测（宽入库，与回测一致）
信号阈值 = 信号阈值_买入      # 兼容旧引用 (3.5%)
需要突破信号 = False          # False=只要>买入阈值就可买; True=要求前日≤阈值(突破信号)
持有天数 = 7                  # 信号后持有7天评估成功
成功标准 = 2.5                # 持有期平均年化 > 2.5% 算成功
最低成功率 = 30.0             # 推荐历史成功率 >= 30%（V5.27: 宽入库，用排名区分优劣）
最少历史信号 = 2              # 至少有2次历史信号
实时窗口 = 10                 # 最近10天内的新信号
缓存有效天数 = 1              # 缓存1天有效（每日更新）

# 轮动策略约束
最大持有天数 = 20             # 预期持有天数上限（V5.27: 从25降至20，加速轮动）
最短持续天数 = 2              # 信号至少持续2天
最长赎回天数 = 14             # 产品赎回周期不超过14天

# 潜力观察池参数
观察池最低信号 = 1            # 至少有1次信号
观察池最低成功率 = 20.0       # 成功率>=20%（低于正式库30%但有潜力）

# 前瞻预测参数 (建议① Alpha加成, 经回测验证: 年化+0.12%, 夏普1.26→1.52)
规律学习窗口天数 = 180
释放识别阈值 = 2.5            # 识别释放事件的收益阈值(%)
最低置信度 = 0.4              # 低于此不生成预测
预测窗口天数 = 10             # 预测未来N天
周期CV阈值 = 0.4              # 变异系数<此值视为有周期
ALPHA_预测加成 = 0.40         # V4修正: 回测最优ALPHA=0.40, Mode B年化+6.13%, 夏普2.45

# 宏观风控（硬限制）
信用利差阈值 = 150          # 信用利差 > 150bp 时全部赎回
CREDIT_SPREAD_FILE = os.path.join(BASE_DIR, "信用利差.json")

# ===== V6.1: 智持参数 (建议②) =====
延迟卖出_最低收益 = 0.0
延迟卖出_预测窗口 = 4
延迟卖出_置信度 = 0.4
延迟卖出_最大持有天数余量 = 5

# ===== V8.1: 闪跌止损参数 (Dynamic Trailing Stop) =====
闪跌止损_回看天数 = 3              # 回看最近N天寻找收益峰值
闪跌止损_阈值 = -0.05              # 从峰值下跌超过此值(%)触发止损建议

# ===== V6.2: 信号新鲜度过滤参数 (统一在 shared_freshness.py) =====
import shared_freshness as sf

# ===== V6.1: 节前买入参数 (建议③) =====
节前买入_实时阈值 = 1.0
节前买入_最低成功率 = 40.0
节前买入_最低均收 = 1.5
节前买入_最大产品数 = 5

# ===== 中国法定假期日历 =====
HOLIDAY_PERIODS = [
    ('端午', '2025-05-31', '2025-06-02'),
    ('国庆', '2025-10-01', '2025-10-08'),
    ('元旦', '2026-01-01', '2026-01-03'),
]
PRE_HOLIDAY_DATES = {
    '2025-09-29': '国庆',
    '2025-09-30': '国庆',
    '2025-12-31': '元旦',
}
HOLIDAY_DATE_SET = set()
for _name, _start, _end in HOLIDAY_PERIODS:
    _s = pd.Timestamp(_start)
    _e = pd.Timestamp(_end)
    _d = _s
    while _d <= _e:
        HOLIDAY_DATE_SET.add(_d.strftime('%Y-%m-%d'))
        _d += pd.Timedelta(days=1)


def is_pre_holiday(date_str):
    """判断是否为节前交易日, 返回 (是否节前, 假期名称)"""
    if date_str in PRE_HOLIDAY_DATES:
        return True, PRE_HOLIDAY_DATES[date_str]
    return False, ''


def is_holiday(date_str):
    """判断是否为假期日, 返回 (是否假期, 假期名称)"""
    if date_str in HOLIDAY_DATE_SET:
        for name, start, end in HOLIDAY_PERIODS:
            if start <= date_str <= end:
                return True, name
        return True, '假期'
    return False, ''


# ============================================================
# 工具函数
# ============================================================

def is_date_col(col_name):
    if not isinstance(col_name, str):
        return False
    return len(col_name) == 10 and col_name[4] == '-' and col_name[7] == '-'


def 判断产品流动性(product_name):
    """从产品名称解析赎回周期，返回 (类型描述, 赎回天数)

    赎回天数越小越适合轮动策略：
    - 0: 日开/天天/每日（最佳）
    - 7: 7天持有期/周开
    - 14~180: 各种持有期
    - 999: 封闭式/无法确定
    """
    name = product_name

    # 日开型（最优先）
    if '日开' in name or '天天' in name or '每日' in name:
        return '日开', 0

    # N天持有期
    m = re.search(r'(\d+)\s*天持有期', name)
    if m:
        days = int(m.group(1))
        return f'{days}天持有期', days

    # 最短持有N天
    m = re.search(r'最短持有\s*(\d+)\s*天', name)
    if m:
        days = int(m.group(1))
        return f'{days}天持有', days

    # 周期N天
    m = re.search(r'周期\s*(\d+)\s*天', name)
    if m:
        days = int(m.group(1))
        return f'{days}天周期', days

    # N个月/N月
    m = re.search(r'(\d+)\s*个?月', name)
    if m and '半年' not in name:
        months = int(m.group(1))
        if months <= 6:
            return f'{months}个月', months * 30

    # 周开
    if '周开' in name:
        return '周开', 7

    # 日申季赎 / 月开 / 季开 / 半年开 / 年开
    if '日申季赎' in name:
        return '季赎', 90
    if '月开' in name:
        return '月开', 30
    if '季开' in name or '季度开' in name:
        return '季开', 90
    if '半年' in name:
        return '半年', 180
    if '年开' in name:
        return '年开', 365
    if '封闭' in name:
        return '封闭', 999
    if '定开' in name:
        return '定开', 999

    # 含有明确年份（如"1Y持有"）
    m = re.search(r'(\d+)\s*Y\s*持有', name)
    if m:
        years = int(m.group(1))
        return f'{years}年持有', years * 365

    # 信颐2030/2035等目标日期基金
    if re.search(r'20[2-5]\d', name) and '信颐' in name:
        return '目标日期', 999

    # 无法判断 → 保守标记为长期
    return '未知', 999


def 获取信用利差():
    """读取当前信用利差(bp)

    数据来源：信用利差.json，格式：
    {"日期": "2026-01-29", "利差bp": 85, "基准": "3Y AA+", "备注": ""}

    返回: (利差bp, 日期, 基准) 或 (None, None, None)
    """
    if not os.path.exists(CREDIT_SPREAD_FILE):
        return None, None, None
    try:
        with open(CREDIT_SPREAD_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data.get('利差bp'), data.get('日期'), data.get('基准', '3Y AA+')
    except Exception as e:
        logger.warning(f"读取信用利差失败: {e}")
        return None, None, None


def 检查信用利差风控():
    """检查信用利差是否触发全部赎回

    Returns:
        (触发, 利差, 日期, 基准): 触发=True时应全部赎回
    """
    利差, 日期, 基准 = 获取信用利差()
    if 利差 is None:
        logger.info("未找到信用利差数据（信用利差.json），跳过风控检查")
        return False, None, None, None

    触发 = 利差 > 信用利差阈值
    if 触发:
        logger.warning(f"!! 信用利差风控触发！当前{基准}利差={利差}bp > 阈值{信用利差阈值}bp，建议全部赎回")
    else:
        logger.info(f"信用利差检查通过: {基准}={利差}bp ≤ {信用利差阈值}bp（{日期}）")

    return 触发, 利差, 日期, 基准


def calc_returns(nav_df, date_cols):
    nav_matrix = nav_df[date_cols].apply(pd.to_numeric, errors='coerce')
    with warnings.catch_warnings():
        warnings.simplefilter('ignore')
        ret_matrix = nav_matrix.pct_change(axis=1) * 365 * 100
    return ret_matrix


def get_product_returns(row, date_cols):
    """获取产品的有效年化收益率序列（跳过数据空缺）

    核心改进：只在有真实净值数据的相邻日期之间计算收益率。
    - 消除NaN→0的虚假信号（空缺日不再被当作0%收益）
    - 正确年化：期间收益 / 实际天数 × 365，适配不同银行数据频率
    - 长间隔保护：间隔超过60天的跳过（跨季度的异常数据）

    Returns:
        dict: {date_str: annualized_return_pct}，仅包含有效数据点
    """
    nav_points = []
    for d in date_cols:
        try:
            v = float(row[d])
            if not np.isnan(v) and v > 0:
                nav_points.append((d, v))
        except (ValueError, TypeError):
            pass

    if len(nav_points) < 2:
        return {}

    rets = {}
    for i in range(1, len(nav_points)):
        d0, n0 = nav_points[i - 1]
        d1, n1 = nav_points[i]
        day_gap = (pd.Timestamp(d1) - pd.Timestamp(d0)).days
        if day_gap > 60:
            continue  # 跨季度数据，跳过
        if day_gap > 0 and n0 > 0:
            # 正确年化：期间收益除以实际天数，再乘365
            ann_ret = (n1 / n0 - 1) / day_gap * 365 * 100
            rets[d1] = ann_ret

    return rets


# ============================================================
# 前瞻预测: 释放规律学习 + 预测 (建议① Alpha加成)
# ============================================================

@dataclass
class ReleasePattern:
    product_key: tuple
    period_days: float = 0.0
    period_cv: float = 1.0
    has_period: bool = False
    phase_dist: list = field(default_factory=lambda: [0.0]*4)
    phase_pvalue: float = 1.0
    top_phase: int = 0
    weekday_dist: list = field(default_factory=lambda: [0.0]*5)
    weekday_pvalue: float = 1.0
    top_weekday: int = 0
    confidence: float = 0.0
    n_events: int = 0
    last_release_date: str = ''
    last_release_window_end: str = ''
    avg_window_days: float = 3.0      # 平均释放窗口持续天数
    # V5: Adaptive Window & Seasonality
    effective_window_days: int = 180
    regime_change_detected: bool = False
    quarterly_dist: list = field(default_factory=lambda: [0.0]*3)
    quarterly_pvalue: float = 1.0
    top_quarter_month: int = 0        # 0=季初月, 1=季中月, 2=季末月


# ── 工作日映射工具函数 ──
_bday_cache = {}


def _build_bday_index(dates):
    """构建工作日索引: calendar_date → business_day_rank
    返回 (bday_list, bday_rank) — bday_list 是排序后的工作日列表,
    bday_rank 是 {date_str: int} 映射 (0-based rank)
    """
    cache_key = (min(dates), max(dates))
    if cache_key in _bday_cache:
        return _bday_cache[cache_key]
    d_min = pd.Timestamp(min(dates)) - pd.Timedelta(days=10)
    d_max = pd.Timestamp(max(dates)) + pd.Timedelta(days=10)
    bdays = pd.bdate_range(d_min, d_max)
    bday_list = [d.strftime('%Y-%m-%d') for d in bdays]
    bday_rank = {d: i for i, d in enumerate(bday_list)}
    _bday_cache[cache_key] = (bday_list, bday_rank)
    return bday_list, bday_rank


def _snap_to_bday(date_str, bday_list):
    """将日期对齐到最近的前一个(含当天)工作日 (T-1 convention)
    周末 → 回退到周五, 用 bisect 查找"""
    idx = bisect.bisect_right(bday_list, date_str)
    if idx == 0:
        return bday_list[0]
    return bday_list[idx - 1]


def _bday_of_month(date_str, bday_list, bday_rank):
    """计算日期在其月份内的工作日序号 (1-based, 范围 1-23)
    月初第一个工作日为 1, 之后递增"""
    ts = pd.Timestamp(date_str)
    month_start_str = f'{ts.year}-{ts.month:02d}-01'
    # 找到该月第一个工作日
    si = bisect.bisect_left(bday_list, month_start_str)
    if si >= len(bday_list):
        return 1
    rank_of_date = bday_rank.get(date_str, 0)
    rank_of_month_start = bday_rank.get(bday_list[si], 0)
    return rank_of_date - rank_of_month_start + 1


def _detect_regime_change(gaps, min_recent=3, min_old=3, deviation_threshold=0.30):
    """检测释放频率的结构性变化 (Regime Change Detection)

    将 gaps 序列按时间分为 "近期" (最后 ~1/4) 和 "早期" (前 ~3/4)，
    比较均值和方差。偏差 > threshold 视为 regime change。

    返回: (detected: bool, split_idx: int or None)
        split_idx — 变化点索引 (gaps 列表中的位置), 之后的数据为新 regime
    """
    if len(gaps) < min_recent + min_old:
        return False, None

    # 分割点: 最近 25% 的 gaps 为 "近期窗口"
    split = max(len(gaps) - max(len(gaps) // 4, min_recent), min_old)
    old_gaps = gaps[:split]
    new_gaps = gaps[split:]

    if len(old_gaps) < min_old or len(new_gaps) < min_recent:
        return False, None

    old_mean = float(np.mean(old_gaps))
    new_mean = float(np.mean(new_gaps))
    old_std = float(np.std(old_gaps))
    new_std = float(np.std(new_gaps))

    if old_mean == 0:
        return False, None

    # 均值偏差
    mean_dev = abs(new_mean - old_mean) / old_mean
    # 方差偏差 (如果旧方差为0则只看均值)
    std_dev = abs(new_std - old_std) / max(old_std, 1.0)

    # 任一偏差超过阈值 → regime change
    detected = mean_dev > deviation_threshold or std_dev > deviation_threshold
    return detected, split if detected else None


class PatternLearner:
    PHASE_RANGES = [(1, 7), (8, 14), (15, 21), (22, 31)]
    BDAY_PHASE_RANGES = [(1, 5), (6, 11), (12, 16), (17, 23)]
    QUARTER_MONTH_NAMES = ['季初月(1/4/7/10)', '季中月(2/5/8/11)', '季末月(3/6/9/12)']

    def learn(self, product_key, ret_dates, rets, as_of_date, window_days=180):
        cutoff = (pd.Timestamp(as_of_date) - pd.Timedelta(days=window_days)).strftime('%Y-%m-%d')
        wdates = [d for d in ret_dates if cutoff <= d < as_of_date]
        if len(wdates) < 10:
            return None

        # 检测全窗口内的所有释放起始点
        starts = []
        prev_below = True
        for d in wdates:
            r = rets.get(d, 0)
            if r > 释放识别阈值:
                if prev_below:
                    starts.append(d)
                prev_below = False
            else:
                prev_below = True

        if len(starts) < 2:
            return None

        bday_list, bday_rank = _build_bday_index(wdates)
        regime_detected = False
        effective_days = (pd.Timestamp(as_of_date) - pd.Timestamp(wdates[0])).days

        pat = ReleasePattern(product_key=product_key)
        pat.n_events = len(starts)
        pat.last_release_date = starts[-1]
        pat.effective_window_days = effective_days
        pat.regime_change_detected = regime_detected

        # 计算每个释放事件的窗口持续天数
        window_durations = []
        for s_date in starts:
            si = wdates.index(s_date)
            ei = si
            for j in range(si, len(wdates)):
                if rets.get(wdates[j], 0) > 释放识别阈值 * 0.5:
                    ei = j
                else:
                    break
            dur = (pd.Timestamp(wdates[ei]) - pd.Timestamp(s_date)).days + 1
            window_durations.append(max(dur, 1))
        pat.avg_window_days = float(np.mean(window_durations)) if window_durations else 3.0

        # 最后一个释放窗口的结束日
        last_si = wdates.index(starts[-1])
        last_ei = last_si
        for j in range(last_si, len(wdates)):
            if rets.get(wdates[j], 0) > 释放识别阈值 * 0.5:
                last_ei = j
            else:
                break
        pat.last_release_window_end = wdates[last_ei]

        # ── 周期 (工作日间隔, 使用 adaptive window 后的 starts) ──
        snapped_starts = [_snap_to_bday(d, bday_list) for d in starts]
        gaps = []
        for i in range(1, len(snapped_starts)):
            r1 = bday_rank.get(snapped_starts[i-1], 0)
            r2 = bday_rank.get(snapped_starts[i], 0)
            g = r2 - r1
            if 2 <= g <= 85:
                gaps.append(g)
        if len(gaps) >= 2:
            med = float(np.median(gaps))
            cv = float(np.std(gaps)) / med if med > 0 else 1.0
            pat.period_days = med
            pat.period_cv = cv
            pat.has_period = cv < 周期CV阈值
        elif len(gaps) == 1:
            pat.period_days = float(gaps[0])
            pat.period_cv = 0.5
            pat.has_period = True

        # ── 阶段 (工作日序号) ──
        pc = [0]*4
        for d in starts:
            snapped = _snap_to_bday(d, bday_list)
            bdom = _bday_of_month(snapped, bday_list, bday_rank)
            for pi, (lo, hi) in enumerate(self.BDAY_PHASE_RANGES):
                if lo <= bdom <= hi:
                    pc[pi] += 1; break
        ne = sum(pc)
        if ne >= 3:
            _, pv = scipy_stats.chisquare(pc, f_exp=[ne/4.0]*4)
            pat.phase_dist = [c/ne for c in pc]
            pat.phase_pvalue = pv
            pat.top_phase = int(np.argmax(pc))
        else:
            pat.phase_dist = [c/max(ne,1) for c in pc]
            pat.phase_pvalue = 1.0
            pat.top_phase = int(np.argmax(pc)) if ne > 0 else 0

        # ── 星期 (snap到工作日后取weekday) ──
        wc = [0]*5
        for d in starts:
            snapped = _snap_to_bday(d, bday_list)
            wd = pd.Timestamp(snapped).weekday()
            if wd < 5: wc[wd] += 1
        nw = sum(wc)
        if nw >= 3:
            _, pw = scipy_stats.chisquare(wc, f_exp=[nw/5.0]*5)
            pat.weekday_dist = [c/nw for c in wc]
            pat.weekday_pvalue = pw
            pat.top_weekday = int(np.argmax(wc))
        else:
            pat.weekday_dist = [c/max(nw,1) for c in wc]
            pat.weekday_pvalue = 1.0
            pat.top_weekday = int(np.argmax(wc)) if nw > 0 else 0

        # ── 置信度 ──
        conf = 0.0
        if pat.has_period and pat.period_cv < 0.3: conf += 0.4
        elif pat.has_period and pat.period_cv < 0.4: conf += 0.25
        if pat.phase_pvalue < 0.05: conf += 0.3
        if pat.weekday_pvalue < 0.05: conf += 0.15
        if pat.n_events >= 5: conf += 0.15
        pat.confidence = min(conf, 1.0)
        return pat


@dataclass
class Prediction:
    product_key: tuple
    predicted_date: str
    confidence: float
    source: str
    td_until: int = 0
    predicted_end_date: str = ''


class ForwardPredictor:

    def predict(self, pattern, current_date, all_dates, date_to_idx):
        if pattern.confidence < 最低置信度:
            return []
        cur_ts = pd.Timestamp(current_date)
        horizon = cur_ts + pd.Timedelta(days=预测窗口天数)
        cur_idx = date_to_idx.get(current_date)
        if cur_idx is None:
            ci = bisect.bisect_left(list(date_to_idx.keys()), current_date)
            if ci < len(date_to_idx):
                cur_idx = ci
            else:
                return []

        preds = {}

        # 路径A: 周期预测 (工作日偏移)
        if pattern.has_period and pattern.last_release_window_end:
            last_ts = pd.Timestamp(pattern.last_release_window_end)
            period_bdays = max(int(pattern.period_days), 3)
            period = pd.tseries.offsets.BDay(period_bdays)
            pt = last_ts + period
            for _ in range(30):
                if pt >= cur_ts - pd.Timedelta(days=2):
                    break
                pt += period
            conf_a = (1 - pattern.period_cv) * 0.7
            for off in range(-2, 3):
                ct = pt + pd.Timedelta(days=off)
                if cur_ts < ct <= horizon:
                    pd_str = ct.strftime('%Y-%m-%d')
                    pi = bisect.bisect_left(all_dates, pd_str)
                    if pi >= len(all_dates): continue
                    td = pi - cur_idx
                    if td < 1: continue
                    if pd_str not in preds:
                        preds[pd_str] = Prediction(pattern.product_key, pd_str, conf_a, 'period', td)
                    else:
                        preds[pd_str].confidence = min(preds[pd_str].confidence + conf_a, 1.0)
                        preds[pd_str].source = 'both'
                        preds[pd_str].td_until = min(preds[pd_str].td_until, td)

        # 路径B: 阶段预测 (工作日中点 + 宽窗口)
        if pattern.phase_pvalue < 0.15:
            tp = pattern.top_phase
            lo, hi = PatternLearner.BDAY_PHASE_RANGES[tp]
            bday_mid = (lo + hi) // 2
            prob = pattern.phase_dist[tp] if tp < len(pattern.phase_dist) else 0.25
            conf_b = prob * 0.6
            for mo in range(2):
                ref = cur_ts + pd.DateOffset(months=mo)
                try:
                    month_start = pd.Timestamp(f'{ref.year}-{ref.month:02d}-01')
                    mt = month_start + pd.tseries.offsets.BDay(bday_mid - 1)
                except (ValueError, KeyError):
                    continue
                ws = mt - pd.Timedelta(days=2)
                we = mt + pd.Timedelta(days=3)
                if we <= cur_ts or ws > horizon: continue
                pd_str = mt.strftime('%Y-%m-%d')
                pi = bisect.bisect_left(all_dates, pd_str)
                if pi >= len(all_dates): continue
                td = pi - cur_idx
                if td < 1: continue
                if pd_str not in preds:
                    preds[pd_str] = Prediction(pattern.product_key, pd_str, conf_b, 'phase', td)
                else:
                    preds[pd_str].confidence = min(preds[pd_str].confidence + conf_b, 1.0)
                    preds[pd_str].source = 'both'
                    preds[pd_str].td_until = min(preds[pd_str].td_until, td)

        # 为每个预测计算释放结束日 = 预测释放日 + 平均窗口天数
        avg_win = max(int(round(pattern.avg_window_days)), 1)
        for p in preds.values():
            end_ts = pd.Timestamp(p.predicted_date) + pd.Timedelta(days=avg_win - 1)
            p.predicted_end_date = end_ts.strftime('%Y-%m-%d')

        results = [p for p in preds.values() if p.confidence >= 0.25]
        results.sort(key=lambda x: x.td_until)
        return results

    def rank_products(self, patterns, current_date, all_dates, date_to_idx):
        """按预测紧迫度排名: score = confidence / sqrt(td_until)"""
        rankings = []
        for key, pat in patterns.items():
            if pat.confidence < 最低置信度: continue
            ps = self.predict(pat, current_date, all_dates, date_to_idx)
            if not ps: continue
            best_s, best_p = 0.0, None
            for p in ps:
                if p.td_until < 1 or p.td_until > 8: continue
                s = p.confidence / max(p.td_until, 1) ** 0.5
                if s > best_s: best_s, best_p = s, p
            if best_p and best_s > 0.1:
                rankings.append((key, best_s, best_p))
        rankings.sort(key=lambda x: -x[1])
        return rankings

    def has_upcoming_release(self, pattern, current_date, all_dates, date_to_idx,
                             max_td=4, min_conf=0.4):
        """检查是否有即将到来的释放 (用于智持/延迟卖出判断)"""
        ps = self.predict(pattern, current_date, all_dates, date_to_idx)
        for p in ps:
            if 1 <= p.td_until <= max_td and p.confidence >= min_conf:
                return True, p
        return False, None


# ============================================================
# 高成功率产品库（带缓存）
# ============================================================

def 加载缓存():
    """加载缓存的高成功率产品库和潜力观察池"""
    if not os.path.exists(CACHE_PATH):
        return None, None

    try:
        with open(CACHE_PATH, 'r', encoding='utf-8') as f:
            cache = json.load(f)

        # 检查缓存是否过期
        cache_date = datetime.strptime(cache['生成日期'], '%Y-%m-%d')
        if (datetime.now() - cache_date).days > 缓存有效天数:
            logger.info("缓存已过期，需要重新计算")
            return None, None

        # 检查参数是否变化
        cached_params = cache.get('参数', {})
        current_params = {'信号阈值_买入': 信号阈值_买入, '信号阈值_卖出': 信号阈值_卖出,
                          '信号阈值_库': 信号阈值_库,
                          '持有天数': 持有天数, '成功标准': 成功标准, '最低成功率': 最低成功率}
        if cached_params != current_params:
            logger.info(f"策略参数已变化，需要重新计算（缓存参数: {cached_params}）")
            return None, None

        产品库 = cache['产品库']
        观察池 = cache.get('潜力观察池', [])
        logger.info(f"加载缓存成功，生成日期: {cache['生成日期']}，"
                    f"产品库{len(产品库)}个，观察池{len(观察池)}个")
        return 产品库, 观察池
    except Exception as e:
        logger.warning(f"加载缓存失败: {e}")
        return None, None


def 保存缓存(产品库, 潜力观察池=None):
    """保存高成功率产品库和潜力观察池到缓存"""
    cache = {
        '生成日期': datetime.now().strftime('%Y-%m-%d'),
        '参数': {
            '信号阈值_买入': 信号阈值_买入,
            '信号阈值_卖出': 信号阈值_卖出,
            '信号阈值_库': 信号阈值_库,
            '持有天数': 持有天数,
            '成功标准': 成功标准,
            '最低成功率': 最低成功率
        },
        '产品库': 产品库,
        '潜力观察池': 潜力观察池 or []
    }

    with open(CACHE_PATH, 'w', encoding='utf-8') as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)

    logger.info(f"缓存已保存: {CACHE_PATH}")


# ============================================================
# Anti-Hook: 脉冲质量检测 (VIP Sniper Edition)
# ============================================================

def _calc_pulse_metrics(rets, ret_dates_valid):
    """Anti-Hook: 计算脉冲质量 + 高收益密度 (V7.1 放宽版)

    两级过滤 — Hard Reject + Soft Penalty:
      Hard Reject (直接跳过):
        - 产品太新 (<60天)
        - 噪声脉冲 (宽度<2天)
        - 粉饰窗口 (宽度>25天 且 密度≤40%)
        - 极端炒作 (炒作比>2.0)
      Soft Penalty (保留但降权 quality_penalty=0.8):
        - 成熟度偏低 (60-90天)
        - 轻度炒作 (1.5<炒作比≤2.0)

    Args:
        rets: dict {date_str: yield_pct}
        ret_dates_valid: sorted list of date strings

    Returns:
        dict with pulse metrics, rejection info, and quality_penalty
    """
    result = {
        'days_since_inception': 0,
        'avg_pulse_width': 0.0,
        'pulse_count': 0,
        'hype_ratio': 0.0,
        'pulse_reject': False,
        'pulse_reject_reason': '',
        '高收益密度30': 0.0,
        'quality_penalty': 1.0,     # 1.0 = 无罚分, <1.0 = 软罚分
    }

    if len(ret_dates_valid) < 2:
        result['pulse_reject'] = True
        result['pulse_reject_reason'] = '数据不足'
        return result

    # 1. 成立天数（首末日历天跨度）
    first_date = pd.Timestamp(ret_dates_valid[0])
    last_date = pd.Timestamp(ret_dates_valid[-1])
    days_since_inception = (last_date - first_date).days
    result['days_since_inception'] = days_since_inception

    # 2. 脉冲宽度算法：扫描连续高收益区间
    pulse_widths = []
    in_pulse = False
    pulse_start = None

    for i, d in enumerate(ret_dates_valid):
        if rets[d] > 信号阈值_库:
            if not in_pulse:
                in_pulse = True
                pulse_start = i
        else:
            if in_pulse:
                # 脉冲结束，计算日历天宽度
                start_ts = pd.Timestamp(ret_dates_valid[pulse_start])
                end_ts = pd.Timestamp(ret_dates_valid[i - 1])
                width_days = (end_ts - start_ts).days + 1
                pulse_widths.append(width_days)
                in_pulse = False

    # 处理末尾仍在脉冲中的情况
    if in_pulse and pulse_start is not None:
        start_ts = pd.Timestamp(ret_dates_valid[pulse_start])
        end_ts = pd.Timestamp(ret_dates_valid[-1])
        width_days = (end_ts - start_ts).days + 1
        pulse_widths.append(width_days)

    result['pulse_count'] = len(pulse_widths)
    result['avg_pulse_width'] = np.mean(pulse_widths) if pulse_widths else 0.0

    # 3. 高收益密度30: 近30个交易日中 yield > 信号阈值_库 的占比
    recent_30 = ret_dates_valid[-30:] if len(ret_dates_valid) >= 30 else ret_dates_valid
    high_count = sum(1 for d in recent_30 if rets[d] > 信号阈值_库)
    yield_density_30 = high_count / len(recent_30) if recent_30 else 0.0
    result['高收益密度30'] = round(yield_density_30, 4)

    # 4. 炒作比: 前30天均收益 / 近90天均收益
    first_30 = ret_dates_valid[:30] if len(ret_dates_valid) >= 30 else ret_dates_valid
    last_90 = ret_dates_valid[-90:] if len(ret_dates_valid) >= 90 else ret_dates_valid
    avg_first_30 = np.mean([rets[d] for d in first_30]) if first_30 else 0
    avg_last_90 = np.mean([rets[d] for d in last_90]) if last_90 else 0
    hype_ratio = (avg_first_30 / avg_last_90) if avg_last_90 > 0 else 0.0
    result['hype_ratio'] = round(hype_ratio, 4)

    # 5. Anti-Hook 拒绝规则 (V7.1: 放宽阈值 + 软罚分机制)
    #    Hard Reject: 绝对不合格 → pulse_reject=True, 直接跳过
    #    Soft Reject: 边缘质量 → pulse_reject=False, quality_penalty=0.8 (精选推荐降权)
    avg_pw = result['avg_pulse_width']

    # ── Hard Reject: 产品太新 (<60天) ──
    if days_since_inception < 60:
        result['pulse_reject'] = True
        result['pulse_reject_reason'] = f'产品太新({days_since_inception}天<60天)'
        return result

    # ── Soft Penalty: 成熟度偏低 (60-90天) ──
    if days_since_inception < 90:
        result['quality_penalty'] = 0.8
        result['pulse_reject_reason'] = f'成熟度偏低({days_since_inception}天<90天,软罚分)'

    # ── Hard Reject: 噪声脉冲 (<2天) ──
    if avg_pw < 2 and len(pulse_widths) > 0:
        result['pulse_reject'] = True
        result['pulse_reject_reason'] = f'噪声信号(脉冲宽度{avg_pw:.1f}天<2天)'
        return result

    # ── Hard Reject: 粉饰窗口 (宽度>25 且 密度≤40%) ──
    if avg_pw > 25 and yield_density_30 <= 0.4:
        result['pulse_reject'] = True
        result['pulse_reject_reason'] = f'粉饰窗口(脉冲{avg_pw:.1f}天>25天,密度{yield_density_30:.0%}≤40%)'
        return result
    # 例外: 密度>40% 的产品属于 "Stable"，即使脉冲宽也保留

    # ── Hard Reject: 极端炒作 (>2.0) ──
    if hype_ratio > 2.0:
        result['pulse_reject'] = True
        result['pulse_reject_reason'] = f'前高后低营销(炒作比{hype_ratio:.2f}>2.0)'
        return result

    # ── Soft Penalty: 轻度炒作 (1.5-2.0) ──
    if hype_ratio > 1.5:
        result['quality_penalty'] = min(result['quality_penalty'], 0.8)
        if not result['pulse_reject_reason']:
            result['pulse_reject_reason'] = f'轻度炒作(炒作比{hype_ratio:.2f},软罚分)'

    return result


def 构建高成功率产品库(强制刷新=False):
    """构建或加载高成功率产品库和潜力观察池

    Returns:
        (产品库, 潜力观察池): 两个列表
    """

    # 尝试加载缓存
    if not 强制刷新:
        产品库, 观察池 = 加载缓存()
        if 产品库 is not None:
            return 产品库, 观察池

    logger.info("开始构建高成功率产品库...")

    with Database() as db:
        bank_names = db.get_bank_names()
    
    产品库 = []
    潜力观察池 = []

    for bank_name in bank_names:
        logger.info(f"  分析 [{bank_name}]...")
        
        df = load_prepared_data(bank_name)
        
        if df.empty:
            continue
            
        date_cols = sorted([c for c in df.columns if is_date_col(c)])
        if len(date_cols) < 10:
            continue

        if len(df) == 0:
            continue

        跳过_流动性 = 0
        for idx, row in df.iterrows():
            code = row.get('产品代码')
            name = row.get('产品名称')
            if pd.isna(code):
                continue

            # 流动性过滤：只保留可在赎回周期内退出的产品
            流动性类型, 赎回天数 = 判断产品流动性(str(name) if pd.notna(name) else '')
            if 赎回天数 > 最长赎回天数:
                跳过_流动性 += 1
                continue

            # 使用间隔感知的收益率计算（正确处理数据空缺）
            rets = get_product_returns(row, date_cols)
            if len(rets) < 10:
                continue

            ret_dates_valid = sorted(rets.keys())

            # Anti-Hook: 计算脉冲质量指标
            pulse_metrics = _calc_pulse_metrics(rets, ret_dates_valid)

            # 找历史信号并计算成功率
            信号列表 = []
            持续天数列表 = []
            收益列表 = []

            for i in range(1, len(ret_dates_valid) - 持有天数):
                curr_date = ret_dates_valid[i]
                prev_date = ret_dates_valid[i - 1]

                # V4修正: 使用前日收益率检测信号(消除前视偏差)
                # 信号含义: 前日收益>库阈值(3.0%), 今日观测到并记录
                # V5.28: 用宽阈值(3.0%)入库，买入推荐用严阈值(3.5%)
                if rets[prev_date] > 信号阈值_库 and (
                    not 需要突破信号 or rets[prev_date] <= 信号阈值_库):
                    # V8.2: 先计算持续天数，用于查询正确的赎回费率
                    持续 = 0
                    for j in range(i, min(i + 20, len(ret_dates_valid))):
                        if rets[ret_dates_valid[j]] > 信号阈值_库 * 0.8:
                            持续 += 1
                        else:
                            break
                    # 转换为日历天数
                    if 持续 > 0:
                        cal_days = (pd.Timestamp(ret_dates_valid[i + 持续 - 1])
                                    - pd.Timestamp(curr_date)).days + 1
                    else:
                        cal_days = 持有天数  # 无持续数据时用默认值
                    持续天数列表.append(cal_days)

                    # 计算持有期收益（下N个有效数据点）
                    持有期收益 = [rets[ret_dates_valid[i + 1 + k]] for k in range(持有天数)]
                    avg_ret = np.mean(持有期收益)

                    # V8.2: 用实际持续天数查询赎回费率（与回测一致）
                    _sig_holding_days = max(cal_days, 持有天数)  # 至少用评估窗口天数
                    _sig_fee_rate = get_fee_rate(sheet, code, _sig_holding_days)
                    avg_ret_net = calc_net_return(avg_ret, _sig_fee_rate, _sig_holding_days)

                    信号列表.append({
                        'idx': i,
                        'date': curr_date,
                        'signal_ret': rets[prev_date],  # V4: 记录触发信号的前日收益率
                        'hold_ret': avg_ret,            # 毛收益
                        'hold_ret_net': avg_ret_net,    # V8.2: 净收益（扣赎回费）
                        'fee_rate': _sig_fee_rate,      # V8.2: 赎回费率
                        'holding_days': _sig_holding_days,  # V8.2: 实际持有天数
                        'success': avg_ret_net > 成功标准  # V8.2: 用净收益判断成功
                    })
                    收益列表.append(avg_ret_net)  # V8.2: 用净收益计算预期

            # 至少有1次信号才进入任何池子
            if len(信号列表) < 观察池最低信号:
                continue

            成功次数 = sum(1 for s in 信号列表 if s['success'])
            成功率 = 成功次数 / len(信号列表) * 100

            # 计算预期指标
            预期持有天数 = int(np.mean(持续天数列表)) if 持续天数列表 else 持有天数
            # V8.2: 收益列表已是净收益（扣赎回费后），无需额外处理
            预期收益率 = np.mean(收益列表) if 收益列表 else 0
            # V8.2: 计算毛收益和平均费率（用于输出参考）
            预期收益率_毛 = np.mean([s['hold_ret'] for s in 信号列表]) if 信号列表 else 0
            平均赎回费率 = np.mean([s.get('fee_rate', 0) for s in 信号列表]) if 信号列表 else 0

            # 查询赎回费信息
            _fee_info = get_fee_info(sheet, code)
            _has_fee = _fee_info.get('has_redemption_fee') if _fee_info else None
            _fee_desc = _fee_info.get('fee_description', '') if _fee_info else ''
            _fee_rate_7 = get_fee_rate(sheet, code, 7)
            _fee_rate_14 = get_fee_rate(sheet, code, 14)

            # 构建产品信息
            产品信息 = {
                '银行': sheet,
                '产品代码': code,
                '产品名称': name,
                '流动性': 流动性类型,
                '赎回天数': 赎回天数,
                '历史信号次数': len(信号列表),
                '成功次数': 成功次数,
                '历史成功率': round(成功率, 1),
                '预期持有天数': 预期持有天数,
                # V8.2: 区分毛收益和净收益
                '预期年化收益_毛': round(预期收益率_毛, 2),  # 扣费前
                '预期年化收益': round(预期收益率, 2),         # 扣费后（净）
                '平均赎回费率': round(平均赎回费率 * 100, 4), # 平均费率%
                '平均信号强度': round(np.mean([s['signal_ret'] for s in 信号列表]), 2),
                '数据天数': len(rets),
                '最新信号日期': 信号列表[-1]['date'] if 信号列表 else '',
                # Anti-Hook 脉冲质量指标 (VIP Sniper Edition)
                '成立天数': pulse_metrics['days_since_inception'],
                '平均脉冲宽度': round(pulse_metrics['avg_pulse_width'], 1),
                '脉冲次数': pulse_metrics['pulse_count'],
                '炒作比': pulse_metrics['hype_ratio'],
                '脉冲拒绝': pulse_metrics['pulse_reject'],
                '脉冲拒绝原因': pulse_metrics['pulse_reject_reason'],
                '高收益密度30': pulse_metrics['高收益密度30'],
                'quality_penalty': pulse_metrics.get('quality_penalty', 1.0),
                # 赎回费信息
                '有赎回费': _has_fee,
                '赎回费描述': _fee_desc,
                '赎回费费率_7天': _fee_rate_7,
                '赎回费费率_14天': _fee_rate_14,
            }

            # Anti-Hook: 仅 Hard Reject 跳过, Soft Penalty 产品保留入库
            if pulse_metrics['pulse_reject']:
                continue

            # 判断：正式入库 vs 潜力观察池 vs 不入
            入库 = False
            观察 = False
            观察原因 = ''

            if len(信号列表) >= 最少历史信号 and 成功率 >= 最低成功率:
                # 满足正式库条件，还需检查持有期约束
                if 预期持有天数 > 最大持有天数:
                    观察 = True
                    观察原因 = f'预期持有{预期持有天数}天>上限{最大持有天数}天'
                elif 预期持有天数 < 最短持续天数:
                    观察 = True
                    观察原因 = f'持续{预期持有天数}天<下限{最短持续天数}天'
                else:
                    入库 = True
            elif len(信号列表) < 最少历史信号:
                # 信号次数不足 → 观察池（数据积累中）
                观察 = True
                观察原因 = f'信号{len(信号列表)}次<{最少历史信号}次（待数据积累）'
            elif 成功率 < 最低成功率 and 成功率 >= 观察池最低成功率:
                # 成功率偏低但有潜力
                观察 = True
                观察原因 = f'成功率{成功率:.0f}%<{最低成功率:.0f}%（待提升）'

            if 入库:
                产品库.append(产品信息)
            elif 观察:
                产品信息['观察原因'] = 观察原因
                潜力观察池.append(产品信息)

        本行正式 = sum(1 for p in 产品库 if p['银行'] == sheet)
        本行观察 = sum(1 for p in 潜力观察池 if p['银行'] == sheet)
        logger.info(f"    跳过{跳过_流动性}个长赎回周期产品，"
                    f"正式库 {本行正式} 个，观察池 {本行观察} 个")

    # 保存缓存
    保存缓存(产品库, 潜力观察池)

    return 产品库, 潜力观察池


# ============================================================
# 实时机会扫描（带预测和跟踪）
# ============================================================

def 扫描实时机会(产品库, 潜力观察池=None, 释放规律=None):
    """在高成功率产品和潜力观察池中扫描实时交易机会"""
    logger.info("扫描实时交易机会...")

    实时机会 = []

    # 按银行分组处理（合并正式库和观察池）
    产品库_by_bank = {}
    for p in 产品库:
        bank = p['银行']
        if bank not in 产品库_by_bank:
            产品库_by_bank[bank] = {}
        产品库_by_bank[bank][p['产品代码']] = p

    # 观察池产品也加入扫描范围（标记来源）
    观察池代码集 = set()
    if 潜力观察池:
        for p in 潜力观察池:
            bank = p['银行']
            if bank not in 产品库_by_bank:
                产品库_by_bank[bank] = {}
            # 不覆盖正式库中已有的产品
            if p['产品代码'] not in 产品库_by_bank[bank]:
                产品库_by_bank[bank][p['产品代码']] = p
                观察池代码集.add(p['产品代码'])

    for bank_name in 产品库_by_bank.keys():
        bank_products = 产品库_by_bank[bank_name]
        正式数 = sum(1 for c in bank_products if c not in 观察池代码集)
        观察数 = sum(1 for c in bank_products if c in 观察池代码集)
        logger.info(f"  扫描 [{bank_name}] ({正式数}个正式+{观察数}个观察池产品)...")

        df = load_prepared_data(bank_name)
        if df.empty:
            continue

        date_cols = sorted([c for c in df.columns if is_date_col(c)])
        if not date_cols:
            continue
            
        最新日期 = date_cols[-1]

        for idx, row in df.iterrows():
            code = row.get('产品代码')
            if code not in bank_products:
                continue

            产品信息 = bank_products[code]

            # 使用间隔感知的收益率计算
            rets = get_product_returns(row, date_cols)
            if not rets:
                continue

            ret_dates_valid = sorted(rets.keys())

            # 确定窗口范围（最近N天内的有效日期）
            窗口起点 = pd.Timestamp(最新日期) - pd.Timedelta(days=实时窗口)
            窗口日期 = [d for d in ret_dates_valid if pd.Timestamp(d) >= 窗口起点]

            # 检查窗口内是否有新信号
            # 从最新日期向前扫描，找最近的信号入口点
            for 当前日期 in reversed(窗口日期):
                当前收益 = rets[当前日期]
                当前索引 = ret_dates_valid.index(当前日期)

                if 当前索引 == 0:
                    continue

                昨日日期 = ret_dates_valid[当前索引 - 1]
                昨日收益 = rets[昨日日期]

                # V4修正: 使用昨日收益率检测信号入口(消除前视偏差)
                # 信号含义: 前日收益>库阈值, 今日观测到信号
                # V5.28: 用宽阈值(3.0%)检测信号入口, 推荐级别由买入阈值(3.5%)区分
                if 昨日收益 > 信号阈值_库 and (
                    not 需要突破信号 or 昨日收益 <= 信号阈值_库):
                    信号距今 = (pd.Timestamp(最新日期) - pd.Timestamp(当前日期)).days

                    # 最新有效收益率 = 最后一个有效数据点
                    最新收益 = rets[ret_dates_valid[-1]]
                    最新有效日期 = ret_dates_valid[-1]

                    仍在强信号区 = 最新收益 > 信号阈值_买入   # >3.5% 强推荐
                    仍在信号区 = 最新收益 > 信号阈值_库     # >3.0% 可关注

                    # 计算信号触发以来的有效收益
                    信号后收益列表 = [rets[d] for d in ret_dates_valid[当前索引:]]

                    信号持续天数 = 0
                    for r in 信号后收益列表:
                        if r > 信号阈值_库 * 0.8:
                            信号持续天数 += 1
                        else:
                            break

                    信号期间平均收益 = np.mean(信号后收益列表) if 信号后收益列表 else 0

                    # 判断操作建议 —— 三级推荐: ★强推(>3.5%) / ☆可关注(3.0~3.5%) / 观望(<3.0%)
                    预期持有 = 产品信息['预期持有天数']
                    预期收益 = 产品信息['预期年化收益']

                    # V13: 流动性成本一票否决 — 防止"为赚4%年化却付5%赎回成本"
                    _is_vetoed, _drag_rate, _veto_reason = check_liquidity_cost_veto(
                        expected_yield=预期收益,
                        days_held=预期持有,
                        product_code=code,
                        bank=sheet,
                        tolerance=0.3  # 磨损率不得超过预期收益的30%
                    )

                    if _is_vetoed:
                        # 赎回磨损过高，一票否决
                        操作建议 = f"⛔ 过滤（{_veto_reason}）"
                    elif 仍在强信号区:
                        # 当前收益>3.5%: 强信号，买入机会
                        剩余窗口 = max(预期持有 - 信号距今, 0)
                        if 信号距今 == 0:
                            操作建议 = "★★★ 今日新信号，强烈推荐买入"
                        elif 信号距今 <= 2:
                            操作建议 = f"★★ 推荐买入（信号第{信号距今+1}天，仍在高收益区）"
                        elif 剩余窗口 > 0:
                            操作建议 = f"★ 可买入（预计还有{剩余窗口}天高收益窗口）"
                        else:
                            操作建议 = "可买入（超预期持续，收益仍高）"
                    elif 仍在信号区:
                        # 当前收益3.0~3.5%: 接近买入门槛，可关注
                        操作建议 = f"☆ 可关注（收益{最新收益:.1f}%接近买入门槛{信号阈值_买入}%）"
                    else:
                        # 当前已跌出信号区 → 窗口已过
                        操作建议 = "观望（高收益窗口已结束）"

                    # 收集信号日起每日年化收益率（含信号日本身）
                    每日收益 = []
                    for d in ret_dates_valid[当前索引:]:
                        每日收益.append((d, round(rets[d], 2)))

                    # 判断产品来源
                    是观察池 = code in 观察池代码集
                    观察原因 = 产品信息.get('观察原因', '') if 是观察池 else ''

                    # 观察池产品的操作建议加标注
                    if 是观察池:
                        操作建议 = f"[观察池] {操作建议}"

                    # 建议④: 计算信号新鲜度
                    _freshness_info = None
                    _freshness_tag = ''
                    _progress_ratio = 0.0
                    _predicted_window = 0.0
                    _yield_velocity = 0.0
                    if not 释放规律:
                        logger.debug(f"  [新鲜度] {code}: 释放规律为空, 跳过新鲜度计算")
                    if sf.信号新鲜度_启用 and 释放规律:
                        _pat_key = (sheet, code)
                        _pat = 释放规律.get(_pat_key)
                        _freshness_info = _calc_signal_freshness_v6(
                            ret_dates_valid, rets, 最新有效日期, _pat)
                        if _freshness_info:
                            _freshness_tag = _freshness_info['freshness_tag']
                            _progress_ratio = _freshness_info['progress_ratio']
                            _predicted_window = _freshness_info['predicted_window']
                            _yield_velocity = _freshness_info['yield_velocity']

                    实时机会.append({
                        '银行': sheet,
                        '产品代码': code,
                        '产品名称': 产品信息['产品名称'],
                        '流动性': 产品信息.get('流动性', ''),
                        '来源': '观察池' if 是观察池 else '正式库',
                        '观察原因': 观察原因,
                        '已持有': '',
                        '历史成功率%': 产品信息['历史成功率'],
                        '历史信号次数': 产品信息['历史信号次数'],
                        '信号日期': 当前日期,
                        '信号距今天数': 信号距今,
                        '每日收益': 每日收益,
                        '信号收益率%': round(当前收益, 2),
                        '最新收益率%': round(最新收益, 2),
                        '预期持有天数': 预期持有,
                        '预期年化收益%': round(预期收益, 2),
                        '信号持续天数': 信号持续天数,
                        '信号期间平均收益%': round(信号期间平均收益, 2),
                        '操作建议': 操作建议,
                        '新鲜度进度': round(_progress_ratio, 3),
                        '预测窗口天数': round(_predicted_window, 1),
                        '新鲜度标签': _freshness_tag,
                        '收益加速度': round(_yield_velocity, 2),
                        # V9.2: Sparkline 数据 — 最近20天收益率 + Deep Disclosure 字段
                        'history_yields': [round(rets[d], 2) for d in ret_dates_valid[-20:]],
                        'max_drawdown': round(min((rets[ret_dates_valid[i]] - max(rets[ret_dates_valid[j]] for j in range(max(0, i-20), i+1))) for i in range(min(5, len(ret_dates_valid)), len(ret_dates_valid))) if len(ret_dates_valid) > 5 else 0, 4),
                        'yield_velocity': round(_yield_velocity, 2),
                        'freshness_val': round(_progress_ratio, 3),
                        'anti_hook_score': max(0, min(100, int(产品信息.get('quality_penalty', 1.0) * 100))),
                        # Anti-Hook 字段传播 (VIP Sniper Edition)
                        '成立天数': 产品信息.get('成立天数', 0),
                        '平均脉冲宽度': 产品信息.get('平均脉冲宽度', 0),
                        '炒作比': 产品信息.get('炒作比', 0),
                        '高收益密度30': 产品信息.get('高收益密度30', 0),
                        '脉冲拒绝': 产品信息.get('脉冲拒绝', False),
                        'quality_penalty': 产品信息.get('quality_penalty', 1.0),
                        # 赎回费信息
                        '有赎回费': 产品信息.get('有赎回费'),
                        '赎回费描述': 产品信息.get('赎回费描述', ''),
                        '赎回费费率_7天': 产品信息.get('赎回费费率_7天', 0),
                        '赎回费费率_14天': 产品信息.get('赎回费费率_14天', 0),
                        # V13: 流动性成本分析
                        '赎回磨损率%': round(_drag_rate, 2),
                        '流动性否决': _is_vetoed,
                    })
                    break  # 每个产品只取最新的信号

        logger.info(f"    发现 {sum(1 for o in 实时机会 if o['银行']==sheet)} 个实时机会")

    # 新鲜度诊断统计
    _n_total = len(实时机会)
    _n_fresh = sum(1 for o in 实时机会 if 'FRESH' in o.get('新鲜度标签', ''))
    _n_stale = sum(1 for o in 实时机会 if 'STALE' in o.get('新鲜度标签', ''))
    _n_nopat = sum(1 for o in 实时机会 if o.get('新鲜度标签', '') == 'NO_PATTERN')
    _n_empty = sum(1 for o in 实时机会 if not o.get('新鲜度标签', ''))
    logger.info(f"  [新鲜度统计] 总={_n_total} FRESH={_n_fresh} STALE={_n_stale} NO_PATTERN={_n_nopat} 空={_n_empty}")
    if 释放规律:
        logger.info(f"  [新鲜度统计] 释放规律: {len(释放规律)}个产品有规律")
    else:
        logger.info(f"  [新鲜度统计] 释放规律: 未传入 (None 或空)")

    return 实时机会


# ============================================================
# 前瞻预测: 学习规律并预测即将释放的产品
# ============================================================

def 学习释放规律并预测(产品库):
    """学习产品收益释放规律，预测即将释放的产品 (建议① Alpha加成)

    基于历史180天收益数据学习每个产品的释放周期、阶段偏好和星期偏好，
    预测未来10天内哪些产品即将释放，用于精选推荐时的Alpha加成。

    Returns:
        dict: {(银行, 产品代码): {'score': float, 'predicted_date': str,
               'confidence': float, 'td_until': int, 'source': str}}
    """
    logger.info("学习释放规律并预测...")

    learner = PatternLearner()
    predictor = ForwardPredictor()

    all_dates_set = set()
    patterns = {}

    with Database() as db:
        bank_names = db.get_bank_names()

    for bank_name in bank_names:
        df = load_prepared_data(bank_name)
        if df.empty:
            continue

        date_cols = sorted([c for c in df.columns if is_date_col(c)])
        if len(date_cols) < 10:
            continue

        all_dates_set.update(date_cols)

        for _, row in df.iterrows():
            code = row.get('产品代码')
            if pd.isna(code):
                continue

            code = str(code).strip()
            rets = get_product_returns(row, date_cols)
            if len(rets) < 10:
                continue

            ret_dates = sorted(rets.keys())
            key = (bank_name, code)

            # as_of_date 需在最新数据之后 (learn 内部用 d < as_of_date 过滤)
            latest = max(ret_dates) if ret_dates else datetime.now().strftime('%Y-%m-%d')
            as_of = (pd.Timestamp(latest) + pd.Timedelta(days=1)).strftime('%Y-%m-%d')
            pat = learner.learn(key, ret_dates, rets, as_of, 规律学习窗口天数)
            if pat and pat.confidence >= 最低置信度:
                patterns[key] = pat

    if not patterns:
        logger.info("  无产品具有可用的释放规律")
        return {'prediction_scores': {}, 'patterns': {}, 'sorted_dates': [], 'date_to_idx': {}}

    # 构建日期索引用于预测
    sorted_dates = sorted(all_dates_set)

    # 使用数据库最新日期作为预测基准日
    nav_latest = sorted_dates[-1] if sorted_dates else datetime.now().strftime('%Y-%m-%d')
    logger.info(f"  数据库最新净值日期: {nav_latest}")

    # 扩展日期列表: 添加未来15个工作日, 否则预测日期超出历史范围会被跳过
    last_date = pd.Timestamp(nav_latest)
    future_bdays = pd.bdate_range(start=last_date + pd.Timedelta(days=1), periods=15)
    for fd in future_bdays:
        fd_str = fd.strftime('%Y-%m-%d')
        if fd_str not in all_dates_set:
            sorted_dates.append(fd_str)
    sorted_dates.sort()
    date_to_idx = {d: i for i, d in enumerate(sorted_dates)}

    # 按预测紧迫度排名 (基于数据库最新日期)
    rankings = predictor.rank_products(patterns, nav_latest, sorted_dates, date_to_idx)

    prediction_scores = {}
    for key, score, pred in rankings:
        prediction_scores[key] = {
            'score': score,
            'predicted_date': pred.predicted_date,
            'predicted_end_date': pred.predicted_end_date,
            'confidence': pred.confidence,
            'td_until': pred.td_until,
            'source': pred.source,
        }

    logger.info(f"  释放规律: {len(patterns)}个产品有规律, {len(prediction_scores)}个有预测")
    return {
        'prediction_scores': prediction_scores,
        'patterns': patterns,
        'sorted_dates': sorted_dates,
        'date_to_idx': date_to_idx,
    }


# ============================================================
# 精选推荐：多维度优中选优，264→10
# ============================================================

# 精选参数
精选_最大推荐数 = 100         # VIP Sniper: 扩大候选池, post_processing 负责分类/切片
精选_单行最大 = 15            # V7.2: 扩大单行容量(5→15)，给 Auto-Replenish 更深储备
精选_最低成功率 = 30.0        # 精选门槛: 成功率≥30%
精选_最低信号次数 = 2         # 精选门槛: 至少2次历史信号
精选_信号最大距今 = 10        # 精选门槛: 信号距今≤10天（匹配扫描窗口）

# 精选评分权重 — 9维度综合评分 (V6.2: +D9收益加速度, 各维度微调腾出0.10)
W_收益强度 = 0.18             # D1: 当日收益 × 历史预期收益 (原0.20)
W_历史可靠 = 0.13             # D2: 成功率 × log2(信号次数) (原0.15)
W_信号时效 = 0.13             # D3: 信号新鲜度 — V6.2改用指数衰减 (原0.15)
W_流动性 = 0.09               # D4: 日开>7天>14天 (原0.10)
W_持续质量 = 0.09             # D5: 信号期间平均收益 (原0.10)
W_收益排名 = 0.13             # D6: 回测验证最优排名: 当日收益×历史平均 (原0.15)
W_收益稳定 = 0.10             # D7: 信号收益的稳定性(低波动)
W_信号密度 = 0.05             # D8: 信号频率(高频=稳定机会)
W_收益加速度 = 0.10           # D9: 收益加速度 (新增, 建议④)
W_赎回费 = 0.08               # D10: 赎回费惩罚 (有赎回费的产品降权)


def _流动性评分(流动性类型):
    """流动性越好分越高（0~1）"""
    if '日开' in str(流动性类型):
        return 1.0
    elif '7天' in str(流动性类型) or '周开' in str(流动性类型):
        return 0.7
    elif '14天' in str(流动性类型):
        return 0.5
    else:
        return 0.3


def _calc_signal_freshness_v6(ret_dates, rets, latest_date, pattern):
    """计算信号新鲜度 — 委托给 shared_freshness.calc_signal_freshness()"""
    return sf.calc_signal_freshness(
        ret_dates, rets, latest_date, None, pattern,
        释放识别阈值 * 0.5)


def _格式化释放规律(pat):
    """将 ReleasePattern 格式化为可读字符串"""
    if pat is None:
        return '', '', '', 0.0
    星期名 = ['周一', '周二', '周三', '周四', '周五']
    阶段名 = ['月初(BD1-5)', '月中(BD6-11)', '月中下(BD12-16)', '月末(BD17-23)']
    # 周期描述 (period_days 现在是工作日数)
    if pat.has_period and pat.period_cv < 0.4:
        周期 = f"每{pat.period_days:.0f}工作日"
    else:
        周期 = ''
    # 阶段描述
    if pat.phase_pvalue < 0.05:
        阶段 = 阶段名[pat.top_phase] if 0 <= pat.top_phase < 4 else ''
    else:
        阶段 = ''
    # 星期描述
    if pat.weekday_pvalue < 0.05:
        星期 = 星期名[pat.top_weekday] if 0 <= pat.top_weekday < 5 else ''
    else:
        星期 = ''
    return 周期, 阶段, 星期, round(pat.confidence, 2)


def 精选推荐(实时机会, 前瞻预测=None, 产品库=None, 是否节前=False, 节前假期名='', 释放规律=None):
    """8维度综合评分优中选优 + 前瞻预测Alpha加成 + 节前推荐: 全量候选 → Top 20精选

    筛选流程:
      1. 硬过滤: 正式库优先 + 信号区(★/☆/可买入) + 收益>3.0% + 成功率≥30% + 距今≤7天
         V5.28: 观察池产品也可入选（标记[观察池]），填充银行多样性
      2. 八维评分: 收益强度/历史可靠/信号时效/流动性/持续质量/收益排名/收益稳定/信号密度
      3. 前瞻预测加成(V6.0): 预测即将释放的产品获得额外Alpha加成
      4. 同系去重: 同一基名产品只保留最高分份额
      5. 银行分散: 同一银行最多N个
      6. 输出 Top 20
      7. V6.1 节前推荐: 节前日追加高收益候选产品

    Args:
        实时机会: 实时机会列表
        前瞻预测: 预测字典 {(银行, 产品代码): {score, predicted_date, confidence, ...}}
        产品库: 产品库列表 (节前推荐使用)
        是否节前: 是否为节前交易日
        节前假期名: 假期名称 (如 '国庆')

    Returns:
        list of dict: 精选产品列表（含评分明细和前瞻预测信息）
    """
    import math

    # ── Step 1: 硬过滤（正式库 + 观察池均可入选） ──
    候选 = []
    _bank_pass = {}  # 诊断: 各银行通过数
    for o in 实时机会:
        来源 = o.get('来源', '')
        if 来源 not in ('正式库', '观察池'):
            continue
        建议 = o.get('操作建议', '')
        # 去掉[观察池]前缀后检查推荐级别
        建议_clean = 建议.replace('[观察池] ', '')
        if '★' not in 建议_clean and '☆' not in 建议_clean and '可买入' not in 建议_clean:
            continue  # ★强推/☆可关注/可买入均纳入
        # 成功率门槛: 正式库≥30%, 观察池≥20%（与观察池入池条件一致）
        最低成功 = 精选_最低成功率 if 来源 == '正式库' else 观察池最低成功率
        if (o.get('历史成功率%', 0) or 0) < 最低成功:
            continue
        # 距今 = 0 表示今日新信号（最佳），不可用 `or 99` 因为 0 is falsy
        距今值 = o.get('信号距今天数')
        if 距今值 is None or 距今值 > 精选_信号最大距今:
            continue
        # V8.2: 用净收益（扣赎回费后）进行硬过滤
        最新收益_毛 = o.get('最新收益率%', 0) or 0
        fee_rate_14 = o.get('赎回费费率_14天', 0) or 0
        预期持有_filter = o.get('预期持有天数', 14) or 14
        最新收益 = calc_net_return(最新收益_毛, fee_rate_14, 预期持有_filter)
        if 最新收益 <= 信号阈值_库:
            continue  # V5.28: 用库阈值(>3.0%)扩大精选候选池，V8.2: 用净收益
        # 正式库: 信号次数≥2; 观察池: 信号次数≥1（放宽以填充多样性）
        最低信号 = 精选_最低信号次数 if 来源 == '正式库' else 1
        if (o.get('历史信号次数', 0) or 0) < 最低信号:
            continue
        # V6.2 → V7.2: 新鲜度 — STALE 标记为软罚分(不淘汰), FRESH/FLASH 正常通过
        is_stale = False
        if sf.信号新鲜度_启用:
            _ftag = o.get('新鲜度标签', '')
            if 'STALE' in _ftag:
                is_stale = True
        o['_is_stale'] = is_stale
        # VIP Sniper V7.1: Anti-Hook — 仅 Hard Reject 跳过, Soft Penalty 保留
        if o.get('脉冲拒绝'):
            continue
        候选.append(o)
        bank = o.get('银行', '?')
        _bank_pass[bank] = _bank_pass.get(bank, 0) + 1

    if not 候选:
        logger.info("精选推荐: 无符合条件的候选")
        return []

    正式库数 = sum(1 for o in 候选 if o.get('来源') == '正式库')
    观察池数 = len(候选) - 正式库数
    logger.info(f"精选推荐: 硬过滤后 {len(候选)} 个候选（正式库{正式库数}+观察池{观察池数}）")

    # ── Step 2: 十维评分 ──
    raw_scores = []
    for o in 候选:
        # V8.2: 获取毛收益、赎回费率和预期持有天数，计算净收益用于评分
        最新收益_毛 = o.get('最新收益率%', 0) or 0
        预期收益_毛 = o.get('预期年化收益%', 0) or 0
        fee_rate = o.get('赎回费费率_14天', 0) or 0
        预期持有 = o.get('预期持有天数', 14) or 14
        # 计算净收益（扣赎回费后，使用预期持有天数）
        最新收益 = calc_net_return(最新收益_毛, fee_rate, 预期持有)
        预期收益 = calc_net_return(预期收益_毛, fee_rate, 预期持有)
        期间均收益_毛 = o.get('信号期间平均收益%', 0) or 0
        期间均收益 = calc_net_return(期间均收益_毛, fee_rate, 预期持有)

        成功率 = o.get('历史成功率%', 0) or 0
        信号次数 = o.get('历史信号次数', 1) or 1
        距今 = o.get('信号距今天数', 0) or 0
        流动性 = o.get('流动性', '')
        信号收益_毛 = o.get('信号收益率%', 0) or 0
        信号收益 = calc_net_return(信号收益_毛, fee_rate, 预期持有)
        持续天数 = o.get('信号持续天数', 0) or 0
        数据天数 = o.get('数据天数', 100) or 100

        # D1: 收益强度 — 当日收益 × 历史预期收益（V8.2: 用净收益）
        d1_收益强度 = 最新收益 * max(预期收益, 0.1)
        # D2: 历史可靠 — 成功率 × log2(信号次数+1)
        d2_历史可靠 = (成功率 / 100) * math.log2(max(信号次数, 1) + 1)
        # D3: 信号时效 — V6.2: 用新鲜度进度的指数衰减
        #   FRESH (有pattern): 10 * exp(-3 * progress_ratio)
        #   FLASH (无pattern, Day 0-1): 线性距今衰减 × 0.5 (无验证规律, 降低排名)
        #   空/其他: 线性距今衰减 (兼容)
        _progress = o.get('新鲜度进度', 0)
        _ftag_d3 = o.get('新鲜度标签', '')
        _has_pattern = 'FRESH' in _ftag_d3  # 有 pattern 的产品标签含 FRESH
        if _has_pattern and _progress >= 0:
            d3_信号时效 = 10 * math.exp(-3 * _progress)
        elif 'FLASH' in _ftag_d3:
            # FLASH: 无规律但信号新鲜 (Day 0-1), D3 打5折排在 FRESH 之后
            d3_信号时效 = max(0, 精选_信号最大距今 - 距今) * 0.5
        else:
            d3_信号时效 = max(0, 精选_信号最大距今 - 距今)
        # D4: 流动性 — 日开>7天>14天
        d4_流动性 = _流动性评分(流动性)
        # D5: 持续质量 — 信号期间的平均收益（V8.2: 用净收益）
        d5_持续质量 = max(期间均收益, 0)
        # D6: 收益排名分 — 回测验证最优: 当日收益 × 历史平均（V8.2: 用净收益）
        d6_收益排名 = 最新收益 * max(预期收益, 0.1)
        # D7: 收益稳定 — 信号收益与期间均收益的一致性（V8.2: 用净收益）
        d7_收益稳定 = max(0, 10 - abs(信号收益 - 期间均收益))
        # D8: 信号密度 — 信号次数/数据天数*1000 (高频=稳定机会源)
        d8_信号密度 = (信号次数 / max(数据天数, 30)) * 1000
        # D9: 收益加速度 — V6.2 建议④: 收益正在加速上升的产品优先
        d9_收益加速度 = max(o.get('收益加速度', 0), 0)
        # D10: 赎回费惩罚 — V8.2改进: 基于费用对收益的实际影响
        # 费用占收益的比例越高，惩罚越大（相对于绝对费率更公平）
        if 最新收益_毛 > 0:
            # fee_impact = 费率 / (毛收益率/100) = 费用占收益的比例
            fee_impact = fee_rate / (最新收益_毛 / 100) if 最新收益_毛 > 0 else 0
            # 影响10%收益 → 0分，影响5% → 5分，影响0% → 10分
            d10_赎回费 = max(0, 10 - fee_impact * 100)
        else:
            d10_赎回费 = 10 if fee_rate == 0 else 0

        raw_scores.append({
            'item': o,
            'd1': d1_收益强度,
            'd2': d2_历史可靠,
            'd3': d3_信号时效,
            'd4': d4_流动性,
            'd5': d5_持续质量,
            'd6': d6_收益排名,
            'd7': d7_收益稳定,
            'd8': d8_信号密度,
            'd9': d9_收益加速度,
            'd10': d10_赎回费,
        })

    # Min-Max 归一化每个维度
    for dim in ['d1', 'd2', 'd3', 'd4', 'd5', 'd6', 'd7', 'd8', 'd9', 'd10']:
        vals = [s[dim] for s in raw_scores]
        vmin, vmax = min(vals), max(vals)
        rng = vmax - vmin if vmax > vmin else 1.0
        for s in raw_scores:
            s[f'{dim}_norm'] = (s[dim] - vmin) / rng

    # 加权综合得分（正式库产品优先于观察池）+ 前瞻预测加成
    n_boosted = 0
    for s in raw_scores:
        base_score = (
            W_收益强度 * s['d1_norm'] +
            W_历史可靠 * s['d2_norm'] +
            W_信号时效 * s['d3_norm'] +
            W_流动性 * s['d4_norm'] +
            W_持续质量 * s['d5_norm'] +
            W_收益排名 * s['d6_norm'] +
            W_收益稳定 * s['d7_norm'] +
            W_信号密度 * s['d8_norm'] +
            W_收益加速度 * s['d9_norm'] +
            W_赎回费 * s['d10_norm']
        )
        # VIP Sniper V7.1: quality_penalty 软罚分 (边缘质量产品降权)
        qp = s['item'].get('quality_penalty', 1.0)
        if qp < 1.0:
            base_score *= qp

        # V7.2: STALE 软罚分 — 保留为备选但大幅降权(70%)
        if s['item'].get('_is_stale'):
            base_score *= 0.30

        # V6.0: 前瞻预测Alpha加成
        s['前瞻加成'] = False
        s['前瞻分'] = 0.0
        s['预测释放日'] = ''
        s['预测释放结束日'] = ''
        s['预测置信度'] = 0.0
        s['预测距今'] = 0
        if 前瞻预测:
            pred_key = (s['item'].get('银行', ''), s['item'].get('产品代码', ''))
            if pred_key in 前瞻预测:
                pred = 前瞻预测[pred_key]
                base_score *= (1 + pred['score'] * ALPHA_预测加成)
                s['前瞻加成'] = True
                s['前瞻分'] = round(pred['score'], 3)
                s['预测释放日'] = pred['predicted_date']
                s['预测释放结束日'] = pred.get('predicted_end_date', '')
                s['预测置信度'] = round(pred['confidence'], 3)
                s['预测距今'] = pred.get('td_until', 0)
                n_boosted += 1

        s['综合得分'] = base_score

    if 前瞻预测:
        logger.info(f"精选推荐: {n_boosted}个候选获得前瞻预测加成")

    # 按得分降序
    raw_scores.sort(key=lambda s: -s['综合得分'])

    # ── Step 3: 同系去重 — V7.2: 按产品代码去重，允许同系不同份额(A/B/E/F/P) ──
    已选代码 = set()
    去重后 = []
    for s in raw_scores:
        code = s['item']['产品代码']
        if code in 已选代码:
            continue
        已选代码.add(code)
        去重后.append(s)

    logger.info(f"精选推荐: 同系去重后 {len(去重后)} 个")

    # ── Step 4: 银行分散 ──
    银行计数 = {}
    分散后 = []
    for s in 去重后:
        bank = s['item']['银行']
        cnt = 银行计数.get(bank, 0)
        if cnt >= 精选_单行最大:
            continue
        银行计数[bank] = cnt + 1
        分散后.append(s)

    # ── Step 5: 取 Top N ──
    精选 = 分散后[:精选_最大推荐数]

    # 构建输出
    结果 = []
    for rank, s in enumerate(精选, 1):
        o = s['item']
        # 释放规律标注
        _pat_key = (o.get('银行', ''), o.get('产品代码', ''))
        _pat = 释放规律.get(_pat_key) if 释放规律 else None
        _周期, _阶段, _星期, _规律置信度 = _格式化释放规律(_pat)
        结果.append({
            '排名': rank,
            '来源': o.get('来源', '正式库'),
            '银行': o['银行'],
            '产品代码': o['产品代码'],
            '产品名称': o['产品名称'],
            '流动性': o.get('流动性', ''),
            '历史成功率%': o['历史成功率%'],
            '历史信号次数': o['历史信号次数'],
            '信号距今天数': o.get('信号距今天数', 0),
            '最新收益率%': o.get('最新收益率%', 0),
            '预期年化收益%': o.get('预期年化收益%', 0),
            '信号期间平均收益%': o.get('信号期间平均收益%', 0),
            '操作建议': o['操作建议'],
            '综合得分': round(s['综合得分'] * 100, 1),
            '收益强度分': round(s['d1_norm'] * 100, 1),
            '可靠性分': round(s['d2_norm'] * 100, 1),
            '时效性分': round(s['d3_norm'] * 100, 1),
            '流动性分': round(s['d4_norm'] * 100, 1),
            '持续质量分': round(s['d5_norm'] * 100, 1),
            '收益排名分': round(s['d6_norm'] * 100, 1),
            '稳定性分': round(s['d7_norm'] * 100, 1),
            '信号密度分': round(s['d8_norm'] * 100, 1),
            '加速度分': round(s['d9_norm'] * 100, 1),
            '赎回费惩罚分': round(s['d10_norm'] * 100, 1),
            # V6.0: 前瞻预测信息
            '前瞻加成': '是' if s.get('前瞻加成') else '',
            '前瞻分': s.get('前瞻分', 0),
            '预测释放日': s.get('预测释放日', ''),
            '预测释放结束日': s.get('预测释放结束日', ''),
            '预测置信度': s.get('预测置信度', 0),
            # 释放规律标注
            '释放周期': _周期,
            '释放阶段': _阶段,
            '释放星期': _星期,
            '规律置信度': _规律置信度,
            # V6.2: 新鲜度信息
            '新鲜度标签': o.get('新鲜度标签', ''),
            '新鲜度进度': o.get('新鲜度进度', 0),
            '预测窗口天数': o.get('预测窗口天数', 0),
            '收益加速度': o.get('收益加速度', 0),
            # VIP Sniper: Anti-Hook 字段
            '成立天数': o.get('成立天数', 0),
            '平均脉冲宽度': o.get('平均脉冲宽度', 0),
            '炒作比': o.get('炒作比', 0),
            '高收益密度30': o.get('高收益密度30', 0),
            'quality_penalty': o.get('quality_penalty', 1.0),
            # V9.2: Deep Disclosure 字段
            'history_yields': o.get('history_yields', []),
            'max_drawdown': o.get('max_drawdown', 0),
            'yield_velocity': o.get('yield_velocity', o.get('收益加速度', 0)),
            'freshness_val': o.get('freshness_val', o.get('新鲜度进度', 0)),
            'anti_hook_score': o.get('anti_hook_score', max(0, min(100, int(o.get('quality_penalty', 1.0) * 100)))),
            # 赎回费信息
            '有赎回费': o.get('有赎回费'),
            '赎回费描述': o.get('赎回费描述', ''),
            '节前推荐': '',
        })

    # ── V6.1 Step 6: 节前推荐 ──
    if 是否节前 and 节前假期名:
        已选代码 = {r['产品代码'] for r in 结果}
        节前候选 = []
        for o in 实时机会:
            code = o.get('产品代码', '')
            if code in 已选代码:
                continue
            最新收益 = o.get('最新收益率%', 0) or 0
            成功率 = o.get('历史成功率%', 0) or 0
            均收 = o.get('信号期间平均收益%', 0) or 0
            if 最新收益 > 节前买入_实时阈值 and 成功率 >= 节前买入_最低成功率 and 均收 >= 节前买入_最低均收:
                节前候选.append(o)
        # 按最新收益降序排序
        节前候选.sort(key=lambda x: -(x.get('最新收益率%', 0) or 0))
        节前追加 = 节前候选[:节前买入_最大产品数]
        for o in 节前追加:
            _pat_key = (o.get('银行', ''), o.get('产品代码', ''))
            _pat = 释放规律.get(_pat_key) if 释放规律 else None
            _周期, _阶段, _星期, _规律置信度 = _格式化释放规律(_pat)
            结果.append({
                '排名': len(结果) + 1,
                '来源': o.get('来源', ''),
                '银行': o['银行'],
                '产品代码': o['产品代码'],
                '产品名称': o['产品名称'],
                '流动性': o.get('流动性', ''),
                '历史成功率%': o.get('历史成功率%', 0),
                '历史信号次数': o.get('历史信号次数', 0),
                '信号距今天数': o.get('信号距今天数', 0),
                '最新收益率%': o.get('最新收益率%', 0),
                '预期年化收益%': o.get('预期年化收益%', 0),
                '信号期间平均收益%': o.get('信号期间平均收益%', 0),
                '操作建议': f'[节前{节前假期名}] {o.get("操作建议", "")}',
                '综合得分': 0,
                '收益强度分': 0, '可靠性分': 0, '时效性分': 0, '流动性分': 0,
                '持续质量分': 0, '收益排名分': 0, '稳定性分': 0, '信号密度分': 0,
                '加速度分': 0, '赎回费惩罚分': 0,
                '前瞻加成': '', '前瞻分': 0, '预测释放日': '', '预测释放结束日': '', '预测置信度': 0,
                '释放周期': _周期, '释放阶段': _阶段, '释放星期': _星期, '规律置信度': _规律置信度,
                '新鲜度标签': o.get('新鲜度标签', ''), '新鲜度进度': o.get('新鲜度进度', 0),
                '预测窗口天数': o.get('预测窗口天数', 0), '收益加速度': o.get('收益加速度', 0),
                '成立天数': o.get('成立天数', 0), '平均脉冲宽度': o.get('平均脉冲宽度', 0),
                '炒作比': o.get('炒作比', 0), '高收益密度30': o.get('高收益密度30', 0),
                'history_yields': o.get('history_yields', []),
                'max_drawdown': o.get('max_drawdown', 0),
                'yield_velocity': o.get('yield_velocity', o.get('收益加速度', 0)),
                'freshness_val': o.get('freshness_val', o.get('新鲜度进度', 0)),
                'anti_hook_score': o.get('anti_hook_score', max(0, min(100, int(o.get('quality_penalty', 1.0) * 100)))),
                '有赎回费': o.get('有赎回费'),
                '赎回费描述': o.get('赎回费描述', ''),
                '节前推荐': f'是({节前假期名})',
            })
        if 节前追加:
            logger.info(f"精选推荐: 节前{节前假期名}追加 {len(节前追加)} 个产品")

    logger.info(f"精选推荐: 最终输出 {len(结果)} 个")
    return 结果


# ============================================================
# 持仓分析
# ============================================================

def 加载持仓():
    """加载持仓交易记录，计算当前持仓"""
    if not os.path.exists(PORTFOLIO_PATH):
        logger.info("未找到持仓文件")
        return None, None

    try:
        df = pd.read_excel(PORTFOLIO_PATH)
        required_cols = {'银行', '产品代码', '产品名称', '交易', '金额', '日期'}
        if not required_cols.issubset(set(df.columns)):
            logger.warning(f"持仓文件列不匹配，需要: {required_cols}")
            return None, None

        df['日期'] = pd.to_datetime(df['日期'])
        df['金额'] = pd.to_numeric(df['金额'], errors='coerce').fillna(0)
        df = df[df['交易'].isin(['买入', '卖出'])].copy()  # 忽略撤单

        logger.info(f"加载持仓记录: {len(df)} 条有效交易")
        return df, _计算当前持仓(df)
    except Exception as e:
        logger.warning(f"加载持仓失败: {e}")
        return None, None


def _计算当前持仓(交易记录):
    """从交易记录计算每个产品的当前持仓"""
    持仓 = {}

    for _, row in 交易记录.iterrows():
        key = (row['银行'], row['产品代码'])
        if key not in 持仓:
            持仓[key] = {
                '银行': row['银行'],
                '产品代码': row['产品代码'],
                '产品名称': row['产品名称'],
                '买入总额': 0,
                '卖出总额': 0,
                '首次买入日': row['日期'],
                '最近交易日': row['日期'],
                '交易次数': 0,
            }

        p = 持仓[key]
        p['交易次数'] += 1
        if row['日期'] < p['首次买入日']:
            p['首次买入日'] = row['日期']
        if row['日期'] > p['最近交易日']:
            p['最近交易日'] = row['日期']

        if row['交易'] == '买入':
            p['买入总额'] += row['金额']
        elif row['交易'] == '卖出':
            p['卖出总额'] += row['金额']

    # 计算净持仓
    结果 = []
    for key, p in 持仓.items():
        p['净持仓'] = p['买入总额'] - p['卖出总额']
        p['持仓状态'] = '持有中' if p['净持仓'] > 0 else '已清仓'
        结果.append(p)

    return 结果


def 分析持仓(当前持仓, 产品库, 实时机会, 预测数据=None):
    """分析持仓产品的收益表现和操作建议 (V6.1: 智持 + 假期保护)"""
    if not 当前持仓:
        return []

    # V6.1: 提取预测基础设施 (用于智持判断)
    _patterns = 预测数据.get('patterns', {}) if 预测数据 else {}
    _sorted_dates = 预测数据.get('sorted_dates', []) if 预测数据 else []
    _date_to_idx = 预测数据.get('date_to_idx', {}) if 预测数据 else {}
    _predictor = ForwardPredictor() if _patterns else None
    today_str = datetime.now().strftime('%Y-%m-%d')
    _is_hol, _hol_name = is_holiday(today_str)

    # 构建产品库和实时机会的索引
    库索引 = {p['产品代码']: p for p in 产品库}
    机会索引 = {o['产品代码']: o for o in 实时机会}

    # 银行数据缓存（懒加载）
    银行数据缓存 = {}
    持仓分析 = []

    for p in 当前持仓:
        bank = p['银行']
        code = p['产品代码']

        分析 = {
            '银行': bank,
            '产品代码': code,
            '产品名称': p['产品名称'],
            '持仓状态': p['持仓状态'],
            '净持仓金额': round(p['净持仓'], 2),
            '买入总额': round(p['买入总额'], 2),
            '卖出总额': round(p['卖出总额'], 2),
            '首次买入日': p['首次买入日'].strftime('%Y-%m-%d'),
            '最近交易日': p['最近交易日'].strftime('%Y-%m-%d'),
            '交易次数': p['交易次数'],
            '是否在高成功率库': '是' if code in 库索引 else '否',
            '最新年化收益%': '',
            '买入以来平均收益%': '',
            '持仓天数': '',
            '持仓建议': '',
            '智持标记': '',
        }

        # 懒加载银行数据
        if bank not in 银行数据缓存:
            logger.info(f"分析持仓: 首次遇到银行 '{bank}'，从数据库加载数据...")
            df = load_prepared_data(bank)
            if not df.empty:
                date_cols = sorted([c for c in df.columns if is_date_col(c)])
                nav_matrix = df[date_cols].apply(pd.to_numeric, errors='coerce')
                with warnings.catch_warnings():
                    warnings.simplefilter('ignore')
                    ret_matrix = nav_matrix.pct_change(axis=1) * 365 * 100
                银行数据缓存[bank] = {
                    'df': df,
                    'date_cols': date_cols,
                    'ret_matrix': ret_matrix
                }
            else:
                银行数据缓存[bank] = None # 标记为已尝试加载但无数据
        
        data = 银行数据缓存.get(bank)
        if not data:
            持仓分析.append(分析)
            continue

        df_code = data['df']
        date_cols = data['date_cols']
        ret_matrix = data['ret_matrix']
        
        mask = df_code['产品代码'] == code
        if mask.sum() > 0:
            idx = df_code[mask].index[0]

            # 最新有效收益率
            最新收益 = 0
            for d in reversed(date_cols):
                v = ret_matrix.loc[idx, d] if d in ret_matrix.columns else np.nan
                if pd.notna(v) and v != 0:
                    最新收益 = v
                    break
            分析['最新年化收益%'] = round(最新收益, 2)

            # 买入以来的平均收益
            买入日str = p['首次买入日'].strftime('%Y-%m-%d')
            买入后日期 = [d for d in date_cols if d >= 买入日str]
            if 买入后日期:
                买入后收益 = []
                for d in 买入后日期:
                    if d in ret_matrix.columns:
                        v = ret_matrix.loc[idx, d]
                        if pd.notna(v) and v != 0:
                            买入后收益.append(v)
                if 买入后收益:
                    分析['买入以来平均收益%'] = round(np.mean(买入后收益), 2)

            分析['持仓天数'] = (datetime.now() - p['首次买入日']).days

        # V8.1: 闪跌止损检测 (Dynamic Trailing Stop)
        _flash_drop = False
        if p['持仓状态'] == '持有中':
            mask = df_code['产品代码'] == code
            if mask.sum() > 0:
                idx = df_code[mask].index[0]
                # 取最近N天有效收益率
                recent_rets = []
                for d in reversed(date_cols[-闪跌止损_回看天数*2:]):
                    if d in ret_matrix.columns:
                        v = ret_matrix.loc[idx, d]
                        if pd.notna(v) and v != 0:
                            recent_rets.append(v)
                    if len(recent_rets) >= 闪跌止损_回看天数:
                        break
                if len(recent_rets) >= 2:
                    current_ret = recent_rets[0]  # 最新
                    peak_ret = max(recent_rets)
                    drawdown = current_ret - peak_ret
                    if drawdown < 闪跌止损_阈值:
                        _flash_drop = True
                        分析['闪跌回撤'] = round(drawdown, 3)
                        分析['近期峰值'] = round(peak_ret, 2)

        # 生成持仓建议 (V8.1: 闪跌止损 + V6.1: 智持 + 假期保护)
        if p['持仓状态'] == '已清仓':
            分析['持仓建议'] = '已清仓'
        elif _flash_drop:
            分析['持仓建议'] = f'止损卖出 (FLASH DROP: 收益从{分析["近期峰值"]}%跌{分析["闪跌回撤"]}%)'
        elif _is_hol:
            # V6.1 假期保护: 假期期间所有卖出建议改为假期勿卖
            分析['持仓建议'] = f'假期勿卖({_hol_name}期间0%非真实信号)'
            分析['智持标记'] = f'假期保护({_hol_name})'
        elif code in 机会索引:
            机会 = 机会索引[code]
            if '买入' in 机会['操作建议']:
                分析['持仓建议'] = '继续持有（仍在高收益区）'
            else:
                # V6.1 智持: 原本建议卖出时, 检查是否有即将释放
                _smart_held = False
                if _predictor:
                    pred_key = (bank, code)
                    if pred_key in _patterns:
                        _has_rel, _pred = _predictor.has_upcoming_release(
                            _patterns[pred_key], today_str, _sorted_dates, _date_to_idx,
                            max_td=延迟卖出_预测窗口, min_conf=延迟卖出_置信度)
                        if _has_rel and _pred:
                            分析['持仓建议'] = f'智持(预测{_pred.td_until}日释放, 置信{_pred.confidence:.2f})'
                            分析['智持标记'] = f'智持({_pred.td_until}日后释放)'
                            _smart_held = True
                if not _smart_held:
                    分析['持仓建议'] = '建议卖出（高收益窗口已结束）'
        elif code in 库索引:
            分析['持仓建议'] = '持有观察（高成功率产品，等待下次信号）'
        else:
            最新 = 分析.get('最新年化收益%', 0)
            if isinstance(最新, (int, float)) and 最新 > 信号阈值_卖出:
                分析['持仓建议'] = '持有（当前收益尚可）'
            else:
                # V6.1 非库产品智持: 检查是否有即将释放
                _smart_held = False
                if _predictor:
                    pred_key = (bank, code)
                    if pred_key in _patterns:
                        _has_rel, _pred = _predictor.has_upcoming_release(
                            _patterns[pred_key], today_str, _sorted_dates, _date_to_idx,
                            max_td=延迟卖出_预测窗口, min_conf=延迟卖出_置信度)
                        if _has_rel and _pred:
                            分析['持仓建议'] = f'智持(预测{_pred.td_until}日释放, 置信{_pred.confidence:.2f})'
                            分析['智持标记'] = f'智持({_pred.td_until}日后释放)'
                            _smart_held = True
                if not _smart_held:
                    分析['持仓建议'] = '关注（非高成功率库产品）'

        持仓分析.append(分析)

    logger.info(f"持仓分析完成: {len(持仓分析)} 个产品")
    return 持仓分析


# ============================================================
# 生成报告
# ============================================================

def 生成报告(产品库, 实时机会, 持仓分析=None, 已持有代码=None, 潜力观察池=None, 精选结果=None):
    """生成Excel报告"""
    logger.info("生成报告...")

    wb = Workbook()

    # 样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    green_fill = PatternFill("solid", fgColor="2E7D32")
    blue_fill = PatternFill("solid", fgColor="1565C0")
    align_center = Alignment(horizontal="center", vertical="center")

    # Sheet 1: 实时交易机会
    ws1 = wb.active
    ws1.title = "实时交易机会"

    # 收集所有信号中出现过的日期（用于动态列头）
    所有日收益日期 = set()
    for o in 实时机会:
        for d, _ in o.get('每日收益', []):
            所有日收益日期.add(d)
    日收益日期排序 = sorted(所有日收益日期)
    最大日数 = len(日收益日期排序)

    # 固定列 + 动态日期列（表头直接显示日期）
    固定列前 = ['银行', '产品代码', '产品名称', '流动性', '来源', '已持有', '历史成功率%', '历史信号次数',
                '信号日期', '信号距今天数']
    固定列后 = ['信号收益率%', '最新收益率%', '预期持有天数', '预期年化收益%',
                '信号持续天数', '信号期间平均收益%', '操作建议']

    headers1 = 固定列前 + 日收益日期排序 + 固定列后

    ws1.append(headers1)
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = green_fill
        cell.alignment = align_center

    # 构建日期→列索引映射
    日期列映射 = {d: i for i, d in enumerate(日收益日期排序)}
    日收益起始列 = len(固定列前) + 1  # 1-based Excel column index

    # 标记已持有的产品
    if 已持有代码:
        for row in 实时机会:
            row['已持有'] = '是' if row['产品代码'] in 已持有代码 else ''

    # 排序：买入优先级 → 收益率复合排名（当日收益×历史预期收益，回测验证最优）
    def 排序键(x):
        建议 = x.get('操作建议', '')
        if '★★★' in 建议:
            优先级 = 0
        elif '★★' in 建议:
            优先级 = 1
        elif '★' in 建议:
            优先级 = 2
        elif '可买入' in 建议:
            优先级 = 3
        elif '☆' in 建议:
            优先级 = 5  # ☆可关注排在★/可买入之后
        else:
            优先级 = 9  # 观望排最后
        # V5.10: 收益率复合排名（当日最新收益 × 历史预期收益）
        最新收益 = x.get('最新收益率%', 0) or 0
        预期收益 = x.get('预期年化收益%', 0) or 0
        复合收益分 = -(最新收益 * 预期收益)
        return (优先级, 复合收益分)

    实时机会_排序 = sorted(实时机会, key=排序键)

    # 高亮样式
    yellow_fill = PatternFill("solid", fgColor="FFF9C4")
    red_font = Font(color="CC0000", bold=True)

    for i, row in enumerate(实时机会_排序):
        # 构建行数据：固定列前 + 按日期对齐的收益值 + 固定列后
        行数据 = [row.get(h, '') for h in 固定列前]
        日收益格子 = [''] * 最大日数
        for d, v in row.get('每日收益', []):
            if d in 日期列映射:
                日收益格子[日期列映射[d]] = v
        行数据 += 日收益格子
        行数据 += [row.get(h, '') for h in 固定列后]

        ws1.append(行数据)
        当前行号 = ws1.max_row

        # 已持有高亮
        if row.get('已持有') == '是':
            for cell in ws1[当前行号]:
                cell.fill = yellow_fill

        # 每日收益 > 4.5% 标红
        for j in range(最大日数):
            col_idx = 日收益起始列 + j
            cell = ws1.cell(row=当前行号, column=col_idx)
            if isinstance(cell.value, (int, float)) and cell.value > 4.5:
                cell.font = red_font

    # 设置列宽
    col_widths1 = {}
    for ci, h in enumerate(固定列前, 1):
        if h == '产品名称':
            col_widths1[ci] = 40
        elif h == '产品代码':
            col_widths1[ci] = 18
        else:
            col_widths1[ci] = 12
    for ci in range(日收益起始列, 日收益起始列 + 最大日数):
        col_widths1[ci] = 11
    for ci, h in enumerate(固定列后, 日收益起始列 + 最大日数):
        if h == '操作建议':
            col_widths1[ci] = 35
        else:
            col_widths1[ci] = 14
    for col, width in col_widths1.items():
        ws1.column_dimensions[get_column_letter(col)].width = width
    ws1.freeze_panes = 'A2'

    # Sheet 精选: 8维度精选推荐 Top 20
    if 精选结果:
        gold_fill = PatternFill("solid", fgColor="FF8F00")
        ws_pick = wb.create_sheet("精选推荐Top20")

        pick_headers = ['排名', '来源', '银行', '产品代码', '产品名称', '流动性',
                        '历史成功率%', '历史信号次数', '信号距今天数',
                        '最新收益率%', '预期年化收益%', '信号期间平均收益%',
                        '操作建议', '综合得分',
                        '收益强度分', '可靠性分', '时效性分', '流动性分',
                        '持续质量分', '收益排名分', '稳定性分', '信号密度分',
                        '加速度分', '赎回费惩罚分',
                        '前瞻加成', '前瞻分', '预测释放日', '预测释放结束日', '预测置信度',
                        '释放周期', '释放阶段', '释放星期', '规律置信度',
                        '新鲜度标签', '新鲜度进度', '预测窗口天数', '收益加速度',
                        '有赎回费', '赎回费描述',
                        '节前推荐']
        ws_pick.append(pick_headers)
        for cell in ws_pick[1]:
            cell.font = header_font
            cell.fill = gold_fill
            cell.alignment = align_center

        for row in 精选结果:
            ws_pick.append([row.get(h, '') for h in pick_headers])

        pick_widths = {1: 6, 2: 8, 3: 12, 4: 18, 5: 40, 6: 10, 7: 12, 8: 10,
                       9: 10, 10: 12, 11: 12, 12: 14, 13: 35,
                       14: 10, 15: 10, 16: 10, 17: 10, 18: 10,
                       19: 10, 20: 10, 21: 10, 22: 10, 23: 10, 24: 10,
                       25: 8, 26: 8, 27: 12, 28: 12, 29: 10,
                       30: 10, 31: 16, 32: 10, 33: 10,
                       34: 18, 35: 10, 36: 10, 37: 10,
                       38: 10, 39: 25, 40: 12}
        for col, width in pick_widths.items():
            ws_pick.column_dimensions[get_column_letter(col)].width = width
        ws_pick.freeze_panes = 'A2'

        # 将精选sheet移到第一个位置（最重要）
        wb.move_sheet(ws_pick, offset=-1)

    # Sheet 2: 高成功率产品库
    ws2 = wb.create_sheet("高成功率产品库")

    headers2 = ['银行', '产品代码', '产品名称', '流动性', '历史成功率', '历史信号次数',
                '成功次数', '预期持有天数', '预期年化收益', '平均信号强度',
                '有赎回费', '赎回费描述']

    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = align_center

    产品库_排序 = sorted(产品库, key=lambda x: -x['历史成功率'])
    for row in 产品库_排序:
        ws2.append([row.get(h, '') for h in headers2])

    col_widths2 = {1: 12, 2: 18, 3: 45, 4: 12, 5: 12, 6: 10, 7: 12, 8: 14, 9: 12,
                   10: 12, 11: 10, 12: 25}
    for col, width in col_widths2.items():
        ws2.column_dimensions[get_column_letter(col)].width = width
    ws2.freeze_panes = 'A2'

    # Sheet 3: 潜力观察池
    if 潜力观察池:
        purple_fill = PatternFill("solid", fgColor="6A1B9A")
        ws_obs = wb.create_sheet("潜力观察池")

        headers_obs = ['银行', '产品代码', '产品名称', '流动性', '观察原因',
                       '历史成功率', '历史信号次数', '成功次数', '数据天数',
                       '预期持有天数', '预期年化收益', '平均信号强度', '最新信号日期']

        ws_obs.append(headers_obs)
        for cell in ws_obs[1]:
            cell.font = header_font
            cell.fill = purple_fill
            cell.alignment = align_center

        # 按银行+成功率排序
        观察池_排序 = sorted(潜力观察池,
                           key=lambda x: (x['银行'], -x['历史成功率']))
        for row in 观察池_排序:
            ws_obs.append([row.get(h, '') for h in headers_obs])

        col_widths_obs = {1: 12, 2: 18, 3: 45, 4: 10, 5: 35, 6: 12,
                          7: 12, 8: 10, 9: 10, 10: 14, 11: 12, 12: 12, 13: 14}
        for col, width in col_widths_obs.items():
            ws_obs.column_dimensions[get_column_letter(col)].width = width
        ws_obs.freeze_panes = 'A2'

        logger.info(f"  潜力观察池: {len(潜力观察池)} 个产品")

    # Sheet 4: 持仓分析
    if 持仓分析:
        orange_fill = PatternFill("solid", fgColor="E65100")
        ws_port = wb.create_sheet("持仓分析")

        headers_port = ['银行', '产品代码', '产品名称', '持仓状态', '净持仓金额',
                        '买入总额', '卖出总额', '首次买入日', '最近交易日', '持仓天数',
                        '交易次数', '是否在高成功率库', '最新年化收益%', '买入以来平均收益%',
                        '持仓建议', '智持标记']

        ws_port.append(headers_port)
        for cell in ws_port[1]:
            cell.font = header_font
            cell.fill = orange_fill
            cell.alignment = align_center

        # 持有中的排前面
        持仓排序 = sorted(持仓分析, key=lambda x: (0 if x['持仓状态'] == '持有中' else 1, -x.get('净持仓金额', 0)))
        red_font = Font(color="C62828")
        green_font = Font(color="2E7D32")
        blue_font = Font(color="1565C0")

        for row in 持仓排序:
            ws_port.append([row.get(h, '') for h in headers_port])
            # 持仓建议颜色标记 (V6.1: 智持=蓝色, 假期=绿色)
            建议col = headers_port.index('持仓建议') + 1
            建议cell = ws_port.cell(row=ws_port.max_row, column=建议col)
            建议内容 = row.get('持仓建议', '')
            if '智持' in 建议内容:
                建议cell.font = blue_font
            elif '假期' in 建议内容:
                建议cell.font = green_font
            elif '卖出' in 建议内容:
                建议cell.font = red_font
            elif '持有' in 建议内容 or '继续' in 建议内容:
                建议cell.font = green_font

        col_widths_port = {1: 12, 2: 18, 3: 35, 4: 10, 5: 16, 6: 16, 7: 16,
                           8: 12, 9: 12, 10: 10, 11: 10, 12: 14, 13: 14, 14: 16, 15: 30, 16: 20}
        for col, width in col_widths_port.items():
            ws_port.column_dimensions[get_column_letter(col)].width = width
        ws_port.freeze_panes = 'A2'

    # Sheet 4: 策略说明
    ws3 = wb.create_sheet("策略说明")
    说明 = [
        ["高成功率短周期轮动策略 V6.1 (智持+节前+Alpha优化)"],
        [""],
        ["【策略目标】"],
        ["捕捉银行理财产品的短期'高收益释放'窗口，6仓集中+高买入门槛+快速轮动+前瞻预测Alpha加成"],
        [""],
        ["【核心逻辑】"],
        ["1. 流动性过滤：只保留赎回周期<=14天的产品（日开/7天/14天持有期）"],
        ["2. 正确年化：期间收益 / 实际天数 × 365，消除数据频率偏差"],
        ["3. 宽入库: 30%成功率入库，用排名区分优劣"],
        ["4. 高买入门槛: 当日年化>3.5%才买入（只抓强信号）"],
        ["5. 6仓集中: 1667万/仓，集中资金于最优信号"],
        ["6. 20天快速轮动: 最多持有20交易日，加速资金周转"],
        ["7. 轮动卖出: 满仓时替换最弱持仓为更强候选"],
        ["8. 按收益率复合排名（当日收益×历史平均收益）优选产品"],
        ["9. V6.0: 前瞻预测Alpha加成 — 学习释放规律(周期/阶段/星期)，预测即将释放"],
        ["   的产品，精选推荐时加成排名（回测验证: 年化+0.12%, 夏普1.26→1.52）"],
        ["10. V6.1 建议① Alpha加成优化: 权重0.25→0.30 (回测Mode B年化+0.25%)"],
        ["11. V6.1 建议② 智持(延迟卖出): 卖出前检查预测释放, 90%命中率避免误卖"],
        ["12. V6.1 建议③ 节前买入: 节前日推荐高收益产品, 含假期卖出保护"],
        [""],
        ["【为什么限制赎回周期<=14天？】"],
        ["- 收益释放窗口通常只有3-10天"],
        ["- 如果产品赎回周期>14天，窗口结束后无法退出"],
        ["- 被迫持有会将高收益拉低至4.5%以下，不符合收益目标"],
        ["- 因此：日开产品最优，7天/14天持有期产品可接受"],
        [""],
        ["【关键指标说明】"],
        [f"买入信号: 当日年化 > {信号阈值_买入}%（高质量信号门槛）{'（需突破: 前日≤阈值）' if 需要突破信号 else '（无需突破）'}"],
        [f"卖出信号: 当日年化 ≤ {信号阈值_卖出}%"],
        [f"成功定义: 持有{持有天数}天平均年化 > {成功标准}%"],
        [f"持仓集中: 6仓×1667万（集中资金于最优信号）"],
        [f"持有期约束: {最短持续天数}天 <= 预期持有天数 <= {最大持有天数}天"],
        [f"赎回周期约束: 产品赎回周期 <= {最长赎回天数}天"],
        ["预期持有天数: 基于历史信号的平均持续时间"],
        ["预期年化收益: 基于历史信号的平均收益"],
        ["流动性: 产品赎回类型（日开/7天持有期/14天持有期）"],
        ["最新收益率: 数据库中最新一天的年化收益"],
        ["信号持续天数: 信号触发后连续维持高收益的天数"],
        ["信号期间平均收益: 信号触发至今的平均年化收益"],
        [""],
        ["【宏观风控硬限制】"],
        [f"信用利差阈值: {信用利差阈值}bp（3Y AA+中票与国债利差）"],
        ["当信用利差 > 阈值时，立即全部赎回，停止一切买入操作"],
        ["数据来源: 信用利差.json（需手动或自动更新）"],
        [""],
        ["【操作建议说明】"],
        ["★★★ 今日新信号: 今日刚触发，最佳买入时机"],
        ["★★ 推荐买入: 信号1-2天内，仍在高收益区"],
        ["★ 可买入: 信号较早但当前收益仍高于阈值，预计仍有剩余窗口"],
        ["可买入（超预期）: 持续时间超预期但收益仍高"],
        ["观望: 高收益窗口已结束，当前收益已跌回阈值以下"],
        [""],
        ["【潜力观察池】"],
        ["数据不足或成功率偏低的产品不直接淘汰，进入观察池跟踪："],
        [f"  入池条件1: 信号次数>=1 但 <{最少历史信号}次（数据积累中）"],
        [f"  入池条件2: 成功率>={观察池最低成功率}% 但 <{最低成功率}%（待提升）"],
        [f"  入池条件3: 预期持有天数超出{最短持续天数}~{最大持有天数}天范围"],
        ["观察池产品仍会被扫描实时信号，但操作建议会标注[观察池]"],
        ["随着数据积累，观察池产品可能自动升入正式库"],
        [""],
        ["【缓存机制】"],
        [f"高成功率产品库每日更新"],
        ["如需强制刷新，使用 --refresh 参数运行"],
    ]
    for row in 说明:
        ws3.append(row)
    ws3.column_dimensions['A'].width = 55

    # 保存
    try:
        wb.save(OUTPUT_PATH)
        logger.info(f"报告已保存: {OUTPUT_PATH}")
    except Exception as e:
        backup = os.path.join(BASE_DIR, f"高成功率交易机会_{datetime.now().strftime('%H%M%S')}.xlsx")
        wb.save(backup)
        logger.info(f"已保存到备用路径: {backup}")


# ============================================================
# 主程序
# ============================================================

def main(强制刷新缓存=False):
    print("=" * 70)
    print("   银行理财产品 - 高成功率交易机会系统 V6.1 (智持+节前+Alpha优化)")
    print("=" * 70)
    print(f"运行时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()

    # V6.1 优化状态显示
    today_str = datetime.now().strftime('%Y-%m-%d')
    _pre_hol, _pre_hol_name = is_pre_holiday(today_str)
    print("【V6.1 优化状态】")
    print(f"  建议① Alpha加成: 权重={ALPHA_预测加成}")
    print(f"  建议② 智持(延迟卖出): 启用")
    if _pre_hol:
        print(f"  建议③ 节前买入: ★ 今日为{_pre_hol_name}节前日!")
    else:
        print(f"  建议③ 节前买入: 非节前")
    print()

    # 步骤0: 信用利差风控检查
    print("【风控检查】信用利差...")
    利差触发, 当前利差, 利差日期, 利差基准 = 检查信用利差风控()
    if 利差触发:
        print(f"  !!!!!! 信用利差风控触发！{利差基准}={当前利差}bp > {信用利差阈值}bp（{利差日期}）")
        print(f"  → 硬限制：全部赎回，停止一切买入操作")
    elif 当前利差 is not None:
        print(f"  {利差基准}={当前利差}bp ≤ {信用利差阈值}bp（{利差日期}） → 正常")
    else:
        print(f"  未配置信用利差数据（请创建 信用利差.json）→ 跳过")
    print()

    # 步骤1: 加载或构建高成功率产品库
    print("【第1步】加载高成功率产品库...")
    产品库, 潜力观察池 = 构建高成功率产品库(强制刷新=强制刷新缓存)
    print(f"  高成功率产品库: {len(产品库)}")
    print(f"  潜力观察池: {len(潜力观察池)}")
    # 按银行统计观察池
    观察池_by_bank = {}
    for p in 潜力观察池:
        b = p['银行']
        观察池_by_bank[b] = 观察池_by_bank.get(b, 0) + 1
    for b, c in 观察池_by_bank.items():
        print(f"    - {b}: {c} 个待验证产品")

    # 步骤2: 前瞻预测（学习释放规律）— V6.2: 提前到扫描之前, 供新鲜度计算使用
    print(f"\n【第2步】学习释放规律并预测...")
    前瞻预测数据 = 学习释放规律并预测(产品库) if not 利差触发 else {'prediction_scores': {}, 'patterns': {}, 'sorted_dates': [], 'date_to_idx': {}}
    前瞻预测 = 前瞻预测数据.get('prediction_scores', {})
    释放规律 = 前瞻预测数据.get('patterns', {})
    if 前瞻预测:
        print(f"  {len(前瞻预测)} 个产品有前瞻预测, {len(释放规律)} 个有释放规律")
    else:
        print(f"  无可用前瞻预测" if not 利差触发 else "  风控触发，跳过预测")

    # 步骤3: 扫描实时机会 (传入释放规律供新鲜度计算)
    print("\n【第3步】扫描实时交易机会...")
    实时机会 = 扫描实时机会(产品库, 潜力观察池, 释放规律=释放规律)

    # 信用利差硬限制：触发时覆盖所有操作建议为全部赎回
    if 利差触发:
        for o in 实时机会:
            o['操作建议'] = f"!! 全部赎回（信用利差{当前利差}bp>{信用利差阈值}bp）"

    # 步骤4: 加载持仓 (不分析, 等分析持仓时再用)
    print("\n【第4步】加载持仓数据...")
    交易记录, 当前持仓 = 加载持仓()
    持仓分析结果 = None
    已持有代码 = set()
    if 当前持仓:
        持有中 = [p for p in 当前持仓 if p['持仓状态'] == '持有中']
        已持有代码 = {p['产品代码'] for p in 持有中}
        print(f"  有效交易: {len(交易记录) if 交易记录 is not None else 0} 条")
        print(f"  持有中: {len(持有中)} 个产品, 已清仓: {len(当前持仓) - len(持有中)} 个")
    else:
        print("  未找到持仓数据")

    # 步骤5: 分析持仓 (传入预测数据供智持判断)
    if 当前持仓:
        print(f"\n【第5步】分析持仓表现...")
        持仓分析结果 = 分析持仓(当前持仓, 产品库, 实时机会, 预测数据=前瞻预测数据)

        # 信用利差硬限制：覆盖所有持仓建议
        if 利差触发 and 持仓分析结果:
            for p in 持仓分析结果:
                if p['持仓状态'] == '持有中':
                    p['持仓建议'] = f"!! 立即赎回（信用利差{当前利差}bp>{信用利差阈值}bp）"

    # 步骤6: 精选推荐（多维度优中选优 + 前瞻加成 + 新鲜度过滤 + 节前推荐）
    精选步骤 = '6' if 当前持仓 else '5'
    print(f"\n【第{精选步骤}步】多维度精选推荐 + 前瞻加成 + 新鲜度过滤...")
    精选结果 = 精选推荐(实时机会, 前瞻预测, 产品库=产品库,
                     是否节前=_pre_hol, 节前假期名=_pre_hol_name,
                     释放规律=释放规律) if not 利差触发 else []

    # 步骤7: 生成报告
    报告步骤 = str(int(精选步骤) + 1)
    print(f"\n【第{报告步骤}步】生成报告...")
    logger.info(f"前瞻预测: {len(前瞻预测)}个产品")
    生成报告(产品库, 实时机会, 持仓分析结果, 已持有代码, 潜力观察池, 精选结果)

    # 汇总
    print("\n" + "=" * 70)
    print("                      分析完成")
    print("=" * 70)

    # 统计
    推荐买入 = [o for o in 实时机会 if '★' in o['操作建议']]
    可买入 = [o for o in 实时机会 if '买入' in o['操作建议']]
    观望 = [o for o in 实时机会 if '观望' in o['操作建议']]

    # 按银行统计
    银行买入 = {}
    for o in 可买入:
        b = o['银行']
        银行买入[b] = 银行买入.get(b, 0) + 1

    # 分离正式库和观察池信号
    正式信号 = [o for o in 实时机会 if o.get('来源') == '正式库']
    观察池信号 = [o for o in 实时机会 if o.get('来源') == '观察池']

    print(f"\n【汇总统计】")

    # 信用利差状态
    if 利差触发:
        print(f"  !! 信用利差风控: {利差基准}={当前利差}bp > {信用利差阈值}bp → 全部赎回！")
        print(f"  所有买入建议已覆盖为「全部赎回」")
    elif 当前利差 is not None:
        print(f"  信用利差: {利差基准}={当前利差}bp（安全，阈值{信用利差阈值}bp）")
    else:
        print(f"  信用利差: 未配置（信用利差.json）")

    print(f"  高成功率产品库: {len(产品库)} 个")
    print(f"  潜力观察池: {len(潜力观察池)} 个")
    print(f"  实时信号总数: {len(实时机会)} 个（正式{len(正式信号)}+观察池{len(观察池信号)}）")
    if not 利差触发:
        print(f"    - 推荐买入(★): {len(推荐买入)} 个")
        print(f"    - 可买入(含★): {len(可买入)} 个")
        print(f"    - 观望(窗口已过): {len(观望)} 个")
        print(f"  各银行买入机会:")
        for b, c in 银行买入.items():
            print(f"    - {b}: {c} 个")
    else:
        print(f"    - 全部标记为赎回: {len(实时机会)} 个")

    # 观察池信号摘要
    if 观察池信号:
        print(f"\n【潜力观察池信号】({len(观察池信号)}个)")
        for o in sorted(观察池信号, key=lambda x: -x['历史成功率%'])[:10]:
            print(f"  {o['银行']} | {o['产品代码']} | "
                  f"成功率{o['历史成功率%']}% | 信号{o['历史信号次数']}次 | "
                  f"最新{o['最新收益率%']}% | {o.get('观察原因', '')}")

    # ====== 精选推荐（核心输出）======
    if 精选结果:
        print(f"\n{'─' * 70}")
        print(f"  【精选推荐 Top {len(精选结果)}】 8维度综合评分优中选优")
        print(f"  筛选: 正式库+观察池 + ★/☆/可买入 + 收益>{信号阈值_库}% + 成功率≥{精选_最低成功率}% + "
              f"距今≤{精选_信号最大距今}天")
        print(f"  评分: 收益强度{int(W_收益强度*100)}% + 可靠性{int(W_历史可靠*100)}% + "
              f"时效{int(W_信号时效*100)}% + 流动性{int(W_流动性*100)}% + "
              f"持续质量{int(W_持续质量*100)}% + 排名{int(W_收益排名*100)}% + "
              f"稳定{int(W_收益稳定*100)}% + 密度{int(W_信号密度*100)}%")
        print(f"  去重: 同系产品仅保留最优份额 | 分散: 同行≤{精选_单行最大}个")
        print(f"{'─' * 70}")
        for r in 精选结果:
            持有标记 = " [已持有]" if r.get('产品代码') in 已持有代码 else ""
            来源标记 = " [观察池]" if r.get('来源') == '观察池' else ""
            预测标记 = " [前瞻]" if r.get('前瞻加成') == '是' else ""
            print(f"  {r['排名']:>2}. {r['银行']} | {r['产品代码']:<14s} | "
                  f"成功率{r['历史成功率%']:>5.1f}% | "
                  f"最新{r['最新收益率%']:>6.2f}% | "
                  f"综合{r['综合得分']:>5.1f}分 | "
                  f"距今{r['信号距今天数']}天{预测标记}{来源标记}{持有标记}")
            print(f"      {r['产品名称'][:35]}")
            预测信息 = ""
            if r.get('前瞻加成') == '是':
                预测信息 = (f"  前瞻: {r['预测释放日']}~{r['预测释放结束日']} "
                          f"置信{r['预测置信度']:.2f}")
            规律信息 = ""
            _parts = []
            if r.get('释放周期'): _parts.append(r['释放周期'])
            if r.get('释放阶段'): _parts.append(r['释放阶段'])
            if r.get('释放星期'): _parts.append(r['释放星期'])
            if _parts:
                规律信息 = f"  规律: {'/'.join(_parts)} (置信{r.get('规律置信度', 0):.2f})"
            print(f"      [收益{r['收益强度分']:.0f} 可靠{r['可靠性分']:.0f} "
                  f"时效{r['时效性分']:.0f} 流动{r['流动性分']:.0f} "
                  f"持续{r['持续质量分']:.0f} 排名{r.get('收益排名分', 0):.0f} "
                  f"稳定{r.get('稳定性分', 0):.0f} 密度{r.get('信号密度分', 0):.0f}]  "
                  f"{r['操作建议']}{预测信息}{规律信息}")
        print(f"{'─' * 70}")
    elif not 利差触发:
        print(f"\n  精选推荐: 暂无符合条件的产品（放宽精选参数或等待新信号）")

    if 推荐买入 and not 精选结果:
        print(f"\n【全量推荐买入 Top 10】（收益率复合排名，未经精选）")
        for i, o in enumerate(sorted(推荐买入,
            key=lambda x: -(x.get('最新收益率%', 0) or 0) * (x.get('预期年化收益%', 0) or 0))[:10]):
            持有标记 = " [已持有]" if o.get('产品代码') in 已持有代码 else ""
            print(f"  {i+1}. {o['银行']} | {o['产品代码']} | "
                  f"成功率{o['历史成功率%']}% | "
                  f"最新收益{o['最新收益率%']}% | "
                  f"{o['操作建议']}{持有标记}")

    # 持仓统计
    if 持仓分析结果:
        持有中 = [p for p in 持仓分析结果 if p['持仓状态'] == '持有中']
        print(f"\n【持仓分析】")
        print(f"  持有产品: {len(持有中)} 个")
        for p in 持有中:
            收益 = p.get('最新年化收益%', '')
            收益str = f"{收益}%" if isinstance(收益, (int, float)) and 收益 != '' else '无数据'
            print(f"    {p['产品代码']} | {p['产品名称'][:15]} | "
                  f"净持仓{p['净持仓金额']:,.0f} | "
                  f"最新收益{收益str} | {p['持仓建议']}")

    print(f"\n报告文件: {OUTPUT_PATH}")
    print("=" * 70)


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--refresh', action='store_true', help='强制刷新产品库缓存')
    args = parser.parse_args()

    # 阻止系统休眠（运行期间保持唤醒）
    ES_CONTINUOUS = 0x80000000
    ES_SYSTEM_REQUIRED = 0x00000001
    ctypes.windll.kernel32.SetThreadExecutionState(ES_CONTINUOUS | ES_SYSTEM_REQUIRED)
    try:
        main(强制刷新缓存=args.refresh)
    finally:
        # 恢复默认休眠策略
        ctypes.windll.kernel32.SetThreadExecutionState(ES_CONTINUOUS)
