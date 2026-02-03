# -*- coding: utf-8 -*-
"""
银行理财产品"收益释放"操纵行为监测与套利分析系统 (V3.0 - 完整版)

新增功能（V3.0）：
1. 风险控制：最大回撤监控、动态止损
2. 收益统计：成功率、平均收益、夏普比率等
3. 可视化：收益分布、累计收益曲线、产品对比图
4. 回测模块：模拟历史交易、计算累计收益

作者：AI-FINANCE
版本：V3.0
"""

import os
import sys
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.drawing.image import Image as XLImage
import logging
import warnings
import json
import matplotlib
matplotlib.use('Agg')  # 使用非交互式后端
import matplotlib.pyplot as plt
from matplotlib.font_manager import FontProperties
import matplotlib.patches as mpatches

# 设置中文字体
plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei']
plt.rcParams['axes.unicode_minus'] = False

# ============================================================
# 日志配置
# ============================================================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('bank_product_strategy.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
日志 = logging.getLogger(__name__)

# ============================================================
# 核心策略参数
# ============================================================
净值数据库路径 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "净值数据库.xlsx")
输出文件路径 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "高收益捕捉分析报告_V3.xlsx")
图表输出目录 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "charts")

# 信号识别参数
收益拉升阈值 = 5.0         # %
收益下滑阈值 = 2.0         # %

# 周期约束
最短持有天数 = 7
最长持有天数 = 28
默认预测周期 = 14

# 分析范围
信号回溯天数 = 30
历史分析回溯 = 365

# 风险控制参数（新增）
最大回撤止损阈值 = 30.0    # %，回撤超过此值触发止损
预期收益折扣系数 = 0.6     # 实际收益低于预期*此系数时触发止损

# 回测参数（新增）
初始资金 = 100000.0        # 初始投资金额
单笔投资比例 = 0.1         # 每笔交易占总资金的比例
无风险收益率 = 0.025       # 年化无风险收益率（用于计算夏普比率）


# ============================================================
# 工具函数
# ============================================================

def 是否日期列(列名):
    if not isinstance(列名, str):
        return False
    if len(列名) == 10 and 列名[4] == '-' and 列名[7] == '-':
        try:
            datetime.strptime(列名, '%Y-%m-%d')
            return True
        except ValueError:
            return False
    return False


def 识别产品类别(产品名称):
    if not isinstance(产品名称, str):
        产品名称 = str(产品名称)

    产品名称原始 = 产品名称
    产品名称 = 产品名称.upper()

    if 'R3' in 产品名称 or 'R4' in 产品名称 or 'R5' in 产品名称:
        return '含权益类'
    if 'R1' in 产品名称 or 'R2' in 产品名称:
        return '不含权益类'

    含权益关键词 = ['增强', '混合', '权益', '股票', '可转债', 'FOF', '偏债',
                    '灵活配置', '进取', '成长', '积极']
    for 关键词 in 含权益关键词:
        if 关键词 in 产品名称原始:
            return '含权益类'

    不含权益关键词 = ['纯债', '货币', '稳健', '固收']
    for 关键词 in 不含权益关键词:
        if 关键词 in 产品名称原始:
            return '不含权益类'

    return '未知类别'


def 计算最大回撤(收益序列):
    """
    计算持有期内的最大回撤
    返回：(最大回撤%, 回撤发生位置)
    """
    if not 收益序列 or len(收益序列) == 0:
        return 0.0, 0

    累计收益 = [sum(收益序列[:i+1]) for i in range(len(收益序列))]
    最大值 = 累计收益[0]
    最大回撤 = 0.0
    回撤位置 = 0

    for i, 当前值 in enumerate(累计收益):
        if 当前值 > 最大值:
            最大值 = 当前值
        回撤 = ((最大值 - 当前值) / 最大值 * 100) if 最大值 > 0 else 0
        if 回撤 > 最大回撤:
            最大回撤 = 回撤
            回撤位置 = i

    return 最大回撤, 回撤位置


# ============================================================
# 数据加载与预处理（复用之前的代码）
# ============================================================

def _加载JSON净值数据(银行名):
    基础目录 = os.path.dirname(os.path.abspath(__file__))
    if '华夏' in 银行名:
        路径 = os.path.join(基础目录, 'huaxia_nav_history.json')
        if os.path.exists(路径):
            try:
                with open(路径, 'r', encoding='utf-8') as f:
                    数据 = json.load(f)
                结果 = {}
                for 代码, 净值列表 in 数据.items():
                    净值字典 = {}
                    for 条目 in 净值列表:
                        净值字典[条目['date']] = str(条目['unit_nav'])
                    结果[代码] = 净值字典
                日志.info(f"    已加载华夏JSON数据: {len(结果)} 个产品")
                return 结果
            except Exception as e:
                日志.warning(f"    加载华夏JSON失败: {e}")
    return None


def 修复产品代码(数据库):
    日志.info("开始修复产品代码（投票匹配法）...")
    修复后数据库 = {}

    for 工作表名, df in 数据库.items():
        缺失数 = df['产品代码'].isna().sum()
        if 缺失数 == 0:
            修复后数据库[工作表名] = df
            continue

        日期列 = sorted([列 for 列 in df.columns if 是否日期列(列)])
        有代码掩码 = df['产品代码'].notna()
        有代码df = df[有代码掩码].copy()
        无代码df = df[~有代码掩码].copy()

        代码名称映射 = {}
        for _, 行 in 有代码df.iterrows():
            代码 = str(行['产品代码']).strip()
            名称 = str(行['产品名称']).strip()
            if 代码 and 名称:
                代码名称映射[代码] = 名称

        日期净值到代码 = {}
        for _, 行 in 有代码df.iterrows():
            代码 = str(行['产品代码']).strip()
            for 日期 in 日期列:
                值 = 行.get(日期)
                if pd.notna(值) and str(值).strip():
                    键 = (日期, str(值).strip())
                    if 键 not in 日期净值到代码:
                        日期净值到代码[键] = set()
                    日期净值到代码[键].add(代码)

        JSON数据 = _加载JSON净值数据(工作表名)
        if JSON数据:
            for 代码, 净值字典 in JSON数据.items():
                for 日期 in 日期列:
                    if 日期 in 净值字典:
                        键 = (日期, 净值字典[日期])
                        if 键 not in 日期净值到代码:
                            日期净值到代码[键] = set()
                        日期净值到代码[键].add(代码)

        已匹配 = 0
        for idx, 行 in 无代码df.iterrows():
            代码投票 = {}
            for 日期 in 日期列:
                值 = 行.get(日期)
                if pd.notna(值) and str(值).strip():
                    键 = (日期, str(值).strip())
                    if 键 in 日期净值到代码:
                        for 候选代码 in 日期净值到代码[键]:
                            代码投票[候选代码] = 代码投票.get(候选代码, 0) + 1
            if 代码投票:
                最佳代码 = max(代码投票, key=代码投票.get)
                无代码df.at[idx, '产品代码'] = 最佳代码
                无代码df.at[idx, '产品名称'] = 代码名称映射.get(最佳代码, 最佳代码)
                已匹配 += 1

        已修复 = pd.concat([有代码df, 无代码df], ignore_index=True)

        合并结果 = []
        按代码分组 = 已修复.groupby('产品代码')
        for 代码, 分组 in 按代码分组:
            if len(分组) == 1:
                合并结果.append(分组.iloc[0])
            else:
                合并行 = 分组.iloc[0].copy()
                for 日期 in 日期列:
                    非空值 = 分组[日期].dropna()
                    if len(非空值) > 0:
                        合并行[日期] = 非空值.iloc[0]
                合并结果.append(合并行)

        修复df = pd.DataFrame(合并结果).reset_index(drop=True)
        修复后数据库[工作表名] = 修复df

    日志.info("产品代码修复完成")
    return 修复后数据库


def 加载净值数据库():
    if not os.path.exists(净值数据库路径):
        日志.error(f"找不到文件: {净值数据库路径}")
        return {}

    日志.info("正在加载净值数据库...")
    xlsx = pd.ExcelFile(净值数据库路径)
    数据库 = {}

    for sheet in xlsx.sheet_names:
        df = pd.read_excel(xlsx, sheet_name=sheet, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]

        重命名映射 = {}
        if 'level_0' in df.columns:
            重命名映射['level_0'] = '产品代码'
        if 'level_1' in df.columns:
            重命名映射['level_1'] = '产品名称'
        if 重命名映射:
            df = df.rename(columns=重命名映射)

        if '产品代码' in df.columns and '产品名称' in df.columns:
            数据库[sheet] = df
            日志.info(f"  {sheet}: {len(df)} 个产品")

    return 数据库


def 计算基础收益率(df, 银行名):
    全部日期 = sorted([列 for 列 in df.columns if 是否日期列(列)])

    if len(全部日期) < 2:
        日志.warning(f"[{银行名}] 日期数量不足")
        return None, [], []

    净值矩阵 = df[全部日期].apply(pd.to_numeric, errors='coerce')

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", FutureWarning)
        收益率矩阵 = 净值矩阵.pct_change(axis=1) * 365 * 100

    收益率日期 = 全部日期[1:]

    收益率数据 = np.round(收益率矩阵[收益率日期].values, 4)
    结果 = pd.DataFrame(收益率数据, columns=收益率日期)
    结果.insert(0, '产品代码', df['产品代码'].values)
    结果.insert(1, '产品名称', df['产品名称'].values)
    结果.insert(2, '银行', 银行名)
    结果.insert(3, '产品类别', df['产品名称'].apply(识别产品类别))

    有效掩码 = 结果[收益率日期].notna().any(axis=1)
    结果 = 结果[有效掩码].reset_index(drop=True)

    日志.info(f"[{银行名}] 收益率计算完成: {len(结果)} 个有效产品")

    return 结果, 全部日期, 收益率日期


# ============================================================
# 历史规律分析模块
# ============================================================

def 分析历史拉升规律(收益率序列, 当前日期, 收益率日期列表):
    try:
        if 当前日期 not in 收益率日期列表:
            return 默认预测周期, 收益拉升阈值, 0

        当前索引 = 收益率日期列表.index(当前日期)
        历史日期 = 收益率日期列表[:当前索引]

        if len(历史日期) < 30:
            return 默认预测周期, 收益拉升阈值, 0

        历史收益 = [收益率序列.get(日期, 0) for 日期 in 历史日期]

        片段列表 = []
        当前片段 = []

        for 收益 in 历史收益:
            if 收益 > (收益拉升阈值 - 0.5):
                当前片段.append(收益)
            else:
                if len(当前片段) >= 3:
                    片段列表.append(当前片段[:])
                当前片段 = []

        if len(当前片段) >= 3:
            片段列表.append(当前片段)

        if not 片段列表:
            return 默认预测周期, 收益拉升阈值, 0

        持续天数集 = [len(p) for p in 片段列表]
        平均收益集 = [np.mean(p) for p in 片段列表]

        预期天数 = int(np.mean(持续天数集))
        预期天数 = max(最短持有天数, min(预期天数, 最长持有天数))
        预期收益 = np.mean(平均收益集)

        return 预期天数, 预期收益, len(片段列表)

    except Exception as e:
        日志.warning(f"分析历史规律出错: {e}")
        return 默认预测周期, 收益拉升阈值, 0


# ============================================================
# 策略执行引擎（增强版：含风险控制）
# ============================================================

def 执行策略分析(收益率df, 全部日期, 收益率日期, 银行名):
    """
    增强版策略执行：
    - 添加最大回撤监控
    - 添加动态止损
    - 记录完整的风险指标
    """
    分析结果 = []
    跟踪详情 = []

    分析起始位 = max(0, len(收益率日期) - 信号回溯天数)

    日志.info(f"[{银行名}] 开始策略分析（{len(收益率df)} 个产品）...")

    处理计数 = 0
    for idx, row in 收益率df.iterrows():
        产品代码 = row['产品代码']
        产品名称 = row['产品名称']
        产品类别 = row['产品类别']

        收益序列 = {}
        for 日期 in 收益率日期:
            try:
                收益 = pd.to_numeric(row[日期], errors='coerce')
                收益序列[日期] = 收益 if pd.notna(收益) else 0
            except:
                收益序列[日期] = 0

        for i in range(分析起始位, len(收益率日期)):
            当前日期 = 收益率日期[i]
            当前收益 = 收益序列[当前日期]

            if i > 0:
                昨日日期 = 收益率日期[i-1]
                昨日收益 = 收益序列[昨日日期]
                是启动信号 = (当前收益 > 收益拉升阈值) and (昨日收益 <= 收益拉升阈值)
            else:
                是启动信号 = 当前收益 > 收益拉升阈值

            if not 是启动信号:
                continue

            预期天数, 预期收益, 历史次数 = 分析历史拉升规律(收益序列, 当前日期, 收益率日期)

            # 持有期跟踪（增强版：含风险控制）
            实际持有天数 = 0
            累计收益和 = 0
            持有期收益列表 = []  # 用于计算回撤
            卖出触发 = False
            卖出原因 = ""
            当前状态 = "持有中"
            最大回撤值 = 0.0
            回撤触发止损 = False

            for j in range(i + 1, len(收益率日期)):
                跟踪日期 = 收益率日期[j]
                跟踪收益 = 收益序列[跟踪日期]

                实际持有天数 += 1
                累计收益和 += 跟踪收益
                持有期收益列表.append(跟踪收益)
                当前持有年化 = 累计收益和 / 实际持有天数

                # 计算最大回撤
                当前回撤, _ = 计算最大回撤(持有期收益列表)
                if 当前回撤 > 最大回撤值:
                    最大回撤值 = 当前回撤

                # 风险控制：回撤止损
                if 当前回撤 > 最大回撤止损阈值:
                    卖出触发 = True
                    卖出原因 = f"回撤止损(回撤{当前回撤:.1f}%)"
                    当前状态 = "止损卖出"
                    回撤触发止损 = True

                # 原有止盈/止损逻辑
                elif 实际持有天数 >= 预期天数:
                    卖出触发 = True
                    卖出原因 = f"达到预期周期({预期天数}天)"
                    当前状态 = "止盈卖出"

                elif 实际持有天数 >= 最短持有天数:
                    if 跟踪收益 < 收益下滑阈值:
                        卖出触发 = True
                        卖出原因 = f"收益下滑(当日{跟踪收益:.2f}%)"
                        当前状态 = "止损卖出"
                    elif 当前持有年化 < (预期收益 * 预期收益折扣系数):
                        卖出触发 = True
                        卖出原因 = f"收益不达预期(实际{当前持有年化:.2f}% < 预期{预期收益*预期收益折扣系数:.2f}%)"
                        当前状态 = "止损卖出"

                跟踪详情.append({
                    '银行': 银行名,
                    '产品代码': 产品代码,
                    '产品名称': 产品名称,
                    '产品类别': 产品类别,
                    '信号日期': 当前日期,
                    '跟踪日期': 跟踪日期,
                    '持有天数': 实际持有天数,
                    '当日收益': round(跟踪收益, 2),
                    '当前持有年化': round(当前持有年化, 2),
                    '当前回撤': round(当前回撤, 2),
                    '卖出触发': '是' if 卖出触发 else ''
                })

                if 卖出触发:
                    break

            if not 卖出触发:
                if 实际持有天数 >= 最长持有天数:
                    当前状态 = "强制卖出(超期)"
                else:
                    当前状态 = "建议持有"

            最终持有年化 = (累计收益和 / 实际持有天数) if 实际持有天数 > 0 else 0

            # 判断交易是否成功（实际收益 >= 预期收益）
            交易成功 = 最终持有年化 >= 预期收益

            分析结果.append({
                '银行': 银行名,
                '产品代码': 产品代码,
                '产品名称': 产品名称,
                '产品类别': 产品类别,
                '信号日期': 当前日期,
                '启动收益率%': round(当前收益, 2),
                '历史拉升次数': 历史次数,
                '预期维持天数': 预期天数,
                '预期年化收益%': round(预期收益, 2),
                '最新状态': 当前状态,
                '已持有天数': 实际持有天数,
                '当前持有年化%': round(最终持有年化, 2),
                '收益差(实际-预期)': round(最终持有年化 - 预期收益, 2),
                '最大回撤%': round(最大回撤值, 2),
                '是否回撤止损': '是' if 回撤触发止损 else '否',
                '交易成功': '成功' if 交易成功 else '失败',
                '卖出原因': 卖出原因,
                '操作建议': '立即买入' if 实际持有天数 == 0 else 当前状态
            })

        处理计数 += 1
        if 处理计数 % 500 == 0:
            日志.info(f"  已处理 {处理计数}/{len(收益率df)} 个产品")

    日志.info(f"[{银行名}] 策略分析完成: 捕获 {len(分析结果)} 个信号")
    return 分析结果, 跟踪详情


# ============================================================
# 收益统计模块（新增）
# ============================================================

def 计算收益统计(分析结果列表):
    """
    计算各类统计指标：
    - 成功率
    - 平均收益
    - 最佳/最差交易
    - 平均持有天数
    - 总收益/总亏损
    - 夏普比率
    """
    if not 分析结果列表:
        return {}

    df = pd.DataFrame(分析结果列表)

    # 基础统计
    总信号数 = len(df)
    成功交易 = df[df['交易成功'] == '成功']
    失败交易 = df[df['交易成功'] == '失败']

    成功率 = len(成功交易) / 总信号数 * 100 if 总信号数 > 0 else 0

    平均收益 = df['当前持有年化%'].mean()
    平均持有天数 = df['已持有天数'].mean()
    平均回撤 = df['最大回撤%'].mean()

    最佳交易 = df.loc[df['当前持有年化%'].idxmax()] if len(df) > 0 else None
    最差交易 = df.loc[df['当前持有年化%'].idxmin()] if len(df) > 0 else None

    # 总盈亏
    总盈利 = 成功交易['当前持有年化%'].sum() if len(成功交易) > 0 else 0
    总亏损 = 失败交易['当前持有年化%'].sum() if len(失败交易) > 0 else 0

    # 夏普比率（简化版：假设每笔交易独立）
    收益率标准差 = df['当前持有年化%'].std()
    夏普比率 = ((平均收益 - 无风险收益率 * 100) / 收益率标准差) if 收益率标准差 > 0 else 0

    # 分类别统计
    类别统计 = {}
    for 类别 in df['产品类别'].unique():
        类别df = df[df['产品类别'] == 类别]
        类别统计[类别] = {
            '信号数': len(类别df),
            '成功率': len(类别df[类别df['交易成功'] == '成功']) / len(类别df) * 100 if len(类别df) > 0 else 0,
            '平均收益': 类别df['当前持有年化%'].mean(),
            '平均回撤': 类别df['最大回撤%'].mean()
        }

    统计结果 = {
        '总信号数': 总信号数,
        '成功交易数': len(成功交易),
        '失败交易数': len(失败交易),
        '成功率%': round(成功率, 2),
        '平均收益%': round(平均收益, 2),
        '平均持有天数': round(平均持有天数, 1),
        '平均回撤%': round(平均回撤, 2),
        '最佳交易收益%': round(最佳交易['当前持有年化%'], 2) if 最佳交易 is not None else 0,
        '最差交易收益%': round(最差交易['当前持有年化%'], 2) if 最差交易 is not None else 0,
        '总盈利%': round(总盈利, 2),
        '总亏损%': round(总亏损, 2),
        '夏普比率': round(夏普比率, 2),
        '分类别统计': 类别统计
    }

    return 统计结果


# ============================================================
# 回测模块（新增）
# ============================================================

def 执行回测模拟(分析结果列表):
    """
    模拟历史交易，计算累计收益
    假设：
    - 初始资金 = 100000
    - 每笔交易投入 = 总资金 * 10%
    - 按信号日期顺序执行交易
    """
    if not 分析结果列表:
        return None, []

    日志.info("开始回测模拟...")

    # 按信号日期排序
    df = pd.DataFrame(分析结果列表)
    df = df.sort_values('信号日期').reset_index(drop=True)

    当前资金 = 初始资金
    持仓记录 = []
    资金曲线 = [{'日期': '初始', '资金': 当前资金, '收益率%': 0}]

    for idx, 交易 in df.iterrows():
        # 计算投资金额
        投资金额 = 当前资金 * 单笔投资比例

        # 计算实际收益（年化收益 * 持有天数 / 365）
        持有天数 = 交易['已持有天数']
        年化收益率 = 交易['当前持有年化%'] / 100
        实际收益率 = 年化收益率 * (持有天数 / 365)

        # 更新资金
        盈亏 = 投资金额 * 实际收益率
        当前资金 += 盈亏

        持仓记录.append({
            '序号': idx + 1,
            '信号日期': 交易['信号日期'],
            '产品代码': 交易['产品代码'],
            '产品名称': 交易['产品名称'],
            '投资金额': round(投资金额, 2),
            '持有天数': 持有天数,
            '年化收益%': round(交易['当前持有年化%'], 2),
            '实际收益率%': round(实际收益率 * 100, 2),
            '盈亏': round(盈亏, 2),
            '当前资金': round(当前资金, 2),
            '累计收益率%': round((当前资金 - 初始资金) / 初始资金 * 100, 2)
        })

        资金曲线.append({
            '日期': 交易['信号日期'],
            '资金': round(当前资金, 2),
            '收益率%': round((当前资金 - 初始资金) / 初始资金 * 100, 2)
        })

    # 计算回测统计
    最终资金 = 当前资金
    累计收益率 = (最终资金 - 初始资金) / 初始资金 * 100

    资金序列 = [x['资金'] for x in 资金曲线]
    最大回撤, _ = 计算最大回撤([x['收益率%'] for x in 资金曲线])

    回测统计 = {
        '初始资金': 初始资金,
        '最终资金': round(最终资金, 2),
        '累计收益': round(最终资金 - 初始资金, 2),
        '累计收益率%': round(累计收益率, 2),
        '交易次数': len(持仓记录),
        '最大回撤%': round(最大回撤, 2)
    }

    日志.info(f"回测完成: 累计收益率 {累计收益率:.2f}%, 最大回撤 {最大回撤:.2f}%")

    return 回测统计, 持仓记录, 资金曲线


# ============================================================
# 可视化模块（新增）
# ============================================================

def 生成可视化图表(分析结果列表, 统计结果, 资金曲线):
    """
    生成各类图表：
    1. 收益分布直方图
    2. 累计收益曲线
    3. 产品类别对比
    4. 回撤分布
    """
    if not 分析结果列表:
        return []

    # 创建图表目录
    if not os.path.exists(图表输出目录):
        os.makedirs(图表输出目录)

    图表文件列表 = []
    df = pd.DataFrame(分析结果列表)

    # 图表1: 收益分布直方图
    plt.figure(figsize=(10, 6))
    plt.hist(df['当前持有年化%'], bins=50, color='steelblue', edgecolor='black', alpha=0.7)
    plt.axvline(df['当前持有年化%'].mean(), color='red', linestyle='--', linewidth=2, label=f'平均收益: {df["当前持有年化%"].mean():.2f}%')
    plt.xlabel('年化收益率 (%)', fontsize=12)
    plt.ylabel('频数', fontsize=12)
    plt.title('持有期年化收益率分布', fontsize=14, fontweight='bold')
    plt.legend()
    plt.grid(axis='y', alpha=0.3)
    图表1路径 = os.path.join(图表输出目录, '收益分布.png')
    plt.savefig(图表1路径, dpi=150, bbox_inches='tight')
    plt.close()
    图表文件列表.append(图表1路径)

    # 图表2: 累计收益曲线
    if 资金曲线 and len(资金曲线) > 1:
        plt.figure(figsize=(12, 6))
        日期列表 = [x['日期'] for x in 资金曲线]
        收益率列表 = [x['收益率%'] for x in 资金曲线]
        plt.plot(range(len(收益率列表)), 收益率列表, linewidth=2, color='green', marker='o', markersize=3)
        plt.axhline(0, color='gray', linestyle='--', linewidth=1)
        plt.xlabel('交易序号', fontsize=12)
        plt.ylabel('累计收益率 (%)', fontsize=12)
        plt.title('回测累计收益率曲线', fontsize=14, fontweight='bold')
        plt.grid(True, alpha=0.3)
        图表2路径 = os.path.join(图表输出目录, '累计收益曲线.png')
        plt.savefig(图表2路径, dpi=150, bbox_inches='tight')
        plt.close()
        图表文件列表.append(图表2路径)

    # 图表3: 产品类别对比
    类别统计 = 统计结果.get('分类别统计', {})
    if 类别统计:
        类别名称 = list(类别统计.keys())
        成功率 = [类别统计[k]['成功率'] for k in 类别名称]
        平均收益 = [类别统计[k]['平均收益'] for k in 类别名称]

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))

        # 子图1: 成功率
        ax1.bar(类别名称, 成功率, color=['#2ecc71', '#e74c3c', '#95a5a6'])
        ax1.set_ylabel('成功率 (%)', fontsize=12)
        ax1.set_title('产品类别成功率对比', fontsize=13, fontweight='bold')
        ax1.grid(axis='y', alpha=0.3)

        # 子图2: 平均收益
        ax2.bar(类别名称, 平均收益, color=['#3498db', '#9b59b6', '#f39c12'])
        ax2.set_ylabel('平均收益 (%)', fontsize=12)
        ax2.set_title('产品类别平均收益对比', fontsize=13, fontweight='bold')
        ax2.grid(axis='y', alpha=0.3)

        plt.tight_layout()
        图表3路径 = os.path.join(图表输出目录, '产品类别对比.png')
        plt.savefig(图表3路径, dpi=150, bbox_inches='tight')
        plt.close()
        图表文件列表.append(图表3路径)

    # 图表4: 回撤分布
    plt.figure(figsize=(10, 6))
    plt.hist(df['最大回撤%'], bins=30, color='coral', edgecolor='black', alpha=0.7)
    plt.axvline(df['最大回撤%'].mean(), color='darkred', linestyle='--', linewidth=2, label=f'平均回撤: {df["最大回撤%"].mean():.2f}%')
    plt.axvline(最大回撤止损阈值, color='red', linestyle='-', linewidth=2, label=f'止损阈值: {最大回撤止损阈值}%')
    plt.xlabel('最大回撤 (%)', fontsize=12)
    plt.ylabel('频数', fontsize=12)
    plt.title('持有期最大回撤分布', fontsize=14, fontweight='bold')
    plt.legend()
    plt.grid(axis='y', alpha=0.3)
    图表4路径 = os.path.join(图表输出目录, '回撤分布.png')
    plt.savefig(图表4路径, dpi=150, bbox_inches='tight')
    plt.close()
    图表文件列表.append(图表4路径)

    日志.info(f"已生成 {len(图表文件列表)} 个图表")
    return 图表文件列表


# ============================================================
# Excel报告生成（增强版）
# ============================================================

def 生成Excel报告(分析结果列表, 跟踪详情列表, 统计结果, 回测统计, 持仓记录, 图表文件列表):
    """生成完整的Excel报告（含统计和图表）"""
    if not 分析结果列表:
        日志.warning("没有任何信号数据，跳过报告生成")
        return

    日志.info("正在生成Excel报告...")

    wb = Workbook()

    # 样式定义
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="4F81BD")
    align_center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def 设置表头(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border

    def 设置列宽(ws, 列宽字典):
        for 列号, 宽度 in 列宽字典.items():
            ws.column_dimensions[get_column_letter(列号)].width = 宽度

    # Sheet 1: 策略总览与统计
    ws1 = wb.active
    ws1.title = "策略总览"

    ws1['A1'] = '银行理财产品收益释放套利策略 - 统计报告'
    ws1['A1'].font = Font(bold=True, size=14)
    ws1.merge_cells('A1:D1')

    行号 = 3
    ws1[f'A{行号}'] = '==== 整体统计 ===='
    ws1[f'A{行号}'].font = Font(bold=True, size=12)
    行号 += 1

    for 键, 值 in 统计结果.items():
        if 键 != '分类别统计':
            ws1[f'A{行号}'] = 键
            ws1[f'B{行号}'] = 值
            行号 += 1

    行号 += 1
    ws1[f'A{行号}'] = '==== 回测统计 ===='
    ws1[f'A{行号}'].font = Font(bold=True, size=12)
    行号 += 1

    if 回测统计:
        for 键, 值 in 回测统计.items():
            ws1[f'A{行号}'] = 键
            ws1[f'B{行号}'] = 值
            行号 += 1

    行号 += 1
    ws1[f'A{行号}'] = '==== 产品类别统计 ===='
    ws1[f'A{行号}'].font = Font(bold=True, size=12)
    行号 += 1

    类别统计 = 统计结果.get('分类别统计', {})
    if 类别统计:
        ws1[f'A{行号}'] = '产品类别'
        ws1[f'B{行号}'] = '信号数'
        ws1[f'C{行号}'] = '成功率%'
        ws1[f'D{行号}'] = '平均收益%'
        ws1[f'E{行号}'] = '平均回撤%'
        行号 += 1

        for 类别, 数据 in 类别统计.items():
            ws1[f'A{行号}'] = 类别
            ws1[f'B{行号}'] = 数据['信号数']
            ws1[f'C{行号}'] = round(数据['成功率'], 2)
            ws1[f'D{行号}'] = round(数据['平均收益'], 2)
            ws1[f'E{行号}'] = round(数据['平均回撤'], 2)
            行号 += 1

    设置列宽(ws1, {1: 25, 2: 20, 3: 15, 4: 15, 5: 15})

    # Sheet 2: 高收益机会监控
    ws2 = wb.create_sheet("高收益机会监控")
    headers2 = ['银行', '产品代码', '产品名称', '产品类别', '信号日期', '启动收益率%',
                '历史拉升次数', '预期维持天数', '预期年化收益%',
                '最新状态', '已持有天数', '当前持有年化%', '收益差(实际-预期)', '最大回撤%',
                '是否回撤止损', '交易成功', '卖出原因', '操作建议']
    设置表头(ws2, headers2)

    分析结果列表 = sorted(分析结果列表, key=lambda x: x['当前持有年化%'], reverse=True)

    for row in 分析结果列表:
        ws2.append([row[k] for k in headers2])

    设置列宽(ws2, {
        1: 12, 2: 22, 3: 40, 4: 14, 5: 14, 6: 14,
        7: 14, 8: 14, 9: 16, 10: 14, 11: 12,
        12: 16, 13: 12, 14: 12, 15: 14, 16: 12, 17: 30, 18: 14
    })
    ws2.freeze_panes = 'A2'

    # Sheet 3: 持有期逐日跟踪
    ws3 = wb.create_sheet("持有期逐日跟踪")
    headers3 = ['银行', '产品代码', '产品名称', '产品类别', '信号日期', '跟踪日期',
                '持有天数', '当日收益', '当前持有年化', '当前回撤', '卖出触发']
    设置表头(ws3, headers3)

    for row in 跟踪详情列表:
        ws3.append([row[k] for k in headers3])

    设置列宽(ws3, {
        1: 12, 2: 22, 3: 40, 4: 14, 5: 14, 6: 14,
        7: 12, 8: 14, 9: 16, 10: 12, 11: 12
    })
    ws3.freeze_panes = 'A2'

    # Sheet 4: 回测交易记录
    if 持仓记录:
        ws4 = wb.create_sheet("回测交易记录")
        headers4 = ['序号', '信号日期', '产品代码', '产品名称', '投资金额', '持有天数',
                    '年化收益%', '实际收益率%', '盈亏', '当前资金', '累计收益率%']
        设置表头(ws4, headers4)

        for row in 持仓记录:
            ws4.append([row[k] for k in headers4])

        设置列宽(ws4, {
            1: 8, 2: 14, 3: 22, 4: 40, 5: 14, 6: 12,
            7: 14, 8: 14, 9: 14, 10: 14, 11: 14
        })
        ws4.freeze_panes = 'A2'

    # Sheet 5: 可视化图表
    if 图表文件列表:
        ws5 = wb.create_sheet("可视化图表")
        行位置 = 1

        for 图表路径 in 图表文件列表:
            if os.path.exists(图表路径):
                try:
                    img = XLImage(图表路径)
                    # 调整图片大小
                    img.width = 600
                    img.height = 360
                    ws5.add_image(img, f'A{行位置}')
                    行位置 += 20  # 为下一张图片留出空间
                except Exception as e:
                    日志.warning(f"插入图表失败 {图表路径}: {e}")

    # 保存
    try:
        wb.save(输出文件路径)
        日志.info(f"报告已生成: {输出文件路径}")
        日志.info(f"  Sheet1 策略总览: 统计数据")
        日志.info(f"  Sheet2 高收益机会监控: {len(分析结果列表)} 行")
        日志.info(f"  Sheet3 持有期逐日跟踪: {len(跟踪详情列表)} 行")
        if 持仓记录:
            日志.info(f"  Sheet4 回测交易记录: {len(持仓记录)} 行")
        if 图表文件列表:
            日志.info(f"  Sheet5 可视化图表: {len(图表文件列表)} 个图表")
    except Exception as e:
        日志.error(f"保存文件失败: {e}")


# ============================================================
# 主程序
# ============================================================

def main():
    print("=" * 70)
    print("   银行理财产品收益释放操纵行为监测与套利分析系统 V3.0")
    print("=" * 70)
    print(f"运行时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"数据来源: {净值数据库路径}")
    print(f"输出文件: {输出文件路径}")
    print(f"收益拉升阈值: {收益拉升阈值}%")
    print(f"收益下滑阈值: {收益下滑阈值}%")
    print(f"持有期约束: {最短持有天数}-{最长持有天数}天")
    print(f"回撤止损阈值: {最大回撤止损阈值}%")
    print(f"初始资金: {初始资金:,.0f}元")
    print()

    # 步骤1: 加载数据
    print("【第1步】加载净值数据库...")
    数据库 = 加载净值数据库()
    if not 数据库:
        return

    # 步骤2: 修复产品代码
    print("\n【第2步】修复产品代码...")
    数据库 = 修复产品代码(数据库)

    # 步骤3: 计算收益率 + 执行策略
    print("\n【第3步】计算收益率并执行策略分析...")
    所有信号事件 = []
    所有跟踪详情 = []

    for 银行名, df in 数据库.items():
        收益率df, 全部日期, 收益率日期 = 计算基础收益率(df, 银行名)
        if 收益率df is None:
            continue

        信号, 跟踪 = 执行策略分析(收益率df, 全部日期, 收益率日期, 银行名)
        所有信号事件.extend(信号)
        所有跟踪详情.extend(跟踪)

    if not 所有信号事件:
        print("\n未发现任何拉升信号，程序结束")
        return

    # 步骤4: 计算统计指标
    print("\n【第4步】计算统计指标...")
    统计结果 = 计算收益统计(所有信号事件)

    # 步骤5: 执行回测
    print("\n【第5步】执行回测模拟...")
    回测统计, 持仓记录, 资金曲线 = 执行回测模拟(所有信号事件)

    # 步骤6: 生成可视化图表
    print("\n【第6步】生成可视化图表...")
    图表文件列表 = 生成可视化图表(所有信号事件, 统计结果, 资金曲线)

    # 步骤7: 生成报告
    print("\n【第7步】生成分析报告...")
    生成Excel报告(所有信号事件, 所有跟踪详情, 统计结果, 回测统计, 持仓记录, 图表文件列表)

    # 统计汇总
    print("\n" + "=" * 70)
    print("                    分析完成")
    print("=" * 70)
    print(f"捕获拉升信号总数: {统计结果['总信号数']}")
    print(f"成功交易数: {统计结果['成功交易数']}")
    print(f"失败交易数: {统计结果['失败交易数']}")
    print(f"交易成功率: {统计结果['成功率%']:.2f}%")
    print(f"平均收益: {统计结果['平均收益%']:.2f}%")
    print(f"平均持有天数: {统计结果['平均持有天数']:.1f}天")
    print(f"平均回撤: {统计结果['平均回撤%']:.2f}%")
    print(f"夏普比率: {统计结果['夏普比率']:.2f}")

    if 回测统计:
        print(f"\n回测结果:")
        print(f"  初始资金: {回测统计['初始资金']:,.0f}元")
        print(f"  最终资金: {回测统计['最终资金']:,.0f}元")
        print(f"  累计收益: {回测统计['累计收益']:,.0f}元")
        print(f"  累计收益率: {回测统计['累计收益率%']:.2f}%")
        print(f"  交易次数: {回测统计['交易次数']}")
        print(f"  最大回撤: {回测统计['最大回撤%']:.2f}%")

    print(f"\n输出文件: {输出文件路径}")
    print(f"图表目录: {图表输出目录}")
    print("=" * 70)


if __name__ == '__main__':
    main()
